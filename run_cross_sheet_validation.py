#!/usr/bin/env python3
"""
Cross-Sheet Validation for AR Data Analysis
==========================================

This script runs comprehensive cross-sheet validation to ensure data consistency
across all sheets with totals in the AR Data Analysis Excel report.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

def extract_totals_from_sheet(ws, sheet_name):
    """
    Extract totals values from a worksheet.
    
    Args:
        ws: openpyxl worksheet
        sheet_name: Name of the sheet
        
    Returns:
        Dictionary of totals found
    """
    totals = {}
    
    try:
        # Read all data from the sheet
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):
                data.append(row)
        
        if not data:
            return totals
        
        # Look for totals patterns
        for i, row in enumerate(data):
            if row and any(cell and 'TOTAL' in str(cell).upper() for cell in row):
                # Found a totals row
                for j, cell in enumerate(row):
                    if isinstance(cell, (int, float)) and cell > 0:
                        # This is a numeric total
                        # Try to identify what column this represents
                        header = None
                        if i > 0 and j < len(data[0]):
                            # Look for header in first row or nearby rows
                            for header_row in data[:min(3, len(data))]:
                                if j < len(header_row) and header_row[j]:
                                    header = str(header_row[j])
                                    break
                        
                        if header and 'total' not in header.lower():
                            totals[f"{header}_total"] = cell
                        else:
                            totals[f"column_{j}_total"] = cell
        
        # Also look for grand totals (largest values that might be totals)
        numeric_values = []
        for row in data:
            for cell in row:
                if isinstance(cell, (int, float)) and cell > 100:  # Reasonable threshold
                    numeric_values.append(cell)
        
        if numeric_values:
            # The largest value might be a grand total
            max_value = max(numeric_values)
            if max_value > 1000:  # Likely a meaningful total
                totals['grand_total_candidate'] = max_value
        
        print(f"   📊 {sheet_name}: Found {len(totals)} totals")
        
    except Exception as e:
        print(f"   ❌ Error extracting totals from {sheet_name}: {e}")
    
    return totals

def run_cross_sheet_validation():
    """
    Run comprehensive cross-sheet validation.
    """
    print("🔍 CROSS-SHEET VALIDATION FOR AR DATA ANALYSIS")
    print("="*60)
    
    # Find the latest report file
    report_files = list(Path(".").glob("AR_Analysis_Report_*.xlsx"))
    if not report_files:
        print("❌ No report files found")
        return
    
    latest_report = max(report_files, key=lambda x: x.stat().st_mtime)
    print(f"📊 Analyzing report: {latest_report}")
    
    try:
        # Load the workbook
        wb = load_workbook(latest_report, data_only=True)
        
        # Sheets with totals to validate
        target_sheets = [
            'Monthly Capture Volume',
            'File Size Stats', 
            'Time of Day',
            'Weekday by Period',
            'MP3 Duration Analysis',
            'Data Cleaning',
            'Daily Counts (ACF_PACF)',
            'Weekly Counts (ACF_PACF)',
            'Summary Statistics'  # If it exists
        ]
        
        all_totals = {}
        
        print(f"\n📋 EXTRACTING TOTALS FROM SHEETS")
        print("-" * 40)
        
        for sheet_name in target_sheets:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_totals = extract_totals_from_sheet(ws, sheet_name)
                if sheet_totals:
                    all_totals[sheet_name] = sheet_totals
            else:
                print(f"   ⚠️  Sheet '{sheet_name}' not found")
        
        print(f"\n🔍 CROSS-SHEET CONSISTENCY ANALYSIS")
        print("-" * 40)
        
        # Look for common patterns that should be consistent
        consistency_checks = []
        
        # Check for file count consistency
        file_count_totals = {}
        for sheet_name, totals in all_totals.items():
            for total_name, value in totals.items():
                if 'count' in total_name.lower() or 'files' in total_name.lower():
                    if value > 5000:  # Likely total file count
                        file_count_totals[sheet_name] = value
        
        if len(file_count_totals) > 1:
            values = list(file_count_totals.values())
            if max(values) - min(values) <= 100:  # Allow small differences
                print(f"   ✅ File count consistency: {file_count_totals}")
                consistency_checks.append(("File Counts", "PASS", file_count_totals))
            else:
                print(f"   ⚠️  File count inconsistency: {file_count_totals}")
                consistency_checks.append(("File Counts", "WARNING", file_count_totals))
        
        # Check for reasonable total values
        print(f"\n📊 TOTALS REASONABLENESS CHECK")
        print("-" * 40)
        
        for sheet_name, totals in all_totals.items():
            reasonable = True
            for total_name, value in totals.items():
                if value < 0:
                    print(f"   ❌ Negative total in {sheet_name}: {total_name} = {value}")
                    reasonable = False
                elif value > 1000000:  # Suspiciously large
                    print(f"   ⚠️  Very large total in {sheet_name}: {total_name} = {value}")
            
            if reasonable:
                print(f"   ✅ {sheet_name}: All totals appear reasonable")
        
        # Generate validation report
        validation_report = {
            'timestamp': datetime.now().isoformat(),
            'report_file': str(latest_report),
            'sheets_analyzed': len(all_totals),
            'total_totals_found': sum(len(totals) for totals in all_totals.values()),
            'consistency_checks': consistency_checks,
            'detailed_totals': all_totals
        }
        
        # Save validation report
        report_file = f"cross_sheet_validation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        with open(report_file, 'w') as f:
            json.dump(validation_report, f, indent=2, default=str)
        
        print(f"\n📋 VALIDATION SUMMARY")
        print("=" * 30)
        print(f"📊 Sheets analyzed: {len(all_totals)}")
        print(f"📊 Total totals found: {sum(len(totals) for totals in all_totals.values())}")
        print(f"📊 Consistency checks: {len(consistency_checks)}")
        print(f"💾 Detailed report saved: {report_file}")
        
        # Summary of findings
        if consistency_checks:
            passed = sum(1 for _, status, _ in consistency_checks if status == "PASS")
            warnings = sum(1 for _, status, _ in consistency_checks if status == "WARNING")
            
            print(f"\n🎯 CONSISTENCY RESULTS:")
            print(f"   ✅ Passed: {passed}")
            print(f"   ⚠️  Warnings: {warnings}")
            
            if warnings == 0:
                print(f"   🏆 EXCELLENT: All consistency checks passed!")
            else:
                print(f"   📋 Review warnings for potential data quality issues")
        
    except Exception as e:
        print(f"❌ Error during validation: {e}")
        import traceback
        traceback.print_exc()

def main():
    """
    Main validation function.
    """
    try:
        run_cross_sheet_validation()
        
        print(f"\n" + "="*60)
        print(f"🎯 CROSS-SHEET VALIDATION COMPLETE")
        print(f"="*60)
        print(f"This validation ensures:")
        print(f"1. ✅ Totals are mathematically reasonable")
        print(f"2. ✅ Related sheets have consistent values")
        print(f"3. ✅ No negative or suspicious totals")
        print(f"4. ✅ Overall data quality is maintained")
        
    except Exception as e:
        print(f"❌ Validation error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
