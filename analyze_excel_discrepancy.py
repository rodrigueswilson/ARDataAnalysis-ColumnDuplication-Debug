#!/usr/bin/env python3
"""
Analyze Excel Sheet Discrepancy

This script examines the actual Excel sheets to understand where the discrepancy
between Data Cleaning (9,731) and Daily Counts (ACF_PACF) (9,372) is coming from.
"""

import sys
import os
import openpyxl
from pathlib import Path
import pandas as pd

def find_latest_report():
    """Find the most recent AR_Analysis_Report file."""
    current_dir = Path(".")
    report_files = list(current_dir.glob("AR_Analysis_Report_*.xlsx"))
    
    if not report_files:
        print("No AR_Analysis_Report files found in current directory")
        return None
    
    # Sort by modification time, get the latest
    latest_file = max(report_files, key=lambda f: f.stat().st_mtime)
    print(f"Found latest report: {latest_file}")
    return latest_file

def analyze_data_cleaning_sheet(workbook):
    """Analyze the Data Cleaning sheet."""
    print("\n" + "="*50)
    print("ANALYZING DATA CLEANING SHEET")
    print("="*50)
    
    try:
        ws = workbook["Data Cleaning"]
        print(f"Sheet found: {ws.title}")
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Look for the total clean files number
        clean_files_found = []
        
        for row in range(1, min(50, ws.max_row + 1)):  # Check first 50 rows
            for col in range(1, min(10, ws.max_column + 1)):  # Check first 10 columns
                cell_value = ws.cell(row=row, column=col).value
                if isinstance(cell_value, (int, float)) and cell_value == 9731:
                    clean_files_found.append(f"Row {row}, Col {col}: {cell_value}")
                elif isinstance(cell_value, str) and "9731" in str(cell_value):
                    clean_files_found.append(f"Row {row}, Col {col}: {cell_value}")
        
        print(f"Found 9731 references: {len(clean_files_found)}")
        for ref in clean_files_found:
            print(f"  {ref}")
            
        return len(clean_files_found) > 0
        
    except KeyError:
        print("Data Cleaning sheet not found!")
        return False

def analyze_daily_counts_sheet(workbook):
    """Analyze the Daily Counts (ACF_PACF) sheet."""
    print("\n" + "="*50)
    print("ANALYZING DAILY COUNTS (ACF_PACF) SHEET")
    print("="*50)
    
    try:
        ws = workbook["Daily Counts (ACF_PACF)"]
        print(f"Sheet found: {ws.title}")
        print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        # Look for totals and sums
        total_files_col = None
        
        # Find the Total_Files column
        for col in range(1, min(20, ws.max_column + 1)):
            header_value = ws.cell(row=1, column=col).value
            if header_value and "Total_Files" in str(header_value):
                total_files_col = col
                print(f"Found Total_Files column at column {col}: {header_value}")
                break
        
        if total_files_col:
            # Calculate sum of Total_Files column
            total_sum = 0
            values_found = []
            
            for row in range(2, ws.max_row + 1):  # Skip header
                cell_value = ws.cell(row=row, column=total_files_col).value
                if isinstance(cell_value, (int, float)) and cell_value > 0:
                    total_sum += cell_value
                    values_found.append(cell_value)
            
            print(f"Sum of Total_Files column: {total_sum}")
            print(f"Number of data rows: {len(values_found)}")
            print(f"First 10 values: {values_found[:10]}")
            
            # Look for the specific 9372 number
            files_9372_found = []
            for row in range(1, ws.max_row + 1):
                for col in range(1, min(20, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if isinstance(cell_value, (int, float)) and cell_value == 9372:
                        files_9372_found.append(f"Row {row}, Col {col}: {cell_value}")
                    elif isinstance(cell_value, str) and "9372" in str(cell_value):
                        files_9372_found.append(f"Row {row}, Col {col}: {cell_value}")
            
            print(f"Found 9372 references: {len(files_9372_found)}")
            for ref in files_9372_found:
                print(f"  {ref}")
            
            return total_sum, len(files_9372_found) > 0
        else:
            print("Total_Files column not found!")
            return 0, False
            
    except KeyError:
        print("Daily Counts (ACF_PACF) sheet not found!")
        return 0, False

def analyze_all_sheets(workbook):
    """Get overview of all sheets."""
    print("\n" + "="*50)
    print("ALL SHEETS OVERVIEW")
    print("="*50)
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        print(f"{sheet_name}: {ws.max_row} rows x {ws.max_column} columns")

def main():
    print("EXCEL SHEET DISCREPANCY ANALYSIS")
    print("="*80)
    
    # Find the latest report
    report_file = find_latest_report()
    if not report_file:
        return
    
    try:
        # Load the workbook
        print(f"Loading workbook: {report_file}")
        workbook = openpyxl.load_workbook(report_file, data_only=True)
        
        # Analyze all sheets
        analyze_all_sheets(workbook)
        
        # Analyze specific sheets
        data_cleaning_found = analyze_data_cleaning_sheet(workbook)
        daily_counts_sum, daily_counts_9372_found = analyze_daily_counts_sheet(workbook)
        
        # Summary
        print("\n" + "="*50)
        print("SUMMARY")
        print("="*50)
        print(f"Data Cleaning sheet has 9731: {'✅ YES' if data_cleaning_found else '❌ NO'}")
        print(f"Daily Counts sheet has 9372: {'✅ YES' if daily_counts_9372_found else '❌ NO'}")
        print(f"Daily Counts Total_Files sum: {daily_counts_sum}")
        
        if daily_counts_sum == 9372:
            print("✅ The 9372 number comes from summing the Total_Files column!")
            print("   This suggests the discrepancy is due to aggregation method.")
        
        workbook.close()
        
    except Exception as e:
        print(f"Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
