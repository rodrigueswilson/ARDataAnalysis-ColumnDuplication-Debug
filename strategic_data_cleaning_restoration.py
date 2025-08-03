#!/usr/bin/env python3
"""
Strategic Data Cleaning Restoration

This script provides a strategic approach to restore Data Cleaning functionality
by working around the structural issues in base.py until they can be fully resolved.
"""

import sys
import os
import openpyxl
import pandas as pd
import traceback
from datetime import datetime

# Add the project root to the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def create_standalone_data_cleaning_sheet(workbook):
    """
    Create a standalone Data Cleaning sheet that works independently of the broken base.py structure.
    This is a temporary solution to restore functionality while the class structure is being fixed.
    """
    print("CREATING STANDALONE DATA CLEANING SHEET")
    print("=" * 50)
    
    try:
        # Import necessary components
        from pymongo import MongoClient
        import yaml
        
        # Load configuration
        with open('config.yaml', 'r') as f:
            config = yaml.safe_load(f)
        
        # Connect to database
        client = MongoClient(config['mongodb']['connection_string'])
        db = client[config['mongodb']['database_name']]
        collection = db['media_records']
        
        print("✅ Database connection established")
        
        # Create worksheet
        ws = workbook.create_sheet(title="Data Cleaning")
        print("✅ Data Cleaning sheet created")
        
        # 1. Title and Introduction
        ws.cell(row=1, column=1, value="AR Data Analysis - Data Cleaning & Filtering Report")
        ws.cell(row=2, column=1, value="This sheet shows the complete breakdown of how both filtering criteria (Collection Days + Non-Outliers) affect the dataset.")
        
        # Apply basic formatting (simplified since we don't have the formatter)
        title_cell = ws['A1']
        title_cell.font = openpyxl.styles.Font(bold=True, size=14)
        
        print("✅ Title and introduction added")
        
        # 2. Run four separate aggregations to get intersection data
        print("🔄 Running intersection analysis with 4 aggregations...")
        
        # Aggregation 1: Raw data (only exclude N/A school years)
        raw_pipeline = [
            {"$match": {"School_Year": {"$ne": "N/A"}}},
            {
                "$group": {
                    "_id": {"school_year": "$School_Year", "file_type": "$file_type"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id.school_year": 1, "_id.file_type": 1}}
        ]
        
        # Aggregation 2: Collection days only (include outliers)
        collection_only_pipeline = [
            {
                "$match": {
                    "School_Year": {"$ne": "N/A"},
                    "is_collection_day": True
                }
            },
            {
                "$group": {
                    "_id": {"school_year": "$School_Year", "file_type": "$file_type"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id.school_year": 1, "_id.file_type": 1}}
        ]
        
        # Aggregation 3: Non-outliers only (include non-collection days)
        non_outliers_pipeline = [
            {
                "$match": {
                    "School_Year": {"$ne": "N/A"},
                    "Outlier_Status": False
                }
            },
            {
                "$group": {
                    "_id": {"school_year": "$School_Year", "file_type": "$file_type"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id.school_year": 1, "_id.file_type": 1}}
        ]
        
        # Aggregation 4: Both criteria (final clean dataset)
        both_criteria_pipeline = [
            {
                "$match": {
                    "School_Year": {"$ne": "N/A"},
                    "is_collection_day": True,
                    "Outlier_Status": False
                }
            },
            {
                "$group": {
                    "_id": {"school_year": "$School_Year", "file_type": "$file_type"},
                    "count": {"$sum": 1}
                }
            },
            {"$sort": {"_id.school_year": 1, "_id.file_type": 1}}
        ]
        
        # Execute all aggregations
        raw_results = list(collection.aggregate(raw_pipeline))
        collection_results = list(collection.aggregate(collection_only_pipeline))
        non_outlier_results = list(collection.aggregate(non_outliers_pipeline))
        both_results = list(collection.aggregate(both_criteria_pipeline))
        
        print(f"✅ Aggregations completed: Raw: {len(raw_results)}, Collection: {len(collection_results)}, Non-outliers: {len(non_outlier_results)}, Both: {len(both_results)}")
        
        # 3. Process results into dictionaries for easy lookup
        def results_to_dict(results):
            return {f"{r['_id']['school_year']}_{r['_id']['file_type']}": r['count'] for r in results}
        
        raw_dict = results_to_dict(raw_results)
        collection_dict = results_to_dict(collection_results)
        non_outlier_dict = results_to_dict(non_outlier_results)
        both_dict = results_to_dict(both_results)
        
        # 4. Create comprehensive analysis table
        categories = []
        for result in raw_results:
            sy = result['_id']['school_year']
            ft = result['_id']['file_type']
            key = f"{sy}_{ft}"
            category_name = f"{sy} {ft} Files"
            
            raw_count = raw_dict.get(key, 0)
            collection_count = collection_dict.get(key, 0)
            non_outlier_count = non_outlier_dict.get(key, 0)
            both_count = both_dict.get(key, 0)
            
            # Calculate intersection breakdown
            collection_only = collection_count - both_count  # Files that pass collection day filter but are outliers
            non_outlier_only = non_outlier_count - both_count  # Files that are non-outliers but on non-collection days
            neither = raw_count - collection_only - non_outlier_only - both_count  # Files that fail both criteria
            
            categories.append({
                'Category': category_name,
                'Total_Raw': raw_count,
                'Collection_Only': collection_only,
                'Non_Outlier_Only': non_outlier_only,
                'Both_Criteria': both_count,
                'Neither': neither,
                'Final_Clean': both_count,
                'Retention_Pct': (both_count / raw_count * 100) if raw_count > 0 else 0
            })
        
        # 5. Calculate totals
        totals = {
            'Category': 'TOTAL',
            'Total_Raw': sum(c['Total_Raw'] for c in categories),
            'Collection_Only': sum(c['Collection_Only'] for c in categories),
            'Non_Outlier_Only': sum(c['Non_Outlier_Only'] for c in categories),
            'Both_Criteria': sum(c['Both_Criteria'] for c in categories),
            'Neither': sum(c['Neither'] for c in categories),
            'Final_Clean': sum(c['Final_Clean'] for c in categories),
            'Retention_Pct': 0
        }
        totals['Retention_Pct'] = (totals['Final_Clean'] / totals['Total_Raw'] * 100) if totals['Total_Raw'] > 0 else 0
        categories.append(totals)
        
        # 6. Write Table 1: Complete Filtering Breakdown
        current_row = 4
        ws.cell(row=current_row, column=1, value="Table 1: Complete Filtering Breakdown (Intersection Analysis)")
        current_row += 2
        
        # Headers
        headers = ['Category', 'Total Raw', 'Collection Only', 'Non-Outlier Only', 'Both Criteria', 'Neither', 'Final Clean', 'Retention %']
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
            ws.cell(row=current_row, column=col).font = openpyxl.styles.Font(bold=True)
        
        current_row += 1
        
        # Data rows
        for category in categories:
            ws.cell(row=current_row, column=1, value=category['Category'])
            ws.cell(row=current_row, column=2, value=category['Total_Raw'])
            ws.cell(row=current_row, column=3, value=category['Collection_Only'])
            ws.cell(row=current_row, column=4, value=category['Non_Outlier_Only'])
            ws.cell(row=current_row, column=5, value=category['Both_Criteria'])
            ws.cell(row=current_row, column=6, value=category['Neither'])
            ws.cell(row=current_row, column=7, value=category['Final_Clean'])
            ws.cell(row=current_row, column=8, value=f"{category['Retention_Pct']:.1f}%")
            
            # Bold the totals row
            if category['Category'] == 'TOTAL':
                for col in range(1, 9):
                    ws.cell(row=current_row, column=col).font = openpyxl.styles.Font(bold=True)
            
            current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        print("✅ Data Cleaning sheet populated successfully")
        print(f"📊 Final clean dataset: {totals['Final_Clean']} files ({totals['Retention_Pct']:.1f}% retention)")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating Data Cleaning sheet: {e}")
        traceback.print_exc()
        return False

def test_standalone_solution():
    """Test the standalone Data Cleaning solution."""
    print("TESTING STANDALONE DATA CLEANING SOLUTION")
    print("=" * 60)
    
    try:
        # Create a test workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create the Data Cleaning sheet
        success = create_standalone_data_cleaning_sheet(wb)
        
        if success:
            # Save test file
            test_file = f"test_data_cleaning_standalone_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(test_file)
            print(f"✅ Test file saved: {test_file}")
            
            # Verify the sheet
            if 'Data Cleaning' in wb.sheetnames:
                ws = wb['Data Cleaning']
                print(f"📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                if ws.max_row > 5:  # Should have substantial content
                    print("🎉 SUCCESS! Standalone Data Cleaning sheet is working!")
                    return True
                else:
                    print("⚠️  Sheet created but has minimal content")
                    return False
            else:
                print("❌ Data Cleaning sheet not found in workbook")
                return False
        else:
            print("❌ Failed to create Data Cleaning sheet")
            return False
            
    except Exception as e:
        print(f"❌ Error in test: {e}")
        traceback.print_exc()
        return False

def main():
    print("STRATEGIC DATA CLEANING RESTORATION")
    print("=" * 70)
    
    try:
        # Test the standalone solution
        success = test_standalone_solution()
        
        print("\n" + "=" * 70)
        print("RESTORATION RESULTS")
        print("=" * 70)
        
        if success:
            print("🎉 STRATEGIC SUCCESS!")
            print("   ✅ Standalone Data Cleaning sheet creation working")
            print("   ✅ Database connectivity functional")
            print("   ✅ Data aggregation and analysis working")
            print("   ✅ Excel output generation successful")
            
            print(f"\n🎯 NEXT STEPS:")
            print("   1. Integrate this standalone solution into the main report generation")
            print("   2. Bypass the broken base.py structure temporarily")
            print("   3. Fix the structural issues in base.py when time permits")
            print("   4. Test full report generation with restored Data Cleaning")
            
        else:
            print("❌ Standalone solution failed")
            print("   Further investigation needed")
        
        return 0 if success else 1
        
    except Exception as e:
        print(f"❌ Error in main: {e}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
