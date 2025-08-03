#!/usr/bin/env python3
"""
Comprehensive Table Inventory Script for AR Data Analysis Excel Reports
Creates a detailed inventory of all tables across all sheets to plan totals implementation.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import json
from datetime import datetime

def analyze_excel_report(excel_path):
    """
    Analyze the Excel report and create a comprehensive table inventory.
    """
    print(f"📊 Analyzing Excel report: {excel_path}")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    inventory = {
        'report_path': str(excel_path),
        'analysis_date': datetime.now().isoformat(),
        'total_sheets': len(wb.sheetnames),
        'sheets': {}
    }
    
    print(f"📋 Found {len(wb.sheetnames)} sheets to analyze")
    
    for sheet_name in wb.sheetnames:
        print(f"\n🔍 Analyzing sheet: {sheet_name}")
        ws = wb[sheet_name]
        
        sheet_info = {
            'sheet_name': sheet_name,
            'max_row': ws.max_row,
            'max_column': ws.max_column,
            'tables': [],
            'current_totals_status': 'unknown',
            'recommended_totals': 'to_be_determined'
        }
        
        # Try to identify tables by looking for data patterns
        tables = identify_tables_in_sheet(ws, sheet_name)
        sheet_info['tables'] = tables
        
        # Analyze current totals status
        totals_status = analyze_current_totals(ws, sheet_name)
        sheet_info['current_totals_status'] = totals_status
        
        # Recommend totals strategy
        recommendations = recommend_totals_strategy(tables, sheet_name)
        sheet_info['recommended_totals'] = recommendations
        
        inventory['sheets'][sheet_name] = sheet_info
        
        print(f"   📈 Found {len(tables)} table(s)")
        print(f"   📊 Current totals: {totals_status}")
        print(f"   💡 Recommendation: {recommendations}")
    
    return inventory

def identify_tables_in_sheet(ws, sheet_name):
    """
    Identify tables within a worksheet by analyzing data patterns.
    """
    tables = []
    
    # Look for header rows (typically bold or different formatting)
    current_table = None
    
    for row_idx in range(1, min(ws.max_row + 1, 100)):  # Limit to first 100 rows for performance
        row = ws[row_idx]
        
        # Check if this looks like a header row
        if is_likely_header_row(row, ws):
            # If we were building a table, finish it
            if current_table:
                current_table['end_row'] = row_idx - 1
                tables.append(current_table)
            
            # Start a new table
            current_table = {
                'table_id': len(tables) + 1,
                'start_row': row_idx,
                'end_row': None,
                'header_row': row_idx,
                'columns': [],
                'data_types': {},
                'estimated_numeric_columns': 0,
                'has_totals_row': False,
                'has_totals_column': False
            }
            
            # Extract column headers
            for cell in row:
                if cell.value:
                    current_table['columns'].append(str(cell.value))
            
            # Analyze data types in next few rows
            analyze_table_data_types(ws, current_table, row_idx)
    
    # Finish the last table if exists
    if current_table:
        current_table['end_row'] = ws.max_row
        tables.append(current_table)
    
    return tables

def is_likely_header_row(row, ws):
    """
    Determine if a row is likely a header row based on formatting and content.
    """
    non_empty_cells = [cell for cell in row if cell.value is not None]
    
    if len(non_empty_cells) < 2:
        return False
    
    # Check for bold formatting (common in headers)
    bold_count = sum(1 for cell in non_empty_cells if cell.font and cell.font.bold)
    
    # Check for text content (headers are usually text)
    text_count = sum(1 for cell in non_empty_cells if isinstance(cell.value, str))
    
    # Heuristic: likely header if mostly bold or mostly text
    return (bold_count / len(non_empty_cells) > 0.5) or (text_count / len(non_empty_cells) > 0.7)

def analyze_table_data_types(ws, table_info, header_row):
    """
    Analyze the data types in a table to determine numeric columns.
    """
    numeric_columns = 0
    data_types = {}
    
    # Look at the next 5 rows after header to determine data types
    for col_idx, column_name in enumerate(table_info['columns']):
        numeric_count = 0
        total_count = 0
        
        for row_offset in range(1, 6):  # Check next 5 rows
            if header_row + row_offset <= ws.max_row:
                cell = ws.cell(row=header_row + row_offset, column=col_idx + 1)
                if cell.value is not None:
                    total_count += 1
                    if isinstance(cell.value, (int, float)):
                        numeric_count += 1
        
        if total_count > 0:
            numeric_ratio = numeric_count / total_count
            if numeric_ratio > 0.7:
                data_types[column_name] = 'numeric'
                numeric_columns += 1
            elif numeric_ratio > 0.3:
                data_types[column_name] = 'mixed'
            else:
                data_types[column_name] = 'text'
        else:
            data_types[column_name] = 'unknown'
    
    table_info['data_types'] = data_types
    table_info['estimated_numeric_columns'] = numeric_columns

def analyze_current_totals(ws, sheet_name):
    """
    Analyze if the sheet currently has any totals implemented.
    """
    # Look for common totals indicators
    totals_indicators = ['total', 'totals', 'sum', 'grand total', 'subtotal']
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value_lower = cell.value.lower()
                if any(indicator in cell_value_lower for indicator in totals_indicators):
                    return 'has_some_totals'
    
    return 'no_totals_detected'

def recommend_totals_strategy(tables, sheet_name):
    """
    Recommend totals strategy based on table analysis and sheet type.
    """
    if not tables:
        return 'no_tables_found'
    
    recommendations = []
    
    for table in tables:
        table_rec = {
            'table_id': table['table_id'],
            'strategy': 'none'
        }
        
        # Determine strategy based on table characteristics
        if table['estimated_numeric_columns'] >= 3:
            # Tables with many numeric columns benefit from both row and column totals
            table_rec['strategy'] = 'full_totals'
            table_rec['rationale'] = f"Table has {table['estimated_numeric_columns']} numeric columns"
        elif table['estimated_numeric_columns'] >= 1:
            # Tables with some numeric columns might benefit from column totals
            if len(table['columns']) > 5:
                table_rec['strategy'] = 'column_totals'
                table_rec['rationale'] = "Multiple rows with numeric data"
            else:
                table_rec['strategy'] = 'row_totals'
                table_rec['rationale'] = "Few rows with numeric data"
        else:
            table_rec['strategy'] = 'none'
            table_rec['rationale'] = "No significant numeric data detected"
        
        # Override based on sheet type
        if 'dashboard' in sheet_name.lower():
            table_rec['strategy'] = 'minimal_or_none'
            table_rec['rationale'] = "Dashboard sheets typically don't need totals"
        elif 'summary' in sheet_name.lower():
            if table_rec['strategy'] == 'none':
                table_rec['strategy'] = 'column_totals'
                table_rec['rationale'] = "Summary sheets benefit from totals"
        
        recommendations.append(table_rec)
    
    return recommendations

def save_inventory(inventory, output_path):
    """
    Save the inventory to a JSON file for reference.
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(inventory, f, indent=2, ensure_ascii=False)
    
    print(f"💾 Inventory saved to: {output_path}")

def print_inventory_summary(inventory):
    """
    Print a human-readable summary of the inventory.
    """
    print("\n" + "="*80)
    print("📊 COMPREHENSIVE TABLE INVENTORY SUMMARY")
    print("="*80)
    
    print(f"📋 Report: {inventory['report_path']}")
    print(f"📅 Analysis Date: {inventory['analysis_date']}")
    print(f"📄 Total Sheets: {inventory['total_sheets']}")
    
    total_tables = sum(len(sheet['tables']) for sheet in inventory['sheets'].values())
    print(f"📈 Total Tables: {total_tables}")
    
    print("\n" + "-"*80)
    print("SHEET-BY-SHEET ANALYSIS")
    print("-"*80)
    
    for sheet_name, sheet_info in inventory['sheets'].items():
        print(f"\n📄 {sheet_name}")
        print(f"   📊 Dimensions: {sheet_info['max_row']} rows × {sheet_info['max_column']} columns")
        print(f"   📈 Tables: {len(sheet_info['tables'])}")
        print(f"   ✅ Current Totals: {sheet_info['current_totals_status']}")
        
        if sheet_info['tables']:
            for table in sheet_info['tables']:
                print(f"      🔸 Table {table['table_id']}: {len(table['columns'])} columns, {table['estimated_numeric_columns']} numeric")
        
        if isinstance(sheet_info['recommended_totals'], list):
            for rec in sheet_info['recommended_totals']:
                print(f"      💡 Table {rec['table_id']}: {rec['strategy']} - {rec['rationale']}")

def main():
    """
    Main function to run the table inventory analysis.
    """
    # Find the most recent Excel report
    report_dir = Path(".")
    excel_files = list(report_dir.glob("AR_Analysis_Report_*.xlsx"))
    
    if not excel_files:
        print("❌ No Excel reports found in current directory")
        return
    
    # Get the most recent report
    latest_report = max(excel_files, key=lambda x: x.stat().st_mtime)
    
    print(f"🎯 Using latest report: {latest_report}")
    
    # Create inventory
    inventory = analyze_excel_report(latest_report)
    
    # Save inventory
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    inventory_path = f"table_inventory_{timestamp}.json"
    save_inventory(inventory, inventory_path)
    
    # Print summary
    print_inventory_summary(inventory)
    
    print("\n" + "="*80)
    print("✅ TABLE INVENTORY ANALYSIS COMPLETE")
    print("="*80)
    print(f"📄 Detailed inventory saved to: {inventory_path}")
    print("🎯 Ready for systematic totals implementation!")

if __name__ == "__main__":
    main()
