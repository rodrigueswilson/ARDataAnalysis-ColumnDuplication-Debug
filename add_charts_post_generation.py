#!/usr/bin/env python3
"""
Post-Generation Chart Adder
Adds charts to completed Excel reports as a separate process
"""

import openpyxl
from openpyxl.chart import LineChart, Reference
import os
import sys

def add_charts_to_latest_report():
    """Add charts to the most recently generated report"""
    
    print("POST-GENERATION CHART ADDER")
    print("=" * 50)
    
    # Find latest report
    report_files = [f for f in os.listdir('.') if f.startswith('AR_Analysis_Report_') and f.endswith('.xlsx')]
    if not report_files:
        print("[ERROR] No report files found!")
        return False
    
    latest_report = max(report_files)
    print(f"[INFO] Adding charts to: {latest_report}")
    
    try:
        # Load workbook
        print("[STEP 1] Loading workbook...")
        wb = openpyxl.load_workbook(latest_report)
        print(f"[SUCCESS] Loaded workbook with {len(wb.sheetnames)} sheets")
        
        # Find ACF/PACF sheets
        acf_pacf_sheets = []
        for sheet_name in wb.sheetnames:
            if any(keyword in sheet_name for keyword in ['ACF', 'PACF']):
                acf_pacf_sheets.append(sheet_name)
        
        print(f"[STEP 2] Found {len(acf_pacf_sheets)} ACF/PACF sheets:")
        for sheet in acf_pacf_sheets:
            print(f"  - {sheet}")
        
        if not acf_pacf_sheets:
            print("[ERROR] No ACF/PACF sheets found!")
            return False
        
        charts_added = 0
        
        # Add charts to each ACF/PACF sheet
        for sheet_name in acf_pacf_sheets:
            print(f"\n[STEP 3] Processing sheet: {sheet_name}")
            
            try:
                ws = wb[sheet_name]
                print(f"  Sheet dimensions: {ws.max_row} x {ws.max_column}")
                
                if ws.max_row < 2:
                    print(f"  [SKIP] No data in sheet")
                    continue
                
                # Find Total_Files column
                total_files_col = None
                for col in range(1, min(ws.max_column + 1, 20)):
                    header = ws.cell(row=1, column=col).value
                    if str(header) == 'Total_Files':
                        total_files_col = col
                        break
                
                if not total_files_col:
                    print(f"  [SKIP] No Total_Files column found")
                    continue
                
                print(f"  Total_Files column at: {total_files_col}")
                
                # Create chart
                chart = LineChart()
                chart.title = f"Time Series - {sheet_name.replace('(ACF_PACF)', '').strip()}"
                chart.height = 8
                chart.width = 12
                chart.y_axis.title = "Total Files"
                chart.x_axis.title = "Time Period"
                
                # Add data (use reasonable range to avoid memory issues)
                max_data_rows = min(ws.max_row, 100)  # Limit to 100 rows max
                data_ref = Reference(ws, min_col=total_files_col, min_row=1, max_row=max_data_rows)
                chart.add_data(data_ref, titles_from_data=True)
                
                # Position chart in a visible location
                # Try column H first, then I, then J if they exist
                chart_position = 'H2'
                if ws.max_column >= 8:  # Column H exists
                    chart_position = 'H2'
                elif ws.max_column >= 9:  # Column I exists
                    chart_position = 'I2'
                else:
                    chart_position = 'A15'  # Fallback below data
                
                ws.add_chart(chart, chart_position)
                charts_added += 1
                print(f"  [SUCCESS] Chart added at {chart_position}")
                
            except Exception as e:
                print(f"  [ERROR] Failed to add chart to {sheet_name}: {e}")
        
        if charts_added > 0:
            # Save enhanced report
            enhanced_filename = f"ENHANCED_{latest_report}"
            print(f"\n[STEP 4] Saving enhanced report...")
            wb.save(enhanced_filename)
            print(f"[SUCCESS] Enhanced report saved: {enhanced_filename}")
            print(f"[RESULT] Added {charts_added} charts successfully!")
            
            wb.close()
            return True
        else:
            print(f"\n[FAILURE] No charts were added")
            wb.close()
            return False
            
    except Exception as e:
        print(f"[ERROR] Chart addition failed: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main execution"""
    print("This script adds charts to completed Excel reports.")
    print("Run this AFTER generating a report with 'python generate_report.py'\n")
    
    success = add_charts_to_latest_report()
    
    if success:
        print(f"\n{'='*50}")
        print("✅ CHARTS SUCCESSFULLY ADDED!")
        print("Open the ENHANCED_*.xlsx file to see your charts.")
        print("Charts are positioned at column H, row 2 in each ACF/PACF sheet.")
    else:
        print(f"\n{'='*50}")
        print("❌ CHART ADDITION FAILED")
        print("Check the error messages above for details.")

if __name__ == "__main__":
    main()
