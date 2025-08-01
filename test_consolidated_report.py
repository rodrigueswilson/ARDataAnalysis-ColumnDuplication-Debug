#!/usr/bin/env python3
"""
Test Consolidated Sheet Creation System

Tests the consolidated sheet creation system by generating a full report
and validating that no duplicate sheets are created.

Generated: 2025-07-31 21:06:50
"""

import sys
import os
import traceback
from datetime import datetime

# Add current directory to path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def test_consolidated_report_generation():
    """Test full report generation with consolidated system."""
    print("[TEST] Testing consolidated report generation...")
    
    try:
        # Import consolidated system
        from report_generator.core_consolidated import ConsolidatedReportGenerator
        from report_generator.sheet_registry import get_sheet_registry
        from report_generator.sheet_validation import get_sheet_validator
        from db_utils import get_db_connection
        
        # Get database connection
        print("[DB] Connecting to database...")
        db = get_db_connection()
        
        # Create consolidated report generator
        print("[INIT] Creating consolidated report generator...")
        output_dir = "./test_output"
        os.makedirs(output_dir, exist_ok=True)
        
        generator = ConsolidatedReportGenerator(db, ".", output_dir)
        
        # Get registry and validator for monitoring
        registry = get_sheet_registry()
        validator = get_sheet_validator()
        
        # Clear any previous registry state
        registry.clear_registry()
        
        # Validate before generation
        print("[VALIDATE] Running pre-generation validation...")
        if not generator.validate_before_generation():
            print("[ERROR] Pre-generation validation failed")
            return False
        
        print("[OK] Pre-generation validation passed")
        
        # Generate report
        print("[GENERATE] Generating report with consolidated system...")
        success = generator.generate_report()
        
        if not success:
            print("[ERROR] Report generation failed")
            return False
        
        print("[OK] Report generation completed")
        
        # Get registry summary
        print("[REGISTRY] Checking sheet creation registry...")
        summary = registry.get_creation_summary()
        print(f"[REGISTRY] Total sheets created: {summary.get('total_created', 0)}")
        print(f"[REGISTRY] Duplicate attempts: {summary.get('duplicate_attempts', 0)}")
        print(f"[REGISTRY] Failed creations: {summary.get('failed_creations', 0)}")
        
        # Print detailed registry report
        print("\n[REGISTRY] Detailed creation report:")
        registry.print_creation_report()
        
        # Validate workbook integrity
        print("\n[VALIDATE] Running post-generation validation...")
        workbook_valid, issues = validator.validate_workbook_integrity(generator.workbook)
        
        if issues:
            print("[VALIDATION] Issues found:")
            validator.print_validation_report(issues)
        else:
            print("[OK] No validation issues found")
        
        # Check for duplicates
        duplicate_attempts = summary.get('duplicate_attempts', 0)
        if duplicate_attempts > 0:
            print(f"[WARNING] {duplicate_attempts} duplicate sheet creation attempts were prevented")
        else:
            print("[OK] No duplicate sheet creation attempts detected")
        
        # Get final workbook info
        if hasattr(generator, 'workbook') and generator.workbook:
            sheet_names = [sheet.title for sheet in generator.workbook.worksheets]
            print(f"[WORKBOOK] Final sheet count: {len(sheet_names)}")
            print(f"[WORKBOOK] Sheet names: {sheet_names}")
        
        return success and workbook_valid and duplicate_attempts == 0
        
    except Exception as e:
        print(f"[ERROR] Exception during test: {e}")
        print(f"[ERROR] Traceback: {traceback.format_exc()}")
        return False

def test_registry_functionality():
    """Test registry functionality in isolation."""
    print("[TEST] Testing registry functionality...")
    
    try:
        from report_generator.sheet_registry import get_sheet_registry
        import openpyxl
        
        registry = get_sheet_registry()
        registry.clear_registry()
        
        workbook = openpyxl.Workbook()
        
        # Test successful creation
        success1 = registry.create_sheet(workbook, "Test Sheet 1", None, "test", "test_method")
        success2 = registry.create_sheet(workbook, "Test Sheet 2", None, "test", "test_method")
        
        # Test duplicate prevention
        duplicate = registry.create_sheet(workbook, "Test Sheet 1", None, "test", "test_method")
        
        # Test invalid names
        invalid1 = registry.create_sheet(workbook, "Invalid/Name", None, "test", "test_method")
        invalid2 = registry.create_sheet(workbook, "", None, "test", "test_method")
        
        summary = registry.get_creation_summary()
        
        print(f"[REGISTRY] Successful creations: {summary.get('total_created', 0)}")
        print(f"[REGISTRY] Duplicate attempts: {summary.get('duplicate_attempts', 0)}")
        print(f"[REGISTRY] Failed creations: {summary.get('failed_creations', 0)}")
        
        expected_success = success1 and success2
        expected_prevention = not duplicate and not invalid1 and not invalid2
        
        if expected_success and expected_prevention:
            print("[OK] Registry functionality test passed")
            return True
        else:
            print("[ERROR] Registry functionality test failed")
            return False
            
    except Exception as e:
        print(f"[ERROR] Registry test exception: {e}")
        return False

def main():
    """Run all consolidated system tests."""
    print("[TESTING] Consolidated Sheet Creation System Tests")
    print("=" * 60)
    
    tests = [
        ("Registry Functionality", test_registry_functionality),
        ("Full Report Generation", test_consolidated_report_generation)
    ]
    
    passed = 0
    total = len(tests)
    
    for test_name, test_func in tests:
        print(f"\n[TEST] Running {test_name}...")
        print("-" * 40)
        
        if test_func():
            print(f"[OK] {test_name} PASSED")
            passed += 1
        else:
            print(f"[ERROR] {test_name} FAILED")
    
    print("\n" + "=" * 60)
    print(f"[RESULTS] Test Results: {passed}/{total} tests passed")
    
    if passed == total:
        print("[SUCCESS] All consolidated system tests passed!")
        print("   The system is ready for production use.")
    else:
        print("[WARNING] Some tests failed. Please review the issues above.")
    
    return passed == total

if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
