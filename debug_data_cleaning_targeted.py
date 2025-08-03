#!/usr/bin/env python3
"""
Targeted debug script to identify the specific error in Data Cleaning sheet creation
"""

import sys
import os
import traceback

# Add the project directory to the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def test_data_cleaning_method():
    """Test the specific Data Cleaning sheet creation method with detailed error tracking"""
    
    print("=== Targeted Data Cleaning Debug Test ===")
    
    try:
        # Import required modules
        print("[1] Importing modules...")
        from report_generator.sheet_creators import SheetCreator
        from db_utils import verify_database_connection
        from report_generator.formatters import ExcelFormatter
        import openpyxl
        
        # Get database connection
        print("[2] Getting database connection...")
        db = verify_database_connection()
        if not db:
            print("❌ Failed to get database connection")
            return
        
        # Create formatter
        print("[3] Creating formatter...")
        formatter = ExcelFormatter()
        
        # Create SheetCreator instance
        print("[4] Creating SheetCreator instance...")
        sheet_creator = SheetCreator(db, formatter)
        
        # Create workbook
        print("[5] Creating workbook...")
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Test the data cleaning sheet creation with detailed logging
        print("[6] Calling create_data_cleaning_sheet method...")
        print("    This should trigger: '[INFO] Running intersection analysis for Data Cleaning sheet...'")
        
        # Call the method and catch any exceptions
        sheet_creator.create_data_cleaning_sheet(wb)
        
        # Check if the sheet was created
        if 'Data Cleaning' in wb.sheetnames:
            ws = wb['Data Cleaning']
            print(f"✅ Data Cleaning sheet created with {ws.max_row} rows and {ws.max_column} columns")
            
            # Print first few rows to see content
            print("\n=== Sheet Content Preview ===")
            for row in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(9, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:50])
                    else:
                        row_data.append("")
                if any(row_data):
                    print(f"Row {row}: {' | '.join(row_data)}")
            
            # Save test file
            test_file = "test_data_cleaning_targeted.xlsx"
            wb.save(test_file)
            print(f"\n✅ Test file saved as: {test_file}")
            
        else:
            print("❌ Data Cleaning sheet was not created")
            print(f"Available sheets: {wb.sheetnames}")
        
    except Exception as e:
        print(f"❌ Error during test: {e}")
        print("\n=== Full Exception Traceback ===")
        traceback.print_exc()
        
        # Try to identify the specific line where the error occurred
        print("\n=== Error Analysis ===")
        if "aggregation" in str(e).lower():
            print("🔍 Error seems related to MongoDB aggregation")
        elif "database" in str(e).lower():
            print("🔍 Error seems related to database connection")
        elif "formatter" in str(e).lower():
            print("🔍 Error seems related to Excel formatting")
        elif "workbook" in str(e).lower():
            print("🔍 Error seems related to Excel workbook operations")
        else:
            print("🔍 Error type unclear - check full traceback above")

if __name__ == "__main__":
    test_data_cleaning_method()
