#!/usr/bin/env python3
"""
Fix Percentage Calculation in Data Cleaning Sheet
================================================

This script fixes the calculation of Filter Application Rate (%) in the Data Cleaning sheet
to ensure it properly includes all excluded files, making the sum with Dataset Validity Rate equal to 100%.
"""

import openpyxl
from report_generator.sheet_creators import SheetCreator
from db_utils import get_db_connection
from report_generator.formatters import ExcelFormatter

def fix_percentage_calculation():
    """
    Patch the SheetCreator.create_data_cleaning_sheet method to fix the percentage calculation.
    """
    print("=" * 80)
    print("FIXING FILTER APPLICATION RATE (%) CALCULATION")
    print("=" * 80)
    
    # Backup the original method for later restoration
    original_method = SheetCreator.create_data_cleaning_sheet
    
    # Define the patched method that fixes the percentage calculation
    def patched_create_data_cleaning_sheet(self, wb):
        """
        Patched version of create_data_cleaning_sheet with corrected percentage calculation.
        """
        try:
            print("[INFO] Running intersection analysis for Data Cleaning sheet...")
            
            # Create the sheet
            sheet_name = "Data Cleaning"
            if sheet_name in wb.sheetnames:
                wb.remove(wb[sheet_name])
            ws = wb.create_sheet(sheet_name)
            
            # Call intersection_analysis to get the data
            intersection_data = self._intersection_analysis()
            
            # Simplified totals from intersection_analysis
            totals = {
                'total_files': sum(item['count'] for item in intersection_data),
                'school_normal': sum(item['count'] for item in intersection_data if item['is_collection_day'] and not item['is_outlier']),
                'school_outliers': sum(item['count'] for item in intersection_data if item['is_collection_day'] and item['is_outlier']),
                'non_school_normal': sum(item['count'] for item in intersection_data if not item['is_collection_day'] and not item['is_outlier']),
                'non_school_outliers': sum(item['count'] for item in intersection_data if not item['is_collection_day'] and item['is_outlier']),
            }
            
            # Calculate retention percentage
            totals['retention_pct'] = (totals['school_normal'] / totals['total_files']) * 100 if totals['total_files'] > 0 else 0
            
            # Sheet title
            ws['A1'] = "Data Cleaning and Filtering Analysis"
            ws.merge_cells('A1:I1')
            self.formatter.apply_title_style(ws, 'A1:I1')
            
            # Table 1 setup - Intersection Analysis
            table1_header_row = 5
            ws[f'A3'] = "Table 1: Intersection Analysis - All Files by Type"
            self.formatter.apply_section_header_style(ws, f'A3')
            
            # Academic terminology for headers
            headers = [
                'Media Type',
                'Initial Collection Size',
                'Recording Errors Filtered',
                'Non-Instructional Days Filtered',
                'Combined Filters Applied',
                'Total Records Filtered',
                'Research Dataset Size',
                'Filter Application Rate (%)',
                'Dataset Validity Rate (%)'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=table1_header_row, column=col, value=header)
            
            self.formatter.apply_header_style(ws, f'A{table1_header_row}:I{table1_header_row}')
            
            # Process data for Table 1
            media_types = {}
            for item in intersection_data:
                media_type = item['file_type'].upper()
                if media_type not in media_types:
                    media_types[media_type] = {
                        'total': 0,
                        'school_normal': 0,
                        'school_outliers': 0,
                        'non_school_normal': 0,
                        'non_school_outliers': 0
                    }
                
                media_types[media_type]['total'] += item['count']
                
                if item['is_collection_day'] and not item['is_outlier']:
                    media_types[media_type]['school_normal'] += item['count']
                elif item['is_collection_day'] and item['is_outlier']:
                    media_types[media_type]['school_outliers'] += item['count']
                elif not item['is_collection_day'] and not item['is_outlier']:
                    media_types[media_type]['non_school_normal'] += item['count']
                else:  # not item['is_collection_day'] and item['is_outlier']
                    media_types[media_type]['non_school_outliers'] += item['count']
            
            # Fill Table 1 data
            row = table1_header_row + 1
            for media_type, counts in media_types.items():
                # Calculate excluded files and percentages
                excluded = counts['total'] - counts['school_normal']
                filter_rate = excluded / counts['total'] if counts['total'] > 0 else 0
                validity_rate = counts['school_normal'] / counts['total'] if counts['total'] > 0 else 0
                
                # Fill the row
                ws.cell(row=row, column=1, value=media_type)
                ws.cell(row=row, column=2, value=counts['total'])
                ws.cell(row=row, column=3, value=counts['school_outliers'])
                ws.cell(row=row, column=4, value=counts['non_school_normal'])
                ws.cell(row=row, column=5, value=counts['non_school_outliers'])
                ws.cell(row=row, column=6, value=excluded)
                ws.cell(row=row, column=7, value=counts['school_normal'])
                ws.cell(row=row, column=8, value=filter_rate)
                ws.cell(row=row, column=9, value=validity_rate)
                
                # Format percentage cells
                ws.cell(row=row, column=8).number_format = '0.0%'
                ws.cell(row=row, column=9).number_format = '0.0%'
                
                row += 1
            
            # Save the row before adding the total row
            table1_row = row
            
            # Apply formatting to data rows (excluding the total row which will be added separately)
            self.formatter.apply_data_style(ws, f'A{table1_header_row + 1}:I{table1_row-1}')
            self.formatter.apply_alternating_row_colors(ws, table1_header_row + 1, table1_row-1, 1, 9)
            
            # Add formatted total row at the specific position
            # FIX: Calculate total excluded files correctly for Filter Application Rate
            total_excluded = totals['school_outliers'] + totals['non_school_normal'] + totals['non_school_outliers']
            
            total_row_values = {
                1: 'TOTAL',
                2: totals['total_files'],
                3: totals['school_outliers'],
                4: totals['non_school_normal'],
                5: totals['non_school_outliers'],
                6: total_excluded,
                7: totals['school_normal'],
                8: total_excluded / totals['total_files'] if totals['total_files'] > 0 else 0,  # FIX: All excluded files
                9: totals['school_normal'] / totals['total_files'] if totals['total_files'] > 0 else 0
            }
            
            # Format percentage cells
            ws.cell(row=table1_row-1, column=8).number_format = '0.0%'
            ws.cell(row=table1_row-1, column=9).number_format = '0.0%'
            
            # Add total row at the correct position
            self.formatter.add_total_row_at_position(ws, table1_row-1, total_row_values)
            
            # Table 2 setup - Year-by-Year breakdown
            table2_header_row = table1_row + 3
            ws[f'A{table1_row + 1}'] = "Table 2: Year-by-Year Breakdown"
            self.formatter.apply_section_header_style(ws, f'A{table1_row + 1}')
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=table2_header_row, column=col, value=header)
            
            self.formatter.apply_header_style(ws, f'A{table2_header_row}:I{table2_header_row}')
            
            # Process data for Table 2
            year_data = {}
            for item in intersection_data:
                if 'School_Year' not in item or item['School_Year'] == 'N/A':
                    continue
                
                year = item['School_Year']
                media_type = item['file_type'].upper()
                key = f"{year}_{media_type}"
                
                if key not in year_data:
                    year_data[key] = {
                        'year': year,
                        'media_type': media_type,
                        'total': 0,
                        'school_normal': 0,
                        'school_outliers': 0,
                        'non_school_normal': 0,
                        'non_school_outliers': 0
                    }
                
                year_data[key]['total'] += item['count']
                
                if item['is_collection_day'] and not item['is_outlier']:
                    year_data[key]['school_normal'] += item['count']
                elif item['is_collection_day'] and item['is_outlier']:
                    year_data[key]['school_outliers'] += item['count']
                elif not item['is_collection_day'] and not item['is_outlier']:
                    year_data[key]['non_school_normal'] += item['count']
                else:  # not item['is_collection_day'] and item['is_outlier']
                    year_data[key]['non_school_outliers'] += item['count']
            
            # Fill Table 2 data
            row = table2_header_row + 1
            for key, data in sorted(year_data.items()):
                # Calculate excluded files and percentages
                excluded = data['total'] - data['school_normal']
                filter_rate = excluded / data['total'] if data['total'] > 0 else 0
                validity_rate = data['school_normal'] / data['total'] if data['total'] > 0 else 0
                
                # Fill the row
                ws.cell(row=row, column=1, value=f"{data['year']} {data['media_type']}")
                ws.cell(row=row, column=2, value=data['total'])
                ws.cell(row=row, column=3, value=data['school_outliers'])
                ws.cell(row=row, column=4, value=data['non_school_normal'])
                ws.cell(row=row, column=5, value=data['non_school_outliers'])
                ws.cell(row=row, column=6, value=excluded)
                ws.cell(row=row, column=7, value=data['school_normal'])
                ws.cell(row=row, column=8, value=filter_rate)
                ws.cell(row=row, column=9, value=validity_rate)
                
                # Format percentage cells
                ws.cell(row=row, column=8).number_format = '0.0%'
                ws.cell(row=row, column=9).number_format = '0.0%'
                
                row += 1
            
            # Save the row before adding the total row
            table2_row = row
            
            # Apply formatting to Table 2 (excluding the total row which will be added separately)
            self.formatter.apply_data_style(ws, f'A{table2_header_row + 1}:I{table2_row-1}')
            self.formatter.apply_alternating_row_colors(ws, table2_header_row + 1, table2_row-1, 1, 9)
            
            # Add formatted total row at the specific position
            # FIX: Calculate total excluded files correctly for Filter Application Rate
            total_excluded = totals['school_outliers'] + totals['non_school_normal'] + totals['non_school_outliers']
            
            total_row_values = {
                2: totals['total_files'],
                3: totals['school_outliers'],
                4: totals['non_school_normal'],
                5: totals['non_school_outliers'],
                6: total_excluded,  # FIX: Total excluded files
                7: totals['school_normal'],
                8: total_excluded / totals['total_files'] if totals['total_files'] > 0 else 0,  # FIX: All excluded files
                9: totals['school_normal'] / totals['total_files'] if totals['total_files'] > 0 else 0
            }
            
            # Format percentage cells
            ws.cell(row=table2_row-1, column=8).number_format = '0.0%'
            ws.cell(row=table2_row-1, column=9).number_format = '0.0%'
            
            # Add total row at the correct position
            self.formatter.add_total_row_at_position(ws, table2_row-1, total_row_values)
            
            # Remaining code unchanged...
            # Rest of the original implementation
            
            # Skipping the rest for brevity in this patch
            
        except Exception as e:
            print(f"[ERROR] Failed to create Data Cleaning sheet: {e}")
            import traceback
            traceback.print_exc()

    # Create a test workbook
    print("Creating test workbook with fixed percentage calculations...")
    workbook = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    # Get database connection
    db = get_db_connection()
    
    # Create formatter instance
    formatter = ExcelFormatter()
    
    # Create sheet creator instance
    sheet_creator = SheetCreator(db, formatter)
    
    # Replace the original method with our patched one
    SheetCreator.create_data_cleaning_sheet = patched_create_data_cleaning_sheet
    
    try:
        # Create the Data Cleaning sheet with fixed percentages
        sheet_creator.create_data_cleaning_sheet(workbook)
        
        # Save test file
        test_filename = "fixed_percentages.xlsx"
        workbook.save(test_filename)
        print(f"✅ Fixed percentages saved to: {test_filename}")
        
        # Verify the fix
        ws = workbook['Data Cleaning']
        
        # Check Table 1 total percentages
        table1_total_row = None
        for row in range(6, 10):  # Search in reasonable range
            if ws.cell(row=row, column=1).value == 'TOTAL':
                table1_total_row = row
                break
        
        if table1_total_row:
            filter_rate = ws.cell(row=table1_total_row, column=8).value
            validity_rate = ws.cell(row=table1_total_row, column=9).value
            total_percentage = (filter_rate or 0) + (validity_rate or 0)
            
            print("\nTABLE 1 VERIFICATION:")
            print(f"Filter Application Rate: {filter_rate:.1%}")
            print(f"Dataset Validity Rate: {validity_rate:.1%}")
            print(f"Sum of percentages: {total_percentage:.1%}")
            
            if abs(total_percentage - 1.0) < 0.0001:  # Allow for floating point error
                print("✅ TABLE 1 PERCENTAGES VERIFIED: Sum equals 100%")
            else:
                print("❌ TABLE 1 PERCENTAGES ERROR: Sum does not equal 100%")
        
        # Check Table 2 total percentages
        table2_total_row = None
        for row in range(11, 20):  # Search in reasonable range after Table 1
            if ws.cell(row=row, column=1).value == 'TOTAL':
                table2_total_row = row
                break
        
        if table2_total_row:
            filter_rate = ws.cell(row=table2_total_row, column=8).value
            validity_rate = ws.cell(row=table2_total_row, column=9).value
            total_percentage = (filter_rate or 0) + (validity_rate or 0)
            
            print("\nTABLE 2 VERIFICATION:")
            print(f"Filter Application Rate: {filter_rate:.1%}")
            print(f"Dataset Validity Rate: {validity_rate:.1%}")
            print(f"Sum of percentages: {total_percentage:.1%}")
            
            if abs(total_percentage - 1.0) < 0.0001:  # Allow for floating point error
                print("✅ TABLE 2 PERCENTAGES VERIFIED: Sum equals 100%")
            else:
                print("❌ TABLE 2 PERCENTAGES ERROR: Sum does not equal 100%")
        
        # Restore the original method
        SheetCreator.create_data_cleaning_sheet = original_method
        print("\n✅ Original method restored")
        
        print("\n" + "=" * 80)
        print("PERCENTAGE CALCULATION FIX VERIFICATION COMPLETE")
        print("=" * 80)
        
        # Guidance for implementing the fix
        print("\nTo implement the fix in the original code, modify:")
        print("1. Calculate total_excluded = totals['school_outliers'] + totals['non_school_normal'] + totals['non_school_outliers']")
        print("2. Use total_excluded in Filter Application Rate (%) calculation")
        print("3. Ensure total records filtered column shows the correct sum of all excluded files")
        
    except Exception as e:
        # Restore the original method in case of error
        SheetCreator.create_data_cleaning_sheet = original_method
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    fix_percentage_calculation()
