#!/usr/bin/env python3
"""
Simple test to verify MP3 Duration Analysis fixes:
1. Both school years (2021-2022 and 2022-2023) appear
2. Headers use standard formatting
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators.specialized import SpecializedSheetCreator
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection
import openpyxl

def test_mp3_fixes():
    """Test MP3 Duration Analysis fixes."""
    
    print("=== Testing MP3 Duration Analysis Fixes ===")
    
    # Initialize components
    db = get_db_connection()
    formatter = ExcelFormatter()
    creator = SpecializedSheetCreator(db, formatter)
    
    # Create a test workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    try:
        # Create the MP3 Duration Analysis sheet
        print("[INFO] Creating MP3 Duration Analysis sheet...")
        creator.create_mp3_duration_analysis_sheet(wb)
        
        # Save test file
        test_filename = "test_mp3_fixes_simple.xlsx"
        wb.save(test_filename)
        print(f"[SUCCESS] Test file saved: {test_filename}")
        
        # Verify the sheet content
        if 'MP3 Duration Analysis' in wb.sheetnames:
            ws = wb['MP3 Duration Analysis']
            print(f"[INFO] Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check for both school years
            print("\n[INFO] Checking for school year data...")
            found_2021_2022 = False
            found_2022_2023 = False
            
            for row in range(1, min(21, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    if '2021-2022' in str(cell_value):
                        found_2021_2022 = True
                        print(f"[FOUND] 2021-2022 data at Row {row}: {cell_value}")
                        # Show the complete row data
                        row_data = []
                        for col in range(1, 7):
                            val = ws.cell(row=row, column=col).value
                            row_data.append(str(val) if val else "None")
                        print(f"  Complete row: {row_data}")
                    elif '2022-2023' in str(cell_value):
                        found_2022_2023 = True
                        print(f"[FOUND] 2022-2023 data at Row {row}: {cell_value}")
                        # Show the complete row data
                        row_data = []
                        for col in range(1, 7):
                            val = ws.cell(row=row, column=col).value
                            row_data.append(str(val) if val else "None")
                        print(f"  Complete row: {row_data}")
            
            # Report results
            if found_2021_2022 and found_2022_2023:
                print("[SUCCESS] Both school years found!")
            elif found_2021_2022 and not found_2022_2023:
                print("[ERROR] Missing 2022-2023 school year data!")
            elif not found_2021_2022 and found_2022_2023:
                print("[ERROR] Missing 2021-2022 school year data!")
            else:
                print("[ERROR] No school year data found!")
                
            # Check header formatting
            print("\n[INFO] Checking header formatting...")
            for row in range(1, min(21, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'School Year' in str(cell_value):
                    cell = ws.cell(row=row, column=1)
                    print(f"[HEADER] Found header at Row {row}: {cell_value}")
                    
                    # Check if header has formatting
                    has_bold = cell.font.bold if cell.font else False
                    has_fill = cell.fill.start_color.index != '00000000' if cell.fill else False
                    has_border = bool(cell.border.top.style) if cell.border else False
                    
                    print(f"  Bold: {has_bold}, Fill: {has_fill}, Border: {has_border}")
                    
                    if has_bold or has_fill or has_border:
                        print("  [SUCCESS] Header has formatting!")
                    else:
                        print("  [WARNING] Header lacks formatting!")
        else:
            print("[ERROR] MP3 Duration Analysis sheet not found!")
            
    except Exception as e:
        print(f"[ERROR] Test failed: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        wb.close()

if __name__ == "__main__":
    test_mp3_fixes()
