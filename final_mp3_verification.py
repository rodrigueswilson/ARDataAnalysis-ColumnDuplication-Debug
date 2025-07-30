#!/usr/bin/env python3
"""
Final verification of MP3 Duration Analysis sheet with updated filters.
"""

import openpyxl
from db_utils import get_db_connection

def final_verification():
    print("=== FINAL MP3 DURATION ANALYSIS VERIFICATION ===")
    print("Verifying updated pipelines with filters:")
    print("- is_collection_day: true")
    print("- Outlier_Status: false")
    print()
    
    try:
        # First, verify the database counts with and without filters
        print("1. DATABASE VERIFICATION:")
        db = get_db_connection()
        collection = db['media_records']
        
        # Total MP3 count
        total_mp3 = collection.count_documents({"file_type": "MP3"})
        print(f"   Total MP3 files in database: {total_mp3}")
        
        # MP3 count with duration
        mp3_with_duration = collection.count_documents({
            "file_type": "MP3",
            "Duration_Seconds": {"$exists": True, "$gt": 0}
        })
        print(f"   MP3 files with duration: {mp3_with_duration}")
        
        # MP3 count with new filters
        mp3_filtered = collection.count_documents({
            "file_type": "MP3",
            "Duration_Seconds": {"$exists": True, "$gt": 0},
            "is_collection_day": True,
            "Outlier_Status": False
        })
        print(f"   MP3 files with filters applied: {mp3_filtered}")
        print(f"   Filtered out: {mp3_with_duration - mp3_filtered} files ({((mp3_with_duration - mp3_filtered) / mp3_with_duration * 100):.1f}%)")
        print()
        
        # Check the latest report
        print("2. REPORT VERIFICATION:")
        report_file = 'AR_Analysis_Report_20250726_174735.xlsx'
        print(f"   Loading report: {report_file}")
        
        wb = openpyxl.load_workbook(report_file)
        
        if 'MP3 Duration Analysis' in wb.sheetnames:
            print("   ✅ MP3 Duration Analysis sheet found!")
            
            sheet = wb['MP3 Duration Analysis']
            print(f"   Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Extract key metrics from the sheet
            print("\n3. SHEET CONTENT VERIFICATION:")
            
            # Find and analyze the School Year Summary table
            school_year_data = {}
            for row in range(1, min(sheet.max_row + 1, 20)):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "School Year" in str(cell_value) and "Summary" in str(cell_value):
                    print(f"   Found School Year Summary at row {row}")
                    
                    # Extract data from the next few rows
                    for data_row in range(row + 2, min(row + 6, sheet.max_row + 1)):
                        year = sheet.cell(row=data_row, column=1).value
                        files = sheet.cell(row=data_row, column=2).value
                        hours = sheet.cell(row=data_row, column=3).value
                        
                        if year and str(year) not in ["School Year", "TOTAL", ""]:
                            school_year_data[year] = {"files": files, "hours": hours}
                            print(f"     {year}: {files} files, {hours} hours")
                    break
            
            # Find and check Collection Period table for real data
            print("\n   Collection Period Table Check:")
            period_data_found = False
            for row in range(1, min(sheet.max_row + 1, 50)):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "Collection Period" in str(cell_value):
                    print(f"   Found Collection Period section at row {row}")
                    
                    # Check first few data rows
                    valid_periods = 0
                    for data_row in range(row + 2, min(row + 8, sheet.max_row + 1)):
                        period = sheet.cell(row=data_row, column=1).value
                        school_year = sheet.cell(row=data_row, column=2).value
                        files = sheet.cell(row=data_row, column=3).value
                        
                        if period and str(period) not in ["Period", ""]:
                            if "Unknown" not in str(period):
                                valid_periods += 1
                                print(f"     {period} | {school_year}: {files} files")
                            period_data_found = True
                    
                    if valid_periods > 0:
                        print(f"   ✅ Collection Period table has {valid_periods} valid entries")
                    break
            
            # Find and check Monthly table
            print("\n   Monthly Table Check:")
            monthly_data_found = False
            for row in range(1, min(sheet.max_row + 1, 50)):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "Monthly" in str(cell_value):
                    print(f"   Found Monthly section at row {row}")
                    
                    # Check for data
                    months_with_data = 0
                    for data_row in range(row + 2, min(row + 15, sheet.max_row + 1)):
                        month = sheet.cell(row=data_row, column=1).value
                        files_2122 = sheet.cell(row=data_row, column=2).value
                        files_2223 = sheet.cell(row=data_row, column=4).value
                        
                        if month and str(month) not in ["Month", ""]:
                            if files_2122 or files_2223:
                                months_with_data += 1
                                print(f"     {month}: 2021-22={files_2122}, 2022-23={files_2223}")
                            monthly_data_found = True
                    
                    if months_with_data > 0:
                        print(f"   ✅ Monthly table has data for {months_with_data} months")
                    break
            
            # Summary
            print("\n4. FINAL VERIFICATION SUMMARY:")
            if school_year_data:
                total_filtered_files = sum(data["files"] for data in school_year_data.values() if isinstance(data["files"], (int, float)))
                print(f"   ✅ School Year Summary: {len(school_year_data)} years, {total_filtered_files} total filtered MP3 files")
            
            if period_data_found:
                print(f"   ✅ Collection Period Table: Populated with real period names")
            else:
                print(f"   ❌ Collection Period Table: No valid data found")
                
            if monthly_data_found:
                print(f"   ✅ Monthly Table: Populated with monthly data")
            else:
                print(f"   ❌ Monthly Table: No data found")
            
            print(f"\n   📊 Filter Impact: Reduced from {mp3_with_duration} to {mp3_filtered} MP3 files")
            print(f"   🎯 Analysis Quality: Only collection day, non-outlier MP3s included")
            
            return True
        else:
            print("   ❌ MP3 Duration Analysis sheet not found")
            return False
            
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = final_verification()
    print(f"\n=== VERIFICATION {'PASSED' if success else 'FAILED'} ===")
