#!/usr/bin/env python3
"""
Debug script to test Data Cleaning sheet creation and identify execution issues.
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from db_utils import get_database_connection
from report_generator.excel_formatter import ExcelFormatter
from report_generator.sheet_creators import SheetCreator
import openpyxl

def test_data_cleaning_creation():
    """Test Data Cleaning sheet creation with detailed logging."""
    
    print("=== Data Cleaning Sheet Creation Debug Test ===")
    
    try:
        # Initialize components
        print("[1] Connecting to database...")
        db = get_database_connection()
        
        print("[2] Initializing formatter...")
        formatter = ExcelFormatter()
        
        print("[3] Creating unified sheet creator...")
        sheet_creator = SheetCreator(db, formatter)
        
        print("[4] Creating workbook...")
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        print("[5] Calling create_data_cleaning_sheet...")
        print("=" * 50)
        
        # Call the method and capture any output
        sheet_creator.create_data_cleaning_sheet(wb)
        
        print("=" * 50)
        print("[6] Method execution completed")
        
        # Check the results
        if 'Data Cleaning' in wb.sheetnames:
            ws = wb['Data Cleaning']
            print(f"[SUCCESS] Data Cleaning sheet created with {ws.max_row} rows and {ws.max_column} columns")
            
            print("\n=== Sheet Content Analysis ===")
            for row in range(1, min(ws.max_row + 1, 20)):
                row_data = []
                for col in range(1, min(ws.max_column + 1, 5)):
                    cell = ws.cell(row=row, column=col)
                    value = cell.value if cell.value is not None else ''
                    row_data.append(str(value))
                if any(row_data):  # Only print non-empty rows
                    print(f'Row {row}: {row_data}')
                    
            # Check for Data Quality Metrics section
            data_quality_found = False
            for row in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'Data Quality Metrics' in str(cell_value):
                    data_quality_found = True
                    print(f"\n[FOUND] Data Quality Metrics section at row {row}")
                    break
            
            if not data_quality_found:
                print("\n[MISSING] Data Quality Metrics section not found in sheet")
                
        else:
            print("[ERROR] Data Cleaning sheet was not created")
            print(f"Available sheets: {wb.sheetnames}")
        
        # Save test workbook
        test_filename = "debug_data_cleaning_test.xlsx"
        wb.save(test_filename)
        print(f"\n[SAVED] Test workbook saved as: {test_filename}")
        
    except Exception as e:
        print(f"[ERROR] Test failed: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_data_cleaning_creation()
