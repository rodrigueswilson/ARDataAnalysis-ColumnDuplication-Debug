#!/usr/bin/env python3
"""
Fix MP3 Duration Analysis issues:
1. Missing 2022-2023 school year row
2. Header formatting inconsistency
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators.specialized import SpecializedSheetCreator
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection
import openpyxl
import pandas as pd

def test_school_year_data():
    """Test the school year aggregation pipeline directly."""
    
    print("=== Testing School Year Data Pipeline ===")
    
    db = get_db_connection()
    
    # Test the exact pipeline used in the MP3 Duration Analysis
    school_year_pipeline = [
        {"$match": {
            "file_type": "MP3",
            "is_collection_day": True,
            "Outlier_Status": False
        }},
        {"$group": {
            "_id": "$School_Year",
            "Total_MP3_Files": {"$sum": 1},
            "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
            "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"},
            "Min_Duration_Seconds": {"$min": "$Duration_Seconds"},
            "Max_Duration_Seconds": {"$max": "$Duration_Seconds"}
        }},
        {"$addFields": {
            "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]}
        }},
        {"$sort": {"_id": 1}}
    ]
    
    # Execute pipeline
    cursor = db.media_records.aggregate(school_year_pipeline, allowDiskUse=True)
    raw_result = list(cursor)
    
    print(f"Raw aggregation result: {len(raw_result)} records")
    for item in raw_result:
        print(f"  {item}")
    
    # Convert to DataFrame and test flattening
    df = pd.DataFrame(raw_result)
    print(f"\nDataFrame shape: {df.shape}")
    print(f"DataFrame columns: {list(df.columns)}")
    
    # Test the flattening logic
    if '_id' in df.columns and df.shape[0] > 0 and isinstance(df['_id'].iloc[0], dict):
        print("Flattening _id field...")
        id_df = pd.json_normalize(df['_id'])
        df = pd.concat([df.drop('_id', axis=1), id_df], axis=1)
        print(f"After flattening - columns: {list(df.columns)}")
    
    print("\nFinal DataFrame content:")
    for _, row in df.iterrows():
        school_year = row.get('School_Year', row.get('_id', 'Unknown'))
        if isinstance(school_year, dict):
            school_year = school_year.get('School_Year', 'Unknown')
        files = row.get('Total_MP3_Files', 0)
        hours = row.get('Total_Duration_Hours', 0)
        print(f"  School Year: {school_year}, Files: {files}, Hours: {hours:.2f}")
    
    return df

def create_fixed_mp3_sheet():
    """Create MP3 Duration Analysis sheet with fixes."""
    
    print("\n=== Creating Fixed MP3 Duration Analysis Sheet ===")
    
    # Test the data first
    df = test_school_year_data()
    
    if df.empty:
        print("ERROR: No data returned from pipeline!")
        return
    
    # Initialize components
    db = get_db_connection()
    formatter = ExcelFormatter()
    creator = SpecializedSheetCreator(db, formatter)
    
    # Create a test workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    try:
        # Create the MP3 Duration Analysis sheet
        print("[INFO] Creating MP3 Duration Analysis sheet with fixes...")
        creator.create_mp3_duration_analysis_sheet(wb)
        
        # Save test file
        test_filename = "test_mp3_duration_fixed.xlsx"
        wb.save(test_filename)
        print(f"[SUCCESS] Test file saved: {test_filename}")
        
        # Verify the sheet content
        if 'MP3 Duration Analysis' in wb.sheetnames:
            ws = wb['MP3 Duration Analysis']
            print(f"[INFO] Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check for both school years
            print("\n[INFO] Checking for school year data...")
            found_2021_2022 = False
            found_2022_2023 = False
            
            for row in range(1, min(21, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    if '2021-2022' in str(cell_value):
                        found_2021_2022 = True
                        print(f"[FOUND] 2021-2022 data at Row {row}: {cell_value}")
                    elif '2022-2023' in str(cell_value):
                        found_2022_2023 = True
                        print(f"[FOUND] 2022-2023 data at Row {row}: {cell_value}")
            
            if found_2021_2022 and found_2022_2023:
                print("[SUCCESS] Both school years found!")
            elif found_2021_2022 and not found_2022_2023:
                print("[ERROR] Missing 2022-2023 school year data!")
            elif not found_2021_2022 and found_2022_2023:
                print("[ERROR] Missing 2021-2022 school year data!")
            else:
                print("[ERROR] No school year data found!")
                
            # Check header formatting
            print("\n[INFO] Checking header formatting...")
            for row in range(1, min(21, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'School Year' in str(cell_value):
                    cell = ws.cell(row=row, column=1)
                    print(f"[HEADER] Found header at Row {row}: {cell_value}")
                    print(f"  Font: {cell.font}")
                    print(f"  Fill: {cell.fill}")
                    print(f"  Border: {cell.border}")
        else:
            print("[ERROR] MP3 Duration Analysis sheet not found!")
            
    except Exception as e:
        print(f"[ERROR] Test failed: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        wb.close()

if __name__ == "__main__":
    create_fixed_mp3_sheet()
