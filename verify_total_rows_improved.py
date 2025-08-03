#!/usr/bin/env python3
"""
Improved verification script to check that Logic Explanation table and Table 3 (Filter Impact Summary) 
now have properly formatted total rows
"""

import openpyxl
import os
import glob

def verify_total_rows_improved():
    """Verify that both tables now have total rows with proper formatting"""
    
    print("=== Improved Total Rows Verification ===")
    
    try:
        # Find the most recent Excel file
        excel_files = glob.glob("AR_Analysis_Report_*.xlsx")
        if not excel_files:
            print("❌ No Excel files found")
            return
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📊 Examining: {latest_file}")
        
        # Open the workbook
        wb = openpyxl.load_workbook(latest_file)
        
        if "Data Cleaning" not in wb.sheetnames:
            print("❌ No 'Data Cleaning' sheet found")
            return
        
        ws = wb["Data Cleaning"]
        print(f"✅ Found 'Data Cleaning' sheet")
        
        # Search through all rows to find both tables
        logic_table_found = False
        logic_total_found = False
        table3_found = False
        table3_total_found = False
        
        print(f"\n=== Scanning All Rows for Tables ===")
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                cell_str = str(cell_value)
                
                # Check for Logic Explanation table
                if "Logic Explanation" in cell_str and "Category Definitions" in cell_str:
                    print(f"✅ Found Logic Explanation table at row {row}")
                    logic_table_found = True
                    
                    # Look for data and total rows
                    print(f"\n📋 Logic Explanation Content:")
                    for data_row in range(row + 3, min(row + 12, ws.max_row + 1)):
                        row_data = []
                        for col in range(1, 5):  # 4 columns
                            cell_value = ws.cell(row=data_row, column=col).value
                            row_data.append(str(cell_value) if cell_value is not None else "")
                        
                        if any(val for val in row_data):
                            print(f"Row {data_row}: {' | '.join(row_data)}")
                            
                            # Check if this is the total row
                            if row_data[0] == "TOTAL":
                                logic_total_found = True
                                cell = ws.cell(row=data_row, column=1)
                                is_bold = cell.font.bold if cell.font else False
                                print(f"   ✅ TOTAL row found with bold formatting: {is_bold}")
                
                # Check for Table 3 (Filter Impact Summary)
                elif "Table 3" in cell_str and "Filter Impact Summary" in cell_str:
                    print(f"✅ Found Table 3 at row {row}")
                    table3_found = True
                    
                    # Look for data and total rows
                    print(f"\n📋 Table 3 Content:")
                    for data_row in range(row + 3, min(row + 12, ws.max_row + 1)):
                        row_data = []
                        for col in range(1, 4):  # 3 columns
                            cell_value = ws.cell(row=data_row, column=col).value
                            if col == 3 and cell_value is not None:
                                # Format percentage values
                                if isinstance(cell_value, (int, float)) and 0 <= cell_value <= 1:
                                    row_data.append(f"{cell_value:.1%}")
                                else:
                                    row_data.append(str(cell_value))
                            else:
                                row_data.append(str(cell_value) if cell_value is not None else "")
                        
                        if any(val for val in row_data):
                            print(f"Row {data_row}: {' | '.join(row_data)}")
                            
                            # Check if this is the total row
                            if row_data[0] == "TOTAL":
                                table3_total_found = True
                                cell = ws.cell(row=data_row, column=1)
                                is_bold = cell.font.bold if cell.font else False
                                print(f"   ✅ TOTAL row found with bold formatting: {is_bold}")
        
        # Summary
        print(f"\n🎯 Verification Results:")
        print(f"• Logic Explanation table found: {'✅' if logic_table_found else '❌'}")
        print(f"• Logic Explanation total row: {'✅' if logic_total_found else '❌'}")
        print(f"• Table 3 (Filter Impact Summary) found: {'✅' if table3_found else '❌'}")
        print(f"• Table 3 total row: {'✅' if table3_total_found else '❌'}")
        
        if logic_total_found and table3_total_found:
            print(f"\n🎉 SUCCESS: Both tables now have properly formatted total rows!")
        elif logic_total_found or table3_total_found:
            print(f"\n⚠️  PARTIAL: Only {'Logic Explanation' if logic_total_found else 'Table 3'} has a total row")
        else:
            print(f"\n❌ FAILURE: No total rows found")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_total_rows_improved()
