#!/usr/bin/env python3
"""
Direct Data Cleaning Sheet Debug

This script tests the create_data_cleaning_sheet method directly to see if our
recent changes broke it.
"""

import sys
import os
import openpyxl
import traceback

# Add the project root to the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def test_data_cleaning_sheet_creation():
    """Test the Data Cleaning sheet creation directly."""
    print("DIRECT DATA CLEANING SHEET DEBUG")
    print("=" * 60)
    
    try:
        # Import the necessary modules
        from report_generator.sheet_creators.base import BaseSheetCreator
        from report_generator.config import Config
        from report_generator.database import DatabaseManager
        
        print("✅ Successfully imported modules")
        
        # Initialize components
        config = Config()
        db_manager = DatabaseManager(config)
        db = db_manager.get_database()
        
        print("✅ Successfully initialized database connection")
        
        # Create a test workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        print("✅ Created test workbook")
        
        # Create the sheet creator
        sheet_creator = BaseSheetCreator(db, config)
        
        print("✅ Created BaseSheetCreator instance")
        
        # Test the create_data_cleaning_sheet method directly
        print("\n🔍 TESTING create_data_cleaning_sheet METHOD:")
        print("=" * 50)
        
        try:
            print("[1] Calling create_data_cleaning_sheet...")
            sheet_creator.create_data_cleaning_sheet(wb)
            print("✅ create_data_cleaning_sheet completed without exception")
            
            # Check if the sheet was created
            if 'Data Cleaning' in wb.sheetnames:
                print("✅ Data Cleaning sheet was created successfully")
                
                # Check sheet content
                ws = wb['Data Cleaning']
                print(f"📊 Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
                
                # Check first few cells
                print("📋 First few cells:")
                for row in range(1, min(6, ws.max_row + 1)):
                    for col in range(1, min(4, ws.max_column + 1)):
                        cell_value = ws.cell(row=row, column=col).value
                        if cell_value:
                            print(f"   Row {row}, Col {col}: {cell_value}")
                
                # Save test file
                test_file = "test_data_cleaning_direct.xlsx"
                wb.save(test_file)
                print(f"✅ Saved test file: {test_file}")
                
                return True
                
            else:
                print("❌ Data Cleaning sheet was NOT created")
                print(f"Available sheets: {wb.sheetnames}")
                return False
                
        except Exception as e:
            print(f"❌ Exception in create_data_cleaning_sheet: {e}")
            print("🔍 Full traceback:")
            traceback.print_exc()
            return False
            
    except Exception as e:
        print(f"❌ Error in setup: {e}")
        print("🔍 Full traceback:")
        traceback.print_exc()
        return False

def check_database_connectivity():
    """Check if database connectivity is working."""
    print(f"\n🔍 CHECKING DATABASE CONNECTIVITY")
    print("=" * 50)
    
    try:
        from report_generator.config import Config
        from report_generator.database import DatabaseManager
        
        config = Config()
        db_manager = DatabaseManager(config)
        db = db_manager.get_database()
        
        # Test a simple query
        count = db.media_records.count_documents({})
        print(f"✅ Database connection successful")
        print(f"📊 Total records in media_records: {count}")
        
        # Test the specific queries used in Data Cleaning sheet
        test_queries = [
            {"School_Year": {"$ne": "N/A"}},
            {"School_Year": {"$ne": "N/A"}, "is_collection_day": True},
            {"School_Year": {"$ne": "N/A"}, "Outlier_Status": False},
            {"School_Year": {"$ne": "N/A"}, "is_collection_day": True, "Outlier_Status": False}
        ]
        
        print(f"\n📋 Testing Data Cleaning queries:")
        for i, query in enumerate(test_queries, 1):
            try:
                count = db.media_records.count_documents(query)
                print(f"   Query {i}: {count} records")
            except Exception as e:
                print(f"   Query {i}: ERROR - {e}")
        
        return True
        
    except Exception as e:
        print(f"❌ Database connectivity error: {e}")
        traceback.print_exc()
        return False

def main():
    print("TARGETED DATA CLEANING SHEET DEBUG")
    print("=" * 70)
    
    try:
        # Check database connectivity first
        db_working = check_database_connectivity()
        
        if not db_working:
            print("❌ Database connectivity failed - cannot proceed")
            return 1
        
        # Test Data Cleaning sheet creation
        sheet_working = test_data_cleaning_sheet_creation()
        
        print("\n" + "=" * 70)
        print("DEBUG RESULTS")
        print("=" * 70)
        
        if sheet_working:
            print("🎉 SUCCESS: Data Cleaning sheet creation is working!")
            print("   The issue may be elsewhere in the report generation flow")
        else:
            print("🚨 FAILURE: Data Cleaning sheet creation is broken")
            print("   Our recent changes likely broke this functionality")
        
        print(f"\n🎯 STATUS:")
        print(f"   Database connectivity: {'✅' if db_working else '❌'}")
        print(f"   Data Cleaning creation: {'✅' if sheet_working else '❌'}")
        
        return 0 if sheet_working else 1
        
    except Exception as e:
        print(f"❌ Error in main: {e}")
        traceback.print_exc()
        return 1

if __name__ == "__main__":
    sys.exit(main())
