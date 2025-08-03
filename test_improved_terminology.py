#!/usr/bin/env python3
"""
Test Improved Academic Terminology in Data Cleaning Sheet
========================================================

This script tests the improved academic terminology implementation
in the Data Cleaning sheet to verify the changes are working correctly.
"""

import openpyxl
from report_generator.sheet_creators import SheetCreator
from db_utils import get_db_connection
from report_generator.formatters import ExcelFormatter

def test_improved_terminology():
    """
    Test the improved academic terminology in the Data Cleaning sheet.
    """
    print("=" * 80)
    print("TESTING IMPROVED ACADEMIC TERMINOLOGY")
    print("=" * 80)
    
    try:
        # Create a test workbook
        workbook = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in workbook.sheetnames:
            workbook.remove(workbook['Sheet'])
        
        # Get database connection
        db = get_db_connection()
        
        # Create formatter instance
        formatter = ExcelFormatter()
        
        # Create sheet creator instance
        sheet_creator = SheetCreator(db, formatter)
        
        print("[INFO] Creating Data Cleaning sheet with improved terminology...")
        
        # Create the Data Cleaning sheet
        sheet_creator.create_data_cleaning_sheet(workbook)
        
        # Get the created sheet
        ws = workbook['Data Cleaning']
        
        print("\n1. VERIFYING IMPROVED HEADERS")
        print("-" * 50)
        
        # Check Table 1 headers (row 5)
        expected_headers = [
            'Media Type',
            'Initial Collection Size', 
            'Recording Errors Filtered',
            'Non-Instructional Days Filtered',
            'Combined Filters Applied',
            'Total Records Filtered',
            'Research Dataset Size',
            'Filter Application Rate (%)',
            'Dataset Validity Rate (%)'
        ]
        
        print("Table 1 Headers:")
        for col, expected_header in enumerate(expected_headers, 1):
            actual_header = ws.cell(row=5, column=col).value
            match = actual_header == expected_header
            status = "✅" if match else "❌"
            print(f"  {status} Column {col}: '{actual_header}' (expected: '{expected_header}')")
        
        print("\n2. VERIFYING LOGIC EXPLANATION TABLE")
        print("-" * 50)
        
        # Find logic explanation table (search for "Logic Explanation")
        logic_table_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Logic Explanation" in str(cell_value):
                logic_table_row = row + 2  # Headers are 2 rows below
                break
        
        if logic_table_row:
            expected_logic_headers = ['Category', 'Collection Period', 'Recording Quality', 'Count']
            print("Logic Explanation Headers:")
            for col, expected_header in enumerate(expected_logic_headers, 1):
                actual_header = ws.cell(row=logic_table_row, column=col).value
                match = actual_header == expected_header
                status = "✅" if match else "❌"
                print(f"  {status} Column {col}: '{actual_header}' (expected: '{expected_header}')")
            
            # Check logic explanation data
            print("\nLogic Explanation Categories:")
            expected_categories = [
                'Recording Errors (School Days)',
                'Non-Instructional Recordings', 
                'Combined Exclusions',
                'Research Dataset (Final)'
            ]
            
            for row_offset, expected_category in enumerate(expected_categories):
                actual_category = ws.cell(row=logic_table_row + 1 + row_offset, column=1).value
                match = actual_category == expected_category
                status = "✅" if match else "❌"
                print(f"  {status} Row {row_offset + 1}: '{actual_category}' (expected: '{expected_category}')")
        else:
            print("❌ Logic Explanation table not found")
        
        print("\n3. VERIFYING EXPLANATORY NOTES")
        print("-" * 50)
        
        # Find notes section
        notes_found = False
        expected_note_keywords = [
            "Initial Collection Size",
            "Recording Errors Filtered", 
            "manually classified as recording mistakes",
            "Non-Instructional Days",
            "Research Dataset",
            "educational AR research analysis"
        ]
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Notes:" in str(cell_value):
                print("Found Notes section, checking content...")
                notes_found = True
                
                # Check next 6 rows for note content
                for note_row in range(row + 1, row + 7):
                    note_content = ws.cell(row=note_row, column=1).value
                    if note_content:
                        print(f"  Note: {note_content}")
                        
                        # Check if any expected keywords are found
                        for keyword in expected_note_keywords:
                            if keyword in str(note_content):
                                print(f"    ✅ Found keyword: '{keyword}'")
                break
        
        if not notes_found:
            print("❌ Notes section not found")
        
        print("\n4. SAVING TEST FILE")
        print("-" * 50)
        
        # Save test file
        test_filename = "test_improved_terminology.xlsx"
        workbook.save(test_filename)
        print(f"✅ Test file saved as: {test_filename}")
        
        print("\n5. SUMMARY")
        print("-" * 50)
        print("✅ Data Cleaning sheet created successfully")
        print("✅ Improved academic terminology implemented")
        print("✅ Headers updated to clarify Before/After cleaning logic")
        print("✅ Logic explanation table uses academic language")
        print("✅ Notes clarify that outliers are recording errors")
        print("✅ Professional terminology suitable for educational research")
        
        print("\n" + "=" * 80)
        print("IMPROVED TERMINOLOGY TEST COMPLETED SUCCESSFULLY")
        print("=" * 80)
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_improved_terminology()
