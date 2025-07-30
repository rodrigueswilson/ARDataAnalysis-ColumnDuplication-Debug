#!/usr/bin/env python3
"""
Verification Script: Check for Duplicate Columns in Period Counts (ACF_PACF) Sheet
===================================================================================

This script verifies that the duplicate columns issue has been resolved in the
latest generated Excel report.
"""

import pandas as pd
import openpyxl
from pathlib import Path
import glob

def find_latest_report():
    """Find the most recent AR_Analysis_Report file."""
    pattern = "AR_Analysis_Report_*.xlsx"
    files = glob.glob(pattern)
    if not files:
        return None
    return max(files, key=lambda x: Path(x).stat().st_mtime)

def check_duplicate_columns(file_path, sheet_name="Period Counts (ACF_PACF)"):
    """Check for duplicate columns in the specified sheet."""
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        if sheet_name not in wb.sheetnames:
            print(f"[ERROR] Sheet '{sheet_name}' not found in {file_path}")
            print(f"Available sheets: {wb.sheetnames}")
            return False
        
        # Get the sheet
        ws = wb[sheet_name]
        
        # Extract column headers (first row)
        headers = []
        for cell in ws[1]:
            if cell.value is not None:
                headers.append(str(cell.value))
            else:
                break
        
        print(f"[INFO] Analyzing sheet: {sheet_name}")
        print(f"[INFO] Total columns found: {len(headers)}")
        print(f"[INFO] Unique columns: {len(set(headers))}")
        
        # Check for duplicates
        duplicates = []
        seen = set()
        for header in headers:
            if header in seen:
                duplicates.append(header)
            else:
                seen.add(header)
        
        if duplicates:
            print(f"[ERROR] Duplicate columns found: {duplicates}")
            print(f"[ERROR] Full column list:")
            for i, header in enumerate(headers, 1):
                marker = " ⚠️ DUPLICATE" if header in duplicates and headers[:i-1].count(header) > 0 else ""
                print(f"  {i:2d}. {header}{marker}")
            return False
        else:
            print(f"[SUCCESS] No duplicate columns found!")
            print(f"[INFO] Column structure:")
            for i, header in enumerate(headers, 1):
                print(f"  {i:2d}. {header}")
            return True
            
    except Exception as e:
        print(f"[ERROR] Failed to analyze file: {e}")
        return False

def main():
    print("=" * 70)
    print("DUPLICATE COLUMNS VERIFICATION")
    print("=" * 70)
    
    # Find latest report
    latest_report = find_latest_report()
    if not latest_report:
        print("[ERROR] No AR_Analysis_Report files found")
        return 1
    
    print(f"[INFO] Analyzing latest report: {latest_report}")
    
    # Check for duplicate columns
    success = check_duplicate_columns(latest_report)
    
    if success:
        print("\n[SUCCESS] ✅ Duplicate columns issue has been RESOLVED!")
        return 0
    else:
        print("\n[FAILURE] ❌ Duplicate columns issue still exists")
        return 1

if __name__ == "__main__":
    exit(main())
