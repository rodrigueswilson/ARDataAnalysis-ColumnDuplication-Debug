#!/usr/bin/env python3
"""
Script to examine the actual contents of the Data Cleaning sheet in the generated Excel file
"""

import openpyxl
import os
import glob

def examine_data_cleaning_sheet():
    """Examine the Data Cleaning sheet contents in the most recent Excel file"""
    
    print("=== Excel Data Cleaning Sheet Examination ===")
    
    try:
        # Find the most recent Excel file
        excel_files = glob.glob("AR_Analysis_Report_*.xlsx")
        if not excel_files:
            print("❌ No Excel files found")
            return
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📊 Examining: {latest_file}")
        
        # Open the workbook
        wb = openpyxl.load_workbook(latest_file)
        
        if "Data Cleaning" not in wb.sheetnames:
            print("❌ No 'Data Cleaning' sheet found")
            print(f"Available sheets: {wb.sheetnames}")
            return
        
        ws = wb["Data Cleaning"]
        print(f"✅ Found 'Data Cleaning' sheet")
        
        # Find all non-empty rows
        print(f"\n=== Sheet Contents ===")
        
        max_row = ws.max_row
        max_col = ws.max_column
        print(f"Sheet dimensions: {max_row} rows × {max_col} columns")
        
        # Print all content
        print(f"\n=== All Cell Contents ===")
        for row in range(1, min(max_row + 1, 50)):  # Limit to first 50 rows
            row_data = []
            has_content = False
            
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value is not None:
                    has_content = True
                row_data.append(str(cell_value) if cell_value is not None else "")
            
            if has_content:
                print(f"Row {row}: {' | '.join(row_data)}")
        
        # Look for specific patterns
        print(f"\n=== Looking for Year-by-Year Data ===")
        found_year_data = False
        
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str):
                    if "2021" in cell_value or "2022" in cell_value:
                        print(f"Found year reference at {row},{col}: {cell_value}")
                        found_year_data = True
                        
                        # Print surrounding context
                        for context_row in range(max(1, row-2), min(max_row+1, row+3)):
                            context_data = []
                            for context_col in range(1, max_col + 1):
                                context_value = ws.cell(row=context_row, column=context_col).value
                                context_data.append(str(context_value) if context_value is not None else "")
                            print(f"  Context Row {context_row}: {' | '.join(context_data)}")
                        print()
        
        if not found_year_data:
            print("❌ No year-by-year data found in Data Cleaning sheet")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    examine_data_cleaning_sheet()
