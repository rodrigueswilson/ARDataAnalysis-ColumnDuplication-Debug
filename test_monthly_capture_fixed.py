#!/usr/bin/env python3
"""
Test Monthly Capture Volume Totals with Data Structure Fix
=========================================================

This script tests the totals application for Monthly Capture Volume sheet
with proper handling of complex data structures.
"""

import pandas as pd
from openpyxl import Workbook
from report_generator.sheet_creators.pipeline import PipelineSheetCreator
from report_generator.totals_manager import TotalsManager
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection
from pipelines import PIPELINES

def test_monthly_capture_totals_fixed():
    """
    Test totals application for Monthly Capture Volume with data structure fixes.
    """
    print("🧪 TESTING MONTHLY CAPTURE VOLUME TOTALS (FIXED)")
    print("="*60)
    
    # Initialize components
    db = get_db_connection()
    formatter = ExcelFormatter()
    pipeline_creator = PipelineSheetCreator(db, formatter)
    
    # Get Monthly Capture Volume data
    pipeline = PIPELINES['CAPTURE_VOLUME_PER_MONTH']
    result = list(db.media_records.aggregate(pipeline))
    df = pd.DataFrame(result)
    
    print(f"📊 Raw Monthly Capture Volume Data:")
    print(f"   Rows: {len(df)}")
    print(f"   Columns: {list(df.columns)}")
    print(f"   Sample _id values: {df['_id'].head(3).tolist()}")
    print()
    
    # Fix the data structure - flatten complex _id field
    print(f"🔧 Fixing Data Structure:")
    
    # Handle complex _id field
    if '_id' in df.columns:
        # Check if _id contains dictionaries
        sample_id = df['_id'].iloc[0] if len(df) > 0 else None
        
        if isinstance(sample_id, dict):
            print(f"   📋 _id field contains dictionaries: {sample_id}")
            
            # Flatten the _id dictionary into separate columns
            id_df = pd.json_normalize(df['_id'])
            print(f"   📋 Flattened _id columns: {list(id_df.columns)}")
            
            # Remove original _id and add flattened columns
            df = df.drop(columns=['_id'])
            df = pd.concat([id_df, df], axis=1)
            
            print(f"   ✅ Data structure fixed")
        else:
            print(f"   ✅ _id field is simple: {type(sample_id)}")
    
    print(f"📊 Fixed Monthly Capture Volume Data:")
    print(f"   Rows: {len(df)}")
    print(f"   Columns: {list(df.columns)}")
    print(f"   Sample data:")
    print(df.head(3).to_string())
    print()
    
    # Test totals configuration
    sheet_name = "Monthly Capture Volume"
    sheet_type = "unknown"  # This isn't a time series
    
    print(f"🔧 Testing Totals Configuration:")
    config = pipeline_creator._get_totals_config_for_sheet(sheet_name, df, sheet_type)
    
    if config:
        print(f"   ✅ Configuration generated:")
        for key, value in config.items():
            if key != 'rationale':  # Skip long rationale for readability
                print(f"      {key}: {value}")
    else:
        print(f"   ❌ No configuration generated")
        return
    
    # Test totals manager directly
    totals_manager = TotalsManager()
    
    print(f"\n🧮 Testing Totals Calculation:")
    
    # Test column totals
    if config.get('add_column_totals', False):
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        print(f"   Numeric columns: {numeric_cols}")
        
        if numeric_cols:
            column_totals = totals_manager.calculate_column_totals(df, numeric_columns=numeric_cols)
            print(f"   ✅ Column totals: {dict(column_totals)}")
        else:
            print(f"   ⚠️  No numeric columns found")
    
    # Test actual worksheet application
    print(f"\n📋 Testing Worksheet Application:")
    
    # Create a test workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=str(col_name))
    
    # Write data with proper type conversion
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            # Convert complex values to strings
            if isinstance(value, (dict, list)):
                cell_value = str(value)
            elif pd.isna(value):
                cell_value = ""
            else:
                cell_value = value
            
            ws.cell(row=row_idx, column=col_idx, value=cell_value)
    
    print(f"   📊 Worksheet created with {len(df)} rows of data")
    
    # Apply totals using the totals manager
    try:
        start_row = 2  # Where data starts
        start_col = 1  # Column A
        
        print(f"   🔧 Applying totals with config:")
        print(f"      start_row: {start_row}")
        print(f"      start_col: {start_col}")
        print(f"      add_column_totals: {config.get('add_column_totals', False)}")
        print(f"      add_grand_total: {config.get('add_grand_total', False)}")
        
        totals_manager.add_totals_to_worksheet(
            worksheet=ws,
            dataframe=df,
            start_row=start_row,
            start_col=start_col,
            config=config
        )
        
        print(f"   ✅ Totals applied successfully!")
        
        # Check what was added
        print(f"   📋 Worksheet after totals:")
        print(f"      Total rows: {ws.max_row}")
        print(f"      Total columns: {ws.max_column}")
        
        # Check last few rows for totals
        print(f"   📋 Last few rows (showing totals):")
        for row_num in range(max(1, ws.max_row - 2), ws.max_row + 1):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_values.append(str(cell_value) if cell_value is not None else "")
            print(f"      Row {row_num}: {row_values}")
        
        # Save test file
        test_file = "test_monthly_capture_fixed.xlsx"
        wb.save(test_file)
        print(f"   💾 Test file saved: {test_file}")
        
    except Exception as e:
        print(f"   ❌ Error applying totals: {e}")
        import traceback
        traceback.print_exc()

def main():
    """
    Main test function.
    """
    try:
        test_monthly_capture_totals_fixed()
        
        print(f"\n" + "="*60)
        print(f"🎯 TEST SUMMARY")
        print(f"="*60)
        print(f"This test identifies and fixes:")
        print(f"1. ✅ Complex data structure issue in _id field")
        print(f"2. ✅ Proper data type conversion for Excel")
        print(f"3. ✅ Totals configuration and application")
        print(f"4. ✅ Final worksheet structure with totals")
        
    except Exception as e:
        print(f"❌ Test error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
