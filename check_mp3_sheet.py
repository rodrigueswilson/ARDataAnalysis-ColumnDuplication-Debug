import openpyxl

# Load the latest report
wb = openpyxl.load_workbook('AR_Analysis_Report_20250726_171918.xlsx')

# Check for MP3 Duration Analysis sheet
sheet_names = wb.sheetnames
print(f"Total sheets: {len(sheet_names)}")

for i, name in enumerate(sheet_names):
    print(f"{i+1:2d}. {name}")

if 'MP3 Duration Analysis' in sheet_names:
    print("\nSUCCESS: MP3 Duration Analysis sheet found!")
    mp3_sheet = wb['MP3 Duration Analysis']
    print(f"Sheet has {mp3_sheet.max_row} rows and {mp3_sheet.max_column} columns")
else:
    print("\nERROR: MP3 Duration Analysis sheet NOT found!")
