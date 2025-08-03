#!/usr/bin/env python3
"""
Investigate Monthly Capture Volume Sheet
=======================================

This script investigates why Monthly Capture Volume totals are still not working
despite data structure fixes.
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

def investigate_monthly_capture_sheet():
    """
    Investigate the Monthly Capture Volume sheet in the latest report.
    """
    print("🔍 INVESTIGATING MONTHLY CAPTURE VOLUME SHEET")
    print("="*60)
    
    # Find the latest report file
    report_files = list(Path(".").glob("AR_Analysis_Report_*.xlsx"))
    if not report_files:
        print("❌ No report files found")
        return
    
    latest_report = max(report_files, key=lambda x: x.stat().st_mtime)
    print(f"📊 Analyzing report: {latest_report}")
    
    try:
        # Load the workbook
        wb = load_workbook(latest_report, data_only=True)
        
        # Check if Monthly Capture Volume sheet exists
        if "Monthly Capture Volume" not in wb.sheetnames:
            print("❌ Monthly Capture Volume sheet not found")
            print(f"   Available sheets: {wb.sheetnames}")
            return
        
        ws = wb["Monthly Capture Volume"]
        print(f"✅ Found Monthly Capture Volume sheet")
        
        # Analyze sheet structure
        print(f"\n📋 Sheet Structure:")
        print(f"   Max row: {ws.max_row}")
        print(f"   Max column: {ws.max_column}")
        
        # Read all data from the sheet
        data = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append(row)
        
        print(f"   Data rows: {len(data)}")
        
        # Show first few rows
        print(f"\n📊 Sheet Content (first 10 rows):")
        for i, row in enumerate(data[:10]):
            print(f"   Row {i+1}: {row}")
        
        # Show last few rows (where totals should be)
        print(f"\n📊 Sheet Content (last 5 rows):")
        for i, row in enumerate(data[-5:], len(data)-4):
            print(f"   Row {i}: {row}")
        
        # Check for totals patterns
        print(f"\n🔍 Searching for Totals:")
        totals_found = False
        
        for i, row in enumerate(data):
            if row and any(cell and 'TOTAL' in str(cell).upper() for cell in row):
                print(f"   ✅ Found totals row {i+1}: {row}")
                totals_found = True
        
        if not totals_found:
            print(f"   ❌ No totals rows found")
        
        # Check for numeric data that should have totals
        print(f"\n🔢 Analyzing Numeric Data:")
        
        if len(data) > 1:  # Has header + data
            headers = data[0] if data else []
            print(f"   Headers: {headers}")
            
            # Find numeric columns
            numeric_cols = []
            for col_idx, header in enumerate(headers):
                if header:
                    # Check if this column has numeric data
                    numeric_values = []
                    for row in data[1:]:  # Skip header
                        if col_idx < len(row) and row[col_idx] is not None:
                            try:
                                val = float(row[col_idx])
                                numeric_values.append(val)
                            except (ValueError, TypeError):
                                pass
                    
                    if numeric_values:
                        numeric_cols.append((col_idx, header, len(numeric_values), sum(numeric_values)))
            
            print(f"   Numeric columns found: {len(numeric_cols)}")
            for col_idx, header, count, total in numeric_cols:
                print(f"      Column {col_idx+1} ({header}): {count} values, sum = {total}")
        
        # Check if sheet creation succeeded but totals failed
        print(f"\n🧪 Totals Analysis:")
        if len(data) > 1 and numeric_cols:
            print(f"   ✅ Sheet has data and numeric columns - totals should be possible")
            print(f"   ❌ But no totals found - indicates totals application failed")
        else:
            print(f"   ⚠️  Sheet may not have proper structure for totals")
        
    except Exception as e:
        print(f"❌ Error analyzing sheet: {e}")
        import traceback
        traceback.print_exc()

def check_pipeline_data():
    """
    Check the raw pipeline data for Monthly Capture Volume.
    """
    print(f"\n🔍 CHECKING RAW PIPELINE DATA")
    print("="*40)
    
    try:
        from db_utils import get_db_connection
        from pipelines import PIPELINES
        
        db = get_db_connection()
        pipeline = PIPELINES['CAPTURE_VOLUME_PER_MONTH']
        result = list(db.media_records.aggregate(pipeline))
        df = pd.DataFrame(result)
        
        print(f"📊 Raw Pipeline Data:")
        print(f"   Rows: {len(df)}")
        print(f"   Columns: {list(df.columns)}")
        
        # Check for complex structures
        if '_id' in df.columns and len(df) > 0:
            sample_id = df['_id'].iloc[0]
            print(f"   Sample _id: {sample_id} (type: {type(sample_id)})")
            
            if isinstance(sample_id, dict):
                print(f"   ⚠️  _id contains complex dictionary - needs flattening")
            else:
                print(f"   ✅ _id is simple structure")
        
        print(f"\n📋 Sample Data:")
        print(df.head(3).to_string())
        
    except Exception as e:
        print(f"❌ Error checking pipeline data: {e}")

def main():
    """
    Main investigation function.
    """
    try:
        investigate_monthly_capture_sheet()
        check_pipeline_data()
        
        print(f"\n" + "="*60)
        print(f"🎯 INVESTIGATION SUMMARY")
        print(f"="*60)
        print(f"This investigation should reveal:")
        print(f"1. Whether Monthly Capture Volume sheet exists and has data")
        print(f"2. Whether the data structure is properly flattened")
        print(f"3. Whether numeric columns exist for totals")
        print(f"4. Whether totals application failed despite proper data")
        
    except Exception as e:
        print(f"❌ Investigation error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
