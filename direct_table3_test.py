#!/usr/bin/env python3
"""
Direct test script to force Table 3 creation in the Data Cleaning sheet
"""

import os
import sys
import pymongo
import openpyxl
from datetime import datetime

# Add the project root to the path
sys.path.insert(0, os.path.abspath('.'))

# Import necessary modules
from report_generator.sheet_creators import SheetCreator
from report_generator.formatters import ExcelFormatter

def force_table3_creation():
    """Force Table 3 creation in an existing Excel report"""
    print("[TEST] Starting direct Table 3 creation test...")
    
    try:
        # Find the most recent Excel file
        import glob
        excel_files = glob.glob("AR_Analysis_Report_*.xlsx")
        if not excel_files:
            print("❌ No Excel files found")
            return
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📊 Opening: {latest_file}")
        
        # Open the workbook
        wb = openpyxl.load_workbook(latest_file)
        
        if "Data Cleaning" not in wb.sheetnames:
            print("❌ No 'Data Cleaning' sheet found")
            return
        
        ws = wb["Data Cleaning"]
        print(f"✅ Found 'Data Cleaning' sheet with {ws.max_row} rows")
        
        # Find the Logic Explanation table
        logic_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Logic Explanation" in str(cell_value):
                logic_row = row
                print(f"✅ Found Logic Explanation at row {row}")
                break
        
        if not logic_row:
            print("❌ Logic Explanation table not found")
            return
        
        # Find the total row of the Logic Explanation table
        total_row = None
        for row in range(logic_row + 1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "TOTAL" in str(cell_value):
                total_row = row
                print(f"✅ Found Logic Explanation total row at row {row}")
                break
        
        if not total_row:
            print("❌ Logic Explanation total row not found")
            return
        
        # Calculate the totals from the Logic Explanation table
        school_outliers = ws.cell(row=logic_row + 3, column=4).value or 0
        non_school_normal = ws.cell(row=logic_row + 4, column=4).value or 0
        non_school_outliers = ws.cell(row=logic_row + 5, column=4).value or 0
        school_normal = ws.cell(row=logic_row + 6, column=4).value or 0
        total_files = ws.cell(row=total_row, column=4).value or 0
        
        totals = {
            'school_outliers': school_outliers,
            'non_school_normal': non_school_normal,
            'non_school_outliers': non_school_outliers,
            'school_normal': school_normal,
            'total_files': total_files,
            'retention_pct': (school_normal / total_files * 100) if total_files > 0 else 0
        }
        
        print(f"✅ Calculated totals: {totals}")
        
        # Create Table 3 directly
        table3_start_row = total_row + 3
        print(f"⚙️ Creating Table 3 at row {table3_start_row}")
        
        # Initialize formatter
        formatter = ExcelFormatter()
        
        # Table 3 header
        ws.cell(row=table3_start_row, column=1, value="Table 3: Filter Impact Summary")
        formatter.apply_section_header_style(ws, f'A{table3_start_row}')
        
        # Headers for summary table
        summary_headers = ['Filter Category', 'Number of Files', '% of All Files']
        summary_header_row = table3_start_row + 2
        
        for col, header in enumerate(summary_headers, 1):
            ws.cell(row=summary_header_row, column=col, value=header)
            ws.cell(row=summary_header_row, column=col).font = openpyxl.styles.Font(bold=True)
        
        # Calculate summary data
        summary_data = [
            ('School Outliers (unusual files from school days)', totals['school_outliers'], 
             (totals['school_outliers'] / totals['total_files'] * 100) if totals['total_files'] > 0 else 0),
            ('Non-School Normal (regular files from non-school days)', totals['non_school_normal'], 
             (totals['non_school_normal'] / totals['total_files'] * 100) if totals['total_files'] > 0 else 0),
            ('School Normal (final research dataset)', totals['school_normal'], 
             (totals['school_normal'] / totals['total_files'] * 100) if totals['total_files'] > 0 else 0),
            ('Non-School Outliers (unusual files from non-school days)', totals['non_school_outliers'], 
             (totals['non_school_outliers'] / totals['total_files'] * 100) if totals['total_files'] > 0 else 0)
        ]
        
        # Fill summary data
        summary_row = summary_header_row + 1
        
        for criterion, count, percentage in summary_data:
            ws.cell(row=summary_row, column=1, value=criterion)
            ws.cell(row=summary_row, column=2, value=count)
            
            # Format percentage
            pct_cell = ws.cell(row=summary_row, column=3, value=percentage / 100)
            pct_cell.number_format = '0.0%'
            
            print(f"  - Added row: {criterion} | {count} | {percentage:.1f}%")
            summary_row += 1
        
        # Add total row
        ws.cell(row=summary_row, column=1, value='TOTAL')
        ws.cell(row=summary_row, column=2, value=totals['total_files'])
        
        # Format total percentage
        total_pct_cell = ws.cell(row=summary_row, column=3, value=1.0)  # 100%
        total_pct_cell.number_format = '0.0%'
        
        # Apply bold formatting to total row
        for col in range(1, 4):
            cell = ws.cell(row=summary_row, column=col)
            cell.font = openpyxl.styles.Font(bold=True)
        
        print(f"  - Added total row: TOTAL | {totals['total_files']} | 100.0%")
        
        # Save the modified workbook with a new name
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        new_filename = f"AR_Analysis_Report_with_Table3_{timestamp}.xlsx"
        wb.save(new_filename)
        print(f"✅ Saved modified workbook as {new_filename}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    force_table3_creation()
