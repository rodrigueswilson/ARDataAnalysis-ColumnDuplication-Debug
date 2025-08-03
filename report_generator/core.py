"""
Core Report Generator Class
==========================

This module contains the main ReportGenerator class that orchestrates the
report generation process. It handles database connections, pipeline execution,
and coordinates with other modules for sheet creation and formatting.
"""

import os
import datetime
import pandas as pd
import openpyxl
from pathlib import Path

from pipelines import PIPELINES  # Now using modular pipelines/ package
from ar_utils import (
    add_acf_pacf_analysis, infer_sheet_type, reorder_with_acf_pacf,
    get_school_calendar, get_non_collection_days, add_arima_forecast_columns,
    reorder_with_forecast_columns
)
# Import chart modules conditionally to avoid import errors
try:
    from acf_pacf_charts import enhance_acf_pacf_visualization, enhance_arima_forecast_visualization
except ImportError:
    enhance_acf_pacf_visualization = None
    enhance_arima_forecast_visualization = None
    
try:
    from dashboard_generator import create_dashboard_summary
except ImportError:
    create_dashboard_summary = None

from .formatters import ExcelFormatter
from .dashboard import DashboardCreator
from .raw_data import RawDataCreator

class ReportGenerator:
    """
    Main report generator class that orchestrates the creation of comprehensive
    Excel reports from MongoDB data.
    
    This class serves as the central coordinator, delegating specific tasks to
    specialized modules while maintaining the overall report generation workflow.
    """
    
    def __init__(self, db, root_dir, output_dir=None):
        """
        Initialize the report generator.
        
        Args:
            db: MongoDB database connection
            root_dir (str): Root directory for the project
            output_dir (str, optional): Output directory for reports. Defaults to root_dir.
        """
        self.db = db
        self.root_dir = root_dir
        self.output_dir = output_dir or root_dir
        self.workbook = openpyxl.Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Initialize specialized modules
        self.formatter = ExcelFormatter()
        self.dashboard_creator = DashboardCreator(self.db, self.formatter)
        self.raw_data_creator = RawDataCreator(self.db, self.formatter)
    
    def _get_school_year(self, date):
        """
        Determines the school year based on the given date.
        
        Args:
            date: Date object or string
            
        Returns:
            str: School year in format "YYYY-YY"
        """
        if hasattr(date, 'month'):
            month = date.month
            year = date.year
        else:
            # Assume it's a string in ISO format
            year, month = int(date[:4]), int(date[5:7])
        
        if month >= 8:  # August or later
            return f"{year}-{str(year + 1)[2:]}"
        else:  # Before August
            return f"{year - 1}-{str(year)[2:]}"
    
    def _seconds_to_hms(self, seconds):
        """
        Converts a duration in seconds to an HH:MM:SS format string.
        
        Args:
            seconds (float): Duration in seconds
            
        Returns:
            str: Formatted time string
        """
        if pd.isna(seconds) or seconds is None:
            return "00:00:00"
        
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        secs = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{secs:02d}"
    
    def _run_aggregation(self, pipeline, use_base_filter=True, collection_name='media_records'):
        """
        Runs a MongoDB aggregation pipeline and returns a DataFrame.
        
        Args:
            pipeline (list): MongoDB aggregation pipeline
            use_base_filter (bool): Whether to apply base filtering
            collection_name (str): Name of the MongoDB collection
            
        Returns:
            pandas.DataFrame: Results of the aggregation
        """
        try:
            collection = self.db[collection_name]
            
            # Apply base filter if requested
            if use_base_filter:
                base_filter = {"$match": {"file_type": {"$in": ["JPG", "MP3"]}}}
                full_pipeline = [base_filter] + pipeline
            else:
                full_pipeline = pipeline
            
            # Execute aggregation
            cursor = collection.aggregate(full_pipeline, allowDiskUse=True)
            results = list(cursor)
            
            if not results:
                print(f"[WARNING] Aggregation returned no results for collection: {collection_name}")
                return pd.DataFrame()
            
            # Convert to DataFrame
            df = pd.DataFrame(results)
            
            # Clean up MongoDB ObjectId columns if present
            if '_id' in df.columns and hasattr(df['_id'].iloc[0], 'inserted_id'):
                df = df.drop('_id', axis=1)
            
            return df
            
        except Exception as e:
            print(f"[ERROR] Aggregation failed for collection {collection_name}: {e}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()
    
    def _zero_fill_daily_counts(self, df):
        """
        Post-processes daily counts DataFrame to include zero entries for all collection days.
        Uses the actual school calendar logic from ar_utils.py for accurate zero-filling.
        
        Args:
            df (pandas.DataFrame): Daily counts DataFrame
            
        Returns:
            pandas.DataFrame: DataFrame with zero-filled collection days
        """
        if df.empty:
            return df
        
        try:
            # Get the school calendar and non-collection days
            school_calendar = get_school_calendar()
            non_collection_days = get_non_collection_days()
            
            # Use the correct precompute_collection_days function
            from ar_utils import precompute_collection_days
            collection_day_map = precompute_collection_days(school_calendar, non_collection_days)
            
            # Get all collection days as a list
            all_collection_days = list(collection_day_map.keys())
            
            # Convert to DataFrame for merging
            all_days_df = pd.DataFrame({'_id': all_collection_days})
            all_days_df['_id'] = pd.to_datetime(all_days_df['_id']).dt.strftime('%Y-%m-%d')
            
            # Merge with existing data
            df['_id'] = pd.to_datetime(df['_id']).dt.strftime('%Y-%m-%d')
            merged_df = all_days_df.merge(df, on='_id', how='left')
            
            # Fill missing values with zeros
            numeric_columns = ['Total_Files', 'MP3_Files', 'JPG_Files', 'Total_Size_MB']
            for col in numeric_columns:
                if col in merged_df.columns:
                    merged_df[col] = merged_df[col].fillna(0)
            
            # Add has_files column
            merged_df['has_files'] = merged_df['Total_Files'] > 0
            
            # Sort by date
            merged_df = merged_df.sort_values('_id')
            
            print(f"[ZERO-FILL] Added {len(all_collection_days) - len(df)} zero-fill days to daily counts")
            print(f"[ZERO-FILL] Total days: {len(merged_df)} (was {len(df)})")
            print(f"[ZERO-FILL] Days with files: {merged_df['has_files'].sum()}")
            print(f"[ZERO-FILL] Days without files: {(~merged_df['has_files']).sum()}")
            
            return merged_df
            
        except Exception as e:
            print(f"[WARNING] Zero-fill failed: {e}. Returning original DataFrame.")
            import traceback
            traceback.print_exc()
            return df
    
    def _fill_missing_collection_days(self, df, pipeline_name):
        """
        Fill in missing collection days with zero counts for complete time series.
        This is critical for ACF/PACF/ARIMA analysis which requires continuous time series.
        
        FIXED: Now includes ALL collection days from the official school year start date
        to resolve left-aligned row issues and ensure consistent totals.
        
        Args:
            df (pandas.DataFrame): DataFrame to fill
            pipeline_name (str): Name of the pipeline being processed
            
        Returns:
            pandas.DataFrame: DataFrame with filled missing days
        """
        if not ('DAILY' in pipeline_name.upper() and 
                ('WITH_ZEROES' in pipeline_name.upper() or 'COLLECTION_ONLY' in pipeline_name.upper())):
            return df
        
        try:
            from ar_utils import get_school_calendar, get_non_collection_days, precompute_collection_days
            import pandas as pd
            
            school_calendar = get_school_calendar()
            non_collection_days = get_non_collection_days()
            collection_day_map = precompute_collection_days(school_calendar, non_collection_days)
            
            # CRITICAL FIX: Ensure ALL collection days are included in zero-fill
            # This resolves the left-aligned row issue by including early September dates
            all_collection_days = []
            for date_obj, info in collection_day_map.items():
                all_collection_days.append({'_id': date_obj.strftime('%Y-%m-%d')})
            
            all_days_df = pd.DataFrame(all_collection_days)
            
            # DEBUG: Log the date range being used
            if all_collection_days:
                min_date = min(day['_id'] for day in all_collection_days)
                max_date = max(day['_id'] for day in all_collection_days)
                print(f"[ZERO_FILL] Including all collection days from {min_date} to {max_date}")
                print(f"[ZERO_FILL] Total collection days: {len(all_collection_days)}")
            
            # Merge with existing data, keeping actual data and filling missing with zeros
            merged_df = pd.merge(all_days_df, df, on='_id', how='left').fillna(0)
            
            # Ensure correct data types after merge
            for col in ['Total_Files', 'MP3_Files', 'JPG_Files']:
                if col in merged_df.columns:
                    merged_df[col] = merged_df[col].astype(int)
            if 'Total_Size_MB' in merged_df.columns:
                 merged_df['Total_Size_MB'] = merged_df['Total_Size_MB'].astype(float)

            # CRITICAL FIX: Sort by date to ensure proper chronological order
            # This ensures early September dates appear in the correct position
            final_df = merged_df.sort_values('_id').reset_index(drop=True)
            
            # DEBUG: Log early September inclusion
            early_sept_dates = [
                "2021-09-13", "2021-09-14", "2021-09-15", "2021-09-16", "2021-09-17",
                "2021-09-20", "2021-09-21", "2021-09-22", "2021-09-23", "2021-09-24", "2021-09-27"
            ]
            
            early_sept_count = 0
            for date_str in early_sept_dates:
                if date_str in final_df['_id'].values:
                    row = final_df[final_df['_id'] == date_str]
                    if not row.empty:
                        count = row.iloc[0]['Total_Files'] if 'Total_Files' in row.columns else 0
                        early_sept_count += count
            
            if early_sept_count > 0:
                print(f"[ZERO_FILL] Early September files included: {early_sept_count}")
                print(f"[ZERO_FILL] This should resolve left-aligned row issues")
            
            total_files = final_df['Total_Files'].sum() if 'Total_Files' in final_df.columns else 0
            print(f"[ZERO_FILL] Total files after zero-fill: {total_files}")
            
            return final_df
            
        except Exception as e:
            print(f"[WARNING] Zero-fill failed, returning original data: {e}")
            return df
    
    def _add_sheet(self, df, sheet_name, position=None, include_total=True):
        """
        Creates a new sheet from a DataFrame, handling formatting, positioning, and data types.
        
        Args:
            df (pandas.DataFrame): Data to add to the sheet
            sheet_name (str): Name of the sheet
            position (int, optional): Position to insert the sheet
            include_total (bool): Whether to include a total row
        """
        if df.empty:
            print(f"[WARNING] Skipping empty sheet: {sheet_name}")
            return
        
        # Create worksheet
        if position is not None:
            ws = self.workbook.create_sheet(title=sheet_name, index=position)
        else:
            ws = self.workbook.create_sheet(title=sheet_name)
        
        # Add headers
        for col_idx, column_name in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=str(column_name))
        
        # Add data
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for col_idx, value in enumerate(row, start=1):
                # Handle different data types appropriately
                if pd.isna(value):
                    cell_value = ""
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Apply formatting
        self.formatter.format_sheet(ws)
        
        # Add total row if requested
        if include_total:
            if sheet_name == "Audio Note Characteristics":
                self.formatter.add_audio_characteristics_total_row(ws, df)
            else:
                self.formatter.add_total_row(ws, df)
        
        print(f"[SUCCESS] Created sheet: {sheet_name} ({len(df)} rows)")
    
    def generate_report(self):
        """
        Generates the full, multi-sheet Excel report.
        
        This is the main orchestration method that coordinates all aspects of
        report generation including data processing, sheet creation, formatting,
        and chart generation.
        """
        print("\n" + "🚨"*80)
        print("🚨 CRITICAL: THIS IS CORE.PY GENERATE_REPORT METHOD")
        print("🚨 IF YOU SEE THIS, THE EXECUTION IS IN CORE.PY")
        print("🚨"*80)
        print("=" * 60)
        print("AR DATA ANALYSIS - REPORT GENERATION")
        print("=" * 60)
        
        # Generate timestamp for output file
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        output_filename = f"AR_Analysis_Report_{timestamp}.xlsx"
        output_path = os.path.join(self.output_dir, output_filename)
        
        print(f"[INFO] Starting report generation...")
        print(f"[INFO] Output file: {output_path}")
        
        try:
            # Create comprehensive dashboard first
            print("\n" + "📊"*50)
            print("📊 EXECUTION TRACE: About to create comprehensive dashboard")
            print("📊"*50)
            print("[INFO] Creating comprehensive dashboard...")
            try:
                self.dashboard_creator.create_comprehensive_dashboard(self.workbook)
                print("📊 EXECUTION TRACE: Dashboard creation completed successfully")
            except Exception as dashboard_error:
                print(f"🚨 EXECUTION TRACE: Dashboard creation failed: {dashboard_error}")
                import traceback
                traceback.print_exc()
                # Continue execution even if dashboard fails
                pass
            
            # Create all non-pipeline sheets using unified sheet creator
            print("\n" + "🔥"*50)
            print("🚀 CRITICAL: STARTING BASIC SHEET CREATION SECTION")
            print("🔥"*50)
            print("[INFO] Creating non-pipeline sheets using unified architecture...")
            from .sheet_creators import SheetCreator
            unified_sheet_creator = SheetCreator(self.db, self.formatter)
            
            # Create summary statistics sheet
            print("[INFO] Creating summary statistics sheet...")
            unified_sheet_creator.create_summary_statistics_sheet(self.workbook)
            
            # Create data cleaning sheet
            print("[INFO] Creating data cleaning sheet...")
            unified_sheet_creator.create_data_cleaning_sheet(self.workbook)
            
            # Create MP3 Duration Analysis sheet
            # NOTE: MP3 Duration Analysis sheet creation is now handled by the configuration-based system
            # in sheet_creators/pipeline.py. This old hardcoded call has been disabled to prevent
            # duplicate creation and positioning conflicts.
            # 
            # print("\n" + "="*80)
            # print("🎯 CRITICAL DEBUG: ABOUT TO CREATE MP3 DURATION ANALYSIS SHEET")
            # print("="*80)
            # print("[INFO] Creating MP3 Duration Analysis sheet...")
            # try:
            #     print("[DEBUG] About to call create_mp3_duration_analysis_sheet method")
            #     unified_sheet_creator.create_mp3_duration_analysis_sheet(self.workbook)
            #     print("[DEBUG] MP3 Duration Analysis sheet creation completed successfully")
            # except Exception as e:
            #     print(f"[ERROR] Failed to create MP3 Duration Analysis sheet: {e}")
            #     import traceback
            #     traceback.print_exc()
            print("[INFO] MP3 Duration sheet creation delegated to configuration-based system")
            
            # Process all configured pipelines from report_config.json using unified architecture
            print("[INFO] Processing pipeline configurations...")
            print("[DEBUG] About to call unified_sheet_creator.process_pipeline_configurations")
            print(f"[DEBUG] unified_sheet_creator type: {type(unified_sheet_creator)}")
            # Each sheet will fetch its own fresh pipeline data to prevent column duplication
            unified_sheet_creator.process_pipeline_configurations(self.workbook)
            print("[DEBUG] Completed unified_sheet_creator.process_pipeline_configurations")
            
            # Create Raw Data sheet (critical for peer reviewers)
            with open("MARKER_1_RAW_DATA.txt", "w") as f: f.write("Reached Raw Data section")
            print("[INFO] Creating Raw Data sheet...")
            try:
                self.raw_data_creator.create_raw_data_sheet(self.workbook)
                print("[SUCCESS] Raw Data sheet created successfully")
            except Exception as e:
                print(f"[ERROR] Could not create Raw Data sheet: {e}")
            
            # Create ACF/PACF Dashboard
            with open("MARKER_4_DASHBOARD.txt", "w") as f: f.write("Reached Dashboard section")
            print("[INFO] Creating ACF/PACF Dashboard...")
            try:
                if create_dashboard_summary:
                    create_dashboard_summary(self.workbook)
                    print("[SUCCESS] ACF/PACF Dashboard created successfully")
                else:
                    print("[WARNING] Dashboard generator module not available")
            except Exception as e:
                print(f"[WARNING] Could not create ACF/PACF Dashboard: {e}")
            
            print("[INFO] ACF/PACF charts integrated during sheet creation")
        
            # Add dashboard-level charts if needed
            try:
                if enhance_acf_pacf_visualization:
                    # Only enhance dashboard and summary sheets that weren't created by PipelineSheetCreator
                    dashboard_sheets = [name for name in self.workbook.sheetnames if 'Dashboard' in name or 'Summary' in name]
                    if dashboard_sheets:
                        enhance_acf_pacf_visualization(self.workbook)
                        print("[SUCCESS] Dashboard charts enhanced successfully")
                else:
                    print("[INFO] Chart enhancement module not available for dashboard")
            except Exception as e:
                print(f"[WARNING] Could not enhance dashboard charts: {e}")
                
            # Add ARIMA forecast charts to sheets with forecast data
            try:
                if enhance_arima_forecast_visualization:
                    # This will add ARIMA charts to all sheets with forecast columns
                    enhanced_sheets = enhance_arima_forecast_visualization(self.workbook)
                    if enhanced_sheets:
                        print(f"[SUCCESS] ARIMA forecast charts added to {len(enhanced_sheets)} sheets: {', '.join(enhanced_sheets)}")
                    else:
                        print("[INFO] No sheets found with ARIMA forecast data")
                else:
                    print("[INFO] ARIMA chart enhancement module not available")
            except Exception as e:
                print(f"[WARNING] Could not add ARIMA forecast charts: {e}")
                import traceback
                traceback.print_exc()
            
            # Post-process: Ensure MP3 Duration sheet is positioned correctly
            self._reorder_mp3_duration_sheet()
            
            # Save the final workbook
            with open("MARKER_5_SAVE.txt", "w") as f: f.write("Reached Save section")
            print("[SAVE] Saving final workbook...")
            self.workbook.save(output_path)
            
            print(f"\n--- Report Generation Complete ---")
            print(f"Successfully saved Excel report to: {output_path}")
            
        except Exception as e:
            print(f"[ERROR] Report generation failed: {e}")
            import traceback
            traceback.print_exc()
            raise
    
    def _reorder_mp3_duration_sheet(self):
        """
        Post-processing method to ensure MP3 Duration sheet is positioned correctly.
        Moves the MP3 Duration sheet to position 3 (immediately after Data Cleaning).
        """
        try:
            # Check if MP3 Duration sheet exists
            if "MP3 Duration" not in self.workbook.sheetnames:
                print("[INFO] MP3 Duration sheet not found, skipping reordering")
                return
            
            # Find current position of MP3 Duration sheet
            current_sheets = self.workbook.sheetnames
            mp3_current_index = current_sheets.index("MP3 Duration")
            
            # Target position: after Data Cleaning (position 3, 0-indexed = 2)
            target_index = 2
            if "Data Cleaning" in current_sheets:
                target_index = current_sheets.index("Data Cleaning") + 1
            
            print(f"[INFO] Moving MP3 Duration sheet from position {mp3_current_index + 1} to position {target_index + 1}")
            
            # Move the sheet to the correct position
            mp3_sheet = self.workbook["MP3 Duration"]
            self.workbook.move_sheet(mp3_sheet, target_index - mp3_current_index)
            
            print(f"[SUCCESS] MP3 Duration sheet repositioned successfully")
            
        except Exception as e:
            print(f"[WARNING] Failed to reorder MP3 Duration sheet: {e}")
            # Don't raise the exception as this is post-processing
