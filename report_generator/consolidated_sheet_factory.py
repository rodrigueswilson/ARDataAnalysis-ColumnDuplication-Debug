"""
Consolidated Sheet Factory
=========================

This module provides a single-source-of-truth for all sheet creation in the AR Data Analysis system.
It eliminates duplications by centralizing sheet creation logic and enforcing the registry system.

Design Principles:
1. Single responsibility: Each sheet type has exactly one creation method
2. Registry integration: All sheet creation goes through the registry
3. Modular delegation: Delegates to specialized creators but controls the flow
4. Error handling: Comprehensive error handling and logging
5. Validation: Prevents duplicate sheet creation
"""

import json
from pathlib import Path
from typing import Optional, Dict, Any
import openpyxl

from .sheet_registry import get_sheet_registry, SheetType
from .sheet_creators.base import BaseSheetCreator
from .sheet_creators.pipeline import PipelineSheetCreator  
from .sheet_creators.specialized import SpecializedSheetCreator
from .dashboard import DashboardCreator
from .raw_data import RawDataCreator

# External modules (conditional imports)
try:
    from dashboard_generator import create_dashboard_summary
except ImportError:
    create_dashboard_summary = None

try:
    from acf_pacf_charts import enhance_acf_pacf_visualization, enhance_arima_forecast_visualization
except ImportError:
    enhance_acf_pacf_visualization = None
    enhance_arima_forecast_visualization = None


class ConsolidatedSheetFactory:
    """
    Single-source-of-truth factory for all sheet creation.
    
    This factory eliminates duplications by:
    1. Providing exactly one method per sheet type
    2. Using the registry to prevent duplicates
    3. Delegating to appropriate specialized creators
    4. Maintaining creation order and dependencies
    """
    
    def __init__(self, db, formatter):
        """
        Initialize the consolidated factory with all required dependencies.
        
        Args:
            db: MongoDB database connection
            formatter: ExcelFormatter instance for styling
        """
        self.db = db
        self.formatter = formatter
        self.registry = get_sheet_registry()
        
        # Initialize all specialized creators
        self.base_creator = BaseSheetCreator(db, formatter)
        self.pipeline_creator = PipelineSheetCreator(db, formatter)
        self.specialized_creator = SpecializedSheetCreator(db, formatter)
        self.dashboard_creator = DashboardCreator(db, formatter)
        self.raw_data_creator = RawDataCreator(db, formatter)
        
        print("[FACTORY] Consolidated Sheet Factory initialized")
    
    def create_all_sheets(self, workbook: openpyxl.Workbook) -> Dict[str, bool]:
        """
        Create all sheets in the correct order with dependency management.
        
        Args:
            workbook: openpyxl workbook object
            
        Returns:
            dict: Results of each sheet creation (sheet_name -> success)
        """
        print("\n" + "="*60)
        print("CONSOLIDATED SHEET FACTORY - CREATING ALL SHEETS")
        print("="*60)
        
        results = {}
        
        # Phase 1: Core sheets (no dependencies)
        print("\n[PHASE 1] Creating core sheets...")
        results.update({
            "Dashboard": self.create_dashboard_sheet(workbook),
            "Summary Statistics": self.create_summary_statistics_sheet(workbook),
            "Data Cleaning": self.create_data_cleaning_sheet(workbook),
            "Raw Data": self.create_raw_data_sheet(workbook)
        })
        
        # Phase 2: Specialized analysis sheets
        print("\n[PHASE 2] Creating specialized analysis sheets...")
        results.update({
            # NOTE: MP3 Duration Analysis sheet creation moved to configuration-based system (Phase 3)
            # "MP3 Duration Analysis": self.create_mp3_duration_analysis_sheet(workbook),
            "Audio Efficiency Details": self.create_audio_efficiency_details_sheet(workbook)
        })
        
        # Phase 3: Pipeline-driven sheets (configuration-based)
        print("\n[PHASE 3] Creating pipeline-driven sheets...")
        pipeline_results = self.create_pipeline_sheets(workbook)
        results.update(pipeline_results)
        
        # Phase 4: Dashboard and chart enhancements
        print("\n[PHASE 4] Creating dashboard enhancements...")
        results.update({
            "ACF_PACF_Dashboard": self.create_acf_pacf_dashboard_sheet(workbook)
        })
        
        # Phase 5: Chart integration (post-processing)
        print("\n[PHASE 5] Adding chart enhancements...")
        chart_results = self.add_chart_enhancements(workbook)
        results.update(chart_results)
        
        # Print summary
        successful = sum(1 for success in results.values() if success)
        total = len(results)
        print(f"\n[SUMMARY] Sheet creation completed: {successful}/{total} successful")
        
        # Print registry report
        self.registry.print_creation_report()
        
        return results
    
    # ========================================================================
    # SINGLE-SOURCE-OF-TRUTH SHEET CREATION METHODS
    # ========================================================================
    
    def create_dashboard_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the main Dashboard sheet.
        
        Eliminates duplication between DashboardCreator and external modules.
        """
        try:
            # Use registry to prevent duplicates
            if not self.registry.create_sheet(
                workbook, "Dashboard", SheetType.DASHBOARD,
                "consolidated_sheet_factory", "create_dashboard_sheet"
            ):
                return False
            
            # Delegate to DashboardCreator for actual content creation
            self.dashboard_creator.create_comprehensive_dashboard(workbook)
            print("[SUCCESS] Dashboard sheet created via DashboardCreator")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create Dashboard sheet: {e}")
            return False
    
    def create_summary_statistics_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the Summary Statistics sheet.
        """
        try:
            if not self.registry.create_sheet(
                workbook, "Summary Statistics", SheetType.SUMMARY_STATISTICS,
                "consolidated_sheet_factory", "create_summary_statistics_sheet"
            ):
                return False
            
            # Delegate to BaseSheetCreator
            self.base_creator.create_summary_statistics_sheet(workbook)
            print("[SUCCESS] Summary Statistics sheet created via BaseSheetCreator")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create Summary Statistics sheet: {e}")
            return False
    
    def create_data_cleaning_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the Data Cleaning sheet.
        """
        try:
            if not self.registry.create_sheet(
                workbook, "Data Cleaning", SheetType.DATA_CLEANING,
                "consolidated_sheet_factory", "create_data_cleaning_sheet"
            ):
                return False
            
            # Delegate to BaseSheetCreator
            self.base_creator.create_data_cleaning_sheet(workbook)
            print("[SUCCESS] Data Cleaning sheet created via BaseSheetCreator")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create Data Cleaning sheet: {e}")
            return False
    
    def create_raw_data_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the Raw Data sheet.
        
        Eliminates duplication between RawDataCreator and BaseSheetCreator.
        Uses RawDataCreator as it's more specialized and feature-complete.
        """
        try:
            if not self.registry.create_sheet(
                workbook, "Raw Data", SheetType.RAW_DATA,
                "consolidated_sheet_factory", "create_raw_data_sheet"
            ):
                return False
            
            # Delegate to RawDataCreator (more specialized than BaseSheetCreator version)
            success = self.raw_data_creator.create_raw_data_sheet(workbook)
            if success:
                print("[SUCCESS] Raw Data sheet created via RawDataCreator")
            return success
            
        except Exception as e:
            print(f"[ERROR] Failed to create Raw Data sheet: {e}")
            return False
    
    def create_mp3_duration_analysis_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the MP3 Duration Analysis sheet.
        """
        try:
            if not self.registry.create_sheet(
                workbook, "MP3 Duration Analysis", SheetType.MP3_DURATION_ANALYSIS,
                "consolidated_sheet_factory", "create_mp3_duration_analysis_sheet"
            ):
                return False
            
            # Delegate to SpecializedSheetCreator
            self.specialized_creator.create_mp3_duration_analysis_sheet(workbook)
            print("[SUCCESS] MP3 Duration Analysis sheet created via SpecializedSheetCreator")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create MP3 Duration Analysis sheet: {e}")
            return False
    
    def create_audio_efficiency_details_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the Audio Efficiency Details sheet.
        """
        try:
            if not self.registry.create_sheet(
                workbook, "Audio Efficiency Details", SheetType.AUDIO_EFFICIENCY_DETAILS,
                "consolidated_sheet_factory", "create_audio_efficiency_details_sheet"
            ):
                return False
            
            # Delegate to SpecializedSheetCreator
            self.specialized_creator.create_audio_efficiency_details_sheet(workbook)
            print("[SUCCESS] Audio Efficiency Details sheet created via SpecializedSheetCreator")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create Audio Efficiency Details sheet: {e}")
            return False
    
    def create_pipeline_sheets(self, workbook: openpyxl.Workbook) -> Dict[str, bool]:
        """
        SINGLE SOURCE: Create all pipeline-driven sheets from configuration.
        
        This replaces direct calls to process_pipeline_configurations.
        """
        results = {}
        
        try:
            # Load configuration
            config_path = Path(__file__).parent.parent / "report_config.json"
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            # Process each configured sheet
            for sheet_config in config.get('sheets', []):
                if not sheet_config.get('enabled', True):
                    print(f"[SKIP] Pipeline sheet '{sheet_config['name']}' is disabled")
                    continue
                
                sheet_name = sheet_config['name']
                
                try:
                    # Determine sheet type based on name
                    sheet_type = self._infer_sheet_type(sheet_name)
                    
                    # Use registry to prevent duplicates
                    if not self.registry.create_sheet(
                        workbook, sheet_name, sheet_type,
                        "consolidated_sheet_factory", "create_pipeline_sheets"
                    ):
                        results[sheet_name] = False
                        continue
                    
                    # Delegate to PipelineSheetCreator for actual content
                    self.pipeline_creator._create_pipeline_sheet(workbook, sheet_config)
                    print(f"[SUCCESS] Pipeline sheet '{sheet_name}' created")
                    results[sheet_name] = True
                    
                except Exception as e:
                    print(f"[ERROR] Failed to create pipeline sheet '{sheet_name}': {e}")
                    results[sheet_name] = False
            
            return results
            
        except Exception as e:
            print(f"[ERROR] Failed to process pipeline configurations: {e}")
            return {}
    
    def create_acf_pacf_dashboard_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        SINGLE SOURCE: Create the ACF/PACF Dashboard sheet.
        
        Uses external dashboard_generator if available.
        """
        try:
            if not create_dashboard_summary:
                print("[SKIP] ACF/PACF Dashboard - external module not available")
                return False
            
            if not self.registry.create_sheet(
                workbook, "ACF_PACF_Dashboard", SheetType.ACF_PACF_DASHBOARD,
                "consolidated_sheet_factory", "create_acf_pacf_dashboard_sheet"
            ):
                return False
            
            # Delegate to external module
            create_dashboard_summary(workbook)
            print("[SUCCESS] ACF/PACF Dashboard sheet created via external module")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create ACF/PACF Dashboard sheet: {e}")
            return False
    
    def add_chart_enhancements(self, workbook: openpyxl.Workbook) -> Dict[str, bool]:
        """
        Add chart enhancements to existing sheets (post-processing).
        
        This doesn't create new sheets but enhances existing ones with charts.
        """
        results = {}
        
        # ACF/PACF chart enhancements
        try:
            if enhance_acf_pacf_visualization:
                dashboard_sheets = [name for name in workbook.sheetnames 
                                  if 'Dashboard' in name or 'Summary' in name]
                if dashboard_sheets:
                    enhance_acf_pacf_visualization(workbook)
                    print("[SUCCESS] ACF/PACF chart enhancements added")
                    results["ACF_PACF_Charts"] = True
                else:
                    results["ACF_PACF_Charts"] = False
            else:
                print("[SKIP] ACF/PACF chart enhancement - module not available")
                results["ACF_PACF_Charts"] = False
        except Exception as e:
            print(f"[ERROR] Failed to add ACF/PACF chart enhancements: {e}")
            results["ACF_PACF_Charts"] = False
        
        # ARIMA forecast chart enhancements
        try:
            if enhance_arima_forecast_visualization:
                enhanced_sheets = enhance_arima_forecast_visualization(workbook)
                if enhanced_sheets:
                    print(f"[SUCCESS] ARIMA forecast charts added to {len(enhanced_sheets)} sheets")
                    results["ARIMA_Charts"] = True
                else:
                    print("[INFO] No sheets found with ARIMA forecast data")
                    results["ARIMA_Charts"] = False
            else:
                print("[SKIP] ARIMA chart enhancement - module not available")
                results["ARIMA_Charts"] = False
        except Exception as e:
            print(f"[ERROR] Failed to add ARIMA chart enhancements: {e}")
            results["ARIMA_Charts"] = False
        
        return results
    
    def _infer_sheet_type(self, sheet_name: str) -> Optional[SheetType]:
        """
        Infer the SheetType enum value from a sheet name.
        
        Args:
            sheet_name: Name of the sheet
            
        Returns:
            SheetType enum value or None if not recognized
        """
        # Direct matches
        for sheet_type in SheetType:
            if sheet_type.value == sheet_name:
                return sheet_type
        
        # Pattern matches for dynamic names
        if "(ACF_PACF)" in sheet_name:
            if "Daily" in sheet_name:
                return SheetType.DAILY_COUNTS_ACF_PACF
            elif "Weekly" in sheet_name:
                return SheetType.WEEKLY_COUNTS_ACF_PACF
            elif "Biweekly" in sheet_name:
                return SheetType.BIWEEKLY_COUNTS_ACF_PACF
            elif "Monthly" in sheet_name:
                return SheetType.MONTHLY_COUNTS_ACF_PACF
            elif "Period" in sheet_name:
                return SheetType.PERIOD_COUNTS_ACF_PACF
        else:
            if "Daily" in sheet_name:
                return SheetType.DAILY_COUNTS
            elif "Weekly" in sheet_name:
                return SheetType.WEEKLY_COUNTS
            elif "Biweekly" in sheet_name:
                return SheetType.BIWEEKLY_COUNTS
            elif "Monthly" in sheet_name:
                return SheetType.MONTHLY_COUNTS
            elif "Period" in sheet_name:
                return SheetType.PERIOD_COUNTS
        
        return None
    
    def validate_workbook(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Validate the final workbook against the registry.
        
        Returns:
            dict: Validation results
        """
        return self.registry.validate_workbook_consistency(workbook)
