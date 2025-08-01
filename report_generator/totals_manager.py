"""
Totals Management Module
========================

This module provides comprehensive totals calculation, formatting, and validation
for all Excel sheets in the AR Data Analysis system. It ensures consistency
across different tables and provides both row and column totals where applicable.
"""

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Tuple, Optional, Any, Union
import json
from pathlib import Path


class TotalsManager:
    """
    Manages totals calculation, formatting, and validation across all Excel sheets.
    Provides both row and column totals with consistency validation.
    """
    
    def __init__(self, formatter=None):
        """
        Initialize the totals manager.
        
        Args:
            formatter: ExcelFormatter instance for consistent styling
        """
        self.formatter = formatter
        self.totals_registry = {}  # Track totals across sheets for validation
        self.validation_rules = self._load_validation_rules()
        
        # Totals styling configuration
        self.totals_style = {
            'font': Font(bold=True, color='FFFFFF', size=11),
            'fill': PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
        }
        
        # Grand totals styling (darker)
        self.grand_totals_style = {
            'font': Font(bold=True, color='FFFFFF', size=12),
            'fill': PatternFill(start_color='2F5F8F', end_color='2F5F8F', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thick', color='000000'),
                right=Side(style='thick', color='000000'),
                top=Side(style='thick', color='000000'),
                bottom=Side(style='thick', color='000000')
            )
        }
    
    def _load_validation_rules(self) -> Dict:
        """Load validation rules for cross-table consistency checking."""
        try:
            config_path = Path(__file__).parent.parent / "totals_validation_rules.json"
            if config_path.exists():
                with open(config_path, 'r') as f:
                    return json.load(f)
        except Exception as e:
            print(f"[WARNING] Could not load totals validation rules: {e}")
        
        # Default validation rules
        return {
            "cross_sheet_validations": [
                {
                    "name": "Total Files Consistency",
                    "sheets": ["Summary Statistics", "Daily Counts", "Weekly Counts"],
                    "field": "Total_Files",
                    "tolerance": 0
                },
                {
                    "name": "MP3 Files Consistency", 
                    "sheets": ["Summary Statistics", "MP3 Duration Analysis"],
                    "field": "MP3_Files",
                    "tolerance": 0
                }
            ],
            "required_totals": {
                "Summary Statistics": ["row_totals", "column_totals"],
                "MP3 Duration Analysis": ["row_totals", "section_totals"],
                "Audio Efficiency Details": ["row_totals"]
            }
        }
    
    def calculate_row_totals(self, df: pd.DataFrame, 
                           numeric_columns: Optional[List[str]] = None,
                           exclude_columns: Optional[List[str]] = None) -> pd.Series:
        """
        Calculate row totals for numeric columns in a DataFrame.
        
        Args:
            df: Input DataFrame
            numeric_columns: Specific columns to total (if None, auto-detect)
            exclude_columns: Columns to exclude from totals
            
        Returns:
            pandas.Series with row totals
        """
        if df.empty:
            return pd.Series(dtype=float)
        
        # Auto-detect numeric columns if not specified
        if numeric_columns is None:
            numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # Remove excluded columns
        if exclude_columns:
            numeric_columns = [col for col in numeric_columns if col not in exclude_columns]
        
        # Calculate row totals
        if numeric_columns:
            return df[numeric_columns].sum(axis=1)
        else:
            return pd.Series([0] * len(df), dtype=float)
    
    def calculate_column_totals(self, df: pd.DataFrame,
                              numeric_columns: Optional[List[str]] = None,
                              exclude_columns: Optional[List[str]] = None) -> pd.Series:
        """
        Calculate column totals for numeric columns in a DataFrame.
        
        Args:
            df: Input DataFrame
            numeric_columns: Specific columns to total (if None, auto-detect)
            exclude_columns: Columns to exclude from totals
            
        Returns:
            pandas.Series with column totals
        """
        if df.empty:
            return pd.Series(dtype=float)
        
        # Auto-detect numeric columns if not specified
        if numeric_columns is None:
            numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # Remove excluded columns
        if exclude_columns:
            numeric_columns = [col for col in numeric_columns if col not in exclude_columns]
        
        # Calculate column totals
        totals = pd.Series(index=df.columns, dtype=object)
        for col in df.columns:
            if col in numeric_columns:
                totals[col] = df[col].sum()
            else:
                totals[col] = "TOTAL" if col == df.columns[0] else ""
        
        return totals
    
    def add_row_totals_to_worksheet(self, ws, df: pd.DataFrame, 
                                  start_row: int = 2, start_col: int = 1,
                                  numeric_columns: Optional[List[str]] = None,
                                  exclude_columns: Optional[List[str]] = None,
                                  totals_column_header: str = "Row Total") -> int:
        """
        Add row totals column to an Excel worksheet.
        
        Args:
            ws: openpyxl worksheet
            df: Source DataFrame
            start_row: Starting row for data (default: 2, assuming row 1 is headers)
            start_col: Starting column for data
            numeric_columns: Columns to include in totals
            exclude_columns: Columns to exclude from totals
            totals_column_header: Header for the totals column
            
        Returns:
            Column number where totals were added
        """
        if df.empty:
            return start_col
        
        # Calculate row totals
        row_totals = self.calculate_row_totals(df, numeric_columns, exclude_columns)
        
        # Find the next available column
        totals_col = ws.max_column + 1
        
        # Add header
        header_cell = ws.cell(row=start_row - 1, column=totals_col, value=totals_column_header)
        if self.formatter:
            self.formatter.apply_header_style(ws, header_cell.coordinate)
        else:
            header_cell.font = self.totals_style['font']
            header_cell.fill = self.totals_style['fill']
        
        # Add row totals
        for i, total in enumerate(row_totals):
            cell = ws.cell(row=start_row + i, column=totals_col, value=total)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='right')
        
        return totals_col
    
    def add_column_totals_to_worksheet(self, ws, df: pd.DataFrame,
                                     start_row: int = 2, start_col: int = 1,
                                     numeric_columns: Optional[List[str]] = None,
                                     exclude_columns: Optional[List[str]] = None,
                                     totals_row_label: str = "TOTAL") -> int:
        """
        Add column totals row to an Excel worksheet.
        
        Args:
            ws: openpyxl worksheet
            df: Source DataFrame
            start_row: Starting row for data
            start_col: Starting column for data
            numeric_columns: Columns to include in totals
            exclude_columns: Columns to exclude from totals
            totals_row_label: Label for the totals row
            
        Returns:
            Row number where totals were added
        """
        if df.empty:
            return start_row
        
        # Calculate column totals
        column_totals = self.calculate_column_totals(df, numeric_columns, exclude_columns)
        
        # Find the next available row
        totals_row = ws.max_row + 1
        
        # Add totals row
        for col_idx, (col_name, total) in enumerate(column_totals.items()):
            cell = ws.cell(row=totals_row, column=start_col + col_idx)
            
            if col_idx == 0:  # First column gets the label
                cell.value = totals_row_label
            else:
                cell.value = total if pd.notna(total) and total != "" else ""
            
            # Apply totals styling
            cell.font = self.totals_style['font']
            cell.fill = self.totals_style['fill']
            cell.alignment = self.totals_style['alignment']
            cell.border = self.totals_style['border']
        
        return totals_row
    
    def add_comprehensive_totals(self, ws, df: pd.DataFrame,
                               start_row: int = 2, start_col: int = 1,
                               numeric_columns: Optional[List[str]] = None,
                               exclude_columns: Optional[List[str]] = None,
                               include_row_totals: bool = True,
                               include_column_totals: bool = True,
                               include_grand_total: bool = True) -> Dict[str, int]:
        """
        Add comprehensive totals (both row and column) to a worksheet.
        
        Args:
            ws: openpyxl worksheet
            df: Source DataFrame
            start_row: Starting row for data
            start_col: Starting column for data
            numeric_columns: Columns to include in totals
            exclude_columns: Columns to exclude from totals
            include_row_totals: Whether to add row totals column
            include_column_totals: Whether to add column totals row
            include_grand_total: Whether to add grand total cell
            
        Returns:
            Dictionary with positions of added totals
        """
        positions = {}
        
        if df.empty:
            return positions
        
        # Add row totals column
        if include_row_totals:
            totals_col = self.add_row_totals_to_worksheet(
                ws, df, start_row, start_col, numeric_columns, exclude_columns
            )
            positions['row_totals_column'] = totals_col
        
        # Add column totals row
        if include_column_totals:
            totals_row = self.add_column_totals_to_worksheet(
                ws, df, start_row, start_col, numeric_columns, exclude_columns
            )
            positions['column_totals_row'] = totals_row
        
        # Add grand total cell (intersection of row and column totals)
        if include_grand_total and include_row_totals and include_column_totals:
            grand_total_row = positions['column_totals_row']
            grand_total_col = positions['row_totals_column']
            
            # Calculate grand total
            row_totals = self.calculate_row_totals(df, numeric_columns, exclude_columns)
            grand_total = row_totals.sum()
            
            # Add grand total cell
            grand_cell = ws.cell(row=grand_total_row, column=grand_total_col, value=grand_total)
            grand_cell.font = self.grand_totals_style['font']
            grand_cell.fill = self.grand_totals_style['fill']
            grand_cell.alignment = self.grand_totals_style['alignment']
            grand_cell.border = self.grand_totals_style['border']
            
            positions['grand_total'] = (grand_total_row, grand_total_col)
        
        return positions
    
    def add_totals_to_worksheet(self, worksheet, dataframe, start_row, start_col, config):
        """
        Add totals to an Excel worksheet based on DataFrame data and configuration.
        
        Args:
            worksheet: openpyxl worksheet object
            dataframe: pandas DataFrame containing the data
            start_row: Starting row for the data (1-indexed)
            start_col: Starting column for the data (1-indexed)
            config: Configuration dictionary with totals settings
        """
        try:
            if dataframe.empty:
                print("[WARNING] Empty DataFrame provided for totals calculation")
                return
            
            # Extract configuration
            add_row_totals = config.get('add_row_totals', False)
            add_column_totals = config.get('add_column_totals', False)
            add_grand_total = config.get('add_grand_total', False)
            totals_label = config.get('totals_label', 'TOTALS')
            row_totals_label = config.get('row_totals_label', 'Row Total')
            
            # Determine which columns to include in totals
            if 'include_columns' in config:
                numeric_columns = [col for col in config['include_columns'] if col in dataframe.columns]
            else:
                # Auto-detect numeric columns
                numeric_columns = dataframe.select_dtypes(include=['number']).columns.tolist()
                
                # Exclude specified columns
                if 'exclude_columns' in config:
                    numeric_columns = [col for col in numeric_columns if col not in config['exclude_columns']]
            
            if not numeric_columns:
                print("[INFO] No numeric columns found for totals calculation")
                return
            
            # Calculate totals
            if add_column_totals:
                # Add column totals row
                totals_row = start_row + len(dataframe)
                
                # Add totals label
                worksheet.cell(row=totals_row, column=start_col, value=totals_label)
                
                # Calculate and add column totals
                for i, col in enumerate(dataframe.columns):
                    col_idx = start_col + i
                    if col in numeric_columns:
                        total_value = dataframe[col].sum()
                        worksheet.cell(row=totals_row, column=col_idx, value=total_value)
                        
                        # Apply totals styling
                        cell = worksheet.cell(row=totals_row, column=col_idx)
                        self._apply_totals_style(cell, is_grand_total=False)
                    else:
                        # Empty cell for non-numeric columns
                        worksheet.cell(row=totals_row, column=col_idx, value="")
                
                # Apply styling to totals label
                label_cell = worksheet.cell(row=totals_row, column=start_col)
                self._apply_totals_style(label_cell, is_grand_total=False)
            
            if add_row_totals:
                # Add row totals column
                totals_col = start_col + len(dataframe.columns)
                
                # Add header for row totals column (only if there's space above data)
                if start_row > 1:
                    worksheet.cell(row=start_row - 1, column=totals_col, value=row_totals_label)
                
                # Calculate and add row totals
                for i, (_, row) in enumerate(dataframe.iterrows()):
                    row_idx = start_row + i
                    row_total = sum(row[col] for col in numeric_columns if pd.notna(row[col]))
                    
                    cell = worksheet.cell(row=row_idx, column=totals_col, value=row_total)
                    self._apply_totals_style(cell, is_grand_total=False)
            
            if add_grand_total and add_row_totals and add_column_totals:
                # Add grand total cell
                grand_total_row = start_row + len(dataframe)
                grand_total_col = start_col + len(dataframe.columns)
                
                grand_total_value = sum(dataframe[col].sum() for col in numeric_columns)
                cell = worksheet.cell(row=grand_total_row, column=grand_total_col, value=grand_total_value)
                self._apply_totals_style(cell, is_grand_total=True)
            
            print(f"[SUCCESS] Added totals to worksheet: {len(numeric_columns)} numeric columns processed")
            
        except Exception as e:
            print(f"[ERROR] Failed to add totals to worksheet: {e}")
            raise
    
    def _apply_totals_style(self, cell, is_grand_total=False):
        """Apply styling to a totals cell."""
        try:
            if is_grand_total:
                cell.font = self.grand_totals_style['font']
                cell.fill = self.grand_totals_style['fill']
                cell.alignment = self.grand_totals_style['alignment']
                cell.border = self.grand_totals_style['border']
            else:
                cell.font = self.totals_style['font']
                cell.fill = self.totals_style['fill']
                cell.alignment = self.totals_style['alignment']
                cell.border = self.totals_style['border']
        except Exception as e:
            print(f"[WARNING] Could not apply totals styling: {e}")

    def register_sheet_totals(self, sheet_name: str, totals_data: Dict[str, Any]):
        """
        Register totals data for a sheet for cross-validation.
        
        Args:
            sheet_name: Name of the sheet
            totals_data: Dictionary containing totals information
        """
        self.totals_registry[sheet_name] = {
            'timestamp': pd.Timestamp.now(),
            'data': totals_data
        }
    
    def register_totals(self, sheet_name: str, totals_data: Dict[str, Any]):
        """
        Register totals data for a sheet for cross-validation.
        Alias for register_sheet_totals for compatibility.
        
        Args:
            sheet_name: Name of the sheet
            totals_data: Dictionary containing totals information
        """
        return self.register_sheet_totals(sheet_name, totals_data)
    
    def validate_cross_sheet_consistency(self) -> Dict[str, List[str]]:
        """
        Validate consistency of totals across different sheets.
        
        Returns:
            Dictionary with validation results (errors and warnings)
        """
        validation_results = {
            'errors': [],
            'warnings': [],
            'info': []
        }
        
        # Check cross-sheet validations
        for validation in self.validation_rules.get('cross_sheet_validations', []):
            name = validation['name']
            sheets = validation['sheets']
            field = validation['field']
            tolerance = validation.get('tolerance', 0)
            
            # Get values from registered sheets
            values = {}
            for sheet in sheets:
                if sheet in self.totals_registry:
                    sheet_data = self.totals_registry[sheet]['data']
                    if field in sheet_data:
                        values[sheet] = sheet_data[field]
            
            # Check consistency
            if len(values) > 1:
                value_list = list(values.values())
                max_val = max(value_list)
                min_val = min(value_list)
                
                if abs(max_val - min_val) > tolerance:
                    validation_results['errors'].append(
                        f"{name}: Inconsistent values across sheets - {values}"
                    )
                else:
                    validation_results['info'].append(
                        f"{name}: Consistent across sheets - {values}"
                    )
        
        return validation_results
    
    def generate_totals_summary_report(self) -> str:
        """
        Generate a summary report of all totals and validation results.
        
        Returns:
            String containing the summary report
        """
        report = ["=" * 60]
        report.append("AR DATA ANALYSIS - TOTALS SUMMARY REPORT")
        report.append("=" * 60)
        report.append(f"Generated: {pd.Timestamp.now()}")
        report.append("")
        
        # Sheet totals summary
        report.append("SHEET TOTALS REGISTRY:")
        report.append("-" * 30)
        for sheet_name, sheet_info in self.totals_registry.items():
            report.append(f"Sheet: {sheet_name}")
            report.append(f"  Timestamp: {sheet_info['timestamp']}")
            for key, value in sheet_info['data'].items():
                report.append(f"  {key}: {value}")
            report.append("")
        
        # Validation results
        validation_results = self.validate_cross_sheet_consistency()
        
        report.append("VALIDATION RESULTS:")
        report.append("-" * 20)
        
        if validation_results['errors']:
            report.append("ERRORS:")
            for error in validation_results['errors']:
                report.append(f"  ❌ {error}")
            report.append("")
        
        if validation_results['warnings']:
            report.append("WARNINGS:")
            for warning in validation_results['warnings']:
                report.append(f"  ⚠️  {warning}")
            report.append("")
        
        if validation_results['info']:
            report.append("INFO:")
            for info in validation_results['info']:
                report.append(f"  ✅ {info}")
            report.append("")
        
        # Summary
        report.append("SUMMARY:")
        report.append(f"  Total sheets processed: {len(self.totals_registry)}")
        report.append(f"  Validation errors: {len(validation_results['errors'])}")
        report.append(f"  Validation warnings: {len(validation_results['warnings'])}")
        report.append("=" * 60)
        
        return "\n".join(report)
    
    def create_validation_rules_template(self):
        """Create a template validation rules file."""
        template = {
            "cross_sheet_validations": [
                {
                    "name": "Total Files Consistency",
                    "description": "Ensure total file counts match across summary and detail sheets",
                    "sheets": ["Summary Statistics", "Daily Counts", "Weekly Counts"],
                    "field": "Total_Files",
                    "tolerance": 0
                },
                {
                    "name": "MP3 Files Consistency",
                    "description": "Ensure MP3 file counts match between summary and MP3 analysis",
                    "sheets": ["Summary Statistics", "MP3 Duration Analysis"],
                    "field": "MP3_Files", 
                    "tolerance": 0
                },
                {
                    "name": "JPG Files Consistency",
                    "description": "Ensure JPG file counts match across sheets",
                    "sheets": ["Summary Statistics", "Daily Counts"],
                    "field": "JPG_Files",
                    "tolerance": 0
                }
            ],
            "required_totals": {
                "Summary Statistics": ["row_totals", "column_totals"],
                "MP3 Duration Analysis": ["row_totals", "section_totals"],
                "Audio Efficiency Details": ["row_totals"],
                "Daily Counts": ["column_totals"],
                "Weekly Counts": ["column_totals"],
                "Monthly Counts": ["column_totals"]
            },
            "totals_formatting": {
                "row_totals_header": "Row Total",
                "column_totals_label": "TOTAL",
                "grand_total_label": "GRAND TOTAL",
                "decimal_places": 2
            }
        }
        
        config_path = Path(__file__).parent.parent / "totals_validation_rules.json"
        try:
            with open(config_path, 'w') as f:
                json.dump(template, f, indent=2)
            print(f"[SUCCESS] Created totals validation rules template: {config_path}")
        except Exception as e:
            print(f"[ERROR] Could not create validation rules template: {e}")


# Convenience functions for easy integration
def add_table_totals(ws, df: pd.DataFrame, totals_manager: TotalsManager = None,
                    start_row: int = 2, **kwargs) -> Dict[str, int]:
    """
    Convenience function to add comprehensive totals to a table.
    
    Args:
        ws: openpyxl worksheet
        df: Source DataFrame
        totals_manager: TotalsManager instance (creates new if None)
        start_row: Starting row for data
        **kwargs: Additional arguments for add_comprehensive_totals
        
    Returns:
        Dictionary with positions of added totals
    """
    if totals_manager is None:
        totals_manager = TotalsManager()
    
    return totals_manager.add_comprehensive_totals(ws, df, start_row=start_row, **kwargs)


def validate_report_totals(totals_manager: TotalsManager) -> str:
    """
    Convenience function to validate all totals in a report.
    
    Args:
        totals_manager: TotalsManager instance with registered sheet data
        
    Returns:
        Validation summary report
    """
    return totals_manager.generate_totals_summary_report()
