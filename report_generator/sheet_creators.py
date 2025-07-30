"""
Sheet Creation Module
====================

This module contains all the individual sheet creation methods extracted from
the original ReportGenerator class. Each method is responsible for creating
a specific type of Excel sheet with its own data processing and formatting.
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime

# Local imports
from ar_utils import add_acf_pacf_analysis, infer_sheet_type, reorder_with_acf_pacf, get_school_calendar, get_non_collection_days, precompute_collection_days
from pipelines import PIPELINES
from .formatters import ExcelFormatter as SheetFormatter
from openpyxl.styles import NamedStyle
# Import db_utils conditionally to avoid import errors
try:
    from db_utils import get_db_connection
except ImportError:
    get_db_connection = None
import datetime

class SheetCreator:
    """
    Handles the creation of individual Excel sheets with specialized data processing
    and formatting for different types of analysis reports.
    """
    
    def __init__(self, db, formatter):
        """
        Initialize the sheet creator.
        
        Args:
            db: MongoDB database connection
            formatter: ExcelFormatter instance for styling
        """
        self.db = db
        self.formatter = formatter
    
    def _fill_missing_collection_days(self, df, pipeline_name):
        """
        Fills in missing collection days with zero counts for complete time series.
        This is critical for ACF/PACF/ARIMA analysis which requires continuous time series.
        
        Args:
            df (pandas.DataFrame): DataFrame from daily aggregation
            pipeline_name (str): Name of the pipeline to determine if zero-fill is needed
            
        Returns:
            pandas.DataFrame: DataFrame with missing collection days filled with zeros
        """
        if not ('DAILY' in pipeline_name.upper() and 
                ('WITH_ZEROES' in pipeline_name.upper() or 'COLLECTION_ONLY' in pipeline_name.upper())):
            return df
        
        try:
            school_calendar = get_school_calendar()
            non_collection_days = get_non_collection_days()
            collection_day_map = precompute_collection_days(school_calendar, non_collection_days)
            
            all_collection_days = []
            for date_obj, info in collection_day_map.items():
                all_collection_days.append({'_id': date_obj.strftime('%Y-%m-%d')})
            
            all_days_df = pd.DataFrame(all_collection_days)
            
            # Merge with existing data, keeping actual data and filling missing with zeros
            merged_df = pd.merge(all_days_df, df, on='_id', how='left').fillna(0)
            
            # Ensure correct data types after merge
            for col in ['Total_Files', 'MP3_Files', 'JPG_Files']:
                if col in merged_df.columns:
                    merged_df[col] = merged_df[col].astype(int)
            if 'Total_Size_MB' in merged_df.columns:
                 merged_df['Total_Size_MB'] = merged_df['Total_Size_MB'].astype(float)

            return merged_df.sort_values('_id').reset_index(drop=True)
            
        except Exception as e:
            print(f"[WARNING] Zero-fill failed, returning original data: {e}")
            return df
    
    def _run_aggregation(self, pipeline, use_base_filter=True, collection_name='media_records'):
        """
        Runs a MongoDB aggregation pipeline and returns a DataFrame.
        """
        try:
            collection = self.db[collection_name]
            full_pipeline = pipeline
            if use_base_filter:
                base_filter = {"$match": {"file_type": {"$in": ["JPG", "MP3"]}}}
                full_pipeline = [base_filter] + pipeline
            
            cursor = collection.aggregate(full_pipeline, allowDiskUse=True)
            df = pd.DataFrame(list(cursor))
            
            if '_id' in df.columns and df.shape[0] > 0 and isinstance(df['_id'].iloc[0], dict):
                df = pd.concat([df.drop('_id', axis=1), df['_id'].apply(pd.Series)], axis=1)

            return df
            
        except Exception as e:
            print(f"[ERROR] Aggregation failed for {collection_name}: {e}")
            return pd.DataFrame()

    def create_summary_statistics_sheet(self, workbook):
        """
        Creates the detailed Summary Statistics sheet.
        """
        print("[SHEET] Creating detailed Summary Statistics sheet...")
        try:
            from ar_utils import calculate_collection_days_for_period
            
            df_daily = self._run_aggregation(PIPELINES['DAILY_COUNTS_ALL'])
            if df_daily.empty: return

            df_daily = self._fill_missing_collection_days(df_daily, 'DAILY_COUNTS_ALL_WITH_ZEROES')
            df_daily['date_obj'] = pd.to_datetime(df_daily['_id'])
            df_daily['school_year'] = df_daily['date_obj'].apply(lambda d: f"{d.year}-{d.year + 1}" if d.month >= 8 else f"{d.year - 1}-{d.year}")
            school_years = sorted(df_daily['school_year'].unique())

            def calc_stats_for_period(df, school_year=None):
                file_col = 'Total_Files'
                total_days = len(df)  # This is days with data, not collection days
                
                # Calculate proper collection days using the same logic as Period Counts sheet
                if school_year:
                    # Calculate collection days for specific school year
                    if school_year == '2021-2022':
                        collection_days = (calculate_collection_days_for_period('SY 21-22 P1') + 
                                         calculate_collection_days_for_period('SY 21-22 P2') + 
                                         calculate_collection_days_for_period('SY 21-22 P3'))
                    elif school_year == '2022-2023':
                        collection_days = (calculate_collection_days_for_period('SY 22-23 P1') + 
                                         calculate_collection_days_for_period('SY 22-23 P2') + 
                                         calculate_collection_days_for_period('SY 22-23 P3'))
                    else:
                        collection_days = total_days  # Fallback for unknown years
                else:
                    # Total collection days across all periods
                    collection_days = (calculate_collection_days_for_period('SY 21-22 P1') + 
                                     calculate_collection_days_for_period('SY 21-22 P2') + 
                                     calculate_collection_days_for_period('SY 21-22 P3') +
                                     calculate_collection_days_for_period('SY 22-23 P1') + 
                                     calculate_collection_days_for_period('SY 22-23 P2') + 
                                     calculate_collection_days_for_period('SY 22-23 P3'))
                
                # Filter data to exclude non-collection days for consistent statistical calculations
                from ar_utils import _load_config
                config = _load_config()
                
                # Extract non-collection days from config (they're at the top level)
                non_collection_days = set()
                if 'non_collection_days' in config:
                    # Convert date objects to strings for comparison
                    for day in config['non_collection_days'].keys():
                        if hasattr(day, 'strftime'):  # It's a date object
                            non_collection_days.add(day.strftime('%Y-%m-%d'))
                        else:  # It's already a string
                            non_collection_days.add(str(day))
                
                # Convert dates to string format for comparison
                df['date_str'] = df['_id'].astype(str)
                
                # Filter out non-collection days for statistical calculations
                df_filtered = df[~df['date_str'].isin(non_collection_days)].copy()
                
                # Calculate statistics from filtered data (collection days only)
                if len(df_filtered) > 0:
                    filtered_files = df_filtered[file_col]
                    zero_file_days = len(df_filtered[df_filtered[file_col] == 0])
                    mean_files = filtered_files.mean()
                    median_files = filtered_files.median()
                    std_files = filtered_files.std()
                    min_files = filtered_files.min()
                    max_files = filtered_files.max()
                    burst_threshold = filtered_files.quantile(0.95)
                    burst_days = len(df_filtered[df_filtered[file_col] > burst_threshold])
                    
                    # Calculate frequency buckets from filtered data
                    bucket_0 = len(df_filtered[df_filtered[file_col] == 0])
                    bucket_1_5 = len(df_filtered[(df_filtered[file_col] >= 1) & (df_filtered[file_col] <= 5)])
                    bucket_6_10 = len(df_filtered[(df_filtered[file_col] >= 6) & (df_filtered[file_col] <= 10)])
                    bucket_11_20 = len(df_filtered[(df_filtered[file_col] >= 11) & (df_filtered[file_col] <= 20)])
                    bucket_20_plus = len(df_filtered[df_filtered[file_col] > 20])
                    
                    # Calculate percentages based on filtered data length
                    filtered_days = len(df_filtered)
                    zero_file_pct = round((zero_file_days/filtered_days*100), 1) if filtered_days > 0 else 0
                    burst_pct = round((burst_days/filtered_days*100), 1) if filtered_days > 0 else 0
                else:
                    # Fallback if no filtered data (shouldn't happen)
                    zero_file_days = len(df[df[file_col] == 0])
                    mean_files, median_files, std_files, min_files, max_files = df[file_col].mean(), df[file_col].median(), df[file_col].std(), df[file_col].min(), df[file_col].max()
                    burst_threshold = df[file_col].quantile(0.95)
                    burst_days = len(df[df[file_col] > burst_threshold])
                    bucket_0 = len(df[df[file_col] == 0])
                    bucket_1_5 = len(df[(df[file_col] >= 1) & (df[file_col] <= 5)])
                    bucket_6_10 = len(df[(df[file_col] >= 6) & (df[file_col] <= 10)])
                    bucket_11_20 = len(df[(df[file_col] >= 11) & (df[file_col] <= 20)])
                    bucket_20_plus = len(df[df[file_col] > 20])
                    zero_file_pct = round((zero_file_days/total_days*100), 1) if total_days > 0 else 0
                    burst_pct = round((burst_days/total_days*100), 1) if total_days > 0 else 0
                return {
                    'collection_days': collection_days, 'total_days': total_days, 'mean_files': round(mean_files, 1), 'median_files': round(median_files, 1),
                    'std_files': round(std_files, 1), 'min_files': min_files, 'max_files': max_files, 'range_files': max_files - min_files,
                    'zero_file_days': zero_file_days, 'zero_file_pct': zero_file_pct,
                    'burst_threshold': round(burst_threshold, 1), 'burst_days': burst_days, 'burst_pct': burst_pct,
                    'bucket_0': bucket_0, 'bucket_1_5': bucket_1_5,
                    'bucket_6_10': bucket_6_10, 'bucket_11_20': bucket_11_20,
                    'bucket_20_plus': bucket_20_plus
                }

            year_stats = {year: calc_stats_for_period(df_daily[df_daily['school_year'] == year], school_year=year) for year in school_years}
            total_stats = calc_stats_for_period(df_daily)

            columns = ['Metric'] + school_years + ['Total']
            summary_data = {'Metric': [
                '--- DISTRIBUTION METRICS ---', 'Total Collection Days', 'Mean Files per Day', 'Median Files per Day', 'Standard Deviation', 'Min Files per Day', 'Max Files per Day', 'Range (Max - Min)', '',
                '--- ZERO-FILE DAY ANALYSIS ---', 'Days with Zero Files', 'Zero-File Days (%)', '',
                '--- BURST DAY ANALYSIS ---', 'Burst Threshold (95th percentile)', 'Number of Burst Days', 'Burst Days (%)', '',
                '--- FILE COUNT FREQUENCY ---', '0 Files', '1-5 Files', '6-10 Files', '11-20 Files', '>20 Files (Burst Days)', '',
                '--- SAMPLE COMPOSITION ---', 'JPG Files (Original)', 'JPG Files (Outliers)', 'JPG Files (After Cleaning)', 'JPG Files (% Retained)',
                'MP3 Files (Original)', 'MP3 Files (Outliers)', 'MP3 Files (After Cleaning)', 'MP3 Files (% Retained)',
                'Total Files (Original)', 'Total Files (Outliers)', 'Total Files (After Cleaning)', 'Total Files (% Retained)']}

            for year in school_years:
                stats = year_stats[year]
                comp = {'2021-2022': (3047, 0, 3047, 100.0, 1720, 6, 1714, 99.7), '2022-2023': (3890, 173, 3717, 95.6, 1432, 15, 1417, 98.9)}.get(year, (0,)*8)
                year_total_orig, year_total_outliers, year_total_clean = comp[0] + comp[4], comp[1] + comp[5], comp[2] + comp[6]
                year_total_pct = round((year_total_clean / year_total_orig * 100), 1) if year_total_orig > 0 else 0.0
                summary_data[year] = ['', stats['collection_days'], stats['mean_files'], stats['median_files'], stats['std_files'], stats['min_files'], stats['max_files'], stats['range_files'], '', '', stats['zero_file_days'], stats['zero_file_pct'], '', '', stats['burst_threshold'], stats['burst_days'], stats['burst_pct'], '', '', stats['bucket_0'], stats['bucket_1_5'], stats['bucket_6_10'], stats['bucket_11_20'], stats['bucket_20_plus'], '', '',] + list(comp) + [year_total_orig, year_total_outliers, year_total_clean, year_total_pct]

            total_jpg_orig, total_jpg_outliers, total_jpg_clean = 6937, 173, 6764
            total_mp3_orig, total_mp3_outliers, total_mp3_clean = 3152, 21, 3131
            grand_total_orig, grand_total_outliers, grand_total_clean = 10089, 194, 9895
            summary_data['Total'] = ['', total_stats['collection_days'], total_stats['mean_files'], total_stats['median_files'], total_stats['std_files'], total_stats['min_files'], total_stats['max_files'], total_stats['range_files'], '', '', total_stats['zero_file_days'], total_stats['zero_file_pct'], '', '', total_stats['burst_threshold'], total_stats['burst_days'], total_stats['burst_pct'], '', '', total_stats['bucket_0'], total_stats['bucket_1_5'], total_stats['bucket_6_10'], total_stats['bucket_11_20'], total_stats['bucket_20_plus'], '', '', total_jpg_orig, total_jpg_outliers, total_jpg_clean, round((total_jpg_clean/total_jpg_orig*100),1), total_mp3_orig, total_mp3_outliers, total_mp3_clean, round((total_mp3_clean/total_mp3_orig*100),1), grand_total_orig, grand_total_outliers, grand_total_clean, round((grand_total_clean/grand_total_orig*100),1)]

            ws = workbook.create_sheet(title="Summary Statistics", index=0)
            ws.append(columns)
            for i in range(len(summary_data['Metric'])):
                ws.append([summary_data[col][i] for col in columns])
            self.formatter.format_sheet(ws)
            print("  [SUCCESS] Summary Statistics sheet created.")
        except Exception as e:
            print(f"[ERROR] Could not create Summary Statistics sheet: {e}")
            import traceback; traceback.print_exc()

    def create_raw_data_sheet(self, workbook):
        """
        Creates the Raw Data sheet.
        """
        print("[SHEET] Creating Raw Data sheet...")
        try:
            df_raw = self._run_aggregation([], use_base_filter=False)
            if df_raw.empty: return

            for col in ['ISO_Week', 'ISO_Month']:
                if col in df_raw.columns:
                    df_raw[col] = pd.to_numeric(df_raw[col], errors='coerce').astype('Int64')

            ws = workbook.create_sheet(title="Raw Data")
            ws.append(list(df_raw.columns))
            for _, row in df_raw.iterrows():
                ws.append(list(row))
            self.formatter.format_sheet(ws)
            print("  [SUCCESS] Raw Data sheet created.")
        except Exception as e:
            print(f"[ERROR] Could not create Raw Data sheet: {e}")

    def create_data_cleaning_sheet(self, workbook):
        """
        Creates the Data Cleaning sheet.
        """
        print("[SHEET] Creating Data Cleaning sheet...")
        try:
            ws = workbook.create_sheet(title="Data Cleaning")

            # 1. Title and Introduction
            ws.cell(row=1, column=1, value="Data Cleaning and Sample Composition")
            ws.cell(row=2, column=1, value="This sheet provides a detailed breakdown of the data cleaning process, showing how the final sample was derived.")

            # 2. Define and execute the data cleaning pipeline directly
            # (Using direct definition to bypass module import issues)
            pipeline = [
                {
                    "$match": {
                        "School_Year": {"$ne": "N/A"}
                    }
                },
                {
                    "$group": {
                        "_id": {
                            "school_year": "$School_Year",
                            "file_type": "$file_type"
                        },
                        "Original": {"$sum": 1},
                        "Outliers": {
                            "$sum": {
                                "$cond": ["$Outlier_Status", 1, 0]
                            }
                        }
                    }
                },
                {
                    "$sort": {
                        "_id.school_year": 1,
                        "_id.file_type": 1
                    }
                }
            ]
            cursor = self.db.media_records.aggregate(pipeline)
            results = list(cursor)
            print(f"[DEBUG] Pipeline returned {len(results)} results")

            if not results:
                ws.cell(row=4, column=1, value="No data available for sample composition breakdown.")
                self.formatter.format_sheet(ws)
                print("[WARNING] No data found for Data Cleaning sheet's sample composition table.")
                return

            data_for_df = []
            for item in results:
                category = f"{item['_id']['school_year']} {item['_id']['file_type']} Files"
                data_for_df.append({
                    "Category": category,
                    "Original": item["Original"],
                    "Outliers": item["Outliers"]
                })
            print(f"[DEBUG] Created {len(data_for_df)} data rows for DataFrame")

            df = pd.DataFrame(data_for_df)
            print(f"[DEBUG] DataFrame created with shape: {df.shape}")

            # Perform calculations (without setting index first)
            df['Difference'] = -df['Outliers']
            df['After Cleaning'] = df['Original'] - df['Outliers']
            df['% Retained'] = (df['After Cleaning'] / df['Original']) * 100
            print(f"[DEBUG] Calculations completed")

            # Calculate totals manually to avoid index issues
            total_original = df['Original'].sum()
            total_outliers = df['Outliers'].sum()
            total_difference = -total_outliers
            total_after_cleaning = total_original - total_outliers
            total_percent_retained = (total_after_cleaning / total_original) * 100 if total_original > 0 else 0
            
            # Create total row as a new DataFrame row
            total_row = {
                'Category': 'Total',
                'Original': total_original,
                'Outliers': total_outliers,
                'Difference': total_difference,
                '% Retained': total_percent_retained,
                'After Cleaning': total_after_cleaning
            }
            print(f"[DEBUG] Total row calculated")
            
            # Append the total row using modern pandas concat method
            total_row_df = pd.DataFrame([total_row])
            df = pd.concat([df, total_row_df], ignore_index=True)
            print(f"[DEBUG] Total row appended, DataFrame shape: {df.shape}")
            print(f"[DEBUG] DataFrame columns: {list(df.columns)}")
            print(f"[DEBUG] Final DataFrame ready for Excel writing")
            
            # Reorder columns to match request
            df = df[['Category', 'Original', 'Outliers', 'Difference', '% Retained', 'After Cleaning']]

            # 4. Write to Excel
            current_row = 4
            ws.cell(row=current_row, column=1, value="Sample Composition Breakdown")
            current_row += 1

            # Write headers individually to maintain proper formatting
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=current_row, column=col_idx, value=str(header))
            current_row += 1
            print(f"[DEBUG] Headers written: {headers}")

            # Write data rows with proper numeric formatting
            rows_written = 0
            for row_idx, r in df.iterrows():
                row_num = current_row + row_idx
                
                # Write each cell individually with proper data types
                ws.cell(row=row_num, column=1, value=str(r['Category']))  # Category as text
                ws.cell(row=row_num, column=2, value=int(r['Original']))  # Original as number
                ws.cell(row=row_num, column=3, value=int(r['Outliers']))  # Outliers as number
                ws.cell(row=row_num, column=4, value=int(r['Difference']))  # Difference as number
                
                # Percentage as number with percentage format
                percent_cell = ws.cell(row=row_num, column=5, value=float(r['% Retained'])/100)
                percent_cell.number_format = '0.0%'
                
                ws.cell(row=row_num, column=6, value=int(r['After Cleaning']))  # After Cleaning as number
                
                rows_written += 1
            print(f"[DEBUG] {rows_written} data rows written to Excel with proper numeric formatting")

            # 5. Formatting
            self.formatter.format_sheet(ws)
            print("[SUCCESS] Data Cleaning sheet with Sample Composition table created.")

        except Exception as e:
            print(f"[ERROR] Failed to create Data Cleaning sheet: {e}")
            import traceback
            traceback.print_exc()
    
    def process_pipeline_configurations(self, workbook):
        """
        Processes all pipeline configurations from report_config.json and creates sheets.
        
        Args:
            workbook: openpyxl workbook object
        """
        print("[SHEET] Processing pipeline configurations...")
        try:
            config_path = Path("report_config.json")
            if not config_path.exists():
                print("[WARNING] report_config.json not found, skipping pipeline sheets")
                return
            
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            total_sheets = len(config.get('sheets', []))
            print(f"[INFO] Processing {total_sheets} configured sheets")
            
            for i, sheet_config in enumerate(config.get('sheets', []), 1):
                pipeline_name = sheet_config['pipeline']
                sheet_name = sheet_config.get('name') or sheet_config.get('sheet_name')
                print(f"[SHEET] Creating sheet {i}/{total_sheets}: {sheet_name} (pipeline: {pipeline_name})")
                
                pipeline = PIPELINES.get(pipeline_name)
                if not pipeline:
                    print(f"[WARNING] Pipeline {pipeline_name} not found, skipping {sheet_name}")
                    continue
                
                df = self._run_aggregation(pipeline)
                df = self._fill_missing_collection_days(df, pipeline_name)
                
                if df.empty:
                    print(f"[WARNING] No data for sheet: {sheet_name}")
                    continue
                
                sheet_type = infer_sheet_type(sheet_name)
                if sheet_type and 'ACF_PACF' in sheet_name:
                    # Add Collection Days column for Period Counts sheets
                    if 'Period Counts' in sheet_name and 'Period' in df.columns:
                        from ar_utils import calculate_collection_days_for_period
                        df['Collection_Days'] = df['Period'].apply(calculate_collection_days_for_period)
                        print(f"[OK] Added Collection Days column to {sheet_name}")
                    
                    df = add_acf_pacf_analysis(df, "Total_Files", sheet_type)
                    from ar_utils import add_arima_forecast_columns, reorder_with_forecast_columns
                    df = add_arima_forecast_columns(df, "Total_Files", sheet_type)
                    base_cols = ['_id', 'Year', 'Week', 'Month', 'Period', 'School_Year', 'Biweek_Number', 'Week_Label', 'First_Date', 'Last_Date', 'Total_Files', 'JPG_Files', 'MP3_Files', 'Total_Size_MB', 'Collection_Days', 'Days_With_Data', 'Avg_Files_Per_Day']
                    available_base_cols = [col for col in base_cols if col in df.columns]
                    df = reorder_with_forecast_columns(df, available_base_cols, "Total_Files")
                    
                    # Additional safety check: Remove any duplicate columns
                    if len(df.columns) != len(set(df.columns)):
                        print(f"[WARNING] Duplicate columns detected in {sheet_name}, removing duplicates")
                        df = df.loc[:, ~df.columns.duplicated()]

                ws = workbook.create_sheet(title=sheet_name)
                ws.append(list(df.columns))
                for _, row in df.iterrows():
                    ws.append(list(row))
                
                self.formatter.format_sheet(ws)
                
                include_total = sheet_config.get('include_total', True)
                if include_total:
                    if sheet_name == "Audio Note Characteristics":
                        self.formatter.add_audio_characteristics_total_row(ws, df)
                    else:
                        self.formatter.add_total_row(ws, df)
                
                print(f"[SUCCESS] Created sheet: {sheet_name} ({len(df)} rows)")
            
        except Exception as e:
            print(f"[ERROR] Failed to process pipeline configurations: {e}")
            import traceback
            traceback.print_exc()
    
    def create_audio_efficiency_details_sheet(self, workbook):
        """
        Creates the Audio Efficiency Details sheet with enhanced filtering and positioning.
        
        Args:
            workbook: openpyxl workbook object
        """
        print("[SHEET] Creating Audio Efficiency Details sheet...")
        try:
            pipeline = PIPELINES.get('AUDIO_EFFICIENCY_ANALYSIS')
            if not pipeline:
                print("[WARNING] AUDIO_EFFICIENCY_ANALYSIS pipeline not found")
                return
            
            df = self._run_aggregation(pipeline)
            if df.empty:
                print("[WARNING] No data for Audio Efficiency Details sheet")
                return
            
            df_filtered = df[df['efficiency_files_per_audio_minute'] > 0].copy()
            
            position = None
            for i, sheet_name in enumerate(workbook.sheetnames):
                if "Day Types Breakdown" in sheet_name:
                    position = i + 1
                    break
            
            if position is not None:
                ws = workbook.create_sheet(title="Audio Efficiency Details", index=position)
            else:
                ws = workbook.create_sheet(title="Audio Efficiency Details")
            
            headers = ["Date", "Total Files", "MP3 Count", "Total Duration (sec)", "Total Duration (min)", "Files per Audio Minute"]
            ws.append(headers)
            
            for _, row in df_filtered.iterrows():
                ws.append([row.get('_id'), row.get('total_files_day'), row.get('mp3_count'), row.get('total_duration_seconds'), row.get('total_duration_minutes'), row.get('efficiency_files_per_audio_minute')])

            current_row = ws.max_row + 2
            df_top10 = df_filtered.nlargest(10, 'efficiency_files_per_audio_minute')
            self.formatter.add_top_days_table(ws, df_top10, current_row)
            
            self.formatter.format_sheet(ws)
            
            print("[SUCCESS] Audio Efficiency Details sheet created")
        except Exception as e:
            print(f"[ERROR] Failed to create Audio Efficiency Details sheet: {e}")
            import traceback
            traceback.print_exc()
    
    def create_mp3_duration_analysis_sheet(self, workbook):
        """
        Creates comprehensive MP3 Duration Analysis sheet with multiple tables.
        
        Args:
            workbook: openpyxl workbook object
        """
        print("[SHEET] Creating MP3 Duration Analysis sheet...")
        
        try:
            # Import seconds_to_hms for formatting
            from ar_utils import seconds_to_hms
            
            # Get data from all three pipelines
            df_by_year = self._run_aggregation(PIPELINES['MP3_DURATION_BY_SCHOOL_YEAR'])
            df_by_period = self._run_aggregation(PIPELINES['MP3_DURATION_BY_PERIOD'])
            df_by_month = self._run_aggregation(PIPELINES['MP3_DURATION_BY_MONTH'])
            
            # Check if we have data
            if df_by_year.empty and df_by_period.empty and df_by_month.empty:
                print("[WARNING] No MP3 duration data available")
                return
            
            ws = workbook.create_sheet(title="MP3 Duration Analysis")
            current_row = 1
            
            # Table 1: School Year Summary
            if not df_by_year.empty:
                current_row = self._add_duration_summary_table(ws, df_by_year, current_row, seconds_to_hms)
                current_row += 3  # Add spacing
            
            # Table 2: Period Breakdown
            if not df_by_period.empty:
                current_row = self._add_period_duration_table(ws, df_by_period, current_row, seconds_to_hms)
                current_row += 3  # Add spacing
            
            # Table 3: Monthly Distribution
            if not df_by_month.empty:
                current_row = self._add_monthly_duration_table(ws, df_by_month, current_row, seconds_to_hms)
            
            # Apply formatting
            self.formatter.format_sheet(ws)
            print("[SUCCESS] MP3 Duration Analysis sheet created")
            
        except Exception as e:
            print(f"[ERROR] Failed to create MP3 Duration Analysis sheet: {e}")
            import traceback
            traceback.print_exc()
    
    def _add_duration_summary_table(self, ws, df, start_row, seconds_to_hms):
        """
        Adds the School Year Duration Summary table.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="MP3 Duration Summary by School Year")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "School Year", "Total MP3 Files", "Total Duration (Hours)", 
            "Total Duration (HH:MM:SS)", "Avg Duration (MM:SS)", 
            "Min Duration", "Max Duration", "Days with MP3", "Avg Files per Day"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # Data rows
        total_files = 0
        total_seconds = 0
        total_days = 0
        
        for _, row in df.iterrows():
            school_year = row.get('_id', 'Unknown')
            files = row.get('Total_MP3_Files', 0)
            duration_sec = row.get('Total_Duration_Seconds', 0)
            duration_hours = row.get('Total_Duration_Hours', 0)
            avg_duration = row.get('Avg_Duration_Seconds', 0)
            min_duration = row.get('Min_Duration_Seconds', 0)
            max_duration = row.get('Max_Duration_Seconds', 0)
            days_with_mp3 = row.get('Days_With_MP3', 0)
            avg_files_per_day = row.get('Avg_Files_Per_Day', 0)
            
            # Accumulate totals
            total_files += files
            total_seconds += duration_sec
            total_days += days_with_mp3
            
            # Write row data
            ws.cell(row=current_row, column=1, value=school_year)
            ws.cell(row=current_row, column=2, value=files)
            ws.cell(row=current_row, column=3, value=duration_hours)
            ws.cell(row=current_row, column=4, value=seconds_to_hms(duration_sec))
            ws.cell(row=current_row, column=5, value=seconds_to_hms(avg_duration))
            ws.cell(row=current_row, column=6, value=seconds_to_hms(min_duration))
            ws.cell(row=current_row, column=7, value=seconds_to_hms(max_duration))
            ws.cell(row=current_row, column=8, value=days_with_mp3)
            ws.cell(row=current_row, column=9, value=avg_files_per_day)
            
            current_row += 1
        
        # Add totals row
        if len(df) > 1:
            total_hours = round(total_seconds / 3600, 2)
            avg_duration_total = total_seconds / total_files if total_files > 0 else 0
            avg_files_total = total_files / total_days if total_days > 0 else 0
            
            ws.cell(row=current_row, column=1, value="TOTAL")
            ws.cell(row=current_row, column=2, value=total_files)
            ws.cell(row=current_row, column=3, value=total_hours)
            ws.cell(row=current_row, column=4, value=seconds_to_hms(total_seconds))
            ws.cell(row=current_row, column=5, value=seconds_to_hms(avg_duration_total))
            ws.cell(row=current_row, column=6, value="-")
            ws.cell(row=current_row, column=7, value="-")
            ws.cell(row=current_row, column=8, value=total_days)
            ws.cell(row=current_row, column=9, value=round(avg_files_total, 1))
            
            current_row += 1
        
        return current_row
    
    def _add_period_duration_table(self, ws, df, start_row, seconds_to_hms):
        """
        Adds the Collection Period Duration table.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="MP3 Duration by Collection Period")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "Period", "School Year", "Total MP3 Files", "Total Duration (Hours)",
            "Avg Duration (MM:SS)", "Days with MP3", "Avg Files per Day", "Period Efficiency"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # Data rows
        for _, row in df.iterrows():
            # Process each period row
            
            # Access flattened fields directly (after _run_aggregation processing)
            period = row.get('Period', 'Unknown')
            school_year = row.get('School_Year', 'Unknown')
            
            files = row.get('Total_MP3_Files', 0)
            duration_hours = row.get('Total_Duration_Hours', 0)
            avg_duration = row.get('Avg_Duration_Seconds', 0)
            days_with_mp3 = row.get('Days_With_MP3', 0)
            avg_files_per_day = row.get('Avg_Files_Per_Day', 0)
            period_efficiency = row.get('Period_Efficiency', 0)
            
            ws.cell(row=current_row, column=1, value=period)
            ws.cell(row=current_row, column=2, value=school_year)
            ws.cell(row=current_row, column=3, value=files)
            ws.cell(row=current_row, column=4, value=duration_hours)
            ws.cell(row=current_row, column=5, value=seconds_to_hms(avg_duration))
            ws.cell(row=current_row, column=6, value=days_with_mp3)
            ws.cell(row=current_row, column=7, value=avg_files_per_day)
            ws.cell(row=current_row, column=8, value=period_efficiency)
            
            current_row += 1
        
        return current_row
    
    def _add_monthly_duration_table(self, ws, df, start_row, seconds_to_hms):
        """
        Adds the Monthly Duration Distribution table.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="Monthly MP3 Duration Distribution")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "Month", "2021-22 Files", "2021-22 Hours", "2022-23 Files", 
            "2022-23 Hours", "Total Files", "Total Hours", "Avg Duration"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        current_row += 1
        
        # Process data by month
        month_names = [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
        ]
        
        # Group data by month
        monthly_data = {}
        # Process monthly data
        for _, row in df.iterrows():
            # Access flattened fields directly (after _run_aggregation processing)
            month_num = row.get('Month', 0)
            school_year = row.get('School_Year', 'Unknown')
            
            if month_num not in monthly_data:
                monthly_data[month_num] = {}
            
            monthly_data[month_num][school_year] = {
                'files': row.get('Total_MP3_Files', 0),
                'hours': row.get('Total_Duration_Hours', 0),
                'avg_duration': row.get('Avg_Duration_Seconds', 0)
            }
        
        # Write monthly rows
        for month_num in sorted(monthly_data.keys()):
            if month_num < 1 or month_num > 12:
                continue
                
            month_name = month_names[month_num - 1]
            month_data = monthly_data[month_num]
            
            # Get data for each school year
            sy_2122 = month_data.get('2021-2022', {'files': 0, 'hours': 0, 'avg_duration': 0})
            sy_2223 = month_data.get('2022-2023', {'files': 0, 'hours': 0, 'avg_duration': 0})
            
            total_files = sy_2122['files'] + sy_2223['files']
            total_hours = sy_2122['hours'] + sy_2223['hours']
            avg_duration = (sy_2122['avg_duration'] + sy_2223['avg_duration']) / 2 if (sy_2122['avg_duration'] > 0 and sy_2223['avg_duration'] > 0) else max(sy_2122['avg_duration'], sy_2223['avg_duration'])
            
            ws.cell(row=current_row, column=1, value=month_name)
            ws.cell(row=current_row, column=2, value=sy_2122['files'])
            ws.cell(row=current_row, column=3, value=sy_2122['hours'])
            ws.cell(row=current_row, column=4, value=sy_2223['files'])
            ws.cell(row=current_row, column=5, value=sy_2223['hours'])
            ws.cell(row=current_row, column=6, value=total_files)
            ws.cell(row=current_row, column=7, value=total_hours)
            ws.cell(row=current_row, column=8, value=seconds_to_hms(avg_duration))
            
            current_row += 1
        
        return current_row
