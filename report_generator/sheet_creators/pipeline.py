"""
Pipeline-Driven Sheet Creator Module
===================================

This module handles the creation of Excel sheets based on pipeline configurations
from report_config.json. It processes all configured pipelines and creates
sheets with ACF/PACF analysis and ARIMA forecasting capabilities.
"""

import json
import pandas as pd
from pathlib import Path
from openpyxl.utils import get_column_letter

# Local imports - CRITICAL: Use explicit imports to avoid namespace collision
# The utils package also exports add_acf_pacf_analysis (problematic version)
# We must ensure ONLY the ar_utils.py version (correct) is used
from ar_utils import add_acf_pacf_analysis, reorder_with_acf_pacf, infer_sheet_type
from column_cleanup_utils import cleanup_duplicate_acf_pacf_columns
from utils.formatting import reorder_with_forecast_columns  # Explicit submodule import
from pipelines import PIPELINES  # Now using modular pipelines/ package
from .base import BaseSheetCreator


class PipelineSheetCreator(BaseSheetCreator):
    """
    Handles creation of sheets based on pipeline configurations with advanced
    time series analysis capabilities.
    """
    
    def process_pipeline_configurations(self, workbook):
        """
        Process all pipeline configurations from report_config.json.
        Creates sheets based on enabled pipeline configurations.
        """
        print("\n" + "="*80)
        print(" CRITICAL DEBUG: process_pipeline_configurations method called!")
        print(" This confirms the execution path is correct")
        print("="*80)
        try:
            # Load report configuration
            config_path = Path(__file__).parent.parent.parent / "report_config.json"
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            # Get enabled sheets and sort by order
            enabled_sheets = [s for s in config.get('sheets', []) if s.get('enabled', True)]
            enabled_sheets.sort(key=lambda x: x.get('order', 999))
            
            print(f"[DEBUG] Found {len(enabled_sheets)} enabled sheets")
            for sheet in enabled_sheets:
                sheet_name = sheet.get('name', sheet.get('sheet_name', 'Unknown'))
                is_specialized = sheet.get('specialized', False)
                order = sheet.get('order', 999)
                print(f"[DEBUG] Sheet: {sheet_name}, Order: {order}, Specialized: {is_specialized}")
            
            # Process each sheet configuration in order
            for sheet_config in enabled_sheets:
                # Clear pipeline cache before each sheet to prevent contamination
                self._pipeline_cache.clear()
                
                sheet_name = sheet_config.get('name', sheet_config.get('sheet_name', 'Unknown'))
                is_specialized = sheet_config.get('specialized', False)
                print(f"[DEBUG] Processing sheet '{sheet_name}': specialized={is_specialized}")
                
                if is_specialized:
                    print(f"[DEBUG] Creating specialized sheet: {sheet_name}")
                    self._create_specialized_sheet(workbook, sheet_config)
                else:
                    print(f"[DEBUG] Creating pipeline sheet: {sheet_name}")
                    self._create_pipeline_sheet(workbook, sheet_config)
            
            print(f"[SUCCESS] Processed {len(enabled_sheets)} sheets (including specialized sheets)")
            
        except Exception as e:
            print(f"[ERROR] Failed to process pipeline configurations: {e}")
    
    def _create_specialized_sheet(self, workbook, sheet_config):
        """
        Creates a specialized sheet that requires custom creation logic.
        
        Args:
            workbook: openpyxl workbook object
            sheet_config: Sheet configuration dictionary with specialized=True
        """
        sheet_name = sheet_config['name']
        pipeline_name = sheet_config['pipeline']
        
        print(f"[INFO] Creating specialized sheet: {sheet_name}")
        print(f"    - Pipeline identifier: {pipeline_name}")
        
        try:
            # Handle different types of specialized sheets
            if pipeline_name == "MP3_DURATION_ANALYSIS":
                print(f"    - Creating {sheet_name} sheet (specialized)")
                from .specialized import SpecializedSheetCreator
                specialized_creator = SpecializedSheetCreator(self.db, self.formatter)
                
                # Create the sheet
                specialized_creator.create_mp3_duration_analysis_sheet(workbook)
                
                # Move the sheet to the correct position based on order
                sheet_order = sheet_config.get('order', 999)
                print("\n" + "*"*100)
                print("*** MP3 POSITIONING LOGIC EXECUTED! ***")
                print("*"*100)
                self._position_sheet_by_order(workbook, sheet_name, sheet_order)
                
                print(f"[SUCCESS] {sheet_name} created and positioned successfully")
            else:
                print(f"[ERROR] Unknown specialized sheet type: {pipeline_name}")
                
        except Exception as e:
            print(f"[ERROR] Failed to create specialized sheet '{sheet_name}': {e}")
            import traceback
            traceback.print_exc()
    
    def _position_sheet_by_order(self, workbook, sheet_name, target_order):
        """
        Moves a sheet to the correct position in the workbook based on its configured order.
        
        Args:
            workbook: openpyxl workbook object
            sheet_name: Name of the sheet to position
            target_order: Target order number for the sheet
        """
        try:
            if sheet_name not in workbook.sheetnames:
                print(f"[WARNING] Sheet '{sheet_name}' not found for positioning")
                return
            
            # Get the sheet to move
            sheet_to_move = workbook[sheet_name]
            
            # Find the correct position based on order
            target_index = 0
            
            # Load configuration to get other sheet orders
            import json
            from pathlib import Path
            config_path = Path(__file__).parent.parent.parent / "report_config.json"
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            # Create a mapping of sheet names to their orders
            sheet_orders = {}
            for sheet_config in config.get('sheets', []):
                if sheet_config.get('enabled', True):
                    name = sheet_config.get('sheet_name', sheet_config.get('name', ''))
                    order = sheet_config.get('order', 999)
                    sheet_orders[name] = order
            
            # Count how many sheets should come before this one
            print(f"[DEBUG] Positioning '{sheet_name}' with target order {target_order}")
            print(f"[DEBUG] Current workbook sheets: {workbook.sheetnames}")
            print(f"[DEBUG] Sheet orders from config: {sheet_orders}")
            
            for existing_sheet_name in workbook.sheetnames:
                if existing_sheet_name == sheet_name:
                    continue
                existing_order = sheet_orders.get(existing_sheet_name, 999)
                print(f"[DEBUG] Sheet '{existing_sheet_name}' has order {existing_order}")
                if existing_order < target_order:
                    target_index += 1
                    print(f"[DEBUG] '{existing_sheet_name}' comes before target, target_index now {target_index}")
            
            # Move the sheet to the correct position
            workbook.move_sheet(sheet_to_move, target_index)
            print(f"[SUCCESS] Positioned '{sheet_name}' at index {target_index} (order {target_order})")
            
        except Exception as e:
            print(f"[ERROR] Failed to position sheet '{sheet_name}': {e}")
    
    def _create_pipeline_sheet(self, workbook, sheet_config, data=None):
        """
        Creates a single sheet based on pipeline configuration.
        
        Args:
            workbook: openpyxl workbook object
            sheet_config: Sheet configuration dictionary
            data: Optional pre-fetched data (for backward compatibility)
        """
        print(f"[DEBUG_TRACE] _create_pipeline_sheet called for: {sheet_config['name']} with pipeline: {sheet_config['pipeline']}")
        sheet_name = sheet_config['name']
        pipeline_name = sheet_config['pipeline']
        
        # Each sheet must use its own fresh pipeline data to prevent column accumulation
        print(f"[INFO] Creating sheet: {sheet_name}")
        print(f"    - Using pipeline: {pipeline_name}")
        
        # Get fresh data from the specific pipeline for this sheet
        from pipelines import PIPELINES
        if pipeline_name not in PIPELINES:
            print(f"[ERROR] Pipeline '{pipeline_name}' not found")
            return
        
        pipeline = PIPELINES[pipeline_name]
        
        # CRITICAL FIX: Use non-cached aggregation to prevent DataFrame contamination
        # The caching mechanism was storing DataFrames that had been mutated with ACF/PACF columns
        # This caused each subsequent sheet to inherit previously added columns
        print(f"[DEBUG] Using fresh (non-cached) pipeline execution to prevent column contamination")
        # COMPREHENSIVE DEBUG LOGGING
        print(f"[PIPELINE_EXEC_DEBUG] ========================================")
        print(f"[PIPELINE_EXEC_DEBUG] Sheet: {sheet_config['name']}")
        print(f"[PIPELINE_EXEC_DEBUG] Pipeline: {sheet_config['pipeline']}")
        print(f"[PIPELINE_EXEC_DEBUG] Module: {sheet_config.get('module', 'unknown')}")
        print(f"[PIPELINE_EXEC_DEBUG] Category: {sheet_config.get('category', 'unknown')}")
        print(f"[PIPELINE_EXEC_DEBUG] About to call _run_aggregation_original...")
        print(f"[PIPELINE_EXEC_DEBUG] ========================================")
        
        # COMPREHENSIVE DEBUG LOGGING
        print(f"[PIPELINE_EXEC_DEBUG] ========================================")
        print(f"[PIPELINE_EXEC_DEBUG] Sheet: {sheet_config['name']}")
        print(f"[PIPELINE_EXEC_DEBUG] Pipeline: {sheet_config['pipeline']}")
        print(f"[PIPELINE_EXEC_DEBUG] Module: {sheet_config.get('module', 'unknown')}")
        print(f"[PIPELINE_EXEC_DEBUG] Category: {sheet_config.get('category', 'unknown')}")
        print(f"[PIPELINE_EXEC_DEBUG] About to call _run_aggregation_original...")
        print(f"[PIPELINE_EXEC_DEBUG] ========================================")
        
        df = self._run_aggregation_original(pipeline)
        print(f"    - Fresh pipeline data: {len(df)} rows × {len(df.columns)} columns")
        
        # ADDITIONAL SAFETY: Verify no ACF/PACF columns exist in fresh data
        acf_pacf_patterns = ['ACF_Lag_', 'PACF_Lag_', '_Significant']
        existing_acf_pacf_cols = [col for col in df.columns 
                                if any(pattern in str(col) for pattern in acf_pacf_patterns)]
        
        if existing_acf_pacf_cols:
            print(f"[CRITICAL ERROR] Fresh pipeline data already contains ACF/PACF columns: {existing_acf_pacf_cols}")
            print(f"[CRITICAL ERROR] This indicates contamination in the MongoDB aggregation pipeline itself")
            # Remove them as emergency fallback
            df = df.drop(columns=existing_acf_pacf_cols)
            print(f"[EMERGENCY FIX] Removed contaminated columns, proceeding with clean data")
        
        if df.empty:
            print(f"[WARNING] No data returned for pipeline '{pipeline_name}'")
            return
        
        # Fix complex data structures before Excel processing
        df = self._fix_complex_data_structures(df, sheet_name)
        
        # Apply zero-fill for daily pipelines if needed
        print(f"[DEBUG] About to call _fill_missing_collection_days for pipeline: {pipeline_name}")
        print(f"[DEBUG] Input data shape: {df.shape}")
        df = self._fill_missing_collection_days(df, pipeline_name)
        print(f"[DEBUG] Output data shape: {df.shape}")
        
        # Determine sheet type and apply appropriate analysis
        sheet_type = infer_sheet_type(sheet_name)
        
        # Apply ACF/PACF analysis based on configuration (not just sheet type)
        from chart_config_helper import should_add_acf_pacf_columns
        
        if should_add_acf_pacf_columns(sheet_name):
            print(f"[INFO] Adding ACF/PACF analysis for {sheet_type} sheet")
            original_columns = df.columns.tolist()
            
            # Match legacy behavior: only analyze Total_Files metric
            if 'Total_Files' in df.columns:
                print(f"    - Analyzing metric: Total_Files (legacy behavior)")
                print(f"    - Before ACF/PACF: {len(df.columns)} columns")
                print(f"    - Existing ACF columns: {[col for col in df.columns if 'ACF_Lag_' in str(col)]}")
                
                # Apply ACF/PACF analysis - FIXED: Proper column deduplication
                # CRITICAL: Ensure we start with a completely fresh DataFrame copy
                df_for_analysis = df.copy(deep=True)
                
                # DEBUG: Log DataFrame state before ACF/PACF analysis
                print(f"    - [DEBUG] DataFrame before ACF/PACF analysis:")
                print(f"      - Shape: {df.shape}")
                print(f"      - Columns: {list(df.columns)}")
                print(f"      - ACF columns already present: {[col for col in df.columns if 'ACF_Lag_' in str(col)]}")
                print(f"      - PACF columns already present: {[col for col in df.columns if 'PACF_Lag_' in str(col)]}")
                
                # Get ACF/PACF results (returns ONLY new columns)
                print(f"    - [DEBUG] Calling add_acf_pacf_analysis...")
                acf_pacf_results = add_acf_pacf_analysis(df_for_analysis, value_col='Total_Files', sheet_type=sheet_type)
                
                # DEBUG: Log ACF/PACF results
                print(f"    - [DEBUG] ACF/PACF analysis results:")
                print(f"      - Shape: {acf_pacf_results.shape}")
                print(f"      - Columns: {list(acf_pacf_results.columns)}")
                
                # CRITICAL DEBUG: Check for PACF columns specifically
                acf_cols_generated = [col for col in acf_pacf_results.columns if 'ACF_Lag_' in col and '_Significant' not in col]
                pacf_cols_generated = [col for col in acf_pacf_results.columns if 'PACF_Lag_' in col and '_Significant' not in col]
                print(f"      - ACF columns generated: {acf_cols_generated}")
                print(f"      - PACF columns generated: {pacf_cols_generated}")
                
                if not pacf_cols_generated:
                    print(f"    - [CRITICAL ERROR] NO PACF COLUMNS GENERATED BY add_acf_pacf_analysis!")
                    print(f"    - [CRITICAL ERROR] This indicates the PACF calculation is failing in the main pipeline")
                    
                if not acf_pacf_results.empty:
                    print(f"    - ACF/PACF analysis returned {len(acf_pacf_results.columns)} new columns")
                    
                    # VERIFICATION: Confirm no ACF/PACF columns exist before adding new ones
                    # With the fix above, this should never trigger, but kept as safety check
                    acf_pacf_patterns = ['ACF_Lag_', 'PACF_Lag_', '_Significant']
                    existing_acf_pacf_cols = [col for col in df.columns 
                                            if any(pattern in str(col) for pattern in acf_pacf_patterns)]
                    
                    if existing_acf_pacf_cols:
                        print(f"    - WARNING: Found unexpected ACF/PACF columns (should not happen with fix): {existing_acf_pacf_cols}")
                        print(f"    - REMOVING them as safety measure")
                        df = df.drop(columns=existing_acf_pacf_cols)
                    
                    # Now safely concatenate the new ACF/PACF columns
                    # Perform concatenation
                    df = pd.concat([df, acf_pacf_results], axis=1)
                    
                    # CRITICAL DEBUG: Check PACF columns after concatenation
                    acf_cols_after_concat = [col for col in df.columns if 'ACF_Lag_' in col and '_Significant' not in col]
                    pacf_cols_after_concat = [col for col in df.columns if 'PACF_Lag_' in col and '_Significant' not in col]
                    print(f"    - [DEBUG] After concatenation:")
                    print(f"      - ACF columns: {acf_cols_after_concat}")
                    print(f"      - PACF columns: {pacf_cols_after_concat}")
                    
                    if not pacf_cols_after_concat:
                        print(f"    - [CRITICAL ERROR] PACF COLUMNS LOST DURING CONCATENATION!")
                    
                    # Check for duplicates immediately after concatenation
                    if len(df.columns) != len(set(df.columns)):
                        print(f"    - WARNING: Duplicate columns detected after concatenation, removing...")
                        df = df.loc[:, ~df.columns.duplicated()]
                    
                    # Clean up any duplicate ACF/PACF columns from inconsistent naming
                    df = cleanup_duplicate_acf_pacf_columns(df, 'Total_Files')
                    
                    print(f"    - After ACF/PACF: {len(df.columns)} columns")
                    print(f"    - New ACF columns: {[col for col in df.columns if 'ACF_Lag_' in str(col)]}")
                    print(f"    - New PACF columns: {[col for col in df.columns if 'PACF_Lag_' in str(col)]}")
                    
                    # CRITICAL DEBUG: Final PACF check after all processing
                    final_acf_cols = [col for col in df.columns if 'ACF_Lag_' in col and '_Significant' not in col]
                    final_pacf_cols = [col for col in df.columns if 'PACF_Lag_' in col and '_Significant' not in col]
                    print(f"    - [FINAL DEBUG] After all processing:")
                    print(f"      - Final ACF columns: {final_acf_cols}")
                    print(f"      - Final PACF columns: {final_pacf_cols}")
                    
                    if not final_pacf_cols:
                        print(f"    - [CRITICAL ERROR] PACF COLUMNS COMPLETELY MISSING FROM FINAL DATAFRAME!")
                        print(f"    - [CRITICAL ERROR] This will result in missing PACF columns in Excel output")
            else:
                print(f"    - Skipping ACF/PACF analysis: Total_Files column not found")
                
            # Ensure no duplicate columns before reordering
            if len(df.columns) != len(set(df.columns)):
                print(f"    - WARNING: Duplicate columns detected before reordering, removing...")
                df = df.loc[:, ~df.columns.duplicated(keep='first')]
                
            df = reorder_with_acf_pacf(df, original_columns)
            
            # Apply ARIMA forecasting based on configuration (not just sheet type)
            from chart_config_helper import should_add_arima_columns
            
            if should_add_arima_columns(sheet_name):
                print(f"[INFO] Adding ARIMA forecasting for {sheet_type} sheet (configuration-driven)")
                # Store columns before forecasting for reordering
                columns_before_forecast = df.columns.tolist()
                df = self._apply_arima_forecasting(df, sheet_type)
                df = reorder_with_forecast_columns(df, columns_before_forecast)
            else:
                print(f"[INFO] Skipping ARIMA forecasting for {sheet_name} (not enabled in configuration)")
        
        # CRITICAL FIX: Smart deduplication that preserves PACF columns
        print(f"    - [DEBUG] Before deduplication: {len(df.columns)} columns")
        print(f"    - [DEBUG] Columns before dedup: {list(df.columns)}")
        original_col_count = len(df.columns)
        
        # Check for duplicates first
        if len(df.columns) != len(set(df.columns)):
            print(f"    - [CRITICAL] Duplicate columns detected before Excel export!")
            duplicate_cols = [col for col in df.columns if list(df.columns).count(col) > 1]
            print(f"    - [CRITICAL] Duplicate columns: {set(duplicate_cols)}")
            
            # SMART DEDUPLICATION: Remove true duplicates while preserving unique columns
            # Use pandas built-in deduplication but with proper validation
            columns_before = list(df.columns)
            df = df.loc[:, ~df.columns.duplicated(keep='first')]
            columns_after = list(df.columns)
            
            # Log which columns were removed
            removed_columns = [col for col in columns_before if col not in columns_after or columns_before.count(col) > columns_after.count(col)]
            for col in set(removed_columns):
                count_before = columns_before.count(col)
                count_after = columns_after.count(col)
                if count_before > count_after:
                    print(f"    - [DEDUP] Removed {count_before - count_after} duplicate(s) of: {col}")
            
            print(f"    - [FIX] After smart deduplication: {len(df.columns)} columns")
            print(f"    - [FIX] Removed {original_col_count - len(df.columns)} duplicate columns")
            print(f"    - [DEBUG] Columns after dedup: {list(df.columns)}")
            
            # VALIDATION: Ensure PACF columns are still present
            pacf_cols_after = [col for col in df.columns if 'PACF_Lag_' in col and '_Significant' not in col]
            acf_cols_after = [col for col in df.columns if 'ACF_Lag_' in col and '_Significant' not in col]
            print(f"    - [VALIDATION] ACF columns after dedup: {acf_cols_after}")
            print(f"    - [VALIDATION] PACF columns after dedup: {pacf_cols_after}")
            
            if not pacf_cols_after and acf_cols_after:
                print(f"    - [CRITICAL ERROR] PACF columns lost during deduplication!")
                print(f"    - [CRITICAL ERROR] This will cause chart overlap issue!")
            elif pacf_cols_after:
                print(f"    - [SUCCESS] PACF columns preserved after deduplication")
        else:
            print(f"    - [OK] No duplicate columns found before Excel export")
        
        # Create the worksheet
        ws = workbook.create_sheet(sheet_name)
        
        # Add title
        ws['A1'] = f"AR Data Analysis - {sheet_name}"
        self.formatter.apply_title_style(ws, 'A1')
        
        # Add headers
        for col, column_name in enumerate(df.columns, 1):
            ws.cell(row=3, column=col, value=str(column_name))
        
        # Apply header formatting with special styling for ACF/PACF columns
        last_col_letter = get_column_letter(len(df.columns))
        header_range = f'A3:{last_col_letter}3'
        self.formatter.apply_header_style(ws, header_range)
        
        # Apply special formatting for ACF/PACF columns
        self._apply_acf_pacf_header_formatting(ws, df.columns, 3)
        
        # Add data rows
        for row_idx, (_, row) in enumerate(df.iterrows(), 4):
            for col_idx, value in enumerate(row, 1):
                # Handle different data types appropriately
                if pd.isna(value):
                    cell_value = ""
                elif isinstance(value, (int, float)):
                    cell_value = value
                else:
                    cell_value = str(value)
                
                ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        # Apply data formatting with optimization for large datasets
        last_col_letter = get_column_letter(len(df.columns))
        data_range = f'A4:{last_col_letter}{3 + len(df)}'
        
        # Calculate total cells to determine if we need optimized approach
        total_cells = len(df) * len(df.columns)
        
        if total_cells > 1000:
            print(f"[INFO] Large dataset detected ({len(df)} rows × {len(df.columns)} cols = {total_cells} cells)")
            print(f"[INFO] Using optimized formatting approach")
            
            # For large datasets, apply minimal formatting to avoid performance issues
            self._apply_minimal_data_formatting(ws, 4, 3 + len(df), len(df.columns))
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, 3 + len(df), 1, len(df.columns))
        else:
            # For smaller datasets, use full formatting
            self.formatter.apply_data_style(ws, data_range)
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, 3 + len(df), 1, len(df.columns))
        
        # Apply special formatting for ACF/PACF data
        self._apply_acf_pacf_data_formatting(ws, df.columns, 4, len(df))
        
        # Auto-adjust column widths
        self.formatter.auto_adjust_columns(ws)
        
        # Add totals to pipeline sheets where appropriate
        try:
            # Configure totals based on sheet type and data characteristics
            totals_config = self._get_totals_config_for_sheet(sheet_name, df, sheet_type)
            
            if totals_config and totals_config.get('add_totals', False):
                # Add totals to the worksheet
                start_row = 4  # Where the data starts (after headers)
                start_col = 1  # Column A
                
                # Register key totals for cross-sheet validation
                self._register_sheet_totals(sheet_name, df, sheet_type)
                
                # Add totals to the worksheet
                self.totals_manager.add_totals_to_worksheet(
                    worksheet=ws,
                    dataframe=df,
                    start_row=start_row,
                    start_col=start_col,
                    config=totals_config
                )
                
                print(f"[SUCCESS] Added totals to sheet '{sheet_name}'")
            else:
                print(f"[INFO] Skipping totals for sheet '{sheet_name}' - not applicable for this sheet type")
                
        except Exception as e:
            print(f"[WARNING] Failed to add totals to sheet '{sheet_name}': {e}")
        
        # Add ACF/PACF charts based on configuration
        from chart_config_helper import should_add_chart
        if sheet_type in ['daily', 'weekly', 'biweekly', 'monthly', 'period'] and should_add_chart(sheet_name, 'acf_pacf'):
            chart_added = self._add_acf_pacf_charts(ws, 4, 3 + len(df), sheet_type)
            if chart_added:
                print(f"[SUCCESS] Added ACF/PACF charts to sheet '{sheet_name}' (config-driven)")
        elif sheet_type in ['daily', 'weekly', 'biweekly', 'monthly', 'period']:
            print(f"[INFO] Skipping ACF/PACF charts for '{sheet_name}' - not enabled in configuration")
        
        print(f"[SUCCESS] Created sheet '{sheet_name}' with {len(df)} rows")
    
    def _add_acf_pacf_charts(self, worksheet, data_start_row, data_end_row, sheet_type):
        """
        Add ACF/PACF charts directly to the worksheet during sheet creation.
        
        Args:
            worksheet: openpyxl worksheet object
            data_start_row: First row of data (after headers)
            data_end_row: Last row of data
            sheet_type: Type of sheet (daily, weekly, etc.)
            
        Returns:
            bool: True if charts were added successfully, False otherwise
        """
        try:
            # Import chart functions from acf_pacf_charts
            from acf_pacf_charts import add_acf_pacf_chart, add_chart_summary_info
            
            # Add ACF/PACF chart directly to the worksheet
            chart = add_acf_pacf_chart(worksheet, data_start_row, data_end_row, sheet_type)
            
            if chart:
                # Count ACF/PACF columns for summary - use more permissive column detection
                headers = [str(cell.value) for cell in worksheet[1] if cell.value is not None]
                
                # More flexible detection of ACF/PACF columns
                acf_cols = [h for h in headers if ('ACF' in h or 'Acf' in h or 'acf' in h) and ('Lag' in h or 'lag' in h) and not ('Significant' in h or 'significant' in h)]
                pacf_cols = [h for h in headers if ('PACF' in h or 'Pacf' in h or 'pacf' in h) and ('Lag' in h or 'lag' in h) and not ('Significant' in h or 'significant' in h)]
                total_lags = len(set([h.split('_Lag')[1].split('_')[0] for h in acf_cols + pacf_cols if '_Lag' in str(h)]))
                computed_lags = len(acf_cols)
                
                # Add summary information
                add_chart_summary_info(worksheet, data_end_row, sheet_type, total_lags, computed_lags)
                
                return True
            else:
                print(f"[INFO] No ACF/PACF charts added to {worksheet.title} - no suitable data found")
                return False
                
        except Exception as e:
            print(f"[WARNING] Could not add ACF/PACF charts to {worksheet.title}: {e}")
            return False
    
    def _should_apply_forecasting(self, sheet_config, sheet_type):
        """
        Determines if ARIMA forecasting should be applied to this sheet.
        
        Args:
            sheet_config: Sheet configuration dictionary
            sheet_type: Inferred sheet type (daily, weekly, etc.)
            
        Returns:
            bool: True if forecasting should be applied
        """
        try:
            # Load forecast configuration
            config_path = Path(__file__).parent.parent.parent / "report_config.json"
            with open(config_path, 'r') as f:
                config = json.load(f)
            
            forecast_options = config.get('forecast_options', {})
            
            # Check if forecasting is globally enabled
            if not forecast_options.get('enabled', False):
                return False
            
            # Check if this time scale is enabled for forecasting
            enabled_scales = forecast_options.get('time_scales', [])
            return sheet_type in enabled_scales
            
        except Exception as e:
            print(f"[WARNING] Could not determine forecasting settings: {e}")
            return False
    
    def _apply_arima_forecasting(self, df, sheet_type):
        """
        Applies ARIMA forecasting to the DataFrame.
        
        Args:
            df: DataFrame with time series data
            sheet_type: Type of time series (daily, weekly, etc.)
            
        Returns:
            DataFrame with forecast columns added
        """
        # ARIMA forecasting is now enabled - use ar_utils implementation
        from ar_utils import add_arima_forecast_columns
        return add_arima_forecast_columns(df, "Total_Files", "daily")
    
    def _apply_acf_pacf_header_formatting(self, ws, columns, header_row):
        """
        Applies special formatting to ACF/PACF column headers with enhanced readability.
        
        Args:
            ws: Worksheet object
            columns: List of column names
            header_row: Row number containing headers
        """
        try:
            from openpyxl.styles import PatternFill, Font
            
            # Enhanced ACF/PACF header styling for better readability and contrast
            # Darker background for better contrast
            acf_pacf_fill = PatternFill(start_color="B8D4F0", end_color="B8D4F0", fill_type="solid")
            
            # Dark blue font for excellent contrast against light blue background
            acf_pacf_font = Font(color="003366", bold=True, size=10)
            
            for col_idx, column_name in enumerate(columns, 1):
                if any(pattern in str(column_name) for pattern in ['ACF_', 'PACF_']):
                    cell = ws.cell(row=header_row, column=col_idx)
                    cell.fill = acf_pacf_fill
                    cell.font = acf_pacf_font
                    
        except Exception as e:
            print(f"[WARNING] Could not apply ACF/PACF header formatting: {e}")
    
    def _apply_acf_pacf_data_formatting(self, ws, columns, start_row, num_rows):
        """
        Applies special formatting to ACF/PACF data cells with enhanced readability.
        
        Args:
            ws: Worksheet object
            columns: List of column names
            start_row: First data row
            num_rows: Number of data rows
        """
        try:
            from openpyxl.styles import PatternFill, Font
            
            # Enhanced ACF/PACF data styling for better readability
            # Subtle background that complements the header styling
            acf_pacf_fill = PatternFill(start_color="F0F6FC", end_color="F0F6FC", fill_type="solid")
            
            # Slightly darker font for better readability
            acf_pacf_font = Font(color="1F4E79", size=9)
            
            for col_idx, column_name in enumerate(columns, 1):
                if any(pattern in str(column_name) for pattern in ['ACF_', 'PACF_']):
                    for row_idx in range(start_row, start_row + num_rows):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = acf_pacf_fill
                        cell.font = acf_pacf_font
                        
        except Exception as e:
            print(f"[WARNING] Could not apply ACF/PACF data formatting: {e}")
    
    def _get_totals_config_for_sheet(self, sheet_name, df, sheet_type):
        """
        Generate totals configuration based on sheet type and data characteristics.
        Implements systematic totals for high-priority sheets identified in analysis.
        
        Args:
            sheet_name: Name of the sheet
            df: DataFrame containing the data
            sheet_type: Type of sheet (daily, weekly, etc.)
            
        Returns:
            dict: Totals configuration or None if totals not appropriate
        """
        try:
            # Import TotalsIntegrationHelper for configuration generation
            from report_generator.totals_integration_guide import TotalsIntegrationHelper
            helper = TotalsIntegrationHelper()
            
            # HIGH-PRIORITY SHEETS: Implement specific totals based on systematic analysis
            high_priority_configs = {
                'Monthly Capture Volume': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'include_columns': ['Count'],  # Only total the Count column, not Year/Month
                    'rationale': 'Count column needs totals, Year/Month are identifiers'
                },
                'File Size Stats': {
                    'add_row_totals': True,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': '4 numeric columns, statistical summary needs comprehensive totals'
                },
                'Time of Day': {
                    'add_row_totals': True,
                    'add_column_totals': False,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': 'Few rows with 1 numeric column, row totals appropriate'
                },
                'Weekday by Period': {
                    'add_row_totals': True,
                    'add_column_totals': False,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': 'Multiple rows with 1 numeric column, row totals appropriate'
                },
                # PHASE 2: HIGH-PRIORITY REMAINING SHEETS
                'Weekly Counts': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '5 numeric columns, weekly analysis benefits from totals'
                },
                'Day of Week Counts': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '4 numeric columns, day-of-week analysis benefits from totals'
                },
                'Activity Counts': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '4 numeric columns, activity analysis benefits from totals'
                },
                'File Size by Day': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '3 numeric columns, daily file size analysis needs totals'
                },
                'SY Days': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '4 numeric columns, school year days analysis benefits from totals'
                },
                'Data Quality': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '11 numeric columns, data quality metrics need comprehensive totals'
                },
                # PHASE 3: MEDIUM-PRIORITY REMAINING SHEETS
                'Collection Periods': {
                    'add_row_totals': True,
                    'add_column_totals': False,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': '2 numeric columns, period analysis benefits from row totals'
                },
                'Camera Usage by Year': {
                    'add_row_totals': False,
                    'add_column_totals': True,
                    'add_grand_total': True,
                    'totals_label': 'TOTAL',
                    'add_totals': True,
                    'rationale': '2 numeric columns, camera usage trends benefit from totals'
                },
                # PHASE 3: LOW-PRIORITY REMAINING SHEETS
                'Camera Usage Dates': {
                    'add_row_totals': True,
                    'add_column_totals': False,
                    'add_grand_total': False,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': 'Multiple tables with 1 numeric column each, row totals appropriate'
                },
                'Camera Usage by Year Range': {
                    'add_row_totals': True,
                    'add_column_totals': False,
                    'add_grand_total': False,
                    'totals_label': 'TOTAL',
                    'row_totals_label': 'Row Total',
                    'add_totals': True,
                    'rationale': 'Multiple tables with 1 numeric column each, row totals appropriate'
                }
            }
            
            # Check if this is a high-priority sheet with specific configuration
            if sheet_name in high_priority_configs:
                config = high_priority_configs[sheet_name].copy()
                print(f"[HIGH-PRIORITY] Applying systematic totals to '{sheet_name}': {config['rationale']}")
                print(f"[CONFIG] {sheet_name}: row_totals={config['add_row_totals']}, column_totals={config['add_column_totals']}, grand_total={config['add_grand_total']}")
                return config
            
            # Generate recommended configuration for other sheets
            config = helper.generate_recommended_config(df, sheet_name)
            
            # Override based on sheet type for pipeline sheets
            if sheet_type in ['daily', 'weekly', 'biweekly', 'monthly', 'period']:
                # Time series sheets typically benefit from column totals
                config['add_column_totals'] = True
                config['add_row_totals'] = False  # Usually not needed for time series
                config['add_grand_total'] = True
                
                # Exclude statistical columns from totals
                exclude_patterns = ['ACF_', 'PACF_', '_Forecast', '_Lower', '_Upper', '_Quality', '_Message', '_Model']
                config['exclude_columns'] = [col for col in df.columns 
                                            if any(pattern in str(col) for pattern in exclude_patterns)]
                
                print(f"[CONFIG] Generated totals config for {sheet_name}: column_totals={config['add_column_totals']}, excluding {len(config['exclude_columns'])} statistical columns")
            
            return config
            
        except Exception as e:
            print(f"[WARNING] Could not generate totals config for {sheet_name}: {e}")
            return None
    
    def _register_sheet_totals(self, sheet_name, df, sheet_type):
        """
        Register key totals from this sheet for cross-sheet validation.
        
        Args:
            sheet_name: Name of the sheet
            df: DataFrame containing the data
            sheet_type: Type of sheet (daily, weekly, etc.)
        """
        try:
            # Register key metrics for cross-sheet validation
            key_columns = ['Total_Files', 'JPG_Files', 'MP3_Files']
            
            for col in key_columns:
                if col in df.columns:
                    # Calculate total for this column
                    total_value = df[col].sum() if pd.api.types.is_numeric_dtype(df[col]) else None
                    
                    if total_value is not None:
                        # Register with validation system
                        validation_key = f"{col}_{sheet_type}"
                        totals_data = {
                            validation_key: {
                                'value': total_value,
                                'source': f"{sheet_type} aggregation",
                                'column': col
                            }
                        }
                        self.totals_manager.register_sheet_totals(sheet_name, totals_data)
                        
                        print(f"[VALIDATION] Registered {validation_key} = {total_value} for cross-sheet validation")
            
        except Exception as e:
            print(f"[WARNING] Could not register totals for validation: {e}")

    def _fix_complex_data_structures(self, df, sheet_name):
        """
        Fix complex data structures (like nested dictionaries) that can't be written to Excel.
        
        Args:
            df: DataFrame with potentially complex data structures
            sheet_name: Name of the sheet for logging
            
        Returns:
            DataFrame with flattened/simplified data structures
        """
        try:
            # Check if _id column contains complex structures
            if '_id' in df.columns and len(df) > 0:
                sample_id = df['_id'].iloc[0]
                
                if isinstance(sample_id, dict):
                    print(f"[FIX] Flattening complex _id structures in '{sheet_name}'")
                    print(f"      Sample _id: {sample_id}")
                    
                    # Flatten the _id dictionary into separate columns
                    id_df = pd.json_normalize(df['_id'])
                    print(f"      Flattened to columns: {list(id_df.columns)}")
                    
                    # Remove original _id and add flattened columns at the beginning
                    df = df.drop(columns=['_id'])
                    df = pd.concat([id_df, df], axis=1)
                    
                    print(f"      ✅ Data structure fixed for Excel compatibility")
            
            # Check for other complex structures in any column
            for col in df.columns:
                if len(df) > 0:
                    sample_value = df[col].iloc[0]
                    if isinstance(sample_value, (dict, list)) and col != '_id':
                        print(f"[FIX] Converting complex values in column '{col}' to strings")
                        df[col] = df[col].astype(str)
            
            return df
            
        except Exception as e:
            print(f"[WARNING] Could not fix complex data structures in '{sheet_name}': {e}")
            return df
    
    def _apply_minimal_data_formatting(self, ws, start_row, end_row, num_cols):
        """
        Applies minimal formatting optimized for large datasets.
        
        Args:
            ws: Worksheet object
            start_row: First data row
            end_row: Last data row (exclusive)
            num_cols: Number of columns
        """
        try:
            from openpyxl.styles import Alignment
            
            print(f"[INFO] Applying minimal formatting to range: A{start_row}:{get_column_letter(num_cols)}{end_row-1}")
            
            # Only set column widths and basic alignment, avoid cell-by-cell styling
            data_alignment = Alignment(horizontal='left', vertical='center')
            
            # Set column widths efficiently
            for col_num in range(1, num_cols + 1):
                col_letter = get_column_letter(col_num)
                if col_letter not in ws.column_dimensions:
                    ws.column_dimensions[col_letter].width = 12
            
            # Apply alignment only to a sample of cells to establish the pattern
            # openpyxl will often inherit this for similar cells
            sample_rows = min(10, end_row - start_row)  # Format first 10 rows as template
            
            for row_num in range(start_row, start_row + sample_rows):
                for col_num in range(1, min(num_cols + 1, 6)):  # First 5 columns as template
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        cell.alignment = data_alignment
                    except Exception:
                        continue
            
            print(f"[INFO] Minimal formatting applied successfully")
            
        except Exception as e:
            print(f"[WARNING] Could not apply minimal data formatting: {e}")
            # Fallback to no formatting rather than crash
            pass
