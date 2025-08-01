"""
Base Sheet Creator Module
========================

This module contains the base SheetCreator class with common functionality
and core sheet creation methods for summary, raw data, and data cleaning sheets.
"""

import json
import pandas as pd
from pathlib import Path
from datetime import datetime

# Local imports
# CRITICAL FIX: Import add_acf_pacf_analysis from ar_utils.py to avoid namespace collision
from ar_utils import add_acf_pacf_analysis, reorder_with_acf_pacf, infer_sheet_type
from utils import (
    get_school_calendar, get_non_collection_days, precompute_collection_days
)
from pipelines import PIPELINES  # Now using modular pipelines/ package

# Import db_utils conditionally to avoid import errors
try:
    from db_utils import get_db_connection
except ImportError:
    get_db_connection = None


class BaseSheetCreator:
    """
    Base class for sheet creation with common functionality and core sheet methods.
    """
    
    def __init__(self, db, formatter):
        """
        Initialize the sheet creator.
        
        Args:
            db: MongoDB database connection
            formatter: ExcelFormatter instance for styling
        """
        self.db = db
        self.formatter = formatter
        # Global pipeline cache to prevent duplicate executions
        self._pipeline_cache = {}
    
    def _fill_missing_collection_days(self, df, pipeline_name):
        """
        Fills in missing collection days with zero counts for complete time series.
        This is critical for ACF/PACF/ARIMA analysis which requires continuous time series.
        
        Args:
            df (pandas.DataFrame): DataFrame from daily aggregation
            pipeline_name (str): Name of the pipeline to determine if zero-fill is needed
            
        Returns:
            pandas.DataFrame: DataFrame with missing collection days filled with zeros
        """
        if not ('DAILY' in pipeline_name.upper() and 
                ('WITH_ZEROES' in pipeline_name.upper() or 'COLLECTION_ONLY' in pipeline_name.upper())):
            return df
        
        try:
            school_calendar = get_school_calendar()
            non_collection_days = get_non_collection_days()
            collection_day_map = precompute_collection_days(school_calendar, non_collection_days)
            
            all_collection_days = []
            for date_obj, info in collection_day_map.items():
                all_collection_days.append({'_id': date_obj.strftime('%Y-%m-%d')})
            
            all_days_df = pd.DataFrame(all_collection_days)
            
            # Merge with existing data, keeping actual data and filling missing with zeros
            merged_df = pd.merge(all_days_df, df, on='_id', how='left').fillna(0)
            
            # Ensure correct data types after merge
            for col in ['Total_Files', 'MP3_Files', 'JPG_Files']:
                if col in merged_df.columns:
                    merged_df[col] = merged_df[col].astype(int)
            if 'Total_Size_MB' in merged_df.columns:
                 merged_df['Total_Size_MB'] = merged_df['Total_Size_MB'].astype(float)

            return merged_df.sort_values('_id').reset_index(drop=True)
            
        except Exception as e:
            print(f"[WARNING] Zero-fill failed, returning original data: {e}")
            return df
    
    def _run_aggregation_cached(self, cache_key, pipeline, use_base_filter=True, collection_name='media_records'):
        """
        Runs a MongoDB aggregation pipeline with caching to prevent duplicate executions.
        """
        # Check cache first
        if cache_key in self._pipeline_cache:
            print(f"[CACHE HIT] BaseSheetCreator: Reusing cached result for {cache_key}")
            return self._pipeline_cache[cache_key].copy(deep=True)
        
        # Execute pipeline and cache result
        print(f"[CACHE MISS] BaseSheetCreator: Executing and caching {cache_key}")
        result = self._run_aggregation_original(pipeline, use_base_filter, collection_name)
        self._pipeline_cache[cache_key] = result.copy(deep=True)
        return result
    
    def _run_aggregation_original(self, pipeline, use_base_filter=True, collection_name='media_records'):
        """
        Original pipeline execution method (renamed to avoid conflicts).
        """
        try:
            collection = self.db[collection_name]
            full_pipeline = pipeline
            if use_base_filter:
                base_filter = {"$match": {"file_type": {"$in": ["JPG", "MP3"]}}}
                full_pipeline = [base_filter] + pipeline
            
            cursor = collection.aggregate(full_pipeline, allowDiskUse=True)
            df = pd.DataFrame(list(cursor))
            
            if '_id' in df.columns and df.shape[0] > 0 and isinstance(df['_id'].iloc[0], dict):
                # Flatten the _id dictionary into separate columns
                id_df = pd.json_normalize(df['_id'])
                df = pd.concat([df.drop('_id', axis=1), id_df], axis=1)
            
            return df
        except Exception as e:
            print(f"[ERROR] Aggregation failed: {e}")
            return pd.DataFrame()
    
    def _run_aggregation(self, pipeline, use_base_filter=True, collection_name='media_records'):
        """
        Runs a MongoDB aggregation pipeline and returns a DataFrame.
        Now uses caching to prevent duplicate executions.
        """
        # Create cache key from pipeline and parameters
        cache_key = f"base_{str(pipeline)}_{use_base_filter}_{collection_name}"
        return self._run_aggregation_cached(cache_key, pipeline, use_base_filter, collection_name)

    def create_summary_statistics_sheet(self, workbook):
        """
        Creates the detailed Summary Statistics sheet with enhanced day analysis.
        """
        try:
            # Get the data for each school year (original statistics)
            pipeline = [
                {"$match": {"file_type": {"$in": ["JPG", "MP3"]}}},
                {"$group": {
                    "_id": "$ISO_Date",
                    "Total_Files": {"$sum": 1},
                    "MP3_Files": {"$sum": {"$cond": [{"$eq": ["$file_type", "MP3"]}, 1, 0]}},
                    "JPG_Files": {"$sum": {"$cond": [{"$eq": ["$file_type", "JPG"]}, 1, 0]}},
                    "Total_Size_MB": {"$sum": "$File_Size_MB"}
                }},
                {"$sort": {"_id": 1}}
            ]
            
            df = self._run_aggregation(pipeline)
            if df.empty:
                print("[WARNING] No data found for Summary Statistics")
                return
            
            # Convert _id to datetime for filtering
            df['Date'] = pd.to_datetime(df['_id'])
            
            # Define school year periods
            periods = {
                '2021-2022': (datetime(2021, 8, 1), datetime(2022, 6, 30)),
                '2022-2023': (datetime(2022, 8, 1), datetime(2023, 6, 30)),
                'Overall': (datetime(2021, 8, 1), datetime(2023, 6, 30))
            }
            
            # Calculate statistics for each period
            def calc_stats_for_period(df, school_year=None):
                """Calculate statistics for a specific period"""
                if school_year and school_year in periods and school_year != 'Overall':
                    start_date, end_date = periods[school_year]
                    period_df = df[(df['Date'] >= start_date) & (df['Date'] <= end_date)].copy()
                else:
                    period_df = df.copy()
                
                if period_df.empty:
                    return {}
                
                # Get non-collection days and convert to strings for filtering
                non_collection_days = get_non_collection_days()
                non_collection_day_strings = []
                for day in non_collection_days:
                    if hasattr(day, 'strftime'):
                        non_collection_day_strings.append(day.strftime('%Y-%m-%d'))
                    else:
                        non_collection_day_strings.append(str(day))
                
                # Filter out non-collection days for statistics calculation
                collection_days_df = period_df[~period_df['_id'].isin(non_collection_day_strings)]
                
                if collection_days_df.empty:
                    return {}
                
                # Calculate collection days using the correct period-based summation approach
                if school_year and school_year != 'Overall':
                    from utils.calendar import calculate_collection_days_for_period
                    if school_year == '2021-2022':
                        # Sum all periods for 2021-2022 school year
                        collection_days = (
                            calculate_collection_days_for_period('SY 21-22 P1') +
                            calculate_collection_days_for_period('SY 21-22 P2') +
                            calculate_collection_days_for_period('SY 21-22 P3')
                        )
                    else:  # 2022-2023
                        # Sum all periods for 2022-2023 school year
                        collection_days = (
                            calculate_collection_days_for_period('SY 22-23 P1') +
                            calculate_collection_days_for_period('SY 22-23 P2') +
                            calculate_collection_days_for_period('SY 22-23 P3')
                        )
                else:
                    # For overall, sum all periods from both school years
                    from utils.calendar import calculate_collection_days_for_period
                    collection_days = (
                        calculate_collection_days_for_period('SY 21-22 P1') +
                        calculate_collection_days_for_period('SY 21-22 P2') +
                        calculate_collection_days_for_period('SY 21-22 P3') +
                        calculate_collection_days_for_period('SY 22-23 P1') +
                        calculate_collection_days_for_period('SY 22-23 P2') +
                        calculate_collection_days_for_period('SY 22-23 P3')
                    )
                
                stats = {
                    'total_files': collection_days_df['Total_Files'].sum(),
                    'mp3_files': collection_days_df['MP3_Files'].sum(),
                    'jpg_files': collection_days_df['JPG_Files'].sum(),
                    'total_size_mb': collection_days_df['Total_Size_MB'].sum(),
                    'collection_days': collection_days,
                    'mean_files_per_day': collection_days_df['Total_Files'].mean(),
                    'median_files_per_day': collection_days_df['Total_Files'].median(),
                    'std_files_per_day': collection_days_df['Total_Files'].std(),
                    'min_files_per_day': collection_days_df['Total_Files'].min(),
                    'max_files_per_day': collection_days_df['Total_Files'].max(),
                    'range_files_per_day': collection_days_df['Total_Files'].max() - collection_days_df['Total_Files'].min()
                }
                
                return stats
            
            # Calculate statistics for each period
            stats_2122 = calc_stats_for_period(df, '2021-2022')
            stats_2223 = calc_stats_for_period(df, '2022-2023')
            stats_overall = calc_stats_for_period(df, 'Overall')
            
            # Create the worksheet
            ws = workbook.create_sheet("Summary Statistics")
            
            # Title
            ws['A1'] = "AR Data Analysis - Summary Statistics"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Headers
            headers = ['Metric', '2021-2022', '2022-2023', 'Overall']
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header)
            self.formatter.apply_header_style(ws, f'A3:D3')
            
            # Data rows
            metrics = [
                ('Total Files', 'total_files'),
                ('MP3 Files', 'mp3_files'),
                ('JPG Files', 'jpg_files'),
                ('Total Size (MB)', 'total_size_mb'),
                ('Collection Days', 'collection_days'),
                ('Mean Files per Day', 'mean_files_per_day'),
                ('Median Files per Day', 'median_files_per_day'),
                ('Std Dev Files per Day', 'std_files_per_day'),
                ('Min Files per Day', 'min_files_per_day'),
                ('Max Files per Day', 'max_files_per_day'),
                ('Range Files per Day', 'range_files_per_day')
            ]
            
            for row, (label, key) in enumerate(metrics, 4):
                ws.cell(row=row, column=1, value=label)
                
                # Format values appropriately
                for col, stats in enumerate([stats_2122, stats_2223, stats_overall], 2):
                    if key in stats:
                        value = stats[key]
                        if key in ['mean_files_per_day', 'median_files_per_day', 'std_files_per_day']:
                            ws.cell(row=row, column=col, value=round(value, 1))
                        elif key == 'total_size_mb':
                            ws.cell(row=row, column=col, value=round(value, 2))
                        else:
                            ws.cell(row=row, column=col, value=value)
                    else:
                        ws.cell(row=row, column=col, value='N/A')
            
            # Apply formatting
            self.formatter.apply_data_style(ws, f'A4:D{3 + len(metrics)}')
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, 3 + len(metrics), 1, 4)
            
            # Add enhanced day analysis tables
            current_row = self._add_day_analysis_tables(ws, 3 + len(metrics) + 3)
            
            self.formatter.auto_adjust_columns(ws)
            
            print("[SUCCESS] Summary Statistics sheet created with day analysis")
            
        except Exception as e:
            print(f"[ERROR] Failed to create Summary Statistics sheet: {e}")
    
    def _add_day_analysis_tables(self, ws, start_row):
        """
        Add comprehensive day analysis tables to the Summary Statistics sheet.
        Only includes cleaned data (is_collection_day=TRUE, Is_outlier=FALSE).
        """
        try:
            # Get cleaned data with proper filtering
            cleaned_pipeline = [
                {"$match": {
                    "file_type": {"$in": ["JPG", "MP3"]},
                    "is_collection_day": True,
                    "Outlier_Status": False
                }},
                {"$group": {
                    "_id": {
                        "date": "$ISO_Date",
                        "school_year": "$School_Year",
                        "period": "$Collection_Period",
                        "month": {"$dateToString": {"format": "%Y-%m", "date": {"$dateFromString": {"dateString": "$ISO_Date"}}}}
                    },
                    "Total_Files": {"$sum": 1},
                    "MP3_Files": {"$sum": {"$cond": [{"$eq": ["$file_type", "MP3"]}, 1, 0]}},
                    "JPG_Files": {"$sum": {"$cond": [{"$eq": ["$file_type", "JPG"]}, 1, 0]}}
                }},
                {"$sort": {"_id.date": 1}}
            ]
            
            cleaned_df = self._run_aggregation(cleaned_pipeline)
            if cleaned_df.empty:
                print("[WARNING] No cleaned data found for day analysis")
                return start_row
            
            # Process the cleaned data
            cleaned_df['Date'] = pd.to_datetime(cleaned_df['date'])
            cleaned_df['School_Year'] = cleaned_df['school_year']
            cleaned_df['Period'] = cleaned_df['period']
            cleaned_df['Month'] = cleaned_df['month']
            
            current_row = start_row
            
            # Table 1: School Year Summary
            current_row = self._create_school_year_summary_table(ws, cleaned_df, current_row)
            current_row += 3  # Add spacing
            
            # Table 2: Period Breakdown
            current_row = self._create_period_breakdown_table(ws, cleaned_df, current_row)
            current_row += 3  # Add spacing
            
            # Table 3: Monthly Breakdown
            current_row = self._create_monthly_breakdown_table(ws, cleaned_df, current_row)
            
            return current_row
            
        except Exception as e:
            print(f"[ERROR] Failed to create day analysis tables: {e}")
            return start_row
    
    def _create_school_year_summary_table(self, ws, df, start_row):
        """
        Create the school year summary table for day analysis.
        """
        # Section title
        ws.cell(row=start_row, column=1, value="Day Analysis - School Year Summary (Collection Days Only, Non-Outliers)")
        self.formatter.apply_title_style(ws, f'A{start_row}')
        start_row += 2
        
        # Headers
        headers = ['Metric', '2021-2022', '2022-2023', 'Overall']
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
        self.formatter.apply_header_style(ws, f'A{start_row}:D{start_row}')
        start_row += 1
        
        # Calculate metrics for each school year
        def calc_year_metrics(df, school_year=None):
            if school_year:
                year_df = df[df['School_Year'] == school_year].copy()
            else:
                year_df = df.copy()
            
            if year_df.empty:
                return {}
            
            # Get collection days for the year by summing all periods
            if school_year:
                from utils.calendar import calculate_collection_days_for_period
                if school_year == '2021-2022':
                    # Sum all periods for 2021-2022 school year
                    collection_days = (
                        calculate_collection_days_for_period('SY 21-22 P1') +
                        calculate_collection_days_for_period('SY 21-22 P2') +
                        calculate_collection_days_for_period('SY 21-22 P3')
                    )
                else:  # 2022-2023
                    # Sum all periods for 2022-2023 school year
                    collection_days = (
                        calculate_collection_days_for_period('SY 22-23 P1') +
                        calculate_collection_days_for_period('SY 22-23 P2') +
                        calculate_collection_days_for_period('SY 22-23 P3')
                    )
            else:
                from utils.calendar import calculate_collection_days_for_period
                # Sum all periods for both school years
                collection_days = (
                    calculate_collection_days_for_period('SY 21-22 P1') +
                    calculate_collection_days_for_period('SY 21-22 P2') +
                    calculate_collection_days_for_period('SY 21-22 P3') +
                    calculate_collection_days_for_period('SY 22-23 P1') +
                    calculate_collection_days_for_period('SY 22-23 P2') +
                    calculate_collection_days_for_period('SY 22-23 P3')
                )
            
            # Calculate metrics
            unique_dates = year_df['Date'].dt.date.unique()
            days_with_data = len(unique_dates)
            days_with_zero = collection_days - days_with_data
            coverage_pct = (days_with_data / collection_days * 100) if collection_days > 0 else 0
            
            total_files = year_df['Total_Files'].sum()
            avg_files_per_collection_day = total_files / collection_days if collection_days > 0 else 0
            avg_files_per_data_day = year_df['Total_Files'].mean() if not year_df.empty else 0
            avg_files_including_zeros = total_files / collection_days if collection_days > 0 else 0
            
            # Calculate consecutive days
            consecutive_data, consecutive_zero = self._calculate_consecutive_days(year_df, collection_days)
            
            return {
                'collection_days': collection_days,
                'days_with_data': days_with_data,
                'days_with_zero': days_with_zero,
                'coverage_pct': coverage_pct,
                'avg_files_per_collection_day': avg_files_per_collection_day,
                'avg_files_per_data_day': avg_files_per_data_day,
                'avg_files_including_zeros': avg_files_including_zeros,
                'max_consecutive_data': consecutive_data,
                'max_consecutive_zero': consecutive_zero
            }
        
        # Calculate for each year
        metrics_2122 = calc_year_metrics(df, '2021-2022')
        metrics_2223 = calc_year_metrics(df, '2022-2023')
        metrics_overall = calc_year_metrics(df)
        
        # Data rows
        metrics = [
            ('Total Collection Days', 'collection_days'),
            ('Days with Data', 'days_with_data'),
            ('Days with Zero Files', 'days_with_zero'),
            ('Data Coverage %', 'coverage_pct'),
            ('Avg Files per Collection Day', 'avg_files_per_collection_day'),
            ('Avg Files per Day with Data', 'avg_files_per_data_day'),
            ('Avg Files per Day (incl. zeros)', 'avg_files_including_zeros'),
            ('Max Consecutive Days with Data', 'max_consecutive_data'),
            ('Max Consecutive Days without Data', 'max_consecutive_zero')
        ]
        
        for row_offset, (label, key) in enumerate(metrics):
            row = start_row + row_offset
            ws.cell(row=row, column=1, value=label)
            
            # Format values appropriately
            for col, stats in enumerate([metrics_2122, metrics_2223, metrics_overall], 2):
                if key in stats and stats[key] is not None:
                    value = stats[key]
                    if key == 'coverage_pct':
                        # Store as numeric value (decimal) and apply percentage formatting
                        cell = ws.cell(row=row, column=col)
                        cell.value = value / 100  # Convert percentage to decimal for Excel
                        cell.number_format = '0.0%'  # Apply Excel percentage formatting
                    elif key in ['avg_files_per_collection_day', 'avg_files_per_data_day', 'avg_files_including_zeros']:
                        ws.cell(row=row, column=col, value=round(value, 1))
                    else:
                        ws.cell(row=row, column=col, value=int(value) if isinstance(value, (int, float)) else value)
                else:
                    ws.cell(row=row, column=col, value='N/A')
        
        # Apply formatting
        end_row = start_row + len(metrics) - 1
        self.formatter.apply_data_style(ws, f'A{start_row}:D{end_row}')
        self.formatter.apply_alternating_row_colors(ws, start_row, end_row, 1, 4)
        
        return end_row + 1

def _create_period_breakdown_table(self, ws, df, start_row):
    """
    Create the period breakdown table for day analysis.
    """
    # Section title
    ws.cell(row=start_row, column=1, value="Day Analysis - Period Breakdown (Collection Days Only, Non-Outliers)")
    self.formatter.apply_title_style(ws, f'A{start_row}')
    start_row += 2
    
    # Headers
    headers = ['Period', 'Collection Days', 'Days with Data', 'Zero Days', 'Coverage %', 'Avg Files/Day', 'Max Consec. Data', 'Max Consec. Zero']
    for col, header in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=header)
    self.formatter.apply_header_style(ws, f'A{start_row}:H{start_row}')
    start_row += 1
    
    # Get all periods
    from utils.calendar import get_all_periods
    periods = get_all_periods()
    
    # Sort periods logically
    period_order = ['SY 21-22 P1', 'SY 21-22 P2', 'SY 21-22 P3', 'SY 22-23 P1', 'SY 22-23 P2', 'SY 22-23 P3']
    periods = [p for p in period_order if p in periods]
    
    for row_offset, period in enumerate(periods):
        row = start_row + row_offset
        period_df = df[df['Period'] == period].copy()
        
        # Calculate period metrics
        from utils.calendar import calculate_collection_days_for_period
        collection_days = calculate_collection_days_for_period(period)
        
        if not period_df.empty:
            unique_dates = period_df['Date'].dt.date.unique()
            days_with_data = len(unique_dates)
            days_with_zero = collection_days - days_with_data
            coverage_pct = (days_with_data / collection_days * 100) if collection_days > 0 else 0
            
            total_files = year_df['Total_Files'].sum()
            avg_files_per_collection_day = total_files / collection_days if collection_days > 0 else 0
            avg_files_per_data_day = year_df['Total_Files'].mean() if not year_df.empty else 0
            avg_files_including_zeros = total_files / collection_days if collection_days > 0 else 0
            
            # Calculate consecutive days
            consecutive_data, consecutive_zero = self._calculate_consecutive_days(year_df, collection_days)
            
            return {
                'collection_days': collection_days,
                'days_with_data': days_with_data,
                'days_with_zero': days_with_zero,
                'coverage_pct': coverage_pct,
                'avg_files_per_collection_day': avg_files_per_collection_day,
                'avg_files_per_data_day': avg_files_per_data_day,
                'avg_files_including_zeros': avg_files_including_zeros,
                'max_consecutive_data': consecutive_data,
                'max_consecutive_zero': consecutive_zero
            }
        
        # Calculate for each year
        metrics_2122 = calc_year_metrics(df, '2021-2022')
        metrics_2223 = calc_year_metrics(df, '2022-2023')
        metrics_overall = calc_year_metrics(df)
        
        # Data rows
        metrics = [
            ('Total Collection Days', 'collection_days'),
            ('Days with Data', 'days_with_data'),
            ('Days with Zero Files', 'days_with_zero'),
            ('Data Coverage %', 'coverage_pct'),
            ('Avg Files per Collection Day', 'avg_files_per_collection_day'),
            ('Avg Files per Day with Data', 'avg_files_per_data_day'),
            ('Avg Files per Day (incl. zeros)', 'avg_files_including_zeros'),
            ('Max Consecutive Days with Data', 'max_consecutive_data'),
            ('Max Consecutive Days without Data', 'max_consecutive_zero')
        ]
        
        for row_offset, (label, key) in enumerate(metrics):
            row = start_row + row_offset
            ws.cell(row=row, column=1, value=label)
            
            # Format values appropriately
            for col, stats in enumerate([metrics_2122, metrics_2223, metrics_overall], 2):
                if key in stats and stats[key] is not None:
                    value = stats[key]
                    if key == 'coverage_pct':
                        # Store as numeric value (decimal) and apply percentage formatting
                        cell = ws.cell(row=row, column=col)
                        cell.value = value / 100  # Convert percentage to decimal for Excel
                        cell.number_format = '0.0%'  # Apply Excel percentage formatting
                    elif key in ['avg_files_per_collection_day', 'avg_files_per_data_day', 'avg_files_including_zeros']:
                        ws.cell(row=row, column=col, value=round(value, 1))
                    else:
                        ws.cell(row=row, column=col, value=int(value) if isinstance(value, (int, float)) else value)
                else:
                    ws.cell(row=row, column=col, value='N/A')
        
        # Apply formatting
        end_row = start_row + len(metrics) - 1
        self.formatter.apply_data_style(ws, f'A{start_row}:D{end_row}')
        self.formatter.apply_alternating_row_colors(ws, start_row, end_row, 1, 4)
        
        return end_row + 1
    
    def _create_period_breakdown_table(self, ws, df, start_row):
        """
        Create the period breakdown table for day analysis.
        """
        # Section title
        ws.cell(row=start_row, column=1, value="Day Analysis - Period Breakdown (Collection Days Only, Non-Outliers)")
        self.formatter.apply_title_style(ws, f'A{start_row}')
        start_row += 2
        
        # Headers
        headers = ['Period', 'Collection Days', 'Days with Data', 'Zero Days', 'Coverage %', 'Avg Files/Day', 'Max Consec. Data', 'Max Consec. Zero']
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
        self.formatter.apply_header_style(ws, f'A{start_row}:H{start_row}')
        start_row += 1
        
        # Get all periods
        from utils.calendar import get_all_periods
        periods = get_all_periods()
        
        # Sort periods logically
        period_order = ['SY 21-22 P1', 'SY 21-22 P2', 'SY 21-22 P3', 'SY 22-23 P1', 'SY 22-23 P2', 'SY 22-23 P3']
        periods = [p for p in period_order if p in periods]
        
        for row_offset, period in enumerate(periods):
            row = start_row + row_offset
            period_df = df[df['Period'] == period].copy()
            
            # Calculate period metrics
            from utils.calendar import calculate_collection_days_for_period
            collection_days = calculate_collection_days_for_period(period)
            
            if not period_df.empty:
                unique_dates = period_df['Date'].dt.date.unique()
                days_with_data = len(unique_dates)
                days_with_zero = collection_days - days_with_data
                coverage_pct = (days_with_data / collection_days * 100) if collection_days > 0 else 0
                avg_files_per_day = period_df['Total_Files'].mean()
                consecutive_data, consecutive_zero = self._calculate_consecutive_days(period_df, collection_days)
            else:
                days_with_data = 0
                days_with_zero = collection_days
                coverage_pct = 0
                avg_files_per_day = 0
                consecutive_data = 0
                consecutive_zero = collection_days
            
            # Fill row data
            ws.cell(row=row, column=1, value=period)
            ws.cell(row=row, column=2, value=collection_days)
            ws.cell(row=row, column=3, value=days_with_data)
            ws.cell(row=row, column=4, value=days_with_zero)
            # Store coverage percentage as numeric value with Excel formatting
            cell = ws.cell(row=row, column=5)
            cell.value = coverage_pct / 100  # Convert percentage to decimal for Excel
            cell.number_format = '0.0%'  # Apply Excel percentage formatting
            ws.cell(row=row, column=6, value=round(avg_files_per_day, 1))
            ws.cell(row=row, column=7, value=consecutive_data)
            ws.cell(row=row, column=8, value=consecutive_zero)
        
        # Apply formatting
        end_row = start_row + len(periods) - 1
        self.formatter.apply_data_style(ws, f'A{start_row}:H{end_row}')
        self.formatter.apply_alternating_row_colors(ws, start_row, end_row, 1, 8)
        
        return end_row + 1
    
    def _create_monthly_breakdown_table(self, ws, df, start_row):
        """
        Create the monthly breakdown table for day analysis.
        """
        # Section title
        ws.cell(row=start_row, column=1, value="Day Analysis - Monthly Breakdown (Collection Days Only, Non-Outliers)")
        self.formatter.apply_title_style(ws, f'A{start_row}')
        start_row += 2
        
        # Headers
        headers = ['Month', 'School Year', 'Collection Days', 'Days with Data', 'Zero Days', 'Coverage %', 'Avg Files/Day']
        for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
        self.formatter.apply_header_style(ws, f'A{start_row}:G{start_row}')
        start_row += 1
        
        # Group by month and school year
        monthly_data = []
        for month in sorted(df['Month'].unique()):
            month_df = df[df['Month'] == month].copy()
            if not month_df.empty:
                school_year = month_df['School_Year'].iloc[0]
                unique_dates = month_df['Date'].dt.date.unique()
                days_with_data = len(unique_dates)
                avg_files_per_day = month_df['Total_Files'].mean()
                
                # Calculate collection days for this month (approximate)
                month_date = pd.to_datetime(month + '-01')
                month_name = month_date.strftime('%b %Y')
                
                # Estimate collection days in month (weekdays excluding holidays)
                import calendar
                year = month_date.year
                month_num = month_date.month
                days_in_month = calendar.monthrange(year, month_num)[1]
                
                # Count weekdays in month (rough estimate)
                weekdays_in_month = 0
                for day in range(1, days_in_month + 1):
                    date_obj = datetime(year, month_num, day)
                    if date_obj.weekday() < 5:  # Monday to Friday
                        weekdays_in_month += 1
                
                # Use actual collection days from data or estimate
                collection_days_estimate = min(weekdays_in_month, 23)  # Max ~23 collection days per month
                days_with_zero = max(0, collection_days_estimate - days_with_data)
                coverage_pct = (days_with_data / collection_days_estimate * 100) if collection_days_estimate > 0 else 0
                
                monthly_data.append({
                    'month': month_name,
                    'school_year': school_year,
                    'collection_days': collection_days_estimate,
                    'days_with_data': days_with_data,
                    'days_with_zero': days_with_zero,
                    'coverage_pct': coverage_pct,
                    'avg_files_per_day': avg_files_per_day
                })
        
        # Fill monthly data
        for row_offset, month_data in enumerate(monthly_data):
            row = start_row + row_offset
            ws.cell(row=row, column=1, value=month_data['month'])
            ws.cell(row=row, column=2, value=month_data['school_year'])
            ws.cell(row=row, column=3, value=month_data['collection_days'])
            ws.cell(row=row, column=4, value=month_data['days_with_data'])
            ws.cell(row=row, column=5, value=month_data['days_with_zero'])
            # Store coverage percentage as numeric value with Excel formatting
            cell = ws.cell(row=row, column=6)
            cell.value = month_data['coverage_pct'] / 100  # Convert percentage to decimal for Excel
            cell.number_format = '0.0%'  # Apply Excel percentage formatting
            ws.cell(row=row, column=7, value=round(month_data['avg_files_per_day'], 1))
        
        # Apply formatting
        end_row = start_row + len(monthly_data) - 1
        self.formatter.apply_data_style(ws, f'A{start_row}:G{end_row}')
        self.formatter.apply_alternating_row_colors(ws, start_row, end_row, 1, 7)
        
        return end_row + 1
    
    def _calculate_consecutive_days(self, df, total_collection_days):
        """
        Calculate maximum consecutive days with and without data.
        """
        try:
            if df.empty:
                return 0, total_collection_days
            
            # Get unique dates with data
            dates_with_data = set(df['Date'].dt.date)
            
            # Create a simple consecutive day calculation
            # This is a simplified version - in a full implementation,
            # you'd want to consider the actual collection day calendar
            
            sorted_dates = sorted(dates_with_data)
            if not sorted_dates:
                return 0, total_collection_days
            
            # Calculate consecutive days with data
            max_consecutive_data = 1
            current_consecutive_data = 1
            
            for i in range(1, len(sorted_dates)):
                if (sorted_dates[i] - sorted_dates[i-1]).days == 1:
                    current_consecutive_data += 1
                    max_consecutive_data = max(max_consecutive_data, current_consecutive_data)
                else:
                    current_consecutive_data = 1
            
            # Estimate consecutive days without data (simplified)
            max_consecutive_zero = max(1, total_collection_days - len(dates_with_data))
            
            return max_consecutive_data, max_consecutive_zero
            
        except Exception as e:
            print(f"[WARNING] Error calculating consecutive days: {e}")
            return 0, 0

    def create_raw_data_sheet(self, workbook):
        """
        Creates the Raw Data sheet.
        """
        try:
            # Get all records
            collection = self.db['media_records']
            cursor = collection.find({"file_type": {"$in": ["JPG", "MP3"]}}).limit(10000)
            df = pd.DataFrame(list(cursor))
            
            if df.empty:
                print("[WARNING] No data found for Raw Data sheet")
                return
            
            # Create worksheet
            ws = workbook.create_sheet("Raw Data")
            
            # Add title
            ws['A1'] = "AR Data Analysis - Raw Data Sample"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Add data starting from row 3
            for col, column_name in enumerate(df.columns, 1):
                ws.cell(row=3, column=col, value=str(column_name))
            
            # Apply header formatting
            from openpyxl.utils import get_column_letter
            last_col_letter = get_column_letter(len(df.columns))
            self.formatter.apply_header_style(ws, f'A3:{last_col_letter}3')
            
            # Add data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), 4):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=str(value))
            
            # Apply data formatting
            self.formatter.apply_data_style(ws, f'A4:{last_col_letter}{3 + len(df)}')
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, 3 + len(df), 1, len(df.columns))
            
            self.formatter.auto_adjust_columns(ws)
            
            print(f"[SUCCESS] Raw Data sheet created with {len(df)} records")
            
        except Exception as e:
            print(f"[ERROR] Failed to create Raw Data sheet: {e}")


    def create_data_cleaning_sheet(self, workbook):
        """
        Creates the Data Cleaning sheet.
        """
        try:
            # Get outlier data
            pipeline = [
                {"$match": {"file_type": {"$in": ["JPG", "MP3"]}}},
                {"$group": {
                    "_id": {
                        "outlier_status": "$Outlier_Status",
                        "file_type": "$file_type"
                    },
                    "count": {"$sum": 1},
                    "total_size_mb": {"$sum": "$File_Size_MB"},
                    "avg_size_mb": {"$avg": "$File_Size_MB"}
                }},
                {"$sort": {"_id.file_type": 1, "_id.outlier_status": 1}}
            ]
            
            df = self._run_aggregation(pipeline)
            if df.empty:
                print("[WARNING] No data found for Data Cleaning sheet")
                return
            
            # Create worksheet
            ws = workbook.create_sheet("Data Cleaning")
            
            # Title
            ws['A1'] = "AR Data Analysis - Data Quality & Cleaning Report"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Section 1: Outlier Analysis
            ws['A3'] = "Outlier Analysis"
            self.formatter.apply_section_header_style(ws, 'A3')
            
            # Headers for outlier table
            outlier_headers = ['File Type', 'Outlier Status', 'Count', 'Total Size (MB)', 'Avg Size (MB)']
            for col, header in enumerate(outlier_headers, 1):
                ws.cell(row=5, column=col, value=header)
            self.formatter.apply_header_style(ws, 'A5:E5')
            
            # Outlier data
            row = 6
            for _, record in df.iterrows():
                ws.cell(row=row, column=1, value=record.get('file_type', 'Unknown'))
                ws.cell(row=row, column=2, value=record.get('outlier_status', 'Unknown'))
                ws.cell(row=row, column=3, value=record.get('count', 0))
                ws.cell(row=row, column=4, value=round(record.get('total_size_mb', 0), 2))
                ws.cell(row=row, column=5, value=round(record.get('avg_size_mb', 0), 2))
                row += 1
            
            # Apply formatting
            self.formatter.apply_data_style(ws, f'A6:E{row-1}')
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 6, row-1, 1, 5)
            
            # Section 2: Data Quality Metrics
            ws[f'A{row+2}'] = "Data Quality Metrics"
            self.formatter.apply_section_header_style(ws, f'A{row+2}')
            
            # Get quality metrics
            total_records = self.db['media_records'].count_documents({"file_type": {"$in": ["JPG", "MP3"]}})
            outlier_records = self.db['media_records'].count_documents({
                "file_type": {"$in": ["JPG", "MP3"]},
                "Outlier_Status": True
            })
            valid_records = total_records - outlier_records
            
            quality_metrics = [
                ('Total Records', total_records),
                ('Valid Records', valid_records),
                ('Outlier Records', outlier_records),
                ('Data Quality Rate', (valid_records/total_records) if total_records > 0 else "N/A")
            ]
            
            metrics_row = row + 4
            for i, (metric, value) in enumerate(quality_metrics):
                ws.cell(row=metrics_row + i, column=1, value=metric)
                cell = ws.cell(row=metrics_row + i, column=2)
                if metric == 'Data Quality Rate' and isinstance(value, (int, float)):
                    # Store as numeric value and apply percentage formatting
                    cell.value = value  # Already a decimal (0.0 to 1.0)
                    cell.number_format = '0.0%'  # Apply Excel percentage formatting
                else:
                    cell.value = value
            
            # Apply formatting
            self.formatter.apply_data_style(ws, f'A{metrics_row}:B{metrics_row + len(quality_metrics) - 1}')
            self.formatter.auto_adjust_columns(ws)
            
            print("[SUCCESS] Data Cleaning sheet created")
            
        except Exception as e:
            print(f"[ERROR] Failed to create Data Cleaning sheet: {e}")
