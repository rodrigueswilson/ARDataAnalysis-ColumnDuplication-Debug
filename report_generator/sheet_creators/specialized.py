"""
Specialized Analysis Sheet Creator Module
========================================

This module contains specialized sheet creators for advanced analysis sheets
such as Audio Efficiency Details and MP3 Duration Analysis.
"""

import pandas as pd
from .base import BaseSheetCreator
from utils.data_cleaning import DataCleaningUtils


class SpecializedSheetCreator(BaseSheetCreator):
    """
    Handles creation of specialized analysis sheets with custom formatting
    and multi-table layouts.
    """
    
    def create_audio_efficiency_details_sheet(self, workbook):
        """
        Creates the Audio Efficiency Details sheet with enhanced filtering and positioning.
        Uses DataCleaningUtils for consistent filtering across the system.
        
        Args:
            workbook: openpyxl workbook object
        """
        try:
            # Initialize DataCleaningUtils if not already done
            if not hasattr(self, 'data_cleaning_utils'):
                try:
                    # Use existing database connection or create a new one
                    if hasattr(self, 'db'):
                        db = self.db
                    elif hasattr(self, 'get_db_connection') and self.get_db_connection is not None:
                        db = self.get_db_connection()
                    else:
                        from db_utils import get_db_connection
                        db = get_db_connection()
                    
                    self.data_cleaning_utils = DataCleaningUtils(db)
                except Exception as e:
                    print(f"[WARNING] Failed to initialize DataCleaningUtils: {e}")
                    # Fallback to original implementation if initialization fails
                    return self._create_audio_efficiency_details_sheet_legacy(workbook)
            
            # Get base pipeline for both filters (is_collection_day=TRUE, Outlier_Status=FALSE)
            # and add MP3 file type filter
            base_pipeline = self.data_cleaning_utils.get_both_pipeline()
            # Add MP3-specific filter
            mp3_filter = [{"$match": {"file_type": "MP3"}}]
            efficiency_pipeline = mp3_filter + base_pipeline + [
                {"$group": {
                    "_id": {
                        "Collection_Period": "$Collection_Period",
                        "School_Year": "$School_Year"
                    },
                    "Total_MP3_Files": {"$sum": 1},
                    "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
                    "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"},
                    "Days_With_MP3": {"$addToSet": "$ISO_Date"}
                }},
                {"$addFields": {
                    "Days_With_MP3": {"$size": "$Days_With_MP3"},
                    "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]},
                    "Avg_Files_Per_Day": {"$divide": ["$Total_MP3_Files", {"$size": "$Days_With_MP3"}]},
                    "Period_Efficiency": {"$multiply": [
                        {"$divide": ["$Total_Duration_Seconds", 3600]}, 100
                    ]}
                }},
                {"$sort": {
                    "_id.School_Year": 1,
                    "_id.Collection_Period": 1
                }}
            ]
            
            df = self._run_aggregation(pipeline)
            if df.empty:
                print("[WARNING] No data found for Audio Efficiency Details")
                return
            
            # Create worksheet
            ws = workbook.create_sheet("Audio Efficiency Details")
            
            # Title
            ws['A1'] = "AR Data Analysis - Audio Efficiency Details"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Headers
            headers = [
                'Collection Period', 'School Year', 'Total MP3 Files', 'Total Duration (Hours)',
                'Avg Duration per File', 'Days with MP3', 'Avg Files per Day', 'Period Efficiency'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header)
            
            self.formatter.apply_header_style(ws, 'A3:H3')
            
            # Helper function to convert seconds to HH:MM:SS format
            def seconds_to_hms(seconds):
                if pd.isna(seconds) or seconds == 0:
                    return "00:00:00"
                hours = int(seconds // 3600)
                minutes = int((seconds % 3600) // 60)
                seconds = int(seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Helper function to convert decimal hours to HH:MM:SS format
            def hours_to_hms(decimal_hours):
                if pd.isna(decimal_hours) or decimal_hours == 0:
                    return "00:00:00"
                total_seconds = decimal_hours * 3600
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Add data rows
            row = 4
            for _, record in df.iterrows():
                period = record.get('Collection_Period', 'Unknown')
                school_year = record.get('School_Year', 'Unknown')
                files = record.get('Total_MP3_Files', 0)
                duration_hours = record.get('Total_Duration_Hours', 0)
                avg_duration = record.get('Avg_Duration_Seconds', 0)
                days_with_mp3 = record.get('Days_With_MP3', 0)
                avg_files_per_day = record.get('Avg_Files_Per_Day', 0)
                period_efficiency = record.get('Period_Efficiency', 0)
                
                ws.cell(row=row, column=1, value=period)
                ws.cell(row=row, column=2, value=school_year)
                ws.cell(row=row, column=3, value=files)
                ws.cell(row=row, column=4, value=hours_to_hms(duration_hours))
                ws.cell(row=row, column=5, value=seconds_to_hms(avg_duration))
                ws.cell(row=row, column=6, value=days_with_mp3)
                ws.cell(row=row, column=7, value=round(avg_files_per_day, 1))
                ws.cell(row=row, column=8, value=round(period_efficiency, 1))
                
                row += 1
            
            # Apply formatting
            self.formatter.apply_data_style(ws, f'A4:H{row-1}')
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, row-1, 1, 8)
            
            self.formatter.auto_adjust_columns(ws)
            
            print("[SUCCESS] Audio Efficiency Details sheet created")
            
        except Exception as e:
            print(f"[ERROR] Failed to create Audio Efficiency Details sheet: {e}")

    def _create_audio_efficiency_details_sheet_legacy(self, workbook):
        """
        Legacy implementation of the Audio Efficiency Details sheet.
        Used as a fallback if DataCleaningUtils initialization fails.
        
        Args:
            workbook: openpyxl workbook object
        """
        try:
            # Get audio efficiency data using legacy pipeline
            pipeline = [
                {"$match": {
                    "file_type": "MP3",
                    "is_collection_day": True,
                    "Outlier_Status": False
                }},
                {"$group": {
                    "_id": {
                        "Collection_Period": "$Collection_Period",
                        "School_Year": "$School_Year"
                    },
                    "Total_MP3_Files": {"$sum": 1},
                    "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
                    "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"},
                    "Days_With_MP3": {"$addToSet": "$ISO_Date"}
                }},
                {"$addFields": {
                    "Days_With_MP3": {"$size": "$Days_With_MP3"},
                    "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]},
                    "Avg_Files_Per_Day": {"$divide": ["$Total_MP3_Files", {"$size": "$Days_With_MP3"}]},
                    "Period_Efficiency": {"$multiply": [
                        {"$divide": ["$Total_Duration_Seconds", 3600]}, 100
                    ]}
                }},
                {"$sort": {
                    "_id.School_Year": 1,
                    "_id.Collection_Period": 1
                }}
            ]
            
            df = self._run_aggregation(pipeline)
            if df.empty:
                print("[WARNING] No data found for Audio Efficiency Details (legacy)")
                return
            
            # Create worksheet
            ws = workbook.create_sheet("Audio Efficiency Details")
            
            # Title
            ws['A1'] = "AR Data Analysis - Audio Efficiency Details (Legacy)"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Headers
            headers = [
                'Collection Period', 'School Year', 'Total MP3 Files', 'Total Duration (Hours)',
                'Avg Duration per File', 'Days with MP3', 'Avg Files per Day', 'Period Efficiency'
            ]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=3, column=col, value=header)
            
            self.formatter.apply_header_style(ws, 'A3:H3')
            
            # Helper function to convert seconds to HH:MM:SS format
            def seconds_to_hms(seconds):
                if pd.isna(seconds) or seconds == 0:
                    return "00:00:00"
                hours = int(seconds // 3600)
                minutes = int((seconds % 3600) // 60)
                seconds = int(seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Helper function to convert decimal hours to HH:MM:SS format
            def hours_to_hms(decimal_hours):
                if pd.isna(decimal_hours) or decimal_hours == 0:
                    return "00:00:00"
                total_seconds = decimal_hours * 3600
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Add data rows
            row = 4
            for _, record in df.iterrows():
                period = record.get('Collection_Period', 'Unknown')
                school_year = record.get('School_Year', 'Unknown')
                files = record.get('Total_MP3_Files', 0)
                duration_hours = record.get('Total_Duration_Hours', 0)
                avg_duration = record.get('Avg_Duration_Seconds', 0)
                days_with_mp3 = record.get('Days_With_MP3', 0)
                avg_files_per_day = record.get('Avg_Files_Per_Day', 0)
                period_efficiency = record.get('Period_Efficiency', 0)
                
                ws.cell(row=row, column=1, value=period)
                ws.cell(row=row, column=2, value=school_year)
                ws.cell(row=row, column=3, value=files)
                ws.cell(row=row, column=4, value=hours_to_hms(duration_hours))
                ws.cell(row=row, column=5, value=seconds_to_hms(avg_duration))
                ws.cell(row=row, column=6, value=days_with_mp3)
                ws.cell(row=row, column=7, value=round(avg_files_per_day, 1))
                ws.cell(row=row, column=8, value=round(period_efficiency, 1))
                
                row += 1
            
            # Apply formatting
            self.formatter.apply_data_style(ws, f'A4:H{row-1}')
            
            # Apply alternating row colors for better readability
            self.formatter.apply_alternating_row_colors(ws, 4, row-1, 1, 8)
            
            self.formatter.auto_adjust_columns(ws)
            
            print("[SUCCESS] Audio Efficiency Details sheet created (legacy)")
            
        except Exception as e:
            print(f"[ERROR] Failed to create Audio Efficiency Details sheet (legacy): {e}")

    def create_mp3_duration_analysis_sheet(self, workbook):
        """
        Creates comprehensive MP3 Duration Analysis sheet with multiple tables.
        Uses DataCleaningUtils for consistent filtering across the system.
        
        Args:
            workbook: openpyxl workbook object
        """
        try:
            # Initialize DataCleaningUtils if not already done
            if not hasattr(self, 'data_cleaning_utils'):
                try:
                    # Use existing database connection or create a new one
                    if hasattr(self, 'db'):
                        db = self.db
                    elif hasattr(self, 'get_db_connection') and self.get_db_connection is not None:
                        db = self.get_db_connection()
                    else:
                        from db_utils import get_db_connection
                        db = get_db_connection()
                    
                    self.data_cleaning_utils = DataCleaningUtils(db)
                except Exception as e:
                    print(f"[WARNING] Failed to initialize DataCleaningUtils: {e}")
                    # Fallback to original implementation if initialization fails
                    return self._create_mp3_duration_analysis_sheet_legacy(workbook)
            
            try:
                # Import our refactored MP3 analysis pipelines
                from pipelines.mp3_analysis import MP3_PIPELINES
                
                print("[DEBUG] Using refactored MP3 analysis pipelines")
                
                # Use our properly structured pipelines
                school_year_pipeline = MP3_PIPELINES["MP3_DURATION_BY_SCHOOL_YEAR"]
                period_pipeline = MP3_PIPELINES["MP3_DURATION_BY_PERIOD"]
                monthly_pipeline = MP3_PIPELINES["MP3_DURATION_BY_MONTH"]
                
                # Execute pipelines with debug information
                print("[DEBUG] Executing MP3 duration pipelines...")
                school_year_df = self._run_aggregation(school_year_pipeline)
                print(f"[DEBUG] School year pipeline returned {len(school_year_df)} rows")
                
                period_df = self._run_aggregation(period_pipeline)
                print(f"[DEBUG] Period pipeline returned {len(period_df)} rows")
                
                monthly_df = self._run_aggregation(monthly_pipeline)
                print(f"[DEBUG] Monthly pipeline returned {len(monthly_df)} rows")
                
                if school_year_df.empty and period_df.empty and monthly_df.empty:
                    print("[WARNING] No data found for MP3 Duration Analysis")
                    print("[DEBUG] Falling back to legacy method...")
                    return self._create_mp3_duration_analysis_sheet_legacy(workbook)
                
                # Create worksheet
                ws = workbook.create_sheet("MP3 Duration Analysis")
                
                # Title
                ws['A1'] = "AR Data Analysis - MP3 Duration Analysis"
                self.formatter.apply_title_style(ws, 'A1')
                
                # Helper function to convert seconds to HH:MM:SS format
                def seconds_to_hms(seconds):
                    if pd.isna(seconds) or seconds == 0:
                        return "00:00:00"
                    hours = int(seconds // 3600)
                    minutes = int((seconds % 3600) // 60)
                    seconds = int(seconds % 60)
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                # Helper function to convert decimal hours to HH:MM:SS format
                def hours_to_hms(decimal_hours):
                    if pd.isna(decimal_hours) or decimal_hours == 0:
                        return "00:00:00"
                    total_seconds = decimal_hours * 3600
                    hours = int(total_seconds // 3600)
                    minutes = int((total_seconds % 3600) // 60)
                    seconds = int(total_seconds % 60)
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                
                current_row = 3
                
                # Add School Year Duration Summary table
                if not school_year_df.empty:
                    current_row = self._add_duration_summary_table(ws, school_year_df, current_row, seconds_to_hms, hours_to_hms)
                    current_row += 3  # Add spacing
                
                # Add Collection Period Duration table
                if not period_df.empty:
                    current_row = self._add_period_duration_table(ws, period_df, current_row, seconds_to_hms, hours_to_hms)
                    current_row += 3  # Add spacing
                
                # Add Monthly Duration Distribution table
                if not monthly_df.empty:
                    current_row = self._add_monthly_duration_table(ws, monthly_df, current_row, seconds_to_hms, hours_to_hms)
                
                # Tables now handle their own totals internally for correct positioning
                print("[SUCCESS] MP3 Duration Analysis tables created with integrated totals")
                
                # Apply formatting and auto-adjust columns
                self.formatter.auto_adjust_columns(ws)
                
                print("[SUCCESS] MP3 Duration Analysis sheet created")
                
            except Exception as e:
                print(f"[ERROR] Failed to create MP3 Duration Analysis sheet: {e}")
                print("[DEBUG] Falling back to legacy method...")
                return self._create_mp3_duration_analysis_sheet_legacy(workbook)
                
        except Exception as e:
            print(f"[ERROR] Failed to create MP3 Duration Analysis sheet: {e}")
            print("[DEBUG] Falling back to legacy method...")
            return self._create_mp3_duration_analysis_sheet_legacy(workbook)
            
    def _create_mp3_duration_analysis_sheet_legacy(self, workbook):
        """
        Legacy implementation of MP3 Duration Analysis sheet.
        Used as a fallback if DataCleaningUtils initialization fails.
        
        Args:
            workbook: openpyxl workbook object
        """
        try:
            # Get MP3 duration data by school year
            school_year_pipeline = [
                {"$match": {
                    "file_type": "MP3",
                    "is_collection_day": True,
                    "Outlier_Status": False
                }},
                {"$group": {
                    "_id": "$School_Year",
                    "Total_MP3_Files": {"$sum": 1},
                    "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
                    "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"},
                    "Min_Duration_Seconds": {"$min": "$Duration_Seconds"},
                    "Max_Duration_Seconds": {"$max": "$Duration_Seconds"}
                }},
                {"$addFields": {
                    "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]}
                }},
                {"$sort": {"_id": 1}}
            ]
            
            # Get MP3 duration data by period
            period_pipeline = [
                {"$match": {
                    "file_type": "MP3",
                    "is_collection_day": True,
                    "Outlier_Status": False
                }},
                {"$group": {
                    "_id": {
                        "Collection_Period": "$Collection_Period",
                        "School_Year": "$School_Year"
                    },
                    "Total_MP3_Files": {"$sum": 1},
                    "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
                    "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"},
                    "Days_With_MP3": {"$addToSet": "$ISO_Date"}
                }},
                {"$addFields": {
                    "Days_With_MP3": {"$size": "$Days_With_MP3"},
                    "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]},
                    "Avg_Files_Per_Day": {"$divide": ["$Total_MP3_Files", {"$size": "$Days_With_MP3"}]},
                    "Period_Efficiency": {"$multiply": [
                        {"$divide": ["$Total_Duration_Seconds", 3600]}, 100
                    ]}
                }},
                {"$sort": {
                    "_id.School_Year": 1,
                    "_id.Collection_Period": 1
                }}
            ]
            
            # Get MP3 duration data by month
            monthly_pipeline = [
                {"$match": {
                    "file_type": "MP3",
                    "is_collection_day": True,
                    "Outlier_Status": False
                }},
                {"$group": {
                    "_id": {
                        "Month": "$ISO_Month",
                        "School_Year": "$School_Year"
                    },
                    "Total_MP3_Files": {"$sum": 1},
                    "Total_Duration_Seconds": {"$sum": "$Duration_Seconds"},
                    "Avg_Duration_Seconds": {"$avg": "$Duration_Seconds"}
                }},
                {"$addFields": {
                    "Total_Duration_Hours": {"$divide": ["$Total_Duration_Seconds", 3600]}
                }},
                {"$sort": {
                    "_id.School_Year": 1,
                    "_id.Month": 1
                }}
            ]
            
            # Execute pipelines
            school_year_df = self._run_aggregation(school_year_pipeline)
            period_df = self._run_aggregation(period_pipeline)
            monthly_df = self._run_aggregation(monthly_pipeline)
            
            if school_year_df.empty and period_df.empty and monthly_df.empty:
                print("[WARNING] No data found for MP3 Duration Analysis (legacy)")
                return
            
            # Create worksheet
            ws = workbook.create_sheet("MP3 Duration Analysis")
            
            # Title
            ws['A1'] = "AR Data Analysis - MP3 Duration Analysis (Legacy)"
            self.formatter.apply_title_style(ws, 'A1')
            
            # Helper function to convert seconds to HH:MM:SS format
            def seconds_to_hms(seconds):
                if pd.isna(seconds) or seconds == 0:
                    return "00:00:00"
                hours = int(seconds // 3600)
                minutes = int((seconds % 3600) // 60)
                seconds = int(seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Helper function to convert decimal hours to HH:MM:SS format
            def hours_to_hms(decimal_hours):
                if pd.isna(decimal_hours) or decimal_hours == 0:
                    return "00:00:00"
                total_seconds = decimal_hours * 3600
                hours = int(total_seconds // 3600)
                minutes = int((total_seconds % 3600) // 60)
                seconds = int(total_seconds % 60)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            current_row = 3
            
            # Add School Year Duration Summary table
            if not school_year_df.empty:
                current_row = self._add_duration_summary_table(ws, school_year_df, current_row, seconds_to_hms, hours_to_hms)
                current_row += 3  # Add spacing
            
            # Add Collection Period Duration table
            if not period_df.empty:
                current_row = self._add_period_duration_table(ws, period_df, current_row, seconds_to_hms, hours_to_hms)
                current_row += 3  # Add spacing
            
            # Add Monthly Duration Distribution table
            if not monthly_df.empty:
                self._add_monthly_duration_table(ws, monthly_df, current_row, seconds_to_hms, hours_to_hms)
            
            # Apply formatting and auto-adjust columns
            self.formatter.auto_adjust_columns(ws)
            
            print("[SUCCESS] MP3 Duration Analysis sheet created (legacy)")
            
        except Exception as e:
            print(f"[ERROR] Failed to create MP3 Duration Analysis sheet (legacy): {e}")

    def _add_duration_summary_table(self, ws, df, start_row, seconds_to_hms, hours_to_hms):
        """
        Adds the School Year Duration Summary table with proper header formatting.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="School Year MP3 Duration Summary")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "School Year", "Total Files", "Total Hours", "Avg Duration", 
            "Min Duration", "Max Duration"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        
        # Apply standard header formatting for consistency
        self.formatter.apply_header_style(ws, f'A{current_row}:F{current_row}')
        current_row += 1
        
        # Debug: Print DataFrame info to understand why 2022-2023 might be missing
        print(f"[DEBUG] Duration Summary DataFrame shape: {df.shape}")
        print(f"[DEBUG] DataFrame columns: {list(df.columns)}")
        for idx, row in df.iterrows():
            print(f"[DEBUG] Row {idx}: {dict(row)}")
        
        # Data rows
        for idx, row in df.iterrows():
            # More robust data field access for flattened MongoDB aggregation results
            # Try multiple possible field names for school year
            school_year = None
            for field in ['School_Year', '_id', '_id.School_Year']:
                if field in row and pd.notna(row[field]):
                    school_year = row[field]
                    break
            
            # Handle nested _id case
            if isinstance(school_year, dict):
                school_year = school_year.get('School_Year', 'Unknown')
            elif school_year is None:
                school_year = 'Unknown'
            
            # Get numeric data with robust field access
            files = row.get('Total_MP3_Files', row.get('total_files', 0))
            hours = row.get('Total_Duration_Hours', row.get('total_hours', 0))
            avg_duration = row.get('Avg_Duration_Seconds', row.get('avg_duration', 0))
            min_duration = row.get('Min_Duration_Seconds', row.get('min_duration', 0))
            max_duration = row.get('Max_Duration_Seconds', row.get('max_duration', 0))
            
            print(f"[DEBUG] Processing row {idx}: School Year={school_year}, Files={files}, Hours={hours}")
            print(f"[DEBUG] Writing to Excel row {current_row}")
            
            # Write data to Excel with explicit value conversion
            ws.cell(row=current_row, column=1, value=str(school_year))
            ws.cell(row=current_row, column=2, value=int(files) if files else 0)
            ws.cell(row=current_row, column=3, value=hours_to_hms(float(hours)) if hours else "00:00:00")
            ws.cell(row=current_row, column=4, value=seconds_to_hms(float(avg_duration)) if avg_duration else "00:00")
            ws.cell(row=current_row, column=5, value=seconds_to_hms(float(min_duration)) if min_duration else "00:00")
            ws.cell(row=current_row, column=6, value=seconds_to_hms(float(max_duration)) if max_duration else "00:00")
            
            print(f"[DEBUG] Successfully wrote row {current_row} for {school_year}")
            current_row += 1
        
        # Apply alternating row colors to this table
        if current_row > start_row + 3:  # Only if we have data rows
            self.formatter.apply_alternating_row_colors(ws, start_row + 3, current_row - 1, 1, 6)
        
        # Add totals row immediately after data rows
        if not df.empty:
            # Calculate totals
            total_files = df['Total_MP3_Files'].sum()
            total_hours = df['Total_Duration_Hours'].sum()
            avg_duration = df['Avg_Duration_Seconds'].mean()
            min_duration = df['Min_Duration_Seconds'].min()
            max_duration = df['Max_Duration_Seconds'].max()
            
            # Add totals row
            ws.cell(row=current_row, column=1, value='Total School Year Summary')
            ws.cell(row=current_row, column=2, value=int(total_files) if total_files else 0)
            ws.cell(row=current_row, column=3, value=hours_to_hms(float(total_hours)) if total_hours else "00:00:00")
            ws.cell(row=current_row, column=4, value=seconds_to_hms(float(avg_duration)) if avg_duration else "00:00")
            ws.cell(row=current_row, column=5, value=seconds_to_hms(float(min_duration)) if min_duration else "00:00")
            ws.cell(row=current_row, column=6, value=seconds_to_hms(float(max_duration)) if max_duration else "00:00")
            
            # Apply professional totals formatting
            from openpyxl.styles import Font, PatternFill
            total_font = Font(bold=True, color='FFFFFF')
            total_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            for col in range(1, 7):
                cell = ws.cell(row=current_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
            
            current_row += 1
        
        return current_row

    def _add_period_duration_table(self, ws, df, start_row, seconds_to_hms, hours_to_hms):
        """
        Adds the Collection Period Duration table.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="Collection Period MP3 Duration Analysis")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "Period", "School Year", "Files", "Total Hours", "Avg Duration",
            "Days with MP3", "Avg Files/Day", "Efficiency"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        
        # Apply standard header formatting for consistency
        self.formatter.apply_header_style(ws, f'A{current_row}:H{current_row}')
        current_row += 1
        
        # Data rows
        for _, row in df.iterrows():
            # Handle both flattened and non-flattened _id field access
            if 'Collection_Period' in row:
                period = row.get('Collection_Period', 'Unknown')
                school_year = row.get('School_Year', 'Unknown')
            else:
                # Handle nested _id structure
                id_data = row.get('_id', {})
                if isinstance(id_data, dict):
                    period = id_data.get('Collection_Period', 'Unknown')
                    school_year = id_data.get('School_Year', 'Unknown')
                else:
                    period = 'Unknown'
                    school_year = 'Unknown'
            
            files = row.get('Total_MP3_Files', 0)
            duration_hours = row.get('Total_Duration_Hours', 0)
            avg_duration = row.get('Avg_Duration_Seconds', 0)
            days_with_mp3 = row.get('Days_With_MP3', 0)
            avg_files_per_day = row.get('Avg_Files_Per_Day', 0)
            period_efficiency = row.get('Period_Efficiency', 0)
            
            ws.cell(row=current_row, column=1, value=period)
            ws.cell(row=current_row, column=2, value=school_year)
            ws.cell(row=current_row, column=3, value=files)
            ws.cell(row=current_row, column=4, value=hours_to_hms(duration_hours))
            ws.cell(row=current_row, column=5, value=seconds_to_hms(avg_duration))
            ws.cell(row=current_row, column=6, value=days_with_mp3)
            ws.cell(row=current_row, column=7, value=avg_files_per_day)
            ws.cell(row=current_row, column=8, value=period_efficiency)
            
            current_row += 1
        
        # Apply alternating row colors to this table
        if current_row > start_row + 3:  # Only if we have data rows
            self.formatter.apply_alternating_row_colors(ws, start_row + 3, current_row - 1, 1, 8)
        
        # Add totals row immediately after data rows
        if not df.empty:
            # Calculate totals
            total_files = df['Total_MP3_Files'].sum()
            total_hours = df['Total_Duration_Hours'].sum()
            avg_duration = df['Avg_Duration_Seconds'].mean()
            total_days = df['Days_With_MP3'].sum()
            avg_files_per_day = df['Avg_Files_Per_Day'].mean()
            avg_efficiency = df['Period_Efficiency'].mean()
            
            # Add totals row
            ws.cell(row=current_row, column=1, value='Total Period Breakdown')
            ws.cell(row=current_row, column=2, value='')  # No school year for totals
            ws.cell(row=current_row, column=3, value=int(total_files) if total_files else 0)
            ws.cell(row=current_row, column=4, value=hours_to_hms(float(total_hours)) if total_hours else "00:00:00")
            ws.cell(row=current_row, column=5, value=seconds_to_hms(float(avg_duration)) if avg_duration else "00:00")
            ws.cell(row=current_row, column=6, value=int(total_days) if total_days else 0)
            ws.cell(row=current_row, column=7, value=round(float(avg_files_per_day), 2) if avg_files_per_day else 0.0)
            ws.cell(row=current_row, column=8, value=round(float(avg_efficiency), 2) if avg_efficiency else 0.0)
            
            # Apply professional totals formatting
            from openpyxl.styles import Font, PatternFill
            total_font = Font(bold=True, color='FFFFFF')
            total_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
            
            current_row += 1
        
        return current_row
    
    def _add_monthly_duration_table(self, ws, df, start_row, seconds_to_hms, hours_to_hms):
        """
        Adds the Monthly Duration Distribution table.
        """
        # Table title
        ws.cell(row=start_row, column=1, value="Monthly MP3 Duration Distribution")
        current_row = start_row + 2
        
        # Headers
        headers = [
            "Month", "2021-22 Files", "2021-22 Hours", "2022-23 Files", 
            "2022-23 Hours", "Total Files", "Total Hours", "Avg Duration"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=current_row, column=col, value=header)
        
        # Apply standard header formatting for consistency
        self.formatter.apply_header_style(ws, f'A{current_row}:H{current_row}')
        current_row += 1
        
        # Process data by month - using school year order (Sep-May)
        month_names = [
            "Jan", "Feb", "Mar", "Apr", "May", "Jun",
            "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"
        ]
        
        # School year month order: Sep(9), Oct(10), Nov(11), Dec(12), Jan(1), Feb(2), Mar(3), Apr(4), May(5)
        school_year_month_order = [9, 10, 11, 12, 1, 2, 3, 4, 5]
        
        # Group data by month
        monthly_data = {}
        # Process monthly data
        for _, row in df.iterrows():
            # Handle both flattened and non-flattened _id field access
            if 'Month' in row:
                month_num = row.get('Month', 0)
                school_year = row.get('School_Year', 'Unknown')
            else:
                # Handle nested _id structure
                id_data = row.get('_id', {})
                if isinstance(id_data, dict):
                    month_num = id_data.get('Month', 0)
                    school_year = id_data.get('School_Year', 'Unknown')
                else:
                    month_num = 0
                    school_year = 'Unknown'
            
            if month_num not in monthly_data:
                monthly_data[month_num] = {}
            
            monthly_data[month_num][school_year] = {
                'files': row.get('Total_MP3_Files', 0),
                'hours': row.get('Total_Duration_Hours', 0),
                'avg_duration': row.get('Avg_Duration_Seconds', 0)
            }
        
        # Write monthly rows in school year order (Sep-May)
        for month_num in school_year_month_order:
            if month_num not in monthly_data:
                continue
                
            month_name = month_names[month_num - 1]
            month_data = monthly_data[month_num]
            
            # Get data for each school year
            sy_2122 = month_data.get('2021-2022', {'files': 0, 'hours': 0, 'avg_duration': 0})
            sy_2223 = month_data.get('2022-2023', {'files': 0, 'hours': 0, 'avg_duration': 0})
            
            total_files = sy_2122['files'] + sy_2223['files']
            total_hours = sy_2122['hours'] + sy_2223['hours']
            avg_duration = (sy_2122['avg_duration'] + sy_2223['avg_duration']) / 2 if (sy_2122['avg_duration'] > 0 and sy_2223['avg_duration'] > 0) else max(sy_2122['avg_duration'], sy_2223['avg_duration'])
            
            ws.cell(row=current_row, column=1, value=month_name)
            ws.cell(row=current_row, column=2, value=sy_2122['files'])
            ws.cell(row=current_row, column=3, value=hours_to_hms(sy_2122['hours']))
            ws.cell(row=current_row, column=4, value=sy_2223['files'])
            ws.cell(row=current_row, column=5, value=hours_to_hms(sy_2223['hours']))
            ws.cell(row=current_row, column=6, value=total_files)
            ws.cell(row=current_row, column=7, value=hours_to_hms(total_hours))
            ws.cell(row=current_row, column=8, value=seconds_to_hms(avg_duration))
            
            current_row += 1
        
        # Apply alternating row colors to this table
        if current_row > start_row + 3:  # Only if we have data rows
            self.formatter.apply_alternating_row_colors(ws, start_row + 3, current_row - 1, 1, 8)
        
        # Add totals row immediately after data rows
        if monthly_data:
            # Calculate totals from monthly_data
            total_2122_files = sum(month_data.get('2021-2022', {}).get('files', 0) for month_data in monthly_data.values())
            total_2122_hours = sum(month_data.get('2021-2022', {}).get('hours', 0) for month_data in monthly_data.values())
            total_2223_files = sum(month_data.get('2022-2023', {}).get('files', 0) for month_data in monthly_data.values())
            total_2223_hours = sum(month_data.get('2022-2023', {}).get('hours', 0) for month_data in monthly_data.values())
            total_files = total_2122_files + total_2223_files
            total_hours = total_2122_hours + total_2223_hours
            
            # Calculate average duration from original df
            avg_duration = df['Avg_Duration_Seconds'].mean() if not df.empty else 0
            
            # Add totals row
            ws.cell(row=current_row, column=1, value='Total Monthly Distribution')
            ws.cell(row=current_row, column=2, value=int(total_2122_files) if total_2122_files else 0)
            ws.cell(row=current_row, column=3, value=hours_to_hms(float(total_2122_hours)) if total_2122_hours else "00:00:00")
            ws.cell(row=current_row, column=4, value=int(total_2223_files) if total_2223_files else 0)
            ws.cell(row=current_row, column=5, value=hours_to_hms(float(total_2223_hours)) if total_2223_hours else "00:00:00")
            ws.cell(row=current_row, column=6, value=int(total_files) if total_files else 0)
            ws.cell(row=current_row, column=7, value=hours_to_hms(float(total_hours)) if total_hours else "00:00:00")
            ws.cell(row=current_row, column=8, value=seconds_to_hms(float(avg_duration)) if avg_duration else "00:00")
            
            # Apply professional totals formatting
            from openpyxl.styles import Font, PatternFill
            total_font = Font(bold=True, color='FFFFFF')
            total_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            
            for col in range(1, 9):
                cell = ws.cell(row=current_row, column=col)
                cell.font = total_font
                cell.fill = total_fill
            
            current_row += 1
        
        return current_row


