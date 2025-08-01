"""
Sheet Validation System
======================

This module provides comprehensive runtime validation for sheet creation,
preventing duplicates and ensuring data integrity throughout the report generation process.
"""

import openpyxl
from typing import Dict, List, Set, Any, Optional, Tuple
from dataclasses import dataclass
from enum import Enum
import re

from .sheet_registry import get_sheet_registry, SheetType


class ValidationSeverity(Enum):
    """Severity levels for validation issues."""
    INFO = "INFO"
    WARNING = "WARNING"
    ERROR = "ERROR"
    CRITICAL = "CRITICAL"


@dataclass
class ValidationIssue:
    """Represents a validation issue found during sheet validation."""
    severity: ValidationSeverity
    category: str
    message: str
    sheet_name: Optional[str] = None
    suggestion: Optional[str] = None


class SheetValidator:
    """
    Comprehensive validation system for sheet creation and workbook integrity.
    
    Provides multiple levels of validation:
    1. Pre-creation validation (prevent duplicates)
    2. Post-creation validation (verify integrity)
    3. Workbook-level validation (overall consistency)
    4. Configuration validation (ensure proper setup)
    """
    
    def __init__(self):
        self.registry = get_sheet_registry()
        self.validation_issues: List[ValidationIssue] = []
        self.reserved_sheet_names = {
            "Sheet", "Sheet1", "Sheet2", "Sheet3"  # Excel default names
        }
    
    def validate_pre_creation(self, workbook: openpyxl.Workbook, 
                            sheet_name: str) -> Tuple[bool, List[ValidationIssue]]:
        """
        Validate before creating a sheet to prevent issues.
        
        Args:
            workbook: openpyxl workbook object
            sheet_name: Proposed name for the new sheet
            
        Returns:
            tuple: (is_valid, list_of_issues)
        """
        issues = []
        
        # Check for duplicate names in workbook
        if sheet_name in workbook.sheetnames:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                category="Duplicate Sheet",
                message=f"Sheet '{sheet_name}' already exists in workbook",
                sheet_name=sheet_name,
                suggestion="Use a different name or check if this is intentional duplication"
            ))
        
        # Check for duplicate names in registry
        if self.registry.is_sheet_created(sheet_name):
            record = self.registry.get_creation_record(sheet_name)
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                category="Registry Duplicate",
                message=f"Sheet '{sheet_name}' already registered by {record.creator_module}.{record.creator_method}",
                sheet_name=sheet_name,
                suggestion="This indicates a code-level duplication that should be fixed"
            ))
        
        # Check for reserved names
        if sheet_name in self.reserved_sheet_names:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Reserved Name",
                message=f"Sheet name '{sheet_name}' conflicts with Excel reserved names",
                sheet_name=sheet_name,
                suggestion="Consider using a more descriptive name"
            ))
        
        # Check name length (Excel limit is 31 characters)
        if len(sheet_name) > 31:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                category="Name Length",
                message=f"Sheet name '{sheet_name}' exceeds Excel's 31-character limit ({len(sheet_name)} chars)",
                sheet_name=sheet_name,
                suggestion="Shorten the sheet name"
            ))
        
        # Check for invalid characters
        invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
        found_invalid = [char for char in invalid_chars if char in sheet_name]
        if found_invalid:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                category="Invalid Characters",
                message=f"Sheet name '{sheet_name}' contains invalid characters: {found_invalid}",
                sheet_name=sheet_name,
                suggestion="Remove invalid characters from the sheet name"
            ))
        
        # Check for leading/trailing spaces
        if sheet_name != sheet_name.strip():
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Whitespace",
                message=f"Sheet name '{sheet_name}' has leading or trailing whitespace",
                sheet_name=sheet_name,
                suggestion="Trim whitespace from sheet name"
            ))
        
        # Check for empty name
        if not sheet_name or sheet_name.strip() == "":
            issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                category="Empty Name",
                message="Sheet name cannot be empty",
                sheet_name=sheet_name,
                suggestion="Provide a valid sheet name"
            ))
        
        # Determine if creation should proceed
        has_errors = any(issue.severity in [ValidationSeverity.ERROR, ValidationSeverity.CRITICAL] 
                        for issue in issues)
        
        return not has_errors, issues
    
    def validate_post_creation(self, workbook: openpyxl.Workbook, 
                             sheet_name: str) -> Tuple[bool, List[ValidationIssue]]:
        """
        Validate after creating a sheet to ensure it was created properly.
        
        Args:
            workbook: openpyxl workbook object
            sheet_name: Name of the sheet that was created
            
        Returns:
            tuple: (is_valid, list_of_issues)
        """
        issues = []
        
        # Verify sheet exists in workbook
        if sheet_name not in workbook.sheetnames:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                category="Creation Failed",
                message=f"Sheet '{sheet_name}' was not found in workbook after creation",
                sheet_name=sheet_name,
                suggestion="Check for exceptions during sheet creation"
            ))
            return False, issues
        
        # Get the worksheet
        ws = workbook[sheet_name]
        
        # Check if sheet is empty (might indicate creation failure)
        if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Empty Sheet",
                message=f"Sheet '{sheet_name}' appears to be empty",
                sheet_name=sheet_name,
                suggestion="Verify that data was properly added to the sheet"
            ))
        
        # Verify registry consistency
        if not self.registry.is_sheet_created(sheet_name):
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Registry Inconsistency",
                message=f"Sheet '{sheet_name}' exists in workbook but not in registry",
                sheet_name=sheet_name,
                suggestion="Sheet may have been created outside the registry system"
            ))
        
        # Check for reasonable data size (basic sanity check)
        if ws.max_row > 100000:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.INFO,
                category="Large Dataset",
                message=f"Sheet '{sheet_name}' has {ws.max_row} rows (very large)",
                sheet_name=sheet_name,
                suggestion="Consider if this data size is expected"
            ))
        
        has_critical_issues = any(issue.severity == ValidationSeverity.CRITICAL 
                                for issue in issues)
        
        return not has_critical_issues, issues
    
    def validate_workbook_integrity(self, workbook: openpyxl.Workbook) -> Tuple[bool, List[ValidationIssue]]:
        """
        Comprehensive validation of the entire workbook.
        
        Args:
            workbook: openpyxl workbook object to validate
            
        Returns:
            tuple: (is_valid, list_of_issues)
        """
        issues = []
        
        # Basic workbook checks
        if len(workbook.sheetnames) == 0:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                category="Empty Workbook",
                message="Workbook contains no sheets",
                suggestion="Ensure at least one sheet is created"
            ))
            return False, issues
        
        # Check for duplicate sheet names (shouldn't happen, but safety check)
        sheet_names = workbook.sheetnames
        duplicate_names = [name for name in sheet_names if sheet_names.count(name) > 1]
        if duplicate_names:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                category="Duplicate Names",
                message=f"Workbook contains duplicate sheet names: {set(duplicate_names)}",
                suggestion="This indicates a serious bug in sheet creation logic"
            ))
        
        # Validate registry consistency
        registry_validation = self.registry.validate_workbook_consistency(workbook)
        if not registry_validation["is_consistent"]:
            if registry_validation["untracked_sheets"]:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.WARNING,
                    category="Untracked Sheets",
                    message=f"Sheets exist in workbook but not in registry: {registry_validation['untracked_sheets']}",
                    suggestion="These sheets were created outside the registry system"
                ))
            
            if registry_validation["missing_sheets"]:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    category="Missing Sheets",
                    message=f"Sheets registered but missing from workbook: {registry_validation['missing_sheets']}",
                    suggestion="Sheet creation may have failed silently"
                ))
        
        # Check for expected core sheets
        expected_core_sheets = ["Dashboard", "Summary Statistics", "Raw Data"]
        missing_core = [sheet for sheet in expected_core_sheets if sheet not in sheet_names]
        if missing_core:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Missing Core Sheets",
                message=f"Expected core sheets not found: {missing_core}",
                suggestion="Verify that core sheet creation completed successfully"
            ))
        
        # Check sheet order (Dashboard should be first)
        if sheet_names and sheet_names[0] != "Dashboard":
            issues.append(ValidationIssue(
                severity=ValidationSeverity.INFO,
                category="Sheet Order",
                message=f"Dashboard is not the first sheet (first is '{sheet_names[0]}')",
                suggestion="Consider reordering sheets for better user experience"
            ))
        
        # Check for reasonable total number of sheets
        if len(sheet_names) > 50:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.WARNING,
                category="Too Many Sheets",
                message=f"Workbook has {len(sheet_names)} sheets (unusually high)",
                suggestion="Consider if all sheets are necessary"
            ))
        elif len(sheet_names) < 5:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.INFO,
                category="Few Sheets",
                message=f"Workbook has only {len(sheet_names)} sheets",
                suggestion="Verify that all expected sheets were created"
            ))
        
        # Check for sheets with suspicious names
        suspicious_patterns = [r"^Sheet\d*$", r"^Worksheet\d*$", r"^Copy of"]
        for sheet_name in sheet_names:
            for pattern in suspicious_patterns:
                if re.match(pattern, sheet_name):
                    issues.append(ValidationIssue(
                        severity=ValidationSeverity.WARNING,
                        category="Suspicious Name",
                        message=f"Sheet '{sheet_name}' has a default/suspicious name",
                        suggestion="Consider using more descriptive sheet names"
                    ))
                    break
        
        has_critical_issues = any(issue.severity == ValidationSeverity.CRITICAL 
                                for issue in issues)
        
        return not has_critical_issues, issues
    
    def validate_configuration(self, config_path: str) -> Tuple[bool, List[ValidationIssue]]:
        """
        Validate the report configuration for potential sheet creation issues.
        
        Args:
            config_path: Path to the report configuration file
            
        Returns:
            tuple: (is_valid, list_of_issues)
        """
        issues = []
        
        try:
            import json
            from pathlib import Path
            
            config_file = Path(config_path)
            if not config_file.exists():
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.CRITICAL,
                    category="Missing Config",
                    message=f"Configuration file not found: {config_path}",
                    suggestion="Ensure the configuration file exists"
                ))
                return False, issues
            
            with open(config_file, 'r') as f:
                config = json.load(f)
            
            # Check for required sections
            if 'sheets' not in config:
                issues.append(ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    category="Missing Config Section",
                    message="Configuration missing 'sheets' section",
                    suggestion="Add sheets configuration section"
                ))
            else:
                sheets = config['sheets']
                
                # Check for duplicate sheet names in config
                sheet_names = [sheet.get('name', '') for sheet in sheets if sheet.get('enabled', True)]
                duplicate_names = [name for name in sheet_names if sheet_names.count(name) > 1]
                if duplicate_names:
                    issues.append(ValidationIssue(
                        severity=ValidationSeverity.ERROR,
                        category="Config Duplicates",
                        message=f"Configuration contains duplicate sheet names: {set(duplicate_names)}",
                        suggestion="Remove duplicate sheet configurations"
                    ))
                
                # Check each sheet configuration
                for i, sheet in enumerate(sheets):
                    if not isinstance(sheet, dict):
                        issues.append(ValidationIssue(
                            severity=ValidationSeverity.ERROR,
                            category="Invalid Config",
                            message=f"Sheet configuration {i} is not a dictionary",
                            suggestion="Fix sheet configuration format"
                        ))
                        continue
                    
                    # Check required fields
                    required_fields = ['name', 'pipeline']
                    for field in required_fields:
                        if field not in sheet:
                            issues.append(ValidationIssue(
                                severity=ValidationSeverity.ERROR,
                                category="Missing Config Field",
                                message=f"Sheet configuration {i} missing required field '{field}'",
                                suggestion=f"Add {field} to sheet configuration"
                            ))
                    
                    # Validate sheet name if present
                    if 'name' in sheet:
                        sheet_name = sheet['name']
                        name_valid, name_issues = self.validate_pre_creation(
                            openpyxl.Workbook(), sheet_name  # Dummy workbook for name validation
                        )
                        issues.extend(name_issues)
        
        except json.JSONDecodeError as e:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.CRITICAL,
                category="Invalid JSON",
                message=f"Configuration file contains invalid JSON: {e}",
                suggestion="Fix JSON syntax in configuration file"
            ))
        except Exception as e:
            issues.append(ValidationIssue(
                severity=ValidationSeverity.ERROR,
                category="Config Error",
                message=f"Error reading configuration: {e}",
                suggestion="Check configuration file format and permissions"
            ))
        
        has_critical_issues = any(issue.severity == ValidationSeverity.CRITICAL 
                                for issue in issues)
        
        return not has_critical_issues, issues
    
    def print_validation_report(self, issues: List[ValidationIssue]) -> None:
        """
        Print a formatted validation report.
        
        Args:
            issues: List of validation issues to report
        """
        if not issues:
            print("✅ No validation issues found")
            return
        
        print("\n" + "="*60)
        print("SHEET VALIDATION REPORT")
        print("="*60)
        
        # Group issues by severity
        by_severity = {}
        for issue in issues:
            if issue.severity not in by_severity:
                by_severity[issue.severity] = []
            by_severity[issue.severity].append(issue)
        
        # Print summary
        print(f"Total Issues: {len(issues)}")
        for severity in ValidationSeverity:
            count = len(by_severity.get(severity, []))
            if count > 0:
                icon = {"CRITICAL": "🔴", "ERROR": "❌", "WARNING": "⚠️", "INFO": "ℹ️"}[severity.value]
                print(f"  {icon} {severity.value}: {count}")
        
        # Print detailed issues
        for severity in [ValidationSeverity.CRITICAL, ValidationSeverity.ERROR, 
                        ValidationSeverity.WARNING, ValidationSeverity.INFO]:
            severity_issues = by_severity.get(severity, [])
            if not severity_issues:
                continue
            
            print(f"\n{severity.value} Issues:")
            for i, issue in enumerate(severity_issues, 1):
                print(f"  {i}. [{issue.category}] {issue.message}")
                if issue.sheet_name:
                    print(f"     Sheet: {issue.sheet_name}")
                if issue.suggestion:
                    print(f"     Suggestion: {issue.suggestion}")
                print()
        
        print("="*60)
    
    def get_validation_summary(self, issues: List[ValidationIssue]) -> Dict[str, Any]:
        """
        Get a summary of validation results.
        
        Args:
            issues: List of validation issues
            
        Returns:
            dict: Summary of validation results
        """
        by_severity = {}
        by_category = {}
        
        for issue in issues:
            # Count by severity
            if issue.severity not in by_severity:
                by_severity[issue.severity] = 0
            by_severity[issue.severity] += 1
            
            # Count by category
            if issue.category not in by_category:
                by_category[issue.category] = 0
            by_category[issue.category] += 1
        
        return {
            "total_issues": len(issues),
            "by_severity": {sev.value: by_severity.get(sev, 0) for sev in ValidationSeverity},
            "by_category": by_category,
            "has_critical": ValidationSeverity.CRITICAL in by_severity,
            "has_errors": ValidationSeverity.ERROR in by_severity,
            "is_valid": not (ValidationSeverity.CRITICAL in by_severity or ValidationSeverity.ERROR in by_severity)
        }


# Global validator instance
_global_validator = SheetValidator()


def get_sheet_validator() -> SheetValidator:
    """Get the global sheet validator instance."""
    return _global_validator


def validate_sheet_creation(workbook: openpyxl.Workbook, sheet_name: str) -> bool:
    """
    Quick validation function for sheet creation.
    
    Args:
        workbook: openpyxl workbook object
        sheet_name: Name of the sheet to validate
        
    Returns:
        bool: True if validation passes, False otherwise
    """
    validator = get_sheet_validator()
    is_valid, issues = validator.validate_pre_creation(workbook, sheet_name)
    
    if issues:
        validator.print_validation_report(issues)
    
    return is_valid
