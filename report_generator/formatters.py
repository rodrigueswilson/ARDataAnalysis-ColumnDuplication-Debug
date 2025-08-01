"""
Excel Formatting and Styling Utilities
======================================

This module contains all Excel formatting functions, styling utilities,
and worksheet appearance management for the report generator.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

class ExcelFormatter:
    """
    Handles all Excel formatting operations including colors, fonts, borders,
    and cell styling for consistent report appearance.
    """
    
    def __init__(self):
        """Initialize formatter with default color scheme."""
        self.color_scheme = {
            'header_bg': '1F497D', 
            'header_font': 'FFFFFF',
            'alt_row_bg': 'DDEBF7', 
            'total_row_bg': '4F81BD',
            'total_row_font': 'FFFFFF', 
            'chart_series1': '4F81BD',
            'chart_series2': 'C0504D'
        }
    
    def format_sheet(self, ws):
        """
        Applies consistent formatting to a worksheet with alternating row colors 
        and proper column widths.
        
        Args:
            ws: openpyxl worksheet object
        """
        if ws.max_row <= 1:
            return  # Skip empty sheets
        
        # Header formatting
        header_font = Font(bold=True, color=self.color_scheme['header_font'])
        header_fill = PatternFill(start_color=self.color_scheme['header_bg'], 
                                 end_color=self.color_scheme['header_bg'], 
                                 fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Alternating row colors (skip header)
        alt_fill = PatternFill(start_color=self.color_scheme['alt_row_bg'], 
                              end_color=self.color_scheme['alt_row_bg'], 
                              fill_type='solid')
        
        for row_num in range(2, ws.max_row + 1):
            if row_num % 2 == 0:  # Even rows get alternate color
                for cell in ws[row_num]:
                    if cell.fill.start_color.index == '00000000':  # Only if not already filled
                        cell.fill = alt_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_total_row(self, ws, df):
        """
        Adds a formatted total row to the worksheet.
        
        Args:
            ws: openpyxl worksheet object
            df: pandas DataFrame with the data
        """
        if df.empty:
            return
        
        total_row_num = ws.max_row + 1
        
        # Total row formatting
        total_font = Font(bold=True, color=self.color_scheme['total_row_font'])
        total_fill = PatternFill(start_color=self.color_scheme['total_row_bg'], 
                                end_color=self.color_scheme['total_row_bg'], 
                                fill_type='solid')
        
        # Add "TOTAL" label in first column
        ws.cell(row=total_row_num, column=1, value="TOTAL")
        ws.cell(row=total_row_num, column=1).font = total_font
        ws.cell(row=total_row_num, column=1).fill = total_fill
        
        # Calculate totals for numeric columns
        for col_idx, column_name in enumerate(df.columns, start=2):
            cell = ws.cell(row=total_row_num, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            
            try:
                # Check if column is numeric and can be summed
                if pd.api.types.is_numeric_dtype(df[column_name]):
                    total_value = df[column_name].sum()
                    if pd.isna(total_value):
                        cell.value = ""
                    else:
                        cell.value = total_value
                else:
                    cell.value = ""
            except Exception as e:
                # Fallback for any unexpected data type issues
                cell.value = ""
    
    def add_audio_characteristics_total_row(self, ws, df):
        """
        Adds a custom total row for Audio Note Characteristics that only sums 'Per Year' rows.
        
        Args:
            ws: openpyxl worksheet object  
            df: pandas DataFrame with the data
        """
        if df.empty:
            return
        
        # Filter for 'Per Year' rows only
        per_year_df = df[df.iloc[:, 0].str.contains('Per Year', na=False)]
        
        if per_year_df.empty:
            return
        
        total_row_num = ws.max_row + 1
        
        # Total row formatting
        total_font = Font(bold=True, color=self.color_scheme['total_row_font'])
        total_fill = PatternFill(start_color=self.color_scheme['total_row_bg'], 
                                end_color=self.color_scheme['total_row_bg'], 
                                fill_type='solid')
        
        # Add "TOTAL (Per Year Only)" label
        ws.cell(row=total_row_num, column=1, value="TOTAL (Per Year Only)")
        ws.cell(row=total_row_num, column=1).font = total_font
        ws.cell(row=total_row_num, column=1).fill = total_fill
        
        # Calculate totals for numeric columns from per_year rows only
        for col_idx, column_name in enumerate(df.columns, start=2):
            cell = ws.cell(row=total_row_num, column=col_idx)
            cell.font = total_font
            cell.fill = total_fill
            
            if per_year_df[column_name].dtype in ['int64', 'float64']:
                try:
                    total_value = per_year_df[column_name].sum()
                    cell.value = total_value
                except:
                    cell.value = ""
            else:
                cell.value = ""
    
    def add_top_days_table(self, ws, df, start_row):
        """
        Adds a table of top efficiency days to the worksheet at the specified starting row.
        
        Args:
            ws: The worksheet to add the table to
            df: DataFrame containing the top efficiency data
            start_row: The row number to start adding the table at
        """
        if df.empty:
            return
        
        # Add table header
        header_font = Font(bold=True, color=self.color_scheme['header_font'])
        header_fill = PatternFill(start_color=self.color_scheme['header_bg'], 
                                 end_color=self.color_scheme['header_bg'], 
                                 fill_type='solid')
        
        ws.cell(row=start_row, column=1, value="Top 10 Days by Efficiency")
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=1).fill = header_fill
        
        # Add column headers
        headers = ["Date", "Total Files", "Audio Minutes", "Files per Audio Minute"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row + 1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Add data rows
        for row_idx, (_, row) in enumerate(df.head(10).iterrows(), start=start_row + 2):
            ws.cell(row=row_idx, column=1, value=row['_id'])
            ws.cell(row=row_idx, column=2, value=row['total_files_day'])
            ws.cell(row=row_idx, column=3, value=round(row['total_duration_minutes'], 1))
            ws.cell(row=row_idx, column=4, value=round(row['efficiency_files_per_audio_minute'], 2))
            
            # Apply alternating row colors
            if row_idx % 2 == 0:
                alt_fill = PatternFill(start_color=self.color_scheme['alt_row_bg'], 
                                     end_color=self.color_scheme['alt_row_bg'], 
                                     fill_type='solid')
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = alt_fill
    
    def apply_title_style(self, ws, cell_range):
        """Apply title styling to a cell or range."""
        title_font = Font(bold=True, size=14, color='1F497D')
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        if ':' in cell_range:
            # Range of cells
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = title_font
                    cell.alignment = title_alignment
        else:
            # Single cell
            cell = ws[cell_range]
            cell.font = title_font
            cell.alignment = title_alignment
    
    def apply_header_style(self, ws, cell_range):
        """Apply header styling to a cell or range."""
        header_font = Font(bold=True, color=self.color_scheme['header_font'])
        header_fill = PatternFill(start_color=self.color_scheme['header_bg'], 
                                 end_color=self.color_scheme['header_bg'], 
                                 fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        if ':' in cell_range:
            # Range of cells
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        else:
            # Single cell
            cell = ws[cell_range]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def apply_data_style(self, ws, cell_range):
        """Apply data styling to a cell or range with optimized performance for large datasets."""
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        if ':' in cell_range:
            # Parse the range to determine size
            start_cell, end_cell = cell_range.split(':')
            start_col_letter = ''.join(filter(str.isalpha, start_cell))
            start_row = int(''.join(filter(str.isdigit, start_cell)))
            end_col_letter = ''.join(filter(str.isalpha, end_cell))
            end_row = int(''.join(filter(str.isdigit, end_cell)))
            
            # Convert column letters to indices for calculation
            from openpyxl.utils import column_index_from_string
            start_col_num = column_index_from_string(start_col_letter)
            end_col_num = column_index_from_string(end_col_letter)
            
            # Calculate total cells in range
            total_cells = (end_row - start_row + 1) * (end_col_num - start_col_num + 1)
            
            # Use optimized approach for large ranges (>1000 cells)
            if total_cells > 1000:
                print(f"[INFO] Applying optimized styling to large range: {cell_range} ({total_cells} cells)")
                self._apply_bulk_data_style(ws, start_row, end_row, start_col_num, end_col_num)
            else:
                # Use original approach for smaller ranges
                for row in ws[cell_range]:
                    for cell in row:
                        cell.alignment = data_alignment
        else:
            # Single cell
            cell = ws[cell_range]
            cell.alignment = data_alignment
    
    def _apply_bulk_data_style(self, ws, start_row, end_row, start_col, end_col):
        """Apply data styling efficiently to large ranges using batch operations."""
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        # Apply alignment in batches to reduce openpyxl overhead
        batch_size = 100  # Process 100 rows at a time
        
        for batch_start in range(start_row, end_row + 1, batch_size):
            batch_end = min(batch_start + batch_size - 1, end_row)
            
            # Apply styling to this batch
            for row_num in range(batch_start, batch_end + 1):
                for col_num in range(start_col, end_col + 1):
                    try:
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.alignment is None or cell.alignment.horizontal != 'left':
                            cell.alignment = data_alignment
                    except Exception:
                        # Skip problematic cells to prevent crashes
                        continue
            
            # Progress indicator for very large ranges
            if end_row - start_row > 500:
                progress = ((batch_end - start_row + 1) / (end_row - start_row + 1)) * 100
                if progress % 25 == 0 or batch_end == end_row:  # Show progress every 25%
                    print(f"[INFO] Styling progress: {progress:.0f}% complete")
    
    def apply_section_header_style(self, ws, cell_range):
        """Apply section header styling."""
        section_font = Font(bold=True, size=12, color='1F497D')
        section_alignment = Alignment(horizontal='left', vertical='center')
        
        if ':' in cell_range:
            # Range of cells
            for row in ws[cell_range]:
                for cell in row:
                    cell.font = section_font
                    cell.alignment = section_alignment
        else:
            # Single cell
            cell = ws[cell_range]
            cell.font = section_font
            cell.alignment = section_alignment
    
    def apply_alternating_row_colors(self, ws, start_row, end_row, start_col=1, end_col=None):
        """
        Applies alternating row colors to a specific range with performance optimization.
        
        Args:
            ws: openpyxl worksheet object
            start_row: First row to apply alternating colors (inclusive)
            end_row: Last row to apply alternating colors (inclusive)
            start_col: First column to apply colors (default: 1)
            end_col: Last column to apply colors (default: worksheet max column)
        """
        try:
            if end_col is None:
                end_col = ws.max_column
            
            if start_row > end_row or start_col > end_col:
                return  # Invalid range
            
            total_rows = end_row - start_row + 1
            total_cells = total_rows * (end_col - start_col + 1)
        
            print(f"[INFO] Applying alternating row colors to range: {get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}")
            print(f"[INFO] Total rows: {total_rows}, Total cells: {total_cells}")
        
            # For very large datasets (>50,000 cells), use a more aggressive optimization
            if total_cells > 50000:
                print(f"[INFO] Very large dataset detected ({total_cells} cells), using aggressive optimization")
                batch_size = 500  # Larger batches for very large datasets
            else:
                batch_size = 100  # Standard batch size
        
            # Use the same color scheme as format_sheet for consistency
            alt_fill = PatternFill(start_color=self.color_scheme['alt_row_bg'], 
                                  end_color=self.color_scheme['alt_row_bg'], 
                                  fill_type='solid')
            
            for batch_start in range(start_row, end_row + 1, batch_size):
                batch_end = min(batch_start + batch_size - 1, end_row)
                
                # Apply alternating colors to this batch
                for row_num in range(batch_start, batch_end + 1):
                    # Apply color to even rows (relative to data start)
                    if (row_num - start_row) % 2 == 1:  # Odd index = even row (since we start at 0)
                        # Apply fill to entire row within the specified column range
                        for col_num in range(start_col, end_col + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            # Only apply if not already formatted (preserve existing formatting)
                            if cell.fill.start_color.index == "00000000":
                                cell.fill = alt_fill
                
                # Progress indicator for very large datasets
                if total_rows > 500 and (batch_end - start_row + 1) % 500 == 0:
                    progress = ((batch_end - start_row + 1) / total_rows) * 100
                    print(f"[INFO] Alternating row colors: {progress:.0f}% complete")
            
            print(f"[INFO] Alternating row colors applied successfully to {total_rows} rows")
            
        except Exception as e:
            print(f"[WARNING] Could not apply alternating row colors: {e}")
            # Fallback to no alternating colors rather than crash
            pass

    def auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
