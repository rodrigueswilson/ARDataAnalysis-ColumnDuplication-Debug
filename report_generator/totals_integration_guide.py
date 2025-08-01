"""
Totals Integration Guide and Helper Functions
=============================================

This module provides integration examples and helper functions for incorporating
the TotalsManager into existing sheet creators in the AR Data Analysis system.
"""

from .totals_manager import TotalsManager, add_table_totals
import pandas as pd
from typing import Dict, List, Optional, Any


class TotalsIntegrationHelper:
    """
    Helper class to demonstrate and facilitate totals integration across
    different sheet types in the AR Data Analysis system.
    """
    
    def __init__(self, formatter=None):
        """Initialize with a TotalsManager instance."""
        self.totals_manager = TotalsManager(formatter)
    
    def generate_recommended_config(self, df: pd.DataFrame, sheet_type: str) -> Dict[str, Any]:
        """
        Generate recommended totals configuration based on DataFrame structure and sheet type.
        
        Args:
            df: DataFrame to analyze
            sheet_type: Type of sheet (daily, weekly, monthly, etc.)
            
        Returns:
            Dictionary with recommended totals configuration
        """
        if df.empty:
            return {'add_totals': False, 'reason': 'Empty DataFrame'}
        
        # Identify numeric columns
        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # Exclude statistical/non-additive columns
        exclude_patterns = ['ACF_', 'PACF_', 'ARIMA', '_Forecast', '_Lower', '_Upper', 
                          'Mean', 'Median', 'Std', 'Correlation', 'P_Value', 'Lag_']
        
        # Filter out non-additive columns
        additive_columns = []
        excluded_columns = []
        
        for col in numeric_columns:
            if any(pattern in str(col) for pattern in exclude_patterns):
                excluded_columns.append(col)
            else:
                additive_columns.append(col)
        
        # Generate configuration based on sheet type and available columns
        if not additive_columns:
            return {
                'add_totals': False,
                'reason': 'No additive numeric columns found',
                'excluded_columns': excluded_columns
            }
        
        # Base configuration
        config = {
            'add_totals': True,
            'add_row_totals': True,
            'add_column_totals': True,
            'add_grand_total': True,
            'include_columns': additive_columns,
            'exclude_columns': excluded_columns,
            'totals_label': 'TOTALS',
            'row_totals_label': 'Row Total'
        }
        
        # Customize based on sheet type
        if sheet_type in ['daily', 'weekly', 'biweekly', 'monthly', 'period']:
            # Time series sheets typically benefit from full totals
            config.update({
                'add_row_totals': True,
                'add_column_totals': True,
                'add_grand_total': True,
                'totals_label': f'{sheet_type.title()} Totals'
            })
        elif sheet_type in ['summary', 'statistics']:
            # Summary sheets may not need row totals
            config.update({
                'add_row_totals': False,
                'add_column_totals': True,
                'add_grand_total': False,
                'totals_label': 'Summary Totals'
            })
        elif sheet_type in ['analysis', 'specialized']:
            # Analysis sheets may have mixed requirements
            config.update({
                'add_row_totals': True,
                'add_column_totals': True,
                'add_grand_total': False,
                'totals_label': 'Analysis Totals'
            })
        
        return config
    
    # ==========================================
    # INTEGRATION EXAMPLES FOR EACH SHEET TYPE
    # ==========================================
    
    def integrate_summary_statistics_totals(self, ws, df: pd.DataFrame, start_row: int = 4) -> Dict[str, Any]:
        """
        Example integration for Summary Statistics sheet.
        
        Args:
            ws: openpyxl worksheet
            df: Summary statistics DataFrame
            start_row: Starting row for data
            
        Returns:
            Dictionary with totals information for validation
        """
        print("[TOTALS] Adding comprehensive totals to Summary Statistics sheet...")
        
        # Define which columns should be included in totals
        numeric_columns = [
            'Total_Files', 'MP3_Files', 'JPG_Files', 'Total_Size_MB',
            'Mean_Files_Per_Day', 'Median_Files_Per_Day', 'Max_Files_Per_Day'
        ]
        
        # Exclude certain columns that shouldn't be totaled
        exclude_columns = ['Std_Files_Per_Day', 'Min_Files_Per_Day']
        
        # Add comprehensive totals
        positions = self.totals_manager.add_comprehensive_totals(
            ws, df, 
            start_row=start_row,
            numeric_columns=numeric_columns,
            exclude_columns=exclude_columns,
            include_row_totals=True,
            include_column_totals=True,
            include_grand_total=True
        )
        
        # Calculate key totals for validation
        totals_data = {
            'Total_Files': df['Total_Files'].sum() if 'Total_Files' in df.columns else 0,
            'MP3_Files': df['MP3_Files'].sum() if 'MP3_Files' in df.columns else 0,
            'JPG_Files': df['JPG_Files'].sum() if 'JPG_Files' in df.columns else 0,
            'sheet_type': 'summary_statistics',
            'positions': positions
        }
        
        # Register for cross-validation
        self.totals_manager.register_sheet_totals('Summary Statistics', totals_data)
        
        print(f"[TOTALS] Summary Statistics totals added at positions: {positions}")
        return totals_data
    
    def integrate_mp3_duration_analysis_totals(self, ws, df: pd.DataFrame, start_row: int = 4) -> Dict[str, Any]:
        """
        Example integration for MP3 Duration Analysis sheet with multiple tables.
        
        Args:
            ws: openpyxl worksheet
            df: MP3 duration DataFrame
            start_row: Starting row for first table
            
        Returns:
            Dictionary with totals information for validation
        """
        print("[TOTALS] Adding totals to MP3 Duration Analysis sheet...")
        
        # For MP3 Duration Analysis, we typically have multiple tables
        # This example shows how to handle the main summary table
        
        numeric_columns = [
            'Total_MP3_Files', 'Total_Duration_Hours', 'Avg_Duration_Seconds',
            'Days_With_MP3', 'Avg_Files_Per_Day'
        ]
        
        # Add row totals (useful for comparing periods/years)
        totals_col = self.totals_manager.add_row_totals_to_worksheet(
            ws, df,
            start_row=start_row,
            numeric_columns=numeric_columns,
            totals_column_header="Period Total"
        )
        
        # Add column totals for overall summary
        totals_row = self.totals_manager.add_column_totals_to_worksheet(
            ws, df,
            start_row=start_row,
            numeric_columns=numeric_columns,
            totals_row_label="OVERALL TOTAL"
        )
        
        # Calculate validation data
        totals_data = {
            'MP3_Files': df['Total_MP3_Files'].sum() if 'Total_MP3_Files' in df.columns else 0,
            'Total_Duration_Hours': df['Total_Duration_Hours'].sum() if 'Total_Duration_Hours' in df.columns else 0,
            'sheet_type': 'mp3_duration_analysis',
            'positions': {'row_totals_column': totals_col, 'column_totals_row': totals_row}
        }
        
        self.totals_manager.register_sheet_totals('MP3 Duration Analysis', totals_data)
        
        print(f"[TOTALS] MP3 Duration Analysis totals added - Row: {totals_col}, Column: {totals_row}")
        return totals_data
    
    def integrate_pipeline_sheet_totals(self, ws, df: pd.DataFrame, sheet_name: str, start_row: int = 2) -> Dict[str, Any]:
        """
        Example integration for pipeline-based sheets (Daily Counts, Weekly Counts, etc.).
        
        Args:
            ws: openpyxl worksheet
            df: Pipeline data DataFrame
            sheet_name: Name of the sheet
            start_row: Starting row for data
            
        Returns:
            Dictionary with totals information for validation
        """
        print(f"[TOTALS] Adding totals to pipeline sheet: {sheet_name}...")
        
        # Auto-detect numeric columns for pipeline sheets
        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        
        # Exclude ACF/PACF and ARIMA columns from totals (they're statistical, not additive)
        exclude_columns = [col for col in df.columns if any(x in col for x in ['ACF_', 'PACF_', 'ARIMA', '_Forecast', '_Lower', '_Upper'])]
        
        # For time series data, column totals are usually more meaningful than row totals
        include_row_totals = len(df) <= 50  # Only add row totals for smaller datasets
        
        positions = self.totals_manager.add_comprehensive_totals(
            ws, df,
            start_row=start_row,
            numeric_columns=numeric_columns,
            exclude_columns=exclude_columns,
            include_row_totals=include_row_totals,
            include_column_totals=True,
            include_grand_total=include_row_totals
        )
        
        # Calculate validation data
        totals_data = {
            'sheet_type': 'pipeline_sheet',
            'sheet_name': sheet_name,
            'positions': positions
        }
        
        # Add specific totals for common fields
        for field in ['Total_Files', 'MP3_Files', 'JPG_Files', 'Total_Size_MB']:
            if field in df.columns:
                totals_data[field] = df[field].sum()
        
        self.totals_manager.register_sheet_totals(sheet_name, totals_data)
        
        print(f"[TOTALS] {sheet_name} totals added at positions: {positions}")
        return totals_data
    
    def integrate_multi_table_sheet_totals(self, ws, tables_data: List[Dict], sheet_name: str) -> Dict[str, Any]:
        """
        Example integration for sheets with multiple tables (like specialized analysis sheets).
        
        Args:
            ws: openpyxl worksheet
            tables_data: List of dictionaries with table information
                        Each dict should have: {'df': DataFrame, 'start_row': int, 'table_name': str}
            sheet_name: Name of the sheet
            
        Returns:
            Dictionary with totals information for validation
        """
        print(f"[TOTALS] Adding totals to multi-table sheet: {sheet_name}...")
        
        all_positions = {}
        combined_totals = {'sheet_type': 'multi_table', 'sheet_name': sheet_name, 'tables': {}}
        
        for i, table_info in enumerate(tables_data):
            df = table_info['df']
            start_row = table_info['start_row']
            table_name = table_info.get('table_name', f'Table_{i+1}')
            
            print(f"[TOTALS] Processing table: {table_name}")
            
            # Add totals for this table
            numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
            
            positions = self.totals_manager.add_comprehensive_totals(
                ws, df,
                start_row=start_row,
                numeric_columns=numeric_columns,
                include_row_totals=True,
                include_column_totals=True,
                include_grand_total=False  # Avoid grand totals in multi-table sheets
            )
            
            all_positions[table_name] = positions
            
            # Calculate table-specific totals
            table_totals = {}
            for col in numeric_columns:
                table_totals[col] = df[col].sum()
            
            combined_totals['tables'][table_name] = table_totals
        
        combined_totals['positions'] = all_positions
        self.totals_manager.register_sheet_totals(sheet_name, combined_totals)
        
        print(f"[TOTALS] {sheet_name} multi-table totals completed for {len(tables_data)} tables")
        return combined_totals
    
    # ==========================================
    # INTEGRATION HELPER METHODS
    # ==========================================
    
    def get_recommended_totals_config(self, df: pd.DataFrame, sheet_type: str) -> Dict[str, Any]:
        """
        Get recommended totals configuration based on DataFrame structure and sheet type.
        
        Args:
            df: Input DataFrame
            sheet_type: Type of sheet ('summary', 'pipeline', 'analysis', 'multi_table')
            
        Returns:
            Dictionary with recommended configuration
        """
        config = {
            'numeric_columns': df.select_dtypes(include=['number']).columns.tolist(),
            'exclude_columns': [],
            'include_row_totals': True,
            'include_column_totals': True,
            'include_grand_total': True
        }
        
        # Adjust based on sheet type
        if sheet_type == 'pipeline':
            # Exclude statistical columns from totals
            config['exclude_columns'] = [col for col in df.columns if any(x in col for x in ['ACF_', 'PACF_', 'ARIMA', '_Forecast'])]
            # For large time series, skip row totals
            config['include_row_totals'] = len(df) <= 50
            
        elif sheet_type == 'summary':
            # Exclude statistical measures that shouldn't be summed
            config['exclude_columns'] = [col for col in df.columns if any(x in col for x in ['Std_', 'Mean_', 'Median_', 'Min_', 'Max_'])]
            
        elif sheet_type == 'analysis':
            # For analysis sheets, usually only column totals make sense
            config['include_row_totals'] = False
            config['include_grand_total'] = False
            
        return config
    
    def validate_and_report(self) -> str:
        """
        Validate all registered totals and generate a comprehensive report.
        
        Returns:
            Validation report string
        """
        return self.totals_manager.generate_totals_summary_report()
    
    def create_totals_configuration_file(self):
        """Create the totals validation rules configuration file."""
        self.totals_manager.create_validation_rules_template()


# ==========================================
# SPECIFIC INTEGRATION FUNCTIONS
# ==========================================

def integrate_totals_into_base_sheet_creator():
    """
    Example of how to integrate totals into the BaseSheetCreator class.
    This shows the code modifications needed.
    """
    integration_code = '''
    # Add to BaseSheetCreator.__init__:
    from .totals_manager import TotalsManager
    
    def __init__(self, db, formatter):
        self.db = db
        self.formatter = formatter
        self.totals_manager = TotalsManager(formatter)  # ADD THIS LINE
        self._pipeline_cache = {}
    
    # Modify create_summary_statistics_sheet method:
    def create_summary_statistics_sheet(self, workbook):
        # ... existing code ...
        
        # ADD TOTALS INTEGRATION BEFORE FINAL FORMATTING:
        if not df.empty:
            totals_data = self.totals_manager.add_comprehensive_totals(
                ws, df, start_row=4,
                numeric_columns=['Total_Files', 'MP3_Files', 'JPG_Files', 'Total_Size_MB'],
                exclude_columns=['Std_Files_Per_Day', 'Min_Files_Per_Day']
            )
            
            # Register for validation
            validation_data = {
                'Total_Files': df['Total_Files'].sum(),
                'MP3_Files': df['MP3_Files'].sum(),
                'JPG_Files': df['JPG_Files'].sum()
            }
            self.totals_manager.register_sheet_totals('Summary Statistics', validation_data)
        
        # ... rest of existing code ...
    '''
    return integration_code


def integrate_totals_into_pipeline_sheet_creator():
    """
    Example of how to integrate totals into the PipelineSheetCreator class.
    """
    integration_code = '''
    # Add to PipelineSheetCreator._create_pipeline_sheet method:
    
    def _create_pipeline_sheet(self, workbook, sheet_config):
        # ... existing code for data processing ...
        
        # ADD TOTALS INTEGRATION BEFORE EXCEL EXPORT:
        if not df.empty and sheet_config.get('add_totals', True):
            print(f"[TOTALS] Adding totals to {sheet_name}")
            
            # Get recommended configuration
            from .totals_integration_guide import TotalsIntegrationHelper
            helper = TotalsIntegrationHelper(self.formatter)
            
            totals_data = helper.integrate_pipeline_sheet_totals(
                ws, df, sheet_name, start_row=2
            )
            
            print(f"[TOTALS] Totals added successfully to {sheet_name}")
        
        # ... rest of existing code ...
    '''
    return integration_code


def integrate_totals_into_specialized_sheet_creator():
    """
    Example of how to integrate totals into the SpecializedSheetCreator class.
    """
    integration_code = '''
    # Add to SpecializedSheetCreator methods:
    
    def create_mp3_duration_analysis_sheet(self, workbook):
        # ... existing code for creating tables ...
        
        # ADD TOTALS TO EACH TABLE:
        from .totals_integration_guide import TotalsIntegrationHelper
        helper = TotalsIntegrationHelper(self.formatter)
        
        # Example for the main summary table
        if not df.empty:
            tables_data = [
                {'df': summary_df, 'start_row': 4, 'table_name': 'Duration Summary'},
                {'df': period_df, 'start_row': 15, 'table_name': 'Period Breakdown'},
                {'df': monthly_df, 'start_row': 30, 'table_name': 'Monthly Distribution'}
            ]
            
            totals_data = helper.integrate_multi_table_sheet_totals(
                ws, tables_data, 'MP3 Duration Analysis'
            )
        
        # ... rest of existing code ...
    '''
    return integration_code


# ==========================================
# CONFIGURATION TEMPLATES
# ==========================================

def create_sheet_totals_configuration():
    """
    Create a configuration template for specifying totals behavior per sheet.
    """
    config_template = {
        "sheet_totals_config": {
            "Summary Statistics": {
                "enabled": True,
                "include_row_totals": True,
                "include_column_totals": True,
                "include_grand_total": True,
                "numeric_columns": ["Total_Files", "MP3_Files", "JPG_Files", "Total_Size_MB"],
                "exclude_columns": ["Std_Files_Per_Day", "Min_Files_Per_Day"]
            },
            "Daily Counts": {
                "enabled": True,
                "include_row_totals": False,
                "include_column_totals": True,
                "include_grand_total": False,
                "exclude_columns": ["ACF_", "PACF_", "ARIMA", "_Forecast"]
            },
            "Weekly Counts": {
                "enabled": True,
                "include_row_totals": True,
                "include_column_totals": True,
                "include_grand_total": True,
                "exclude_columns": ["ACF_", "PACF_", "ARIMA", "_Forecast"]
            },
            "MP3 Duration Analysis": {
                "enabled": True,
                "multi_table": True,
                "table_configs": {
                    "Duration Summary": {"include_row_totals": True, "include_column_totals": True},
                    "Period Breakdown": {"include_row_totals": True, "include_column_totals": False},
                    "Monthly Distribution": {"include_row_totals": False, "include_column_totals": True}
                }
            }
        }
    }
    
    return config_template


# ==========================================
# TESTING AND VALIDATION HELPERS
# ==========================================

def test_totals_integration(sample_df: pd.DataFrame, sheet_name: str = "Test Sheet"):
    """
    Test the totals integration with a sample DataFrame.
    
    Args:
        sample_df: Sample DataFrame for testing
        sheet_name: Name for the test sheet
    """
    from openpyxl import Workbook
    
    print(f"[TEST] Testing totals integration with {sheet_name}")
    print(f"[TEST] DataFrame shape: {sample_df.shape}")
    print(f"[TEST] Numeric columns: {sample_df.select_dtypes(include=['number']).columns.tolist()}")
    
    # Create test workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Add headers
    for col_idx, col_name in enumerate(sample_df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Add data
    for row_idx, (_, row) in enumerate(sample_df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Test totals integration
    helper = TotalsIntegrationHelper()
    totals_data = helper.integrate_pipeline_sheet_totals(ws, sample_df, sheet_name)
    
    print(f"[TEST] Totals integration completed: {totals_data}")
    
    # Generate validation report
    report = helper.validate_and_report()
    print(f"[TEST] Validation report:\n{report}")
    
    return wb, totals_data


if __name__ == "__main__":
    # Example usage and testing
    print("Totals Integration Guide - Example Usage")
    print("=" * 50)
    
    # Create sample data for testing
    sample_data = pd.DataFrame({
        'Date': pd.date_range('2023-01-01', periods=10),
        'Total_Files': [25, 30, 15, 40, 35, 20, 45, 30, 25, 35],
        'MP3_Files': [10, 15, 8, 20, 18, 12, 22, 15, 12, 18],
        'JPG_Files': [15, 15, 7, 20, 17, 8, 23, 15, 13, 17],
        'Total_Size_MB': [125.5, 180.2, 95.3, 220.1, 195.7, 110.8, 245.9, 165.4, 135.2, 185.6]
    })
    
    # Test the integration
    wb, totals_data = test_totals_integration(sample_data, "Sample Test Sheet")
    
    print("\nIntegration test completed successfully!")
    print("Check the generated workbook and validation report above.")
