"""
Sheet Creation Registry
======================

This module provides a centralized registry system to track sheet creation,
prevent duplicates, and enforce single-source-of-truth for each sheet type.
"""

from typing import Dict, List, Optional, Callable, Any
from dataclasses import dataclass
from enum import Enum
import openpyxl


class SheetType(Enum):
    """Enumeration of all possible sheet types in the system."""
    DASHBOARD = "Dashboard"
    ACF_PACF_DASHBOARD = "ACF_PACF_Dashboard"
    SUMMARY_STATISTICS = "Summary Statistics"
    RAW_DATA = "Raw Data"
    DATA_CLEANING = "Data Cleaning"
    MP3_DURATION_ANALYSIS = "MP3 Duration Analysis"
    AUDIO_EFFICIENCY_DETAILS = "Audio Efficiency Details"
    
    # Pipeline-driven sheets (dynamic names)
    DAILY_COUNTS = "Daily Counts"
    WEEKLY_COUNTS = "Weekly Counts"
    BIWEEKLY_COUNTS = "Biweekly Counts"
    MONTHLY_COUNTS = "Monthly Counts"
    PERIOD_COUNTS = "Period Counts"
    
    # ACF/PACF analysis sheets (dynamic names with suffix)
    DAILY_COUNTS_ACF_PACF = "Daily Counts (ACF_PACF)"
    WEEKLY_COUNTS_ACF_PACF = "Weekly Counts (ACF_PACF)"
    BIWEEKLY_COUNTS_ACF_PACF = "Biweekly Counts (ACF_PACF)"
    MONTHLY_COUNTS_ACF_PACF = "Monthly Counts (ACF_PACF)"
    PERIOD_COUNTS_ACF_PACF = "Period Counts (ACF_PACF)"


@dataclass
class SheetCreationRecord:
    """Record of a sheet creation event."""
    sheet_name: str
    sheet_type: Optional[SheetType]
    creator_module: str
    creator_method: str
    position: Optional[int]
    created_successfully: bool
    error_message: Optional[str] = None


class SheetRegistry:
    """
    Centralized registry for tracking sheet creation and preventing duplicates.
    
    This registry enforces single-source-of-truth by:
    1. Tracking which sheets have been created
    2. Preventing duplicate sheet names
    3. Recording creation metadata for debugging
    4. Providing validation and reporting capabilities
    """
    
    def __init__(self):
        self._created_sheets: Dict[str, SheetCreationRecord] = {}
        self._registered_creators: Dict[SheetType, Callable] = {}
        self._creation_order: List[str] = []
    
    def register_creator(self, sheet_type: SheetType, creator_func: Callable, 
                        module_name: str, method_name: str) -> None:
        """
        Register the official creator function for a sheet type.
        
        Args:
            sheet_type: Type of sheet this creator handles
            creator_func: Function that creates the sheet
            module_name: Name of the module containing the creator
            method_name: Name of the creator method
        """
        if sheet_type in self._registered_creators:
            existing = self._registered_creators[sheet_type]
            print(f"[WARNING] Overriding existing creator for {sheet_type.value}")
            print(f"  Previous: {existing.__module__}.{existing.__name__}")
            print(f"  New: {module_name}.{method_name}")
        
        self._registered_creators[sheet_type] = creator_func
        print(f"[REGISTRY] Registered creator for {sheet_type.value}: {module_name}.{method_name}")
    
    def create_sheet(self, workbook: openpyxl.Workbook, sheet_name: str, 
                    sheet_type: Optional[SheetType] = None, 
                    creator_module: str = "unknown", 
                    creator_method: str = "unknown",
                    position: Optional[int] = None,
                    **kwargs) -> bool:
        """
        Create a sheet through the registry with duplicate prevention.
        
        Args:
            workbook: openpyxl workbook object
            sheet_name: Name of the sheet to create
            sheet_type: Type of sheet (if known)
            creator_module: Module name for tracking
            creator_method: Method name for tracking
            position: Position to insert sheet (optional)
            **kwargs: Additional arguments for sheet creation
            
        Returns:
            bool: True if sheet was created successfully, False if duplicate or error
        """
        # Check for duplicates
        if self.is_sheet_created(sheet_name):
            existing = self._created_sheets[sheet_name]
            print(f"[ERROR] Duplicate sheet creation attempted!")
            print(f"  Sheet: {sheet_name}")
            print(f"  Previous creator: {existing.creator_module}.{existing.creator_method}")
            print(f"  Attempted creator: {creator_module}.{creator_method}")
            return False
        
        # Check if sheet already exists in workbook
        if sheet_name in workbook.sheetnames:
            print(f"[ERROR] Sheet '{sheet_name}' already exists in workbook!")
            self._record_creation(sheet_name, sheet_type, creator_module, 
                                creator_method, position, False, 
                                "Sheet already exists in workbook")
            return False
        
        try:
            # Create the sheet
            if position is not None:
                ws = workbook.create_sheet(title=sheet_name, index=position)
            else:
                ws = workbook.create_sheet(title=sheet_name)
            
            # Record successful creation
            self._record_creation(sheet_name, sheet_type, creator_module, 
                                creator_method, position, True)
            
            print(f"[SUCCESS] Sheet '{sheet_name}' created by {creator_module}.{creator_method}")
            return True
            
        except Exception as e:
            # Record failed creation
            self._record_creation(sheet_name, sheet_type, creator_module, 
                                creator_method, position, False, str(e))
            print(f"[ERROR] Failed to create sheet '{sheet_name}': {e}")
            return False
    
    def is_sheet_created(self, sheet_name: str) -> bool:
        """Check if a sheet has been created through the registry."""
        return sheet_name in self._created_sheets
    
    def get_creation_record(self, sheet_name: str) -> Optional[SheetCreationRecord]:
        """Get the creation record for a sheet."""
        return self._created_sheets.get(sheet_name)
    
    def get_all_created_sheets(self) -> List[str]:
        """Get list of all sheets created through the registry."""
        return self._creation_order.copy()
    
    def get_creation_summary(self) -> Dict[str, Any]:
        """Get summary of sheet creation activity."""
        successful = sum(1 for record in self._created_sheets.values() if record.created_successfully)
        failed = len(self._created_sheets) - successful
        
        return {
            "total_attempts": len(self._created_sheets),
            "successful_creations": successful,
            "failed_creations": failed,
            "created_sheets": self._creation_order.copy(),
            "registered_creators": len(self._registered_creators)
        }
    
    def clear_registry(self) -> None:
        """Clear all registry data. Useful for testing and cleanup."""
        self._created_sheets.clear()
        self._creation_order.clear()
        print("[REGISTRY] Registry cleared")
    
    def validate_workbook_consistency(self, workbook: openpyxl.Workbook) -> Dict[str, Any]:
        """
        Validate that the workbook state matches the registry records.
        
        Returns:
            dict: Validation results with any inconsistencies found
        """
        registry_sheets = set(self.get_all_created_sheets())
        workbook_sheets = set(workbook.sheetnames)
        
        # Find sheets in workbook but not in registry
        untracked_sheets = workbook_sheets - registry_sheets
        
        # Find sheets in registry but not in workbook
        missing_sheets = registry_sheets - workbook_sheets
        
        # Check for failed creations that somehow exist
        failed_but_exist = []
        for sheet_name, record in self._created_sheets.items():
            if not record.created_successfully and sheet_name in workbook_sheets:
                failed_but_exist.append(sheet_name)
        
        return {
            "is_consistent": len(untracked_sheets) == 0 and len(missing_sheets) == 0,
            "untracked_sheets": list(untracked_sheets),
            "missing_sheets": list(missing_sheets),
            "failed_but_exist": failed_but_exist,
            "total_workbook_sheets": len(workbook_sheets),
            "total_registry_sheets": len(registry_sheets)
        }
    
    def print_creation_report(self) -> None:
        """Print a detailed report of all sheet creation activity."""
        print("\n" + "="*60)
        print("SHEET CREATION REGISTRY REPORT")
        print("="*60)
        
        summary = self.get_creation_summary()
        print(f"Total Creation Attempts: {summary['total_attempts']}")
        print(f"Successful Creations: {summary['successful_creations']}")
        print(f"Failed Creations: {summary['failed_creations']}")
        print(f"Registered Creators: {summary['registered_creators']}")
        
        print(f"\nSheet Creation Order:")
        for i, sheet_name in enumerate(self._creation_order, 1):
            record = self._created_sheets[sheet_name]
            status = "✅" if record.created_successfully else "❌"
            print(f"  {i:2d}. {status} {sheet_name}")
            print(f"      Creator: {record.creator_module}.{record.creator_method}")
            if record.position is not None:
                print(f"      Position: {record.position}")
            if not record.created_successfully and record.error_message:
                print(f"      Error: {record.error_message}")
        
        print("="*60)
    
    def _record_creation(self, sheet_name: str, sheet_type: Optional[SheetType],
                        creator_module: str, creator_method: str, 
                        position: Optional[int], success: bool,
                        error_message: Optional[str] = None) -> None:
        """Record a sheet creation attempt."""
        record = SheetCreationRecord(
            sheet_name=sheet_name,
            sheet_type=sheet_type,
            creator_module=creator_module,
            creator_method=creator_method,
            position=position,
            created_successfully=success,
            error_message=error_message
        )
        
        self._created_sheets[sheet_name] = record
        if success:
            self._creation_order.append(sheet_name)


# Global registry instance
_global_registry = SheetRegistry()


def get_sheet_registry() -> SheetRegistry:
    """Get the global sheet registry instance."""
    return _global_registry


def reset_sheet_registry() -> None:
    """Reset the global sheet registry (useful for testing)."""
    global _global_registry
    _global_registry = SheetRegistry()
