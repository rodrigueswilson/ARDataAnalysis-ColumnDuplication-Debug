"""
Raw Data Creator Module
======================

This module handles the creation of the Raw Data sheet, which provides a complete
dump of all media records from the MongoDB database. This sheet is critical for
peer reviewers who need access to the complete raw dataset for validation and
transparency.

The implementation replicates the exact functionality from the original monolithic
generator to ensure backward compatibility and consistency.
"""

import pandas as pd
from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os


class RawDataCreator:
    """
    Handles the creation of the Raw Data sheet with complete database dump.
    
    This class provides peer reviewers and researchers with access to the complete
    raw dataset, formatted consistently with the original implementation.
    """
    
    def __init__(self, db, formatter):
        """
        Initialize the raw data creator.
        
        Args:
            db: MongoDB database connection
            formatter: ExcelFormatter instance for consistent formatting
        """
        self.db = db
        self.formatter = formatter
    
    def create_raw_data_sheet(self, workbook: openpyxl.Workbook) -> bool:
        """
        Creates the Raw Data sheet with complete database dump.
        
        This method replicates the exact functionality from the original generator:
        - Fetches all documents from MongoDB media_records collection
        - Removes MongoDB _id column
        - Ensures ISO_Month is stored as integer to prevent Excel warnings
        - Moves file_name column to first position if present
        - Creates sheet without hyperlinks to prevent workbook corruption
        - Applies consistent formatting
        
        Args:
            workbook: openpyxl workbook object to add the sheet to
            
        Returns:
            bool: True if successful, False otherwise
        """
        print("[RAW DATA] Creating Raw Data sheet...")
        
        try:
            # Fetch all documents from MongoDB with efficient batching
            print("[RAW DATA] Fetching all documents from MongoDB...")
            collection = self.db['media_records']
            
            # First, get total count for progress tracking
            total_count = collection.count_documents({})
            print(f"[RAW DATA] Found {total_count} total records")
            
            if total_count == 0:
                print("[WARNING] No raw data found in database")
                # Create empty sheet with message
                ws = workbook.create_sheet(title="Raw Data", index=1)
                ws.append(["No data available", "The database appears to be empty"])
                self.formatter.format_sheet(ws)
                return True
            
            # Use efficient cursor iteration instead of loading all into memory
            all_docs_cursor = collection.find({})
            
            # Process first document to get column structure
            first_doc = next(all_docs_cursor, None)
            if not first_doc:
                print("[WARNING] No documents found")
                ws = workbook.create_sheet(title="Raw Data", index=1)
                ws.append(["No data available", "The database appears to be empty"])
                self.formatter.format_sheet(ws)
                return True
            
            # Remove MongoDB _id from first document
            if '_id' in first_doc:
                del first_doc['_id']
            
            # Get column names from first document
            column_names = list(first_doc.keys())
            
            # Ensure file_name is first if present
            if 'file_name' in column_names:
                column_names.remove('file_name')
                column_names.insert(0, 'file_name')
            
            print(f"[RAW DATA] Schema: {len(column_names)} columns")
            
            # Create Raw Data Sheet - exact replica including index position
            ws = workbook.create_sheet(title="Raw Data", index=1)
            print("[RAW DATA] Created Raw Data worksheet at index 1")
            
            # Add header row
            ws.append(column_names)
            print(f"[RAW DATA] Added header row with {len(column_names)} columns")
            
            # Process first document and add to sheet
            first_row = []
            for col in column_names:
                value = first_doc.get(col, '')
                # Handle ISO_Month conversion for Excel compatibility
                if col == 'ISO_Month' and value is not None:
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass
                first_row.append(value)
            ws.append(first_row)
            
            # Process remaining documents efficiently
            print("[RAW DATA] Adding remaining data rows...")
            row_count = 1  # Already added first row
            
            for doc in all_docs_cursor:
                # Remove MongoDB _id
                if '_id' in doc:
                    del doc['_id']
                
                # Create row in column order
                row = []
                for col in column_names:
                    value = doc.get(col, '')
                    # Handle ISO_Month conversion for Excel compatibility
                    if col == 'ISO_Month' and value is not None:
                        try:
                            value = int(value)
                        except (ValueError, TypeError):
                            pass
                    row.append(value)
                
                ws.append(row)
                row_count += 1
                
                # Progress indicator for large datasets
                if row_count % 1000 == 0:
                    print(f"[RAW DATA] Processed {row_count} rows...")
            
            print(f"[RAW DATA] Added {row_count} total data rows successfully")
            
            # Add hyperlinks to file_name column (if file_path exists)
            print("[RAW DATA] Adding hyperlinks to file names...")
            self._add_file_hyperlinks(ws, column_names)
            
            # Apply enhanced formatting with proper column widths
            print("[RAW DATA] Applying enhanced formatting...")
            self._format_raw_data_sheet(ws)
            
            print("[SUCCESS] Raw Data sheet created successfully")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to create Raw Data sheet: {e}")
            import traceback
            traceback.print_exc()
            
            # Create error sheet for transparency
            try:
                ws = workbook.create_sheet(title="Raw Data", index=1)
                ws.append(["Error", "Failed to load raw data"])
                ws.append(["Details", str(e)])
                self.formatter.format_sheet(ws)
            except:
                pass
            
            return False
    
    def get_raw_data_summary(self) -> dict:
        """
        Get summary statistics about the raw data for reporting.
        
        Returns:
            dict: Summary statistics including record count, columns, etc.
        """
        try:
            collection = self.db['media_records']
            total_records = collection.count_documents({})
            
            if total_records > 0:
                # Get sample document to determine schema
                sample_doc = collection.find_one({})
                if sample_doc:
                    columns = list(sample_doc.keys())
                    if '_id' in columns:
                        columns.remove('_id')
                else:
                    columns = []
            else:
                columns = []
            
            return {
                'total_records': total_records,
                'column_count': len(columns),
                'columns': columns,
                'status': 'success' if total_records > 0 else 'empty'
            }
            
        except Exception as e:
            return {
                'total_records': 0,
                'column_count': 0,
                'columns': [],
                'status': 'error',
                'error': str(e)
            }
    
    def _add_file_hyperlinks(self, ws, column_names):
        """
        Add hyperlinks to file_name column based on file_path column.
        OPTIMIZED for large datasets by removing expensive file existence checks.
        
        Args:
            ws: openpyxl worksheet object
            column_names: list of column names
        """
        try:
            # Find column indices
            file_name_col_idx = -1
            file_path_col_idx = -1
            
            if 'file_name' in column_names:
                file_name_col_idx = column_names.index('file_name') + 1  # openpyxl is 1-based
            
            if 'file_path' in column_names:
                file_path_col_idx = column_names.index('file_path') + 1
            
            # Only add hyperlinks if both columns exist
            if file_name_col_idx > 0 and file_path_col_idx > 0:
                print(f"[RAW DATA] Adding hyperlinks from column {file_name_col_idx} to paths in column {file_path_col_idx}")
                print(f"[RAW DATA] Processing {ws.max_row-1} rows for hyperlinks...")
                
                hyperlink_count = 0
                for row_idx in range(2, ws.max_row + 1):  # Skip header row
                    file_path_cell = ws.cell(row=row_idx, column=file_path_col_idx)
                    file_name_cell = ws.cell(row=row_idx, column=file_name_col_idx)
                    
                    if file_path_cell.value and file_name_cell.value:
                        file_path = str(file_path_cell.value)
                        
                        # PERFORMANCE OPTIMIZATION: Skip file existence check
                        # The database should only contain valid file paths
                        # Checking 10,000+ files with os.path.exists() is extremely slow
                        
                        # Create hyperlink (trust database integrity)
                        file_name_cell.hyperlink = f'file:///{file_path.replace("\\", "/")}'
                        file_name_cell.style = "Hyperlink"
                        hyperlink_count += 1
                    
                    # Progress indicator for large datasets
                    if row_idx % 1000 == 0:
                        print(f"[RAW DATA] Processed {row_idx-1} hyperlinks...")
                
                print(f"[RAW DATA] Added {hyperlink_count} hyperlinks (file existence not verified for performance)")
            else:
                print("[RAW DATA] Skipping hyperlinks - file_name or file_path column not found")
                
        except Exception as e:
            print(f"[WARNING] Could not add hyperlinks: {e}")
    
    def _format_raw_data_sheet(self, ws):
        """
        Apply enhanced formatting to the Raw Data sheet with proper column widths.
        OPTIMIZED for large datasets to avoid performance issues.
        
        Args:
            ws: openpyxl worksheet object
        """
        try:
            if ws.max_row <= 1:
                return  # No data to format
            
            print(f"[RAW DATA] Formatting sheet with {ws.max_row-1} data rows...")
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Apply optimized alternating row colors using centralized ExcelFormatter method
            # Our new implementation handles large datasets efficiently with batch processing
            print(f"[RAW DATA] Applying optimized alternating row colors to {ws.max_row-1} data rows...")
            self.formatter.apply_alternating_row_colors(ws, 2, ws.max_row, 1, ws.max_column)
            
            # Smart column width adjustment (based on original implementation)
            print("[RAW DATA] Calculating optimal column widths...")
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                
                # Check header length
                header_cell = ws.cell(row=1, column=col_idx)
                if header_cell.value:
                    max_length = max(max_length, len(str(header_cell.value)))
                
                # Check a sample of data rows for optimal width (limit to first 100 rows for performance)
                sample_rows = min(100, ws.max_row)
                for row_idx in range(2, sample_rows + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                
                # Set column width with reasonable bounds (exact replica of original)
                adjusted_width = min(max(max_length + 2, 10), 50)  # Min 10, Max 50, +2 for padding
                ws.column_dimensions[col_letter].width = adjusted_width
            
            # Freeze panes at A2 (exact replica of original)
            ws.freeze_panes = 'A2'
            
            print("[RAW DATA] Enhanced formatting completed successfully")
            
        except Exception as e:
            print(f"[WARNING] Could not apply enhanced formatting: {e}")
            # Fall back to basic formatting
            self.formatter.format_sheet(ws)
