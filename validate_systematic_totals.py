#!/usr/bin/env python3
"""
Systematic Totals Validation Script
==================================

This script validates that the systematic totals implementation has been
successfully applied to all high-priority sheets identified in our analysis.
"""

import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import json

def find_latest_report():
    """Find the most recent Excel report."""
    report_files = list(Path(".").glob("AR_Analysis_Report_*.xlsx"))
    if not report_files:
        raise FileNotFoundError("No Excel reports found")
    
    latest_report = max(report_files, key=lambda x: x.stat().st_mtime)
    print(f"📊 Validating report: {latest_report}")
    return latest_report

def validate_sheet_totals(ws, sheet_name):
    """
    Validate that totals have been properly applied to a sheet.
    
    Returns:
        dict: Validation results for the sheet
    """
    validation_result = {
        'sheet_name': sheet_name,
        'has_totals': False,
        'totals_type': 'none',
        'totals_found': [],
        'data_rows': ws.max_row,
        'data_columns': ws.max_column,
        'validation_status': 'unknown'
    }
    
    # Look for totals indicators
    totals_indicators = ['total', 'totals', 'sum', 'grand total', 'subtotal', 'row total']
    
    totals_found = []
    
    # Check all cells for totals indicators
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                cell_value_lower = cell.value.lower()
                for indicator in totals_indicators:
                    if indicator in cell_value_lower:
                        totals_found.append({
                            'cell': f"{cell.column_letter}{cell.row}",
                            'value': cell.value,
                            'type': indicator
                        })
    
    if totals_found:
        validation_result['has_totals'] = True
        validation_result['totals_found'] = totals_found
        
        # Determine totals type based on what was found
        total_types = [t['type'] for t in totals_found]
        if 'row total' in total_types and any('total' in t for t in total_types if t != 'row total'):
            validation_result['totals_type'] = 'comprehensive'
        elif 'row total' in total_types:
            validation_result['totals_type'] = 'row_totals'
        elif any('total' in t for t in total_types):
            validation_result['totals_type'] = 'column_totals'
        
        validation_result['validation_status'] = 'success'
    else:
        validation_result['validation_status'] = 'no_totals_found'
    
    return validation_result

def validate_high_priority_sheets(excel_path):
    """
    Validate totals implementation for high-priority sheets.
    """
    # High-priority sheets from our systematic analysis
    high_priority_sheets = {
        'Monthly Capture Volume': {
            'expected_totals': 'column_totals',
            'rationale': '3 numeric columns, perfect for column totals'
        },
        'File Size Stats': {
            'expected_totals': 'comprehensive',
            'rationale': '4 numeric columns, statistical summary needs comprehensive totals'
        },
        'Time of Day': {
            'expected_totals': 'row_totals',
            'rationale': 'Few rows with 1 numeric column, row totals appropriate'
        },
        'Weekday by Period': {
            'expected_totals': 'row_totals',
            'rationale': 'Multiple rows with 1 numeric column, row totals appropriate'
        }
    }
    
    print("\n" + "="*80)
    print("📊 VALIDATING HIGH-PRIORITY TOTALS IMPLEMENTATION")
    print("="*80)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    validation_results = {}
    success_count = 0
    total_count = len(high_priority_sheets)
    
    for sheet_name, expected in high_priority_sheets.items():
        print(f"\n🔍 Validating: {sheet_name}")
        print(f"   📋 Expected: {expected['expected_totals']}")
        print(f"   💡 Rationale: {expected['rationale']}")
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = validate_sheet_totals(ws, sheet_name)
            
            # Check if implementation matches expectations
            if result['has_totals']:
                if (expected['expected_totals'] == 'comprehensive' and result['totals_type'] == 'comprehensive') or \
                   (expected['expected_totals'] == 'column_totals' and result['totals_type'] in ['column_totals', 'comprehensive']) or \
                   (expected['expected_totals'] == 'row_totals' and result['totals_type'] in ['row_totals', 'comprehensive']):
                    result['implementation_status'] = 'success'
                    success_count += 1
                    print(f"   ✅ SUCCESS: Found {result['totals_type']} as expected")
                else:
                    result['implementation_status'] = 'partial'
                    print(f"   ⚠️  PARTIAL: Found {result['totals_type']}, expected {expected['expected_totals']}")
            else:
                result['implementation_status'] = 'failed'
                print(f"   ❌ FAILED: No totals found")
            
            # Show details of found totals
            if result['totals_found']:
                print(f"   📊 Totals found: {len(result['totals_found'])}")
                for total in result['totals_found'][:3]:  # Show first 3
                    print(f"      • {total['cell']}: {total['value']}")
                if len(result['totals_found']) > 3:
                    print(f"      ... and {len(result['totals_found']) - 3} more")
            
            validation_results[sheet_name] = result
        else:
            print(f"   ❌ ERROR: Sheet not found in workbook")
            validation_results[sheet_name] = {
                'sheet_name': sheet_name,
                'validation_status': 'sheet_not_found'
            }
    
    return validation_results, success_count, total_count

def validate_existing_totals_sheets(excel_path):
    """
    Check existing sheets that already had some totals to ensure they're still working.
    """
    print("\n" + "-"*60)
    print("🔍 VALIDATING EXISTING TOTALS SHEETS")
    print("-"*60)
    
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Sheets that should already have totals based on our inventory
    existing_totals_sheets = [
        'Summary Statistics', 'MP3 Duration Analysis', 'Data Cleaning',
        'Daily Counts (ACF_PACF)', 'Weekly Counts (ACF_PACF)', 
        'Biweekly Counts (ACF_PACF)', 'Monthly Counts (ACF_PACF)',
        'Period Counts (ACF_PACF)'
    ]
    
    validation_results = {}
    success_count = 0
    
    for sheet_name in existing_totals_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            result = validate_sheet_totals(ws, sheet_name)
            
            if result['has_totals']:
                print(f"   ✅ {sheet_name}: {result['totals_type']} ({len(result['totals_found'])} totals)")
                success_count += 1
            else:
                print(f"   ⚠️  {sheet_name}: No totals detected")
            
            validation_results[sheet_name] = result
    
    return validation_results, success_count, len(existing_totals_sheets)

def generate_validation_summary(high_priority_results, existing_results, hp_success, hp_total, ex_success, ex_total):
    """
    Generate a comprehensive validation summary.
    """
    print("\n" + "="*80)
    print("📋 SYSTEMATIC TOTALS VALIDATION SUMMARY")
    print("="*80)
    
    print(f"📅 Validation Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    print("\n🎯 HIGH-PRIORITY IMPLEMENTATION RESULTS:")
    print(f"   ✅ Successfully implemented: {hp_success}/{hp_total} sheets ({hp_success/hp_total*100:.1f}%)")
    
    if hp_success == hp_total:
        print("   🏆 PERFECT SCORE: All high-priority sheets have appropriate totals!")
    elif hp_success >= hp_total * 0.75:
        print("   🎯 EXCELLENT: Most high-priority sheets implemented successfully")
    elif hp_success >= hp_total * 0.5:
        print("   ⚠️  GOOD: Majority of high-priority sheets implemented")
    else:
        print("   ❌ NEEDS WORK: Several high-priority sheets still need totals")
    
    print(f"\n📊 EXISTING TOTALS VALIDATION:")
    print(f"   ✅ Sheets with totals preserved: {ex_success}/{ex_total} sheets ({ex_success/ex_total*100:.1f}%)")
    
    print(f"\n📈 OVERALL TOTALS COVERAGE:")
    total_success = hp_success + ex_success
    total_sheets = hp_total + ex_total
    print(f"   📊 Total sheets with totals: {total_success}/{total_sheets} ({total_success/total_sheets*100:.1f}%)")
    
    # Detailed breakdown
    print("\n" + "-"*60)
    print("DETAILED IMPLEMENTATION STATUS")
    print("-"*60)
    
    for sheet_name, result in high_priority_results.items():
        status = result.get('implementation_status', 'unknown')
        totals_type = result.get('totals_type', 'none')
        status_icon = {'success': '✅', 'partial': '⚠️', 'failed': '❌'}.get(status, '❓')
        print(f"{status_icon} {sheet_name}: {totals_type}")
    
    print("\n🎯 RECOMMENDATIONS:")
    
    failed_sheets = [name for name, result in high_priority_results.items() 
                    if result.get('implementation_status') == 'failed']
    
    if not failed_sheets:
        print("   🏆 Excellent work! All high-priority totals have been implemented.")
        print("   📋 Consider reviewing existing sheets for enhancement opportunities.")
    else:
        print(f"   🔧 Focus on implementing totals for: {', '.join(failed_sheets)}")
        print("   📊 Review configuration and ensure TotalsManager is properly integrated.")
    
    print("\n✅ SYSTEMATIC TOTALS VALIDATION COMPLETE")

def save_validation_report(high_priority_results, existing_results):
    """
    Save detailed validation results to JSON file.
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_file = f"totals_validation_report_{timestamp}.json"
    
    validation_report = {
        'validation_date': datetime.now().isoformat(),
        'high_priority_sheets': high_priority_results,
        'existing_totals_sheets': existing_results,
        'summary': {
            'high_priority_success_rate': len([r for r in high_priority_results.values() 
                                             if r.get('implementation_status') == 'success']) / len(high_priority_results),
            'existing_totals_preserved_rate': len([r for r in existing_results.values() 
                                                 if r.get('has_totals', False)]) / len(existing_results),
            'total_sheets_validated': len(high_priority_results) + len(existing_results)
        }
    }
    
    with open(report_file, 'w') as f:
        json.dump(validation_report, f, indent=2)
    
    print(f"\n💾 Detailed validation report saved to: {report_file}")
    return report_file

def main():
    """
    Main validation function.
    """
    try:
        print("🚀 Starting systematic totals validation...")
        
        # Find latest report
        excel_path = find_latest_report()
        
        # Validate high-priority sheets
        hp_results, hp_success, hp_total = validate_high_priority_sheets(excel_path)
        
        # Validate existing totals sheets
        ex_results, ex_success, ex_total = validate_existing_totals_sheets(excel_path)
        
        # Generate summary
        generate_validation_summary(hp_results, ex_results, hp_success, hp_total, ex_success, ex_total)
        
        # Save detailed report
        report_file = save_validation_report(hp_results, ex_results)
        
        print(f"\n🎯 Next Steps:")
        if hp_success == hp_total:
            print("1. ✅ High-priority implementation complete!")
            print("2. 📋 Consider enhancing existing sheets with additional totals")
            print("3. 🔍 Run cross-sheet validation to ensure consistency")
            print("4. 📊 Document totals implementation for future reference")
        else:
            print("1. 🔧 Address any failed high-priority implementations")
            print("2. 📊 Review TotalsManager configuration for failed sheets")
            print("3. 🧪 Test individual sheet totals implementation")
            print("4. 🔄 Re-run validation after fixes")
        
    except Exception as e:
        print(f"❌ Validation error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
