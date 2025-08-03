#!/usr/bin/env python3
"""
Test script to verify Monthly MP3 Duration Distribution table uses school year month order
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators.specialized import SpecializedSheetCreator
from report_generator.formatters import ExcelFormatter
from openpyxl import Workbook
from db_utils import get_db_connection

def test_month_order():
    """Test the Monthly MP3 Duration Distribution table month order"""
    try:
        print("=== Monthly MP3 Duration Distribution Month Order Test ===")
        
        # Create test workbook
        workbook = Workbook()
        formatter = ExcelFormatter()
        
        # Create specialized sheet creator
        creator = SpecializedSheetCreator(get_db_connection(), formatter)
        
        # Create the MP3 Duration Analysis sheet
        print("\n[TEST] Creating MP3 Duration Analysis sheet...")
        creator.create_mp3_duration_analysis_sheet(workbook)
        
        # Get the worksheet
        ws = workbook['MP3 Duration Analysis']
        
        # Find the Monthly Distribution table
        print("\n=== Finding Monthly Distribution Table ===")
        monthly_table_start = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "Monthly MP3 Duration Distribution":
                monthly_table_start = row
                print(f"[INFO] Monthly Distribution table starts at row {row}")
                break
        
        if not monthly_table_start:
            print("[ERROR] Monthly Distribution table not found!")
            return False
        
        # Extract month order from the table
        print("\n=== Extracting Month Order ===")
        months_found = []
        
        # Start from header row + 1 to get data rows
        header_row = monthly_table_start + 2
        data_start_row = header_row + 1
        
        for row in range(data_start_row, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and isinstance(cell_value, str) and len(cell_value) == 3:
                # This looks like a month abbreviation
                months_found.append(cell_value)
                print(f"[INFO] Found month at row {row}: {cell_value}")
            elif cell_value and "Total" in str(cell_value):
                # We've reached the totals row, stop
                print(f"[INFO] Reached totals row at {row}: {cell_value}")
                break
        
        # Expected school year order
        expected_order = ["Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May"]
        
        print(f"\n=== Month Order Comparison ===")
        print(f"Expected (School Year): {expected_order}")
        print(f"Found in Table:         {months_found}")
        
        # Check if the order matches
        if months_found == expected_order:
            print("[SUCCESS] ✅ Month order is correct! Using school year order (Sep-May)")
            success = True
        else:
            print("[ERROR] ❌ Month order does not match expected school year order")
            success = False
            
            # Show differences
            for i, (expected, found) in enumerate(zip(expected_order, months_found)):
                if expected != found:
                    print(f"  Position {i+1}: Expected '{expected}', Found '{found}'")
        
        # Save test workbook
        test_filename = "test_month_order.xlsx"
        workbook.save(test_filename)
        print(f"\n[INFO] Test workbook saved as: {test_filename}")
        
        return success
        
    except Exception as e:
        print(f"[ERROR] Test failed with exception: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_month_order()
    sys.exit(0 if success else 1)
