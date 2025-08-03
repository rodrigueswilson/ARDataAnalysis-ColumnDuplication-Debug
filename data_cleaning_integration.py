"""
Data Cleaning Integration Example

This script demonstrates how to integrate the DataCleaningUtils class
with the existing SheetCreator.create_data_cleaning_sheet method.

This is a proof of concept that shows how to use the modular approach
without changing the existing implementation.
"""

from utils.data_cleaning import DataCleaningUtils
from db_utils import get_db_connection
import pandas as pd
from report_generator.sheet_creators import SheetCreator
from report_generator.formatters import ExcelFormatter
import openpyxl

def create_modular_data_cleaning_sheet(workbook, sheet_creator, data_cleaning):
    """
    Creates a Data Cleaning sheet using the modular DataCleaningUtils class.
    
    Args:
        workbook: openpyxl workbook object
        sheet_creator: SheetCreator instance for formatting and utils
        data_cleaning: DataCleaningUtils instance for data calculations
    """
    try:
        # Create worksheet
        ws = workbook.create_sheet("Data Cleaning (Modular)")
        
        # Title
        ws['A1'] = "AR Data Analysis - Data Cleaning & Filtering Report (Modular Implementation)"
        sheet_creator.formatter.apply_title_style(ws, 'A1')
        
        print("[INFO] Running intersection analysis for Data Cleaning sheet using modular approach...")
        
        # Get complete cleaning data using the utility class
        result = data_cleaning.get_complete_cleaning_data()
        intersection_data = result['intersection_data']
        totals = result['totals']
        
        # Table 1: Complete Filtering Breakdown
        ws['A3'] = "Table 1: Complete Filtering Breakdown (Intersection Analysis)"
        sheet_creator.formatter.apply_section_header_style(ws, 'A3')
        
        # Headers for 2x2 matrix table (academic terminology with before/after cleaning clarity)
        headers = ['Media Type', 'Initial Collection Size', 'Recording Errors Filtered', 'Non-Instructional Days Filtered', 
                'Combined Filters Applied', 'Total Records Filtered', 'Research Dataset Size', 
                'Filter Application Rate (%)', 'Dataset Validity Rate (%)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=5, column=col, value=header)
        sheet_creator.formatter.apply_header_style(ws, 'A5:I5')
        
        # Data rows
        row = 6
        for item in intersection_data[:-1]:  # Skip the last item which is totals
            ws.cell(row=row, column=1, value=item['file_type'])
            ws.cell(row=row, column=2, value=item['total_files'])
            ws.cell(row=row, column=3, value=item['school_outliers'])
            ws.cell(row=row, column=4, value=item['non_school_normal'])
            ws.cell(row=row, column=5, value=item['non_school_outliers'])
            ws.cell(row=row, column=6, value=item['total_excluded'])
            ws.cell(row=row, column=7, value=item['school_normal'])
            
            # Format exclusion percentage
            exclusion_cell = ws.cell(row=row, column=8, value=item['exclusion_pct'] / 100)
            exclusion_cell.number_format = '0.0%'
            
            # Format retention percentage
            retention_cell = ws.cell(row=row, column=9, value=item['retention_pct'] / 100)
            retention_cell.number_format = '0.0%'
            
            row += 1
        
        # Apply formatting to Table 1 (excluding the total row which will be added separately)
        sheet_creator.formatter.apply_data_style(ws, f'A6:I{row-1}')
        sheet_creator.formatter.apply_alternating_row_colors(ws, 6, row-1, 1, 9)
        
        # Add formatted total row at the specific position
        total_row_values = {
            1: 'TOTAL',
            2: totals['total_files'],
            3: totals['school_outliers'],
            4: totals['non_school_normal'],
            5: totals['non_school_outliers'],
            6: totals['total_excluded'],
            7: totals['school_normal'],
            8: totals['exclusion_pct'] / 100,
            9: totals['retention_pct'] / 100
        }
        
        # Add total row at the correct position
        sheet_creator.formatter.add_total_row_at_position(ws, row, total_row_values)
        
        # Table 2: Year-by-Year Breakdown
        table2_start_row = row + 3
        ws[f'A{table2_start_row}'] = "Table 2: Year-by-Year Breakdown"
        sheet_creator.formatter.apply_section_header_style(ws, f'A{table2_start_row}')
        
        # Headers for Table 2 (same academic terminology as Table 1)
        table2_headers = ['Category', 'Initial Collection Size', 'Recording Errors Filtered', 
                        'Non-Instructional Days Filtered', 'Combined Filters Applied', 
                        'Total Records Filtered', 'Research Dataset Size', 
                        'Filter Application Rate (%)', 'Dataset Validity Rate (%)']
        table2_header_row = table2_start_row + 2
        for col, header in enumerate(table2_headers, 1):
            ws.cell(row=table2_header_row, column=col, value=header)
        sheet_creator.formatter.apply_header_style(ws, f'A{table2_header_row}:I{table2_header_row}')
        
        # Get year breakdown data
        year_data = data_cleaning.get_year_breakdown_data()
        
        # Fill Table 2 data
        table2_row = table2_header_row + 1
        for item in year_data[:-1]:  # Skip the last item which is totals
            ws.cell(row=table2_row, column=1, value=item['category'])
            ws.cell(row=table2_row, column=2, value=item['total_files'])
            ws.cell(row=table2_row, column=3, value=item['outliers'])
            ws.cell(row=table2_row, column=4, value=item['non_school_days'])
            ws.cell(row=table2_row, column=5, value=item['non_school_outliers'])
            ws.cell(row=table2_row, column=6, value=item['total_excluded'])
            ws.cell(row=table2_row, column=7, value=item['school_days'])
            
            # Format exclusion percentage
            exclusion_cell = ws.cell(row=table2_row, column=8, value=item['exclusion_pct'] / 100)
            exclusion_cell.number_format = '0.0%'
            
            # Format retention percentage
            retention_cell = ws.cell(row=table2_row, column=9, value=item['retention_pct'] / 100)
            retention_cell.number_format = '0.0%'
            
            table2_row += 1
        
        # Apply formatting to Table 2 (excluding the total row which will be added separately)
        sheet_creator.formatter.apply_data_style(ws, f'A{table2_header_row + 1}:I{table2_row-1}')
        sheet_creator.formatter.apply_alternating_row_colors(ws, table2_header_row + 1, table2_row-1, 1, 9)
        
        # Get the totals from year breakdown
        year_totals = year_data[-1]
        
        # Add formatted total row at the specific position
        year_total_row_values = {
            1: 'TOTAL',
            2: year_totals['total_files'],
            3: year_totals['outliers'],
            4: year_totals['non_school_days'],
            5: year_totals['non_school_outliers'],
            6: year_totals['total_excluded'],
            7: year_totals['school_days'],
            8: year_totals['exclusion_pct'] / 100,
            9: year_totals['retention_pct'] / 100
        }
        
        # Add total row at the correct position
        sheet_creator.formatter.add_total_row_at_position(ws, table2_row, year_total_row_values)
        
        # Logic Explanation Table
        logic_table_start = table2_row + 3
        ws[f'A{logic_table_start}'] = "Logic Explanation: Category Definitions"
        sheet_creator.formatter.apply_section_header_style(ws, f'A{logic_table_start}')
        
        # Headers for logic explanation table (academic terminology)
        logic_headers = ['Category', 'Collection Period', 'Recording Quality', 'Count']
        logic_header_row = logic_table_start + 2
        for col, header in enumerate(logic_headers, 1):
            ws.cell(row=logic_header_row, column=col, value=header)
        sheet_creator.formatter.apply_header_style(ws, f'A{logic_header_row}:D{logic_header_row}')
        
        # Get logic explanation data
        logic_data = data_cleaning.get_logic_explanation_data(totals)
        
        # Fill logic explanation data
        logic_row = logic_header_row + 1
        for category, collection_day, outlier_status, count in logic_data:
            ws.cell(row=logic_row, column=1, value=category)
            ws.cell(row=logic_row, column=2, value=collection_day)
            ws.cell(row=logic_row, column=3, value=outlier_status)
            ws.cell(row=logic_row, column=4, value=count)
            logic_row += 1
        
        # Add formatted total row at the specific position
        logic_total_values = {
            1: 'TOTAL',
            2: '-',
            3: '-',
            4: totals['total_files']
        }
        
        # Add total row at the correct position
        sheet_creator.formatter.add_total_row_at_position(ws, logic_row, logic_total_values)
        
        logic_row += 1
        
        # Apply formatting to logic table
        sheet_creator.formatter.apply_data_style(ws, f'A{logic_header_row + 1}:D{logic_row - 1}')
        sheet_creator.formatter.apply_alternating_row_colors(ws, logic_header_row + 1, logic_row - 1, 1, 4)
        
        # Create Table 3: Filter Impact Summary if _create_filter_impact_summary method is available
        if hasattr(sheet_creator, '_create_filter_impact_summary'):
            next_row = sheet_creator._create_filter_impact_summary(ws, logic_row + 3, totals)
        else:
            next_row = logic_row + 3
        
        # Add explanatory notes
        notes_start_row = next_row + 2
        ws.cell(row=notes_start_row, column=1, value="Notes:")
        sheet_creator.formatter.apply_section_header_style(ws, f'A{notes_start_row}')
        
        notes = [
            "• Initial Collection Size: Complete raw dataset before quality assessment",
            "• Recording Errors Filtered: Files manually classified as recording mistakes (too short, too long,",
            "  recorder not stopped properly, or otherwise irrelevant for AR collection analysis)",
            "• Non-Instructional Days: Files captured outside designated school collection periods",
            "• Research Dataset: High-quality files meeting both technical validity and temporal criteria",
            "• This preprocessing ensures dataset integrity for educational AR research analysis"
        ]
        
        for i, note in enumerate(notes):
            note_row = notes_start_row + 1 + i
            ws[f'A{note_row}'] = note
            sheet_creator.formatter.apply_data_style(ws, f'A{note_row}')
        
        # Auto-adjust columns
        sheet_creator.formatter.auto_adjust_columns(ws)
        
        print("[SUCCESS] Modular Data Cleaning sheet created with 2x2 matrix showing all four mutually exclusive categories")
        print(f"[INFO] Final clean dataset: {totals['school_normal']} files ({totals['retention_pct']:.1f}% retention)")
        
    except Exception as e:
        print(f"[ERROR] Failed to create Modular Data Cleaning sheet: {e}")
        import traceback
        traceback.print_exc()


def compare_original_and_modular_approaches():
    """
    Run both the original and modular data cleaning implementations and compare the results.
    """
    print("=" * 80)
    print("Comparing Original and Modular Data Cleaning Approaches")
    print("=" * 80)
    
    # Get database connection
    db = get_db_connection()
    
    # Create formatter
    formatter = ExcelFormatter()
    
    # Create sheet creator
    sheet_creator = SheetCreator(db, formatter)
    
    # Create data cleaning utils
    data_cleaning = DataCleaningUtils(db)
    
    # Create a workbook for comparison
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    print("\nCreating original Data Cleaning sheet...")
    try:
        # Create original sheet
        sheet_creator.create_data_cleaning_sheet(wb)
        print("✓ Original Data Cleaning sheet created")
    except Exception as e:
        print(f"✗ Failed to create original Data Cleaning sheet: {e}")
        import traceback
        traceback.print_exc()
    
    print("\nCreating modular Data Cleaning sheet...")
    try:
        # Create modular sheet
        create_modular_data_cleaning_sheet(wb, sheet_creator, data_cleaning)
        print("✓ Modular Data Cleaning sheet created")
    except Exception as e:
        print(f"✗ Failed to create modular Data Cleaning sheet: {e}")
        import traceback
        traceback.print_exc()
    
    # Save the workbook
    output_file = "data_cleaning_comparison.xlsx"
    wb.save(output_file)
    print(f"\nComparison saved to {output_file}")
    print("You can now open this file to compare both implementations side by side.")
    
    print("\n=== Comparison Complete ===")
    

if __name__ == "__main__":
    compare_original_and_modular_approaches()
