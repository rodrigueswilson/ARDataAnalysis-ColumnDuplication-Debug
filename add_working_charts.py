#!/usr/bin/env python3
"""
Working solution to add visible charts to Excel reports
"""

import openpyxl
from openpyxl.chart import LineChart, Reference

def add_working_charts():
    """Add charts that are guaranteed to be visible"""
    
    # Load the latest report
    report_file = 'AR_Analysis_Report_20250730_155346.xlsx'
    
    print(f"Adding charts to: {report_file}")
    
    wb = openpyxl.load_workbook(report_file)
    
    # Target sheets that should have charts
    target_sheets = [
        'Daily Counts (ACF_PACF)',
        'Weekly Counts (ACF_PACF)', 
        'Monthly Counts (ACF_PACF)',
        'Biweekly Counts (ACF_PACF)',
        'Period Counts (ACF_PACF)'
    ]
    
    charts_added = 0
    
    for sheet_name in target_sheets:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"Processing: {sheet_name}")
            
            # Find Total_Files column
            headers = [cell.value for cell in ws[1]]
            total_files_col = None
            
            for i, header in enumerate(headers, 1):
                if str(header) == 'Total_Files':
                    total_files_col = i
                    break
            
            if total_files_col and ws.max_row > 1:
                # Create visible chart
                chart = LineChart()
                chart.title = f"Time Series Data - {sheet_name.replace('(ACF_PACF)', '')}"
                chart.height = 10
                chart.width = 16
                chart.y_axis.title = "Total Files"
                chart.x_axis.title = "Time Period"
                
                # Add data
                data_ref = Reference(ws, min_col=total_files_col, min_row=1, max_row=ws.max_row)
                chart.add_data(data_ref, titles_from_data=True)
                
                # Position at column I (after typical data ends)
                ws.add_chart(chart, 'I2')
                charts_added += 1
                print(f"  ✓ Added chart at I2")
    
    # Save with charts
    output_file = 'Report_With_Working_Charts.xlsx'
    wb.save(output_file)
    
    print(f"\nSUCCESS!")
    print(f"Charts added: {charts_added}")
    print(f"File saved: {output_file}")
    print(f"\nTo see charts:")
    print(f"1. Open {output_file}")
    print(f"2. Go to any ACF_PACF sheet")
    print(f"3. Look at column I, row 2")
    print(f"4. Charts are positioned to the RIGHT of your data")

if __name__ == "__main__":
    add_working_charts()
