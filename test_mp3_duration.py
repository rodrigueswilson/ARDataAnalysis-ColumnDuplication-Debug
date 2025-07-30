#!/usr/bin/env python3
"""
Test script to debug MP3 Duration Analysis sheet creation.
"""

import openpyxl
from db_utils import get_db_connection
from report_generator.sheet_creators import SheetCreator
from report_generator.formatters import ExcelFormatter
from pipelines import PIPELINES

def test_mp3_duration_sheet():
    print("=== MP3 DURATION ANALYSIS DEBUG TEST ===")
    
    try:
        # Connect to database
        print("[1] Connecting to database...")
        db = get_db_connection()
        print("    ✅ Database connected")
        
        # Check if our pipelines exist
        print("[2] Checking pipelines...")
        required_pipelines = ['MP3_DURATION_BY_SCHOOL_YEAR', 'MP3_DURATION_BY_PERIOD', 'MP3_DURATION_BY_MONTH']
        for pipeline_name in required_pipelines:
            if pipeline_name in PIPELINES:
                print(f"    ✅ {pipeline_name} found")
            else:
                print(f"    ❌ {pipeline_name} NOT found")
                return False
        
        # Test pipeline execution
        print("[3] Testing pipeline execution...")
        formatter = ExcelFormatter()
        sheet_creator = SheetCreator(db, formatter)
        
        for pipeline_name in required_pipelines:
            try:
                df = sheet_creator._run_aggregation(PIPELINES[pipeline_name])
                print(f"    ✅ {pipeline_name}: {len(df)} rows")
                if len(df) > 0:
                    print(f"       Columns: {list(df.columns)}")
            except Exception as e:
                print(f"    ❌ {pipeline_name}: Error - {e}")
                return False
        
        # Test sheet creation
        print("[4] Testing sheet creation...")
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        try:
            sheet_creator.create_mp3_duration_analysis_sheet(wb)
            print("    ✅ Sheet creation method executed")
            
            # Check if sheet was created
            if 'MP3 Duration Analysis' in wb.sheetnames:
                print("    ✅ MP3 Duration Analysis sheet found in workbook")
                mp3_sheet = wb['MP3 Duration Analysis']
                print(f"    ✅ Sheet dimensions: {mp3_sheet.max_row} rows x {mp3_sheet.max_column} columns")
                return True
            else:
                print("    ❌ MP3 Duration Analysis sheet NOT found in workbook")
                print(f"    Available sheets: {wb.sheetnames}")
                return False
                
        except Exception as e:
            print(f"    ❌ Sheet creation failed: {e}")
            import traceback
            traceback.print_exc()
            return False
            
    except Exception as e:
        print(f"❌ Test failed: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_mp3_duration_sheet()
    print(f"\n=== TEST {'PASSED' if success else 'FAILED'} ===")
