#!/usr/bin/env python3
"""
Test Monthly Capture Volume Totals Specifically
==============================================

This script tests the totals application for Monthly Capture Volume sheet specifically
to understand why totals aren't being applied.
"""

import pandas as pd
from openpyxl import Workbook
from report_generator.sheet_creators.pipeline import PipelineSheetCreator
from report_generator.totals_manager import TotalsManager
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection
from pipelines import PIPELINES

def test_monthly_capture_totals():
    """
    Test totals application for Monthly Capture Volume specifically.
    """
    print("🧪 TESTING MONTHLY CAPTURE VOLUME TOTALS")
    print("="*50)
    
    # Initialize components
    db = get_db_connection()
    formatter = ExcelFormatter()
    pipeline_creator = PipelineSheetCreator(db, formatter)
    
    # Get Monthly Capture Volume data
    pipeline = PIPELINES['CAPTURE_VOLUME_PER_MONTH']
    result = list(db.media_records.aggregate(pipeline))
    df = pd.DataFrame(result)
    
    print(f"📊 Monthly Capture Volume Data:")
    print(f"   Rows: {len(df)}")
    print(f"   Columns: {list(df.columns)}")
    print(f"   Sample data:")
    print(df.head(3).to_string())
    print()
    
    # Test totals configuration
    sheet_name = "Monthly Capture Volume"
    sheet_type = "unknown"  # This isn't a time series
    
    print(f"🔧 Testing Totals Configuration:")
    config = pipeline_creator._get_totals_config_for_sheet(sheet_name, df, sheet_type)
    
    if config:
        print(f"   ✅ Configuration generated:")
        for key, value in config.items():
            print(f"      {key}: {value}")
    else:
        print(f"   ❌ No configuration generated")
        return
    
    # Test totals manager directly
    totals_manager = TotalsManager()
    
    print(f"\n🧮 Testing Totals Calculation:")
    
    # Test column totals
    if config.get('add_column_totals', False):
        numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
        print(f"   Numeric columns: {numeric_cols}")
        
        if numeric_cols:
            column_totals = totals_manager.calculate_column_totals(df, numeric_columns=numeric_cols)
            print(f"   ✅ Column totals: {dict(column_totals)}")
        else:
            print(f"   ⚠️  No numeric columns found")
    
    # Test row totals
    if config.get('add_row_totals', False):
        print(f"   Testing row totals...")
        # Row totals would be calculated per row, not tested here
    
    # Test actual worksheet application
    print(f"\n📋 Testing Worksheet Application:")
    
    # Create a test workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write data
    for row_idx, (_, row) in enumerate(df.iterrows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    print(f"   📊 Worksheet created with {len(df)} rows of data")
    
    # Apply totals using the totals manager
    try:
        start_row = 2  # Where data starts
        start_col = 1  # Column A
        
        print(f"   🔧 Applying totals with config:")
        print(f"      start_row: {start_row}")
        print(f"      start_col: {start_col}")
        print(f"      config: {config}")
        
        totals_manager.add_totals_to_worksheet(
            worksheet=ws,
            dataframe=df,
            start_row=start_row,
            start_col=start_col,
            config=config
        )
        
        print(f"   ✅ Totals applied successfully!")
        
        # Check what was added
        print(f"   📋 Worksheet after totals:")
        print(f"      Total rows: {ws.max_row}")
        print(f"      Total columns: {ws.max_column}")
        
        # Check last few rows for totals
        for row_num in range(max(1, ws.max_row - 2), ws.max_row + 1):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_values.append(str(cell_value) if cell_value is not None else "")
            print(f"      Row {row_num}: {row_values}")
        
        # Save test file
        test_file = "test_monthly_capture_totals.xlsx"
        wb.save(test_file)
        print(f"   💾 Test file saved: {test_file}")
        
    except Exception as e:
        print(f"   ❌ Error applying totals: {e}")
        import traceback
        traceback.print_exc()

def main():
    """
    Main test function.
    """
    try:
        test_monthly_capture_totals()
        
        print(f"\n" + "="*50)
        print(f"🎯 TEST SUMMARY")
        print(f"="*50)
        print(f"This test should help identify:")
        print(f"1. Whether totals configuration is generated correctly")
        print(f"2. Whether numeric columns are detected properly")
        print(f"3. Whether totals manager can apply totals to worksheet")
        print(f"4. What the final worksheet looks like with totals")
        
    except Exception as e:
        print(f"❌ Test error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
