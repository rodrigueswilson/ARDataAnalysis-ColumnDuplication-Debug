#!/usr/bin/env python3
"""
Direct Excel Post-Processing Fix

Since all our attempts to patch the sheet creation logic have failed,
this script directly modifies the Excel file after generation to apply
the zero-fill logic and fix the ACF/PACF sheets.
"""

import sys
import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta
import yaml

# Add the project root to the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

def load_collection_day_map():
    """Load the collection day map from config.yaml."""
    try:
        with open('config/config.yaml', 'r') as f:
            config = yaml.safe_load(f)
        return config.get('collection_day_map', {})
    except Exception as e:
        print(f"Error loading config: {e}")
        return {}

def generate_complete_date_range():
    """Generate the complete date range that should be in ACF/PACF sheets."""
    collection_day_map = load_collection_day_map()
    
    # Extract all dates from the collection day map
    all_dates = []
    for week_data in collection_day_map.values():
        if isinstance(week_data, dict):
            for date_str in week_data.get('dates', []):
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                    all_dates.append(date_obj)
                except ValueError:
                    continue
    
    if not all_dates:
        print("No dates found in collection day map")
        return []
    
    # Sort dates and create complete range
    all_dates.sort()
    start_date = all_dates[0]
    end_date = all_dates[-1]
    
    print(f"Date range: {start_date} to {end_date}")
    
    # Generate complete date range
    complete_dates = []
    current_date = start_date
    while current_date <= end_date:
        complete_dates.append(current_date)
        current_date += timedelta(days=1)
    
    return complete_dates

def fix_daily_counts_sheet(file_path):
    """Fix the Daily Counts (ACF_PACF) sheet by applying zero-fill logic."""
    print(f"Fixing Daily Counts sheet in {file_path}")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        if 'Daily Counts (ACF_PACF)' not in wb.sheetnames:
            print("Daily Counts (ACF_PACF) sheet not found")
            return False
        
        ws = wb['Daily Counts (ACF_PACF)']
        
        # Read existing data
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Skip empty rows
                data.append(list(row))
        
        print(f"Original data has {len(data)} rows")
        
        # Convert to DataFrame for easier manipulation
        df = pd.DataFrame(data)
        if df.empty:
            print("No data found in sheet")
            return False
        
        # Identify date column (should be first column)
        date_col = 0
        
        # Convert dates to datetime
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
        df = df.dropna(subset=[date_col])
        
        if df.empty:
            print("No valid dates found")
            return False
        
        # Generate complete date range
        complete_dates = generate_complete_date_range()
        if not complete_dates:
            print("Could not generate complete date range")
            return False
        
        # Create complete date DataFrame
        complete_df = pd.DataFrame({
            date_col: pd.to_datetime(complete_dates)
        })
        
        # Merge with existing data to fill gaps
        merged_df = complete_df.merge(df, on=date_col, how='left')
        
        # Fill missing values with 0 for numeric columns
        for col in merged_df.columns:
            if col != date_col and merged_df[col].dtype in ['int64', 'float64']:
                merged_df[col] = merged_df[col].fillna(0)
        
        print(f"After zero-fill: {len(merged_df)} rows")
        
        # Clear the worksheet
        ws.delete_rows(2, ws.max_row)
        
        # Write the fixed data back
        for idx, row in merged_df.iterrows():
            ws.append(row.tolist())
        
        # Save the workbook
        wb.save(file_path)
        print(f"✅ Successfully fixed Daily Counts sheet")
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing Daily Counts sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def fix_weekly_counts_sheet(file_path):
    """Fix the Weekly Counts (ACF_PACF) sheet by applying zero-fill logic."""
    print(f"Fixing Weekly Counts sheet in {file_path}")
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path)
        
        if 'Weekly Counts (ACF_PACF)' not in wb.sheetnames:
            print("Weekly Counts (ACF_PACF) sheet not found")
            return False
        
        ws = wb['Weekly Counts (ACF_PACF)']
        
        # Read existing data
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:  # Skip empty rows
                data.append(list(row))
        
        print(f"Original weekly data has {len(data)} rows")
        
        # Convert to DataFrame
        df = pd.DataFrame(data)
        if df.empty:
            print("No weekly data found in sheet")
            return False
        
        # Generate complete weekly range (47 weeks)
        complete_weeks = list(range(1, 48))  # Weeks 1-47
        
        # Create complete weekly DataFrame
        complete_df = pd.DataFrame({
            0: complete_weeks  # Week column
        })
        
        # Merge with existing data
        merged_df = complete_df.merge(df, on=0, how='left')
        
        # Fill missing values with 0 for numeric columns
        for col in merged_df.columns:
            if col != 0 and merged_df[col].dtype in ['int64', 'float64']:
                merged_df[col] = merged_df[col].fillna(0)
        
        print(f"After weekly zero-fill: {len(merged_df)} rows")
        
        # Clear the worksheet
        ws.delete_rows(2, ws.max_row)
        
        # Write the fixed data back
        for idx, row in merged_df.iterrows():
            ws.append(row.tolist())
        
        # Save the workbook
        wb.save(file_path)
        print(f"✅ Successfully fixed Weekly Counts sheet")
        
        return True
        
    except Exception as e:
        print(f"❌ Error fixing Weekly Counts sheet: {e}")
        import traceback
        traceback.print_exc()
        return False

def find_latest_excel_file():
    """Find the latest Excel report file."""
    excel_files = [f for f in os.listdir('.') if f.startswith('AR_Analysis_Report_') and f.endswith('.xlsx')]
    if not excel_files:
        return None
    
    # Sort by modification time
    excel_files.sort(key=lambda x: os.path.getmtime(x), reverse=True)
    return excel_files[0]

def main():
    print("DIRECT EXCEL POST-PROCESSING FIX")
    print("=" * 50)
    
    # Find the latest Excel file
    excel_file = find_latest_excel_file()
    if not excel_file:
        print("❌ No Excel report files found")
        return 1
    
    print(f"📊 Processing: {excel_file}")
    
    # Create backup
    backup_file = excel_file.replace('.xlsx', '_backup.xlsx')
    import shutil
    shutil.copy2(excel_file, backup_file)
    print(f"📋 Backup created: {backup_file}")
    
    # Apply fixes
    success_count = 0
    
    if fix_daily_counts_sheet(excel_file):
        success_count += 1
    
    if fix_weekly_counts_sheet(excel_file):
        success_count += 1
    
    print(f"\n{'=' * 50}")
    print(f"RESULTS: {success_count}/2 sheets fixed successfully")
    
    if success_count > 0:
        print(f"✅ Fixed Excel file: {excel_file}")
        print(f"📋 Original backed up as: {backup_file}")
        
        # Verify the fix
        print(f"\n🔍 Verifying fix...")
        os.system("python verify_excel_fix.py")
    else:
        print(f"❌ No sheets were fixed")
    
    return 0 if success_count > 0 else 1

if __name__ == "__main__":
    sys.exit(main())
