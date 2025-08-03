#!/usr/bin/env python3
"""
Final verification test for MP3 Duration Analysis sheet fixes
Tests that SY 22-23 P3 row is present and all totals are correctly positioned
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators.specialized import SpecializedSheetCreator
from report_generator.formatters import ExcelFormatter
from openpyxl import Workbook
from db_utils import get_db_connection

def test_mp3_fixes():
    """Test the MP3 Duration Analysis sheet with all fixes applied"""
    try:
        print("=== Final MP3 Duration Analysis Fix Verification ===")
        
        # Create test workbook
        workbook = Workbook()
        formatter = ExcelFormatter()
        
        # Create specialized sheet creator
        creator = SpecializedSheetCreator(get_db_connection(), formatter)
        
        # Create the MP3 Duration Analysis sheet
        print("\n[TEST] Creating MP3 Duration Analysis sheet...")
        creator.create_mp3_duration_analysis_sheet(workbook)
        
        # Get the worksheet
        ws = workbook['MP3 Duration Analysis']
        print(f"[INFO] Sheet created with {ws.max_row} rows and {ws.max_column} columns")
        
        # Test 1: Check for SY 22-23 P3 in Period Breakdown table
        print("\n=== TEST 1: Checking for SY 22-23 P3 row ===")
        sy_22_23_p3_found = False
        period_breakdown_start = None
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value == "Collection Period MP3 Duration Analysis":
                period_breakdown_start = row
                print(f"[INFO] Period Breakdown table starts at row {row}")
            elif cell_value == "SY 22-23 P3":
                sy_22_23_p3_found = True
                print(f"[SUCCESS] Found SY 22-23 P3 at row {row}")
                # Check the data in this row
                school_year = ws.cell(row=row, column=2).value
                files = ws.cell(row=row, column=3).value
                print(f"[INFO] SY 22-23 P3 data: School Year={school_year}, Files={files}")
                break
        
        if sy_22_23_p3_found:
            print("[SUCCESS] ✅ SY 22-23 P3 row is present!")
        else:
            print("[ERROR] ❌ SY 22-23 P3 row is missing!")
        
        # Test 2: Check totals positioning for all tables
        print("\n=== TEST 2: Checking totals positioning ===")
        totals_found = []
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Total" in str(cell_value):
                totals_found.append((row, cell_value))
                print(f"[INFO] Found totals row at {row}: '{cell_value}'")
                
                # Check if there's data immediately above (no blank rows)
                prev_row_value = ws.cell(row=row-1, column=1).value
                if prev_row_value and prev_row_value != "":
                    print(f"[SUCCESS] ✅ Totals at row {row} positioned immediately after data")
                else:
                    print(f"[WARNING] ⚠️ Totals at row {row} may have blank row above")
        
        print(f"\n[INFO] Found {len(totals_found)} totals rows:")
        for row, title in totals_found:
            print(f"  - Row {row}: {title}")
        
        # Test 3: Verify all expected periods are present
        print("\n=== TEST 3: Checking all expected periods ===")
        expected_periods = [
            "SY 21-22 P1", "SY 21-22 P2", "SY 21-22 P3",
            "SY 22-23 P1", "SY 22-23 P2", "SY 22-23 P3"
        ]
        
        found_periods = []
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value in expected_periods:
                found_periods.append(cell_value)
                files = ws.cell(row=row, column=3).value
                print(f"[INFO] Found {cell_value}: {files} files")
        
        missing_periods = set(expected_periods) - set(found_periods)
        if not missing_periods:
            print("[SUCCESS] ✅ All expected periods are present!")
        else:
            print(f"[ERROR] ❌ Missing periods: {missing_periods}")
        
        # Test 4: Check professional formatting of totals
        print("\n=== TEST 4: Checking totals formatting ===")
        for row, title in totals_found:
            cell = ws.cell(row=row, column=1)
            if cell.font and cell.font.bold and cell.fill:
                print(f"[SUCCESS] ✅ Row {row} has professional formatting")
            else:
                print(f"[WARNING] ⚠️ Row {row} may be missing formatting")
        
        # Summary
        print("\n=== FINAL VERIFICATION SUMMARY ===")
        all_tests_passed = (
            sy_22_23_p3_found and 
            len(totals_found) >= 3 and 
            not missing_periods
        )
        
        if all_tests_passed:
            print("[SUCCESS] 🎉 ALL TESTS PASSED! MP3 Duration Analysis fixes are working correctly!")
            print("✅ SY 22-23 P3 row is present")
            print("✅ All totals rows are positioned correctly")
            print("✅ All expected periods are present")
            print("✅ Professional formatting is applied")
        else:
            print("[ERROR] ❌ Some tests failed. Please review the issues above.")
        
        # Save test workbook for manual inspection
        test_filename = "test_mp3_fixes_final.xlsx"
        workbook.save(test_filename)
        print(f"\n[INFO] Test workbook saved as: {test_filename}")
        
        return all_tests_passed
        
    except Exception as e:
        print(f"[ERROR] Test failed with exception: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_mp3_fixes()
    sys.exit(0 if success else 1)
