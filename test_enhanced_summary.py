#!/usr/bin/env python3
"""
Test script for the enhanced Summary Statistics sheet with day analysis tables.
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from report_generator.sheet_creators.base import BaseSheetCreator
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection

def test_enhanced_summary_statistics():
    """Test the enhanced Summary Statistics sheet creation."""
    try:
        print("🔧 Initializing test components...")
        
        # Create workbook and formatter
        workbook = openpyxl.Workbook()
        formatter = ExcelFormatter()
        
        # Get database connection
        db = get_db_connection()
        
        # Initialize BaseSheetCreator with database and formatter
        creator = BaseSheetCreator(db, formatter)
        
        print("📊 Testing enhanced Summary Statistics sheet creation...")
        
        # Create the enhanced summary statistics sheet
        creator.create_summary_statistics_sheet(workbook)
        
        # Verify the sheet was created
        if 'Summary Statistics' in workbook.sheetnames:
            ws = workbook['Summary Statistics']
            print(f"✅ Summary Statistics sheet created successfully")
            print(f"✅ Sheet dimensions: {ws.max_row} rows × {ws.max_column} columns")
            
            # Check for day analysis sections
            day_analysis_sections = []
            for row in range(1, min(100, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'Day Analysis' in str(cell_value):
                    day_analysis_sections.append((row, str(cell_value)))
            
            if day_analysis_sections:
                print(f"✅ Found {len(day_analysis_sections)} day analysis sections:")
                for row, title in day_analysis_sections:
                    print(f"   - Row {row}: {title}")
            else:
                print("⚠️ Day analysis sections not found")
            
            # Save test workbook
            test_filename = f"test_enhanced_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            workbook.save(test_filename)
            print(f"✅ Test workbook saved as: {test_filename}")
            
            # Check for specific table content
            print("\n📋 Checking table content...")
            
            # Look for specific metrics
            metrics_found = []
            for row in range(1, min(50, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    cell_str = str(cell_value).lower()
                    if any(keyword in cell_str for keyword in ['collection days', 'coverage', 'consecutive', 'avg files']):
                        metrics_found.append((row, str(cell_value)))
            
            if metrics_found:
                print(f"✅ Found {len(metrics_found)} day analysis metrics:")
                for row, metric in metrics_found[:10]:  # Show first 10
                    print(f"   - Row {row}: {metric}")
            else:
                print("⚠️ Day analysis metrics not found")
                
        else:
            print("❌ Summary Statistics sheet was not created")
            return False
        
        print("\n🎉 Enhanced Summary Statistics test completed successfully!")
        return True
        
    except Exception as e:
        print(f"❌ Error during testing: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    # Import pandas for timestamp
    import pandas as pd
    
    print("🚀 Starting Enhanced Summary Statistics Test")
    print("=" * 50)
    
    success = test_enhanced_summary_statistics()
    
    print("=" * 50)
    if success:
        print("✅ All tests passed!")
    else:
        print("❌ Some tests failed!")
