#!/usr/bin/env python3
"""
Debug script to test Data Cleaning totals functionality
"""

import pandas as pd
from openpyxl import Workbook
from report_generator.totals_manager import TotalsManager
from report_generator.formatters import ExcelFormatter

def test_data_cleaning_totals():
    """Test the totals functionality for Data Cleaning sheet"""
    
    # Create sample data similar to what the Data Cleaning sheet uses
    sample_data = [
        {'file_type': 'JPG', 'outlier_status': False, 'count': 6764, 'total_size_mb': 37183.35, 'avg_size_mb': 5.5},
        {'file_type': 'JPG', 'outlier_status': True, 'count': 293, 'total_size_mb': 2159.86, 'avg_size_mb': 7.37},
        {'file_type': 'MP3', 'outlier_status': False, 'count': 3134, 'total_size_mb': 3289.35, 'avg_size_mb': 1.05},
        {'file_type': 'MP3', 'outlier_status': True, 'count': 50, 'total_size_mb': 661.6, 'avg_size_mb': 13.23}
    ]
    
    # Create DataFrame for totals calculation with proper column names
    totals_df = pd.DataFrame()
    for record in sample_data:
        totals_df = pd.concat([totals_df, pd.DataFrame({
            'File Type': [record['file_type']],
            'Outlier Status': [record['outlier_status']],
            'Count': [record['count']],
            'Total Size (MB)': [round(record['total_size_mb'], 2)],
            'Avg Size (MB)': [round(record['avg_size_mb'], 2)]
        })], ignore_index=True)
    
    print("DataFrame created:")
    print(totals_df)
    print(f"\nDataFrame shape: {totals_df.shape}")
    print(f"DataFrame columns: {list(totals_df.columns)}")
    
    # Test totals calculation
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Data Cleaning"
    
    # Add headers
    headers = ['File Type', 'Outlier Status', 'Count', 'Total Size (MB)', 'Avg Size (MB)']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Add data
    for row_idx, (_, record) in enumerate(totals_df.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=record['File Type'])
        ws.cell(row=row_idx, column=2, value=record['Outlier Status'])
        ws.cell(row=row_idx, column=3, value=record['Count'])
        ws.cell(row=row_idx, column=4, value=record['Total Size (MB)'])
        ws.cell(row=row_idx, column=5, value=record['Avg Size (MB)'])
    
    print(f"\nData added to Excel. Last data row: {len(totals_df) + 1}")
    
    # Initialize totals manager
    formatter = ExcelFormatter()
    totals_manager = TotalsManager(formatter)
    
    # Configure totals
    totals_config = {
        'include_column_totals': True,
        'include_row_totals': False,
        'include_grand_total': False,
        'numeric_columns': ['Count', 'Total Size (MB)'],  # Exclude Avg Size as it's not summable
        'totals_row_label': 'TOTAL'
    }
    
    print(f"\nTotals config: {totals_config}")
    
    try:
        print("\nAttempting to add totals...")
        totals_positions = totals_manager.add_totals_to_worksheet(
            ws, totals_df, start_row=2, start_col=1, config=totals_config
        )
        
        print(f"Totals added successfully!")
        print(f"Totals positions: {totals_positions}")
        
        # Check the totals row
        if 'totals_row' in totals_positions:
            totals_row = totals_positions['totals_row']
            print(f"\nTotals row {totals_row} contents:")
            for col in range(1, 6):
                cell_value = ws.cell(row=totals_row, column=col).value
                print(f"  Column {col}: {cell_value}")
        
        # Save test file
        wb.save("test_data_cleaning_totals.xlsx")
        print(f"\nTest file saved: test_data_cleaning_totals.xlsx")
        
        return True
        
    except Exception as e:
        print(f"Error adding totals: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    test_data_cleaning_totals()
