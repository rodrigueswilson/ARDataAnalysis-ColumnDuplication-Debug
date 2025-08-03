#!/usr/bin/env python3
"""
Quick verification script to check if Table 2 (Year-by-Year Breakdown) is now properly populated
"""

import openpyxl
import os
import glob

def verify_table2_fix():
    """Verify that Table 2 now has data populated"""
    
    print("=== Table 2 Fix Verification ===")
    
    try:
        # Find the most recent Excel file
        excel_files = glob.glob("AR_Analysis_Report_*.xlsx")
        if not excel_files:
            print("❌ No Excel files found")
            return
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📊 Examining: {latest_file}")
        
        # Open the workbook
        wb = openpyxl.load_workbook(latest_file)
        
        if "Data Cleaning" not in wb.sheetnames:
            print("❌ No 'Data Cleaning' sheet found")
            return
        
        ws = wb["Data Cleaning"]
        print(f"✅ Found 'Data Cleaning' sheet")
        
        # Look for Table 2 content
        print(f"\n=== Searching for Table 2 Data ===")
        
        table2_found = False
        table2_data_rows = 0
        
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Table 2: Year-by-Year Breakdown" in str(cell_value):
                print(f"✅ Found Table 2 header at row {row}")
                table2_found = True
                
                # Look for data rows after the header
                print(f"\n📋 Table 2 Content:")
                for data_row in range(row + 3, min(row + 15, ws.max_row + 1)):  # Check next 12 rows
                    row_data = []
                    has_content = False
                    
                    for col in range(1, 10):  # 9 columns
                        cell_value = ws.cell(row=data_row, column=col).value
                        if cell_value is not None:
                            has_content = True
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    
                    if has_content:
                        # Check if this is a data row (not just headers)
                        if any("Files" in str(val) or str(val).isdigit() or "%" in str(val) for val in row_data):
                            table2_data_rows += 1
                            print(f"Row {data_row}: {' | '.join(row_data)}")
                    elif table2_data_rows > 0:
                        # Stop when we hit empty rows after finding data
                        break
                
                break
        
        if not table2_found:
            print("❌ Table 2 header not found")
            return
        
        print(f"\n🎯 Results:")
        print(f"• Table 2 header found: ✅")
        print(f"• Data rows found: {table2_data_rows}")
        
        if table2_data_rows >= 4:  # Expecting 4 year/type combinations + total
            print(f"✅ SUCCESS: Table 2 is properly populated with {table2_data_rows} data rows!")
            print("✅ The collection variable scope issue has been fixed!")
        elif table2_data_rows > 0:
            print(f"⚠️  PARTIAL: Table 2 has some data ({table2_data_rows} rows) but may be incomplete")
        else:
            print(f"❌ FAILURE: Table 2 still has no data rows")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    verify_table2_fix()
