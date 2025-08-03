#!/usr/bin/env python3
"""
Simple debug script to test Data Cleaning sheet creation
"""

import openpyxl
import sys
import os

# Add the project directory to the path
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators import SheetCreator

def test_data_cleaning_sheet():
    """Test the Data Cleaning sheet creation directly"""
    
    print("=== Simple Data Cleaning Sheet Debug Test ===")
    
    try:
        # Create a SheetCreator instance
        print("[1] Creating SheetCreator...")
        sheet_creator = SheetCreator()
        
        # Create a workbook
        print("[2] Creating workbook...")
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Test the data cleaning sheet creation
        print("[3] Creating Data Cleaning sheet...")
        sheet_creator.create_data_cleaning_sheet(wb)
        
        # Check if the sheet was created
        if 'Data Cleaning' in wb.sheetnames:
            ws = wb['Data Cleaning']
            print(f"✅ Data Cleaning sheet created with {ws.max_row} rows and {ws.max_column} columns")
            
            # Print first few rows to see content
            print("\n=== Sheet Content Preview ===")
            for row in range(1, min(11, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(9, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None:
                        row_data.append(str(cell_value)[:30])
                    else:
                        row_data.append("")
                if any(row_data):
                    print(f"Row {row}: {' | '.join(row_data)}")
            
            # Save test file
            test_file = "test_data_cleaning_debug.xlsx"
            wb.save(test_file)
            print(f"\n✅ Test file saved as: {test_file}")
            
        else:
            print("❌ Data Cleaning sheet was not created")
            print(f"Available sheets: {wb.sheetnames}")
        
    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_data_cleaning_sheet()
