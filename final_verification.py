#!/usr/bin/env python3
"""
Final verification of MP3 Duration Analysis sheet in the latest report.
"""

import openpyxl
import sys

def verify_mp3_sheet():
    try:
        print("=== FINAL MP3 DURATION ANALYSIS VERIFICATION ===")
        
        # Load the latest report
        report_file = 'AR_Analysis_Report_20250726_173817.xlsx'
        print(f"Loading report: {report_file}")
        
        wb = openpyxl.load_workbook(report_file)
        sheet_names = wb.sheetnames
        
        print(f"\nTotal sheets in report: {len(sheet_names)}")
        print("\nAll sheets in the report:")
        for i, name in enumerate(sheet_names, 1):
            marker = " *** MP3 DURATION ANALYSIS ***" if name == 'MP3 Duration Analysis' else ""
            print(f"{i:2d}. {name}{marker}")
        
        # Check specifically for MP3 Duration Analysis sheet
        if 'MP3 Duration Analysis' in sheet_names:
            print(f"\n🎉 SUCCESS: MP3 Duration Analysis sheet FOUND!")
            
            # Get sheet details
            mp3_sheet = wb['MP3 Duration Analysis']
            print(f"   Sheet dimensions: {mp3_sheet.max_row} rows x {mp3_sheet.max_column} columns")
            
            # Show some sample content
            print(f"\n   Sample content from first few cells:")
            for row in range(1, min(6, mp3_sheet.max_row + 1)):
                for col in range(1, min(4, mp3_sheet.max_column + 1)):
                    cell_value = mp3_sheet.cell(row=row, column=col).value
                    if cell_value:
                        print(f"   [{row},{col}]: {cell_value}")
            
            return True
        else:
            print(f"\n❌ ERROR: MP3 Duration Analysis sheet NOT FOUND!")
            print(f"   Available sheets: {sheet_names}")
            return False
            
    except FileNotFoundError:
        print(f"❌ ERROR: Report file not found: {report_file}")
        return False
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = verify_mp3_sheet()
    print(f"\n=== VERIFICATION {'PASSED' if success else 'FAILED'} ===")
    sys.exit(0 if success else 1)
