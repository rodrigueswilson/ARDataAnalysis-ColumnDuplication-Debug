#!/usr/bin/env python3
"""
Test the Summary Statistics sheet creation to identify the error.
"""

import sys
import traceback
from report_generator.sheet_creators import SheetCreator
from openpyxl import Workbook

def test_summary_stats_creation():
    """Test creating the Summary Statistics sheet to identify errors."""
    print("Testing Summary Statistics sheet creation...")
    
    try:
        # Create a test workbook
        workbook = Workbook()
        
        # Create sheet creator instance
        sheet_creator = SheetCreator()
        
        print("Attempting to create Summary Statistics sheet...")
        
        # Try to create the Summary Statistics sheet
        sheet_creator.create_summary_statistics_sheet(workbook)
        
        print("✅ Summary Statistics sheet created successfully!")
        
        # Check if the sheet was actually created
        if "Summary Statistics" in workbook.sheetnames:
            print("✅ Summary Statistics sheet found in workbook")
            ws = workbook["Summary Statistics"]
            print(f"   Sheet has {ws.max_row} rows and {ws.max_column} columns")
        else:
            print("❌ Summary Statistics sheet not found in workbook")
            print(f"   Available sheets: {workbook.sheetnames}")
        
    except Exception as e:
        print(f"❌ Error creating Summary Statistics sheet:")
        print(f"   Error: {e}")
        print(f"   Error type: {type(e).__name__}")
        print("\nFull traceback:")
        traceback.print_exc()
        
        # Try to identify the specific issue
        if "non_collection_days" in str(e):
            print("\n🔍 Likely issue: Problem with non_collection_days filtering")
        elif "_load_config" in str(e):
            print("\n🔍 Likely issue: Problem with config loading")
        elif "date_str" in str(e):
            print("\n🔍 Likely issue: Problem with date string conversion")
        else:
            print(f"\n🔍 Unknown issue: {e}")

if __name__ == "__main__":
    test_summary_stats_creation()
