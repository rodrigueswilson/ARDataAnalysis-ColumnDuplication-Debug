#!/usr/bin/env python3
"""
Quick verification script to check if MP3 Duration Analysis sheet was created successfully.
"""

import openpyxl
import sys

def verify_mp3_sheet():
    try:
        # Load the latest report
        wb = openpyxl.load_workbook('AR_Analysis_Report_20250726_165833.xlsx')
        
        print("=== REPORT VERIFICATION ===")
        print(f"Total sheets: {len(wb.sheetnames)}")
        print("\nAll sheets:")
        for i, name in enumerate(wb.sheetnames, 1):
            print(f"  {i:2d}. {name}")
        
        # Check if MP3 Duration Analysis sheet exists
        if 'MP3 Duration Analysis' in wb.sheetnames:
            print("\n✅ MP3 Duration Analysis sheet FOUND!")
            
            # Get sheet details
            mp3_sheet = wb['MP3 Duration Analysis']
            print(f"   Dimensions: {mp3_sheet.max_row} rows x {mp3_sheet.max_column} columns")
            
            # Show first few rows to verify content
            print("\n   First 10 rows of content:")
            for row_num in range(1, min(11, mp3_sheet.max_row + 1)):
                row_data = []
                for col_num in range(1, min(6, mp3_sheet.max_column + 1)):  # Show first 5 columns
                    cell_value = mp3_sheet.cell(row=row_num, column=col_num).value
                    if cell_value is None:
                        cell_value = ""
                    row_data.append(str(cell_value)[:20])  # Truncate long values
                print(f"   Row {row_num:2d}: {' | '.join(row_data)}")
            
            return True
        else:
            print("\n❌ MP3 Duration Analysis sheet NOT FOUND!")
            return False
            
    except FileNotFoundError:
        print("❌ Report file not found!")
        return False
    except Exception as e:
        print(f"❌ Error verifying sheet: {e}")
        return False

if __name__ == "__main__":
    success = verify_mp3_sheet()
    sys.exit(0 if success else 1)
