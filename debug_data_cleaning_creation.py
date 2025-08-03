#!/usr/bin/env python3
"""
Debug script to test Data Cleaning sheet creation specifically
"""

from openpyxl import Workbook
from report_generator.sheet_creators.base import BaseSheetCreator
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection

def test_data_cleaning_creation():
    """Test the Data Cleaning sheet creation with totals"""
    
    try:
        # Get database connection
        db = get_db_connection()
        print("✅ Database connection established")
        
        # Create formatter
        formatter = ExcelFormatter()
        print("✅ Formatter created")
        
        # Create sheet creator
        sheet_creator = BaseSheetCreator(db, formatter)
        print("✅ BaseSheetCreator initialized")
        
        # Create workbook
        wb = Workbook()
        print("✅ Workbook created")
        
        # Test the data cleaning sheet creation
        print("\n🔍 Creating Data Cleaning sheet...")
        sheet_creator.create_data_cleaning_sheet(wb)
        
        # Check if the sheet was created
        if 'Data Cleaning' in wb.sheetnames:
            ws = wb['Data Cleaning']
            print(f"✅ Data Cleaning sheet created with {ws.max_row} rows and {ws.max_column} columns")
            
            # Check the content
            print("\n📋 Sheet content:")
            for row in range(1, min(15, ws.max_row + 1)):
                row_data = []
                for col in range(1, min(6, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(str(cell_value) if cell_value is not None else 'None')
                print(f"Row {row}: {row_data}")
                
            # Save test file
            wb.save("debug_data_cleaning_creation.xlsx")
            print(f"\n💾 Test file saved: debug_data_cleaning_creation.xlsx")
            
        else:
            print("❌ Data Cleaning sheet was not created")
            print(f"Available sheets: {wb.sheetnames}")
            
    except Exception as e:
        print(f"❌ Error during test: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    test_data_cleaning_creation()
