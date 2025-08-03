"""
ACF/PACF Dashboard Generator
Creates comparative visualizations across all time scales in the dashboard sheet.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_dashboard_summary(wb):
    """
    Create a comprehensive ACF/PACF dashboard with comparative analysis
    across all time scales.
    """
    print("[DEBUG] Starting create_dashboard_summary function...")
    try:
        
        # Handle dashboard sheet creation and cleanup
        # PRESERVE the original "Dashboard" sheet - it contains important high-level summary data
        # Only handle ACF_PACF_Dashboard deduplication
        
        print(f"[DEBUG] Current sheets in workbook: {wb.sheetnames}")
        
        # Remove any duplicate ACF_PACF_Dashboard sheets (e.g., ACF_PACF_Dashboard1, ACF_PACF_Dashboard2, etc.)
        duplicate_sheets = [name for name in wb.sheetnames if name.startswith('ACF_PACF_Dashboard') and name != 'ACF_PACF_Dashboard']
        for duplicate in duplicate_sheets:
            wb.remove(wb[duplicate])
            print(f"[REMOVE] Removed duplicate dashboard sheet: {duplicate}")
        
        # Find the correct position for the dashboard (after ACF/PACF sheets)
        acf_pacf_sheet_names = [
            'Daily Counts (ACF_PACF)',
            'Weekly Counts (ACF_PACF)', 
            'Biweekly Counts (ACF_PACF)',
            'Monthly Counts (ACF_PACF)',
            'Period Counts (ACF_PACF)'
        ]
        
        print(f"[SHEETS] Current sheet order: {wb.sheetnames}")
        
        # Find the position after the last ACF/PACF sheet
        target_position = None
        last_acf_pacf_sheet = None
        
        for i, sheet_name in enumerate(wb.sheetnames):
            if sheet_name in acf_pacf_sheet_names:
                target_position = i + 1
                last_acf_pacf_sheet = sheet_name
                print(f"[FOUND] Found ACF/PACF sheet '{sheet_name}' at position {i + 1}")
        
        if target_position is not None:
            print(f"[TARGET] Target position for dashboard: {target_position + 1} (after '{last_acf_pacf_sheet}')")
        else:
            # If no ACF/PACF sheets found, place at end
            target_position = len(wb.sheetnames)
            print(f"[TARGET] No ACF/PACF sheets found, placing dashboard at end (position {target_position + 1})")
        
        # Handle existing dashboard sheet
        if 'ACF_PACF_Dashboard' in wb.sheetnames:
            # Remove existing dashboard to avoid duplication
            wb.remove(wb['ACF_PACF_Dashboard'])
            print("[REMOVE] Removed existing ACF_PACF_Dashboard to prevent duplication")
            # Adjust target position if we removed a sheet before the target
            current_dashboard_pos = wb.sheetnames.index('ACF_PACF_Dashboard') if 'ACF_PACF_Dashboard' in wb.sheetnames else -1
            if current_dashboard_pos >= 0 and current_dashboard_pos < target_position:
                target_position -= 1
        
        # Create the dashboard sheet at the correct position from the start
        dashboard_ws = wb.create_sheet('ACF_PACF_Dashboard')
        print(f"[CREATE] Created ACF_PACF_Dashboard at position {target_position + 1} (no movement needed)")
        
        print("[DEBUG] Dashboard sheet prepared, extracting ACF/PACF data...")
        
        # Extract ACF/PACF data from all sheets
        try:
            acf_pacf_data = extract_acf_pacf_from_sheets(wb)
            print(f"[DEBUG] Extracted data for {len(acf_pacf_data)} time scales: {list(acf_pacf_data.keys())}")
        except Exception as e:
            print(f"[ERROR] Failed to extract ACF/PACF data: {e}")
            import traceback
            traceback.print_exc()
            return
        
        if not acf_pacf_data:
            print("[WARNING] No ACF/PACF data found. Adding 'no data' message.")
            add_no_data_message(dashboard_ws)
            return
        
        # Create dashboard components with individual error handling
        print("[DEBUG] Creating dashboard header...")
        try:
            create_dashboard_header(dashboard_ws)
            print("[DEBUG] Dashboard header created successfully")
        except Exception as e:
            print(f"[ERROR] Failed to create dashboard header: {e}")
            import traceback
            traceback.print_exc()
        
        print("[DEBUG] Creating summary table...")
        try:
            create_summary_table(dashboard_ws, acf_pacf_data)
            print("[DEBUG] Summary table created successfully")
        except Exception as e:
            print(f"[ERROR] Failed to create summary table: {e}")
            import traceback
            traceback.print_exc()
        
        print("[DEBUG] Creating comparative charts...")
        try:
            create_comparative_charts(dashboard_ws, acf_pacf_data)
            print("[DEBUG] Comparative charts created successfully")
        except Exception as e:
            print(f"[ERROR] Failed to create comparative charts: {e}")
            import traceback
            traceback.print_exc()
        
        print("[DEBUG] Creating interpretation notes...")
        try:
            create_interpretation_notes(dashboard_ws)
            print("[DEBUG] Interpretation notes created successfully")
        except Exception as e:
            print(f"[ERROR] Failed to create interpretation notes: {e}")
            import traceback
            traceback.print_exc()
        
        print("[DEBUG] Styling dashboard...")
        try:
            style_dashboard(dashboard_ws)
            print("[DEBUG] Dashboard styling completed successfully")
        except Exception as e:
            print(f"[ERROR] Failed to style dashboard: {e}")
            import traceback
            traceback.print_exc()
        
        print("[INFO] ACF/PACF Dashboard creation process completed!")
        print(f"[DEBUG] Final dashboard sheet has {dashboard_ws.max_row} rows and {dashboard_ws.max_column} columns")
        
    except Exception as e:
        print(f"[ERROR] Critical failure in create_dashboard_summary: {e}")
        import traceback
        traceback.print_exc()
        # Add fallback message to dashboard sheet if it exists
        try:
            if 'ACF_PACF_Dashboard' in wb.sheetnames:
                dashboard_ws = wb['ACF_PACF_Dashboard']
                dashboard_ws['A1'] = 'Dashboard Creation Failed'
                dashboard_ws['A2'] = f'Error: {str(e)}'
                print("[DEBUG] Added error message to dashboard sheet")
        except:
            print("[ERROR] Could not even add error message to dashboard sheet")

def extract_acf_pacf_from_sheets(wb):
    """Extract ACF/PACF values from all relevant sheets."""
    acf_pacf_data = {}
    
    # Define sheets to analyze (including Daily sheet that was previously omitted)
    target_sheets = [
        'Daily Counts (ACF_PACF)',
        'Weekly Counts (ACF_PACF)', 
        'Biweekly Counts (ACF_PACF)',
        'Monthly Counts (ACF_PACF)',
        'Period Counts (ACF_PACF)'
    ]
    
    for sheet_name in target_sheets:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        
        # Convert sheet to DataFrame for easier processing
        data = []
        headers = []
        
        # Get headers from first row
        for cell in ws[1]:
            if cell.value:
                headers.append(cell.value)
        
        # Get first data row (assuming global statistics)
        if ws.max_row > 1:
            first_data_row = []
            for cell in ws[2]:
                first_data_row.append(cell.value)
            
            # Create dictionary of ACF/PACF values
            sheet_data = {}
            for i, header in enumerate(headers):
                if i < len(first_data_row) and ('ACF_' in str(header) or 'PACF_' in str(header)):
                    value = first_data_row[i]
                    # Include both numeric values and meaningful placeholders
                    if isinstance(value, (int, float)):
                        sheet_data[header] = value
                    elif isinstance(value, str) and value.strip():
                        # Include non-empty string placeholders for dashboard display
                        sheet_data[header] = value
            
            if sheet_data:
                # Determine time scale
                time_scale = determine_time_scale(sheet_name)
                acf_pacf_data[time_scale] = sheet_data
    
    return acf_pacf_data

def determine_time_scale(sheet_name):
    """Determine time scale from sheet name."""
    if 'Daily' in sheet_name:
        return 'Daily'
    elif 'Biweekly' in sheet_name:
        return 'Biweekly'
    elif 'Weekly' in sheet_name:
        return 'Weekly'
    elif 'Monthly' in sheet_name:
        return 'Monthly'
    elif 'Period' in sheet_name:
        return 'Period'
    else:
        return 'Unknown'

def create_dashboard_header(ws):
    """Create dashboard title and description."""
    # Title
    ws['A1'] = 'ACF/PACF Analysis Dashboard'
    ws['A1'].font = Font(size=16, bold=True, color='003366')
    ws.merge_cells('A1:F1')
    
    # Description
    ws['A3'] = 'This dashboard provides a comparative view of autocorrelation and partial autocorrelation patterns across different time scales.'
    ws['A3'].font = Font(size=10)
    ws.merge_cells('A3:F3')
    
    ws['A4'] = 'Use this dashboard to identify consistent patterns and differences in temporal dependencies based on aggregation level.'
    ws['A4'].font = Font(size=10)
    ws.merge_cells('A4:F4')
    
    # Navigation instructions
    ws['A6'] = 'Click on any time scale name to navigate to the corresponding sheet.'
    ws['A6'].font = Font(size=10, italic=True, color='666666')
    ws.merge_cells('A6:F6')
    
    # Timestamp
    from datetime import datetime
    ws['A4'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A4'].font = Font(size=10, color='999999')
    ws.merge_cells('A4:F4')
    ws['A4'].alignment = Alignment(horizontal='center')

def create_summary_table(ws, acf_pacf_data):
    """Create summary table of ACF/PACF values across time scales."""
    start_row = 7
    
    # Table header
    ws[f'A{start_row}'] = 'Summary: ACF/PACF Values by Time Scale'
    ws[f'A{start_row}'].font = Font(size=14, bold=True, color='003366')
    ws.merge_cells(f'A{start_row}:F{start_row}')
    
    # Column headers
    headers = ['Time Scale', 'ACF Lag 1', 'PACF Lag 1', 'Strongest ACF', 'Strongest PACF', 'Pattern Type']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(start_row + 2, i, header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Map time scales to sheet names
    sheet_map = {
        'Daily': 'Daily Counts (ACF_PACF)',
        'Weekly': 'Weekly Counts (ACF_PACF)',
        'Biweekly': 'Biweekly Counts (ACF_PACF)',
        'Monthly': 'Monthly Counts (ACF_PACF)',
        'Period': 'Period Counts (ACF_PACF)'
    }
    
    # Data rows
    row = start_row + 3
    for time_scale, data in acf_pacf_data.items():
        # Create cell with time scale name
        time_scale_cell = ws.cell(row, 1, time_scale)
        
        # Add hyperlink to corresponding sheet if it exists
        target_sheet = sheet_map.get(time_scale)
        if target_sheet and target_sheet in ws.parent.sheetnames:
            # Create hyperlink formula
            time_scale_cell.hyperlink = f"#{target_sheet}!A1"
            time_scale_cell.font = Font(color="0563C1", underline="single")
            print(f"[LINK] Added hyperlink from dashboard to '{target_sheet}'")
        else:
            print(f"[WARNING] Could not create link for '{time_scale}', sheet '{target_sheet}' not found")
        
        # ACF Lag 1
        acf_lag1 = data.get('ACF_Lag1', 'N/A')
        ws.cell(row, 2, acf_lag1 if isinstance(acf_lag1, (int, float)) else acf_lag1)
        
        # PACF Lag 1
        pacf_lag1 = data.get('PACF_Lag1', 'N/A')
        ws.cell(row, 3, pacf_lag1 if isinstance(pacf_lag1, (int, float)) else pacf_lag1)
        
        # Strongest ACF (handle both numeric and string values)
        acf_values = {k: v for k, v in data.items() if 'ACF_' in k and isinstance(v, (int, float))}
        acf_strings = {k: v for k, v in data.items() if 'ACF_' in k and isinstance(v, str)}
        
        if acf_values:
            strongest_acf = max(acf_values.items(), key=lambda x: abs(x[1]))
            ws.cell(row, 4, f"{strongest_acf[0]}: {strongest_acf[1]:.3f}")
        elif acf_strings:
            # Show first string placeholder if no numeric values
            first_acf_string = list(acf_strings.items())[0]
            ws.cell(row, 4, f"{first_acf_string[0]}: {first_acf_string[1]}")
        else:
            ws.cell(row, 4, 'N/A')
        
        # Strongest PACF (handle both numeric and string values)
        pacf_values = {k: v for k, v in data.items() if 'PACF_' in k and isinstance(v, (int, float))}
        pacf_strings = {k: v for k, v in data.items() if 'PACF_' in k and isinstance(v, str)}
        
        if pacf_values:
            strongest_pacf = max(pacf_values.items(), key=lambda x: abs(x[1]))
            ws.cell(row, 5, f"{strongest_pacf[0]}: {strongest_pacf[1]:.3f}")
        elif pacf_strings:
            # Show first string placeholder if no numeric values
            first_pacf_string = list(pacf_strings.items())[0]
            ws.cell(row, 5, f"{first_pacf_string[0]}: {first_pacf_string[1]}")
        else:
            ws.cell(row, 5, 'N/A')
        
        # Pattern interpretation
        pattern_type = interpret_pattern(acf_values, pacf_values)
        ws.cell(row, 6, pattern_type)
        
        row += 1

def interpret_pattern(acf_values, pacf_values):
    """Interpret the overall pattern type based on ACF/PACF values."""
    if not acf_values and not pacf_values:
        return 'Insufficient Data'
    
    # Get lag 1 values if available
    acf_lag1 = acf_values.get('ACF_Lag1', 0)
    pacf_lag1 = pacf_values.get('PACF_Lag1', 0)
    
    if isinstance(acf_lag1, (int, float)) and isinstance(pacf_lag1, (int, float)):
        if abs(acf_lag1) > 0.7:
            return 'Strong Autocorrelation'
        elif abs(acf_lag1) > 0.3:
            return 'Moderate Autocorrelation'
        elif abs(pacf_lag1) > 0.3:
            return 'Direct Correlation'
        else:
            return 'Random/Weak Pattern'
    else:
        return 'Mixed/Complex Pattern'

def create_comparative_charts(ws, acf_pacf_data):
    """Create comparative charts for ACF/PACF values."""
    chart_start_row = 15
    
    # Prepare data for charting
    chart_data = prepare_chart_data(acf_pacf_data)
    
    if not chart_data:
        ws[f'A{chart_start_row}'] = 'No sufficient data for comparative charts'
        return
    
    # Create data table for charts
    ws[f'A{chart_start_row}'] = 'Comparative ACF/PACF Values'
    ws[f'A{chart_start_row}'].font = Font(size=14, bold=True, color='003366')
    
    # Write chart data
    data_start_row = chart_start_row + 2
    for i, (header, values) in enumerate(chart_data.items()):
        ws.cell(data_start_row, i + 1, header)
        for j, value in enumerate(values):
            ws.cell(data_start_row + j + 1, i + 1, value)
    
    # Create chart
    chart = LineChart()
    chart.title = "ACF/PACF Comparison Across Time Scales"
    chart.style = 10
    chart.y_axis.title = 'Correlation Coefficient'
    chart.x_axis.title = 'Time Scale'
    
    # Add data series
    data_range = Reference(ws, min_col=2, min_row=data_start_row, 
                          max_col=len(chart_data), max_row=data_start_row + len(list(chart_data.values())[0]))
    categories = Reference(ws, min_col=1, min_row=data_start_row + 1, 
                          max_row=data_start_row + len(list(chart_data.values())[0]))
    
    chart.add_data(data_range, titles_from_data=True)
    chart.set_categories(categories)
    
    # Position chart
    ws.add_chart(chart, f'H{chart_start_row}')

def prepare_chart_data(acf_pacf_data):
    """Prepare data structure for comparative charting."""
    chart_data = {'Time Scale': []}
    
    # Collect all unique lag types
    all_lags = set()
    for data in acf_pacf_data.values():
        all_lags.update(data.keys())
    
    # Sort lags logically
    sorted_lags = sorted([lag for lag in all_lags if isinstance(lag, str) and ('ACF_' in lag or 'PACF_' in lag)])
    
    # Initialize data structure
    for lag in sorted_lags[:6]:  # Limit to first 6 lags for readability
        chart_data[lag] = []
    
    # Populate data
    for time_scale, data in acf_pacf_data.items():
        chart_data['Time Scale'].append(time_scale)
        for lag in sorted_lags[:6]:
            value = data.get(lag, None)
            if isinstance(value, (int, float)):
                chart_data[lag].append(value)
            else:
                chart_data[lag].append(0)  # Use 0 for non-numeric values in charts
    
    return chart_data

def create_interpretation_notes(ws):
    """Add interpretation notes to the dashboard."""
    notes_start_row = 30
    
    ws[f'A{notes_start_row}'] = 'Interpretation Guide'
    ws[f'A{notes_start_row}'].font = Font(size=14, bold=True, color='003366')
    
    notes = [
        "* Strong Autocorrelation (|r| > 0.7): Clear temporal patterns, predictable behavior",
        "* Moderate Autocorrelation (0.3 < |r| < 0.7): Some temporal structure, partial predictability", 
        "* Weak/Random Pattern (|r| < 0.3): Little temporal structure, mostly random variation",
        "* ACF shows total correlation including indirect effects",
        "* PACF shows direct correlation after removing indirect effects",
        "* Compare patterns across time scales to understand system behavior at different levels"
    ]
    
    for i, note in enumerate(notes):
        ws[f'A{notes_start_row + i + 2}'] = note
        ws[f'A{notes_start_row + i + 2}'].font = Font(size=10, color='666666')

def add_no_data_message(ws):
    """Add message when no ACF/PACF data is available."""
    ws['A1'] = 'ACF/PACF Dashboard'
    ws['A1'].font = Font(size=18, bold=True, color='003366')
    
    ws['A3'] = 'No ACF/PACF data available for dashboard creation.'
    ws['A3'].font = Font(size=12, color='CC0000')
    
    ws['A4'] = 'Please ensure ACF/PACF analysis has been run on the data sheets.'
    ws['A4'].font = Font(size=10, color='666666')

def style_dashboard(ws):
    """Apply consistent styling to the dashboard."""
    # Set column widths
    column_widths = {'A': 15, 'B': 12, 'C': 12, 'D': 18, 'E': 18, 'F': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add borders to summary table
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Apply borders to data cells (approximate range)
    for row in range(9, 15):  # Adjust based on actual data
        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.border = thin_border
            if row > 9:  # Data rows
                cell.alignment = Alignment(horizontal='center')

if __name__ == "__main__":
    # Test with latest report
    import glob
    
    # Find the most recent report
    reports = glob.glob("d:/ARDataAnalysis/*Report*.xlsx")
    if reports:
        latest_report = max(reports, key=os.path.getctime)
        print(f"Creating dashboard for: {latest_report}")
        create_dashboard_summary(latest_report)
    else:
        print("No Excel reports found to create dashboard for.")
