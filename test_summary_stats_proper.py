#!/usr/bin/env python3
"""
Proper test of Summary Statistics sheet creation with correct initialization.
"""

import sys
import traceback
from report_generator.sheet_creators import SheetCreator
from report_generator.formatter import ExcelFormatter
from db_utils import get_database
from openpyxl import Workbook

def test_summary_stats_creation():
    """Test creating the Summary Statistics sheet with proper initialization."""
    print("Testing Summary Statistics sheet creation with proper setup...")
    
    try:
        # Get database connection
        print("Connecting to database...")
        db = get_database()
        
        # Create formatter
        print("Creating formatter...")
        formatter = ExcelFormatter()
        
        # Create a test workbook
        workbook = Workbook()
        
        # Create sheet creator instance with proper arguments
        print("Creating SheetCreator instance...")
        sheet_creator = SheetCreator(db, formatter)
        
        print("Attempting to create Summary Statistics sheet...")
        
        # Try to create the Summary Statistics sheet
        sheet_creator.create_summary_statistics_sheet(workbook)
        
        print("✅ Summary Statistics sheet created successfully!")
        
        # Check if the sheet was actually created
        if "Summary Statistics" in workbook.sheetnames:
            print("✅ Summary Statistics sheet found in workbook")
            ws = workbook["Summary Statistics"]
            print(f"   Sheet has {ws.max_row} rows and {ws.max_column} columns")
            
            # Try to read some sample data
            if ws.max_row > 1:
                print("   Sample data:")
                for row_num in range(1, min(6, ws.max_row + 1)):  # First 5 rows
                    row_data = []
                    for col_num in range(1, ws.max_column + 1):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    print(f"     Row {row_num}: {row_data}")
        else:
            print("❌ Summary Statistics sheet not found in workbook")
            print(f"   Available sheets: {workbook.sheetnames}")
        
    except Exception as e:
        print(f"❌ Error creating Summary Statistics sheet:")
        print(f"   Error: {e}")
        print(f"   Error type: {type(e).__name__}")
        print("\nFull traceback:")
        traceback.print_exc()
        
        # Try to identify the specific issue
        error_str = str(e)
        if "non_collection_days" in error_str:
            print("\n🔍 Likely issue: Problem with non_collection_days filtering")
        elif "_load_config" in error_str:
            print("\n🔍 Likely issue: Problem with config loading")
        elif "date_str" in error_str:
            print("\n🔍 Likely issue: Problem with date string conversion")
        elif "KeyError" in str(type(e)):
            print(f"\n🔍 Likely issue: Missing key in data structure: {e}")
        elif "AttributeError" in str(type(e)):
            print(f"\n🔍 Likely issue: Missing attribute or method: {e}")
        else:
            print(f"\n🔍 Issue details: {e}")

if __name__ == "__main__":
    test_summary_stats_creation()
