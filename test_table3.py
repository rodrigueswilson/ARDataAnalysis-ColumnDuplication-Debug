#!/usr/bin/env python3
"""
Test script to isolate and fix Table 3 creation in the Data Cleaning sheet
"""

import os
import sys
import pymongo
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

# Add the project root to the path
sys.path.insert(0, os.path.abspath('.'))

# Import necessary modules
from report_generator.sheet_creators import SheetCreator
from report_generator.formatters import ExcelFormatter

class Table3Tester:
    """Test class to isolate and fix Table 3 creation"""
    
    def __init__(self):
        """Initialize the tester"""
        self.db = pymongo.MongoClient()['ARDataAnalysis']
        self.formatter = ExcelFormatter()
        
    def create_test_workbook(self):
        """Create a test workbook with just Table 3"""
        print("[TEST] Creating test workbook for Table 3...")
        
        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data Cleaning"
        
        # Get the totals data (same as in the actual sheet creator)
        totals = self._calculate_totals()
        
        # Create Table 3 directly
        self._create_table3(ws, totals)
        
        # Save the workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Table3_Test_{timestamp}.xlsx"
        wb.save(filename)
        print(f"[TEST] Test workbook saved as {filename}")
        
    def _calculate_totals(self):
        """Calculate the totals for the intersection analysis"""
        print("[TEST] Calculating totals for intersection analysis...")
        
        # Get the counts from MongoDB (same as in the actual sheet creator)
        collection = self.db['media_records']
        
        # Calculate the counts for each category
        school_outliers = collection.count_documents({
            'is_collection_day': True,
            'outlier_status': True,
            'school_year': {'$ne': 'N/A'}
        })
        
        non_school_normal = collection.count_documents({
            'is_collection_day': False,
            'outlier_status': False,
            'school_year': {'$ne': 'N/A'}
        })
        
        non_school_outliers = collection.count_documents({
            'is_collection_day': False,
            'outlier_status': True,
            'school_year': {'$ne': 'N/A'}
        })
        
        school_normal = collection.count_documents({
            'is_collection_day': True,
            'outlier_status': False,
            'school_year': {'$ne': 'N/A'}
        })
        
        total_files = school_outliers + non_school_normal + non_school_outliers + school_normal
        
        # Calculate retention percentage
        retention_pct = (school_normal / total_files * 100) if total_files > 0 else 0
        
        # Return the totals
        return {
            'school_outliers': school_outliers,
            'non_school_normal': non_school_normal,
            'non_school_outliers': non_school_outliers,
            'school_normal': school_normal,
            'total_files': total_files,
            'retention_pct': retention_pct
        }
    
    def _create_table3(self, ws, totals):
        """Create Table 3: Filter Impact Summary"""
        print("[TEST] Creating Table 3: Filter Impact Summary...")
        
        # Start at row 1 for this test
        table3_start_row = 1
        
        # Table 3 header
        print(f"[TEST] Creating Table 3 header at row {table3_start_row}")
        ws.cell(row=table3_start_row, column=1, value="Table 3: Filter Impact Summary")
        ws.cell(row=table3_start_row, column=1).font = Font(bold=True)
        
        # Headers for summary table
        summary_headers = ['Filter Category', 'Number of Files', '% of All Files']
        summary_header_row = table3_start_row + 2
        print(f"[TEST] Creating Table 3 headers at row {summary_header_row}")
        
        for col, header in enumerate(summary_headers, 1):
            ws.cell(row=summary_header_row, column=col, value=header)
            ws.cell(row=summary_header_row, column=col).font = Font(bold=True)
        
        # Calculate summary data
        total_files = totals['total_files']
        summary_data = [
            ('School Outliers (unusual files from school days)', totals['school_outliers'], 
             (totals['school_outliers'] / total_files * 100) if total_files > 0 else 0),
            ('Non-School Normal (regular files from non-school days)', totals['non_school_normal'], 
             (totals['non_school_normal'] / total_files * 100) if total_files > 0 else 0),
            ('School Normal (final research dataset)', totals['school_normal'], 
             (totals['school_normal'] / total_files * 100) if total_files > 0 else 0),
            ('Non-School Outliers (unusual files from non-school days)', totals['non_school_outliers'], 
             (totals['non_school_outliers'] / total_files * 100) if total_files > 0 else 0)
        ]
        
        # Fill summary data
        summary_row = summary_header_row + 1
        print(f"[TEST] Filling Table 3 data starting at row {summary_row}")
        
        for criterion, count, percentage in summary_data:
            ws.cell(row=summary_row, column=1, value=criterion)
            ws.cell(row=summary_row, column=2, value=count)
            
            # Format percentage
            pct_cell = ws.cell(row=summary_row, column=3, value=percentage / 100)
            pct_cell.number_format = '0.0%'
            
            print(f"[TEST] Added Table 3 row: {criterion} | {count} | {percentage:.1f}%")
            summary_row += 1
        
        # Add total row
        print(f"[TEST] Adding Table 3 total row at row {summary_row}")
        ws.cell(row=summary_row, column=1, value='TOTAL')
        ws.cell(row=summary_row, column=2, value=totals['total_files'])
        
        # Format total percentage
        total_pct_cell = ws.cell(row=summary_row, column=3, value=1.0)  # 100%
        total_pct_cell.number_format = '0.0%'
        
        # Apply bold formatting to total row
        for col in range(1, 4):
            cell = ws.cell(row=summary_row, column=col)
            cell.font = Font(bold=True)
        
        print(f"[TEST] Table 3 total row added: TOTAL | {totals['total_files']} | 100.0%")
        
        print("[TEST] Table 3 creation complete")

if __name__ == "__main__":
    tester = Table3Tester()
    tester.create_test_workbook()
