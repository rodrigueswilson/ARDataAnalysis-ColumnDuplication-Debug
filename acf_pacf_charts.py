#!/usr/bin/env python3
"""
ACF/PACF Chart Visualization Utilities

This module provides functions to create professional ACF/PACF charts in Excel
using openpyxl, enhancing the interpretability of time series analysis results.
"""

import math
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.series import Series, SeriesLabel
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.styles import Font, PatternFill, Alignment
import numpy as np
import math


def add_acf_pacf_chart(worksheet, data_start_row, data_end_row, sheet_type="daily"):
    """
    Add an ACF/PACF line chart to an Excel worksheet.
    
    Args:
        worksheet: openpyxl worksheet object
        data_start_row: First row of data (after headers)
        data_end_row: Last row of data
        sheet_type: Type of sheet (daily, weekly, monthly, etc.)
    
    Returns:
        Chart object that was added to the worksheet
    """
    
    # Find ACF and PACF columns - headers are in row 3 based on diagnostic analysis
    headers = [cell.value for cell in worksheet[3]]  # Headers are in row 3
    
    # Filter for ACF/PACF columns (exclude significance columns for charting)
    # FIXED: Ensure ACF detection excludes PACF columns by checking PACF is NOT present
    acf_cols = [(i+1, h) for i, h in enumerate(headers) if h and '_ACF_Lag_' in str(h) and '_PACF_Lag_' not in str(h) and '_Significant' not in str(h)]
    pacf_cols = [(i+1, h) for i, h in enumerate(headers) if h and '_PACF_Lag_' in str(h) and '_Significant' not in str(h)]
    
    print(f"[CHART] Found {len(acf_cols)} ACF columns and {len(pacf_cols)} PACF columns in {worksheet.title}")
    if acf_cols:
        print(f"[CHART] ACF columns: {[h for _, h in acf_cols[:3]]}")
    if pacf_cols:
        print(f"[CHART] PACF columns: {[h for _, h in pacf_cols[:3]]}")
    
    if not acf_cols and not pacf_cols:
        print(f"[WARNING] No ACF/PACF columns found in {worksheet.title}")
        return None
    
    # Dynamically find the row containing actual ACF/PACF numeric values
    # ACF/PACF values are global statistics and only appear in one specific row
    sample_row = None
    
    # Search for the row with numeric ACF/PACF values
    test_col = acf_cols[0][0] if acf_cols else pacf_cols[0][0]
    for row_num in range(4, min(15, worksheet.max_row + 1)):  # Check rows 4-14
        test_value = worksheet.cell(row=row_num, column=test_col).value
        if isinstance(test_value, (int, float)):
            sample_row = row_num
            break
    
    if sample_row is None:
        print(f"[WARNING] No numeric ACF/PACF values found in {worksheet.title}")
        return None
    
    print(f"[CHART] Found ACF/PACF data in row {sample_row}")
    
    # Extract ACF/PACF values and lag numbers
    acf_data = []
    pacf_data = []
    lag_numbers = []
    
    # Process ACF columns
    for col_idx, header in acf_cols:
        lag_part = str(header).split('_Lag_')[1]
        if lag_part.isdigit():
            lag_num = int(lag_part)
            acf_value = worksheet.cell(row=sample_row, column=col_idx).value
            print(f"[CHART] ACF Lag {lag_num}: {acf_value} (type: {type(acf_value).__name__})")
            if isinstance(acf_value, (int, float)):
                lag_numbers.append(lag_num)
                acf_data.append((lag_num, acf_value))
    
    # Process PACF columns
    for col_idx, header in pacf_cols:
        lag_part = str(header).split('_Lag_')[1]
        if lag_part.isdigit():
            lag_num = int(lag_part)
            pacf_value = worksheet.cell(row=sample_row, column=col_idx).value
            print(f"[CHART] PACF Lag {lag_num}: {pacf_value} (type: {type(pacf_value).__name__})")
            if isinstance(pacf_value, (int, float)):
                if lag_num not in lag_numbers:
                    lag_numbers.append(lag_num)
                pacf_data.append((lag_num, pacf_value))
    
    if not acf_data and not pacf_data:
        print(f"[WARNING] No valid ACF/PACF values found in {worksheet.title}")
        return None
    
    print(f"[CHART] Found {len(acf_data)} ACF values and {len(pacf_data)} PACF values")
    
    # Create line chart
    chart = LineChart()
    chart.title = f"ACF/PACF Analysis - {sheet_type.title()} Total Files"
    chart.style = 10  # Professional style
    chart.height = 10  # Reasonable height
    chart.width = 15   # Good width for readability
    
    # Configure axes
    chart.y_axis.title = "Correlation Coefficient"
    chart.x_axis.title = "Lag"
    chart.y_axis.scaling.min = -1.0
    chart.y_axis.scaling.max = 1.0
    
    # Sort and organize data by lag
    all_lags = sorted(set(lag_numbers))
    print(f"[CHART] Processing lags: {all_lags}")
    
    # Create organized data structure
    acf_by_lag = {lag: val for lag, val in acf_data}
    pacf_by_lag = {lag: val for lag, val in pacf_data}
    
    # Create temporary data area for chart (below main data)
    chart_data_start_row = data_end_row + 5
    
    # Write headers
    worksheet.cell(row=chart_data_start_row - 1, column=1, value="Lag")
    worksheet.cell(row=chart_data_start_row - 1, column=2, value="ACF")
    worksheet.cell(row=chart_data_start_row - 1, column=3, value="PACF")
    
    # Calculate confidence intervals for statistical significance
    n_observations = max(data_end_row - data_start_row + 1, 10)  # Minimum 10 for safety
    confidence_level = 1.96 / math.sqrt(n_observations)  # 95% confidence interval
    
    # Write headers including confidence intervals
    worksheet.cell(row=chart_data_start_row - 1, column=4, value="Upper_CI")
    worksheet.cell(row=chart_data_start_row - 1, column=5, value="Lower_CI")
    
    # Write organized data with confidence intervals
    for i, lag in enumerate(all_lags):
        row = chart_data_start_row + i
        worksheet.cell(row=row, column=1, value=lag)
        
        # Write ACF value if available
        if lag in acf_by_lag:
            worksheet.cell(row=row, column=2, value=acf_by_lag[lag])
        
        # Write PACF value if available  
        if lag in pacf_by_lag:
            worksheet.cell(row=row, column=3, value=pacf_by_lag[lag])
        
        # Write confidence interval bounds
        worksheet.cell(row=row, column=4, value=confidence_level)   # Upper CI
        worksheet.cell(row=row, column=5, value=-confidence_level)  # Lower CI
    
    # Create chart references
    chart_data_end_row = chart_data_start_row + len(all_lags) - 1
    x_axis_data = Reference(worksheet, min_col=1, min_row=chart_data_start_row, max_row=chart_data_end_row)
    
    chart_series = []
    if acf_data:
        acf_values = Reference(worksheet, min_col=2, min_row=chart_data_start_row - 1, max_row=chart_data_end_row)
        chart_series.append(('ACF', acf_values, "4F81BD"))  # Blue
    
    if pacf_data:
        pacf_values = Reference(worksheet, min_col=3, min_row=chart_data_start_row - 1, max_row=chart_data_end_row)
        chart_series.append(('PACF', pacf_values, "C0504D"))  # Red
    
    # Add confidence interval bands
    upper_ci_values = Reference(worksheet, min_col=4, min_row=chart_data_start_row - 1, max_row=chart_data_end_row)
    lower_ci_values = Reference(worksheet, min_col=5, min_row=chart_data_start_row - 1, max_row=chart_data_end_row)
    chart_series.append(('Upper CI', upper_ci_values, "808080"))  # Gray
    chart_series.append(('Lower CI', lower_ci_values, "808080"))  # Gray
    
    print(f"[CHART] Chart data for {worksheet.title}: {len(chart_series)} series, n~{n_observations}, CI=+/-{confidence_level:.3f}")
    
    if not chart_series:
        print(f"[WARNING] No valid ACF/PACF values found in {worksheet.title}")
        
        # Add user-friendly message for insufficient data
        insufficient_data_row = data_end_row + 3
        worksheet.cell(row=insufficient_data_row, column=1, value="[CHART] ACF/PACF Chart Information")
        worksheet.cell(row=insufficient_data_row, column=1).font = openpyxl.styles.Font(bold=True, size=12)
        
        if "Period" in worksheet.title:
            message = "ACF/PACF chart omitted due to insufficient data (only 6 periods)"
        else:
            message = "ACF/PACF chart omitted due to insufficient statistical data"
            
        worksheet.cell(row=insufficient_data_row + 1, column=1, value=message)
        worksheet.cell(row=insufficient_data_row + 1, column=1).font = openpyxl.styles.Font(italic=True)
        
        print(f"[OK] Added explanatory message to {worksheet.title}")
        return None

    # Add data series to chart with enhanced styling
    for i, (series_name, series_values, color) in enumerate(chart_series):
        chart.add_data(series_values, titles_from_data=False)
        if len(chart.series) > 0:
            current_series = chart.series[-1]
            # Use SeriesLabel for proper OpenPyXL compatibility
            from openpyxl.chart.series import SeriesLabel
            current_series.title = SeriesLabel(v=series_name)
            
            # Apply enhanced styling to the series
            if hasattr(current_series, 'graphicalProperties'):
                current_series.graphicalProperties.line.solidFill = color
                
                # Different styling for ACF/PACF vs Confidence Intervals
                if 'CI' in series_name:
                    # Confidence intervals: thinner, dashed gray lines
                    current_series.graphicalProperties.line.width = 15000
                    current_series.graphicalProperties.line.dashStyle = "dash"
                else:
                    # ACF/PACF: thicker, solid lines
                    current_series.graphicalProperties.line.width = 30000
                    current_series.smooth = True  # Smooth lines for ACF/PACF only
    
    # Set categories (x-axis labels) - use row numbers as lag indicators
    chart.set_categories(x_axis_data)
    
    # Enhanced chart styling (ChatGPT recommendations)
    chart.legend.position = 'b'  # Bottom legend positioning
    # Note: Gridlines require ChartLines objects in openpyxl, simplified for compatibility
    
    # Position chart to the right of the data
    chart_position = f"F{data_start_row}"
    worksheet.add_chart(chart, chart_position)
    
    print(f"[SUCCESS] Added ACF/PACF chart to {worksheet.title} at {chart_position}")
    return chart


def add_arima_forecast_chart(worksheet, data_start_row, data_end_row, sheet_type="daily"):
    """
    Add an ARIMA forecast chart to an Excel worksheet if forecast columns exist.
    
    Args:
        worksheet: openpyxl worksheet object
        data_start_row: First row of data (after headers)
        data_end_row: Last row of data
        sheet_type: Type of sheet (daily, weekly, monthly, etc.)
    
    Returns:
        Chart object that was added to the worksheet, or None if no forecast data
    """
    
    # Find Total_Files and forecast columns
    headers = [cell.value for cell in worksheet[1]]
    total_files_col = None
    forecast_col = None
    
    for i, header in enumerate(headers):
        if str(header) == 'Total_Files':
            total_files_col = i + 1
        elif str(header) == 'Total_Files_Forecast':
            forecast_col = i + 1
    
    if not total_files_col or not forecast_col:
        print(f"No ARIMA forecast data found in {worksheet.title}")
        return None
    
    # Check if forecast column contains numeric data (not error messages)
    sample_forecast = worksheet.cell(row=data_start_row, column=forecast_col).value
    if not isinstance(sample_forecast, (int, float)):
        print(f"ARIMA forecast contains non-numeric data in {worksheet.title}: {sample_forecast}")
        return None
    
    # Create line chart for ARIMA forecast
    chart = LineChart()
    chart.title = f"ARIMA Forecast vs. Actual - {sheet_type.title()} Total Files"
    chart.style = 10
    chart.height = 10
    chart.width = 15
    
    # Configure axes
    chart.y_axis.title = "File Count"
    chart.x_axis.title = "Time Period"
    
    # Add historical data series
    historical_data = Reference(worksheet, min_col=total_files_col, min_row=1, max_row=data_end_row)
    chart.add_data(historical_data, titles_from_data=True)
    
    # Add forecast data series
    forecast_data = Reference(worksheet, min_col=forecast_col, min_row=1, max_row=data_end_row)
    chart.add_data(forecast_data, titles_from_data=True)
    
    # Position chart to the right of ACF/PACF chart
    chart_position = f"Q{data_end_row + 5}"
    worksheet.add_chart(chart, chart_position)
    
    print(f"Added ARIMA forecast chart to {worksheet.title} at {chart_position}")
    return chart


def add_chart_summary_info(worksheet, data_end_row, sheet_type, total_lags, computed_lags):
    """
    Add a compact summary panel below the chart data.
    
    Args:
        worksheet: openpyxl worksheet object
        data_end_row: Last row of main data
        sheet_type: Type of sheet (daily, weekly, etc.)
        total_lags: Total number of requested lags
        computed_lags: Number of successfully computed lags
    """
    
    summary_start_row = data_end_row + 8  # Below chart data
    
    # Add summary headers with formatting
    summary_data = [
        ("[CHART] ACF/PACF Analysis Summary", ""),
        ("Lag Range Used:", f"{computed_lags} of {total_lags} requested"),
        ("Computed on Column:", "Total_Files"),
        ("Analysis Type:", f"{sheet_type.title()} Time Series"),
        ("Status:", "[OK] Complete" if computed_lags > 0 else "[EMOJI] Limited Data")
    ]
    
    for i, (label, value) in enumerate(summary_data):
        worksheet.cell(row=summary_start_row + i, column=1, value=label)
        worksheet.cell(row=summary_start_row + i, column=2, value=value)
        
        # Bold the first row (title)
        if i == 0:
            worksheet.cell(row=summary_start_row + i, column=1).font = openpyxl.styles.Font(bold=True)
    
    print(f"[SUCCESS] Added summary panel to {worksheet.title}")


def create_acf_pacf_dashboard_sheet(workbook, acf_pacf_sheets, target_position=None):
    """
    Create a comprehensive dashboard sheet with ACF/PACF analysis summary,
    cross-sheet navigation, statistical interpretation, and mini-charts.
    
    Args:
        workbook: openpyxl workbook object
        acf_pacf_sheets: List of sheet names containing ACF/PACF data
        target_position: Optional position index for sheet placement
    
    Returns:
        Dashboard worksheet object
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Create new sheet with proper positioning
    if target_position is not None:
        dashboard = workbook.create_sheet(title="ACF_PACF_Dashboard")
        print(f"[FIX] Created ACF_PACF_Dashboard at position {target_position + 1}")
    else:
        dashboard = workbook.create_sheet(title="ACF_PACF_Dashboard")
        print(f"[CREATE] Created ACF_PACF_Dashboard at default position")
    
    # Define styles
    title_font = Font(size=18, bold=True, color="2F4F4F")
    header_font = Font(size=12, bold=True, color="4682B4")
    subheader_font = Font(size=10, bold=True)
    normal_font = Font(size=10)
    italic_font = Font(size=10, italic=True, color="696969")
    
    header_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    summary_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # === HEADER SECTION ===
    dashboard.cell(row=1, column=1, value="ACF/PACF Analysis Dashboard")
    dashboard.cell(row=1, column=1).font = title_font
    dashboard.merge_cells('A1:H1')
    
    dashboard.cell(row=2, column=1, value="Comprehensive Time Series Autocorrelation Analysis")
    dashboard.cell(row=2, column=1).font = italic_font
    dashboard.merge_cells('A2:H2')
    
    # === SUMMARY STATISTICS SECTION ===
    current_row = 4
    dashboard.cell(row=current_row, column=1, value="Analysis Summary")
    dashboard.cell(row=current_row, column=1).font = header_font
    dashboard.cell(row=current_row, column=1).fill = header_fill
    dashboard.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    # Summary table headers
    headers = ["Time Scale", "Data Points", "ACF Lags", "PACF Lags", "Max |ACF|", "Max |PACF|", "Confidence Level", "Sheet Link"]
    for col, header in enumerate(headers, 1):
        cell = dashboard.cell(row=current_row, column=col, value=header)
        cell.font = subheader_font
        cell.fill = summary_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Analyze each ACF/PACF sheet and populate summary
    for sheet_name in acf_pacf_sheets:
        if sheet_name in workbook.sheetnames:
            source_sheet = workbook[sheet_name]
            
            # Extract summary statistics from the sheet
            try:
                # Find headers in row 3
                headers_row = [cell.value for cell in source_sheet[3]]
                acf_cols = [h for h in headers_row if h and '_ACF_Lag_' in str(h) and '_Significant' not in str(h)]
                pacf_cols = [h for h in headers_row if h and '_PACF_Lag_' in str(h) and '_Significant' not in str(h)]
                
                # Count data rows (excluding headers)
                data_rows = 0
                for row in range(4, source_sheet.max_row + 1):
                    if source_sheet.cell(row=row, column=1).value:
                        data_rows += 1
                
                # Find maximum absolute ACF/PACF values
                max_acf = 0
                max_pacf = 0
                
                # Dynamic detection of data row (same logic as chart creation)
                sample_row = None
                if acf_cols:
                    test_col = None
                    for col_idx, header in enumerate(headers_row, 1):
                        if header == acf_cols[0]:
                            test_col = col_idx
                            break
                    
                    if test_col:
                        for row_num in range(4, min(15, source_sheet.max_row + 1)):
                            test_value = source_sheet.cell(row=row_num, column=test_col).value
                            if isinstance(test_value, (int, float)):
                                sample_row = row_num
                                break
                
                if sample_row:
                    # Extract ACF/PACF values from the data row
                    for col_idx, header in enumerate(headers_row, 1):
                        if header in acf_cols:
                            value = source_sheet.cell(row=sample_row, column=col_idx).value
                            if isinstance(value, (int, float)):
                                max_acf = max(max_acf, abs(value))
                        elif header in pacf_cols:
                            value = source_sheet.cell(row=sample_row, column=col_idx).value
                            if isinstance(value, (int, float)):
                                max_pacf = max(max_pacf, abs(value))
                
                # Calculate confidence level (95% CI)
                confidence_level = f"±{1.96/math.sqrt(max(data_rows, 1)):.3f}" if data_rows > 0 else "N/A"
                
                # Time scale extraction
                time_scale = sheet_name.replace(" (ACF_PACF)", "").replace(" Counts", "")
                
                # Populate summary row with proper data types
                row_data = [
                    time_scale,  # Text
                    data_rows,   # Number
                    len(acf_cols),   # Number
                    len(pacf_cols),  # Number
                    max_acf if max_acf > 0 else 0.0,  # Number
                    max_pacf if max_pacf > 0 else 0.0,  # Number
                    confidence_level,  # Text (contains ± symbol)
                    f"→ {sheet_name}"  # Text
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = dashboard.cell(row=current_row, column=col, value=value)
                    cell.font = normal_font
                    cell.border = thin_border
                    if col == 8:  # Sheet link column
                        cell.font = Font(size=10, color="0000FF", underline="single")
                    elif col in [5, 6]:  # Max ACF/PACF columns
                        cell.alignment = Alignment(horizontal='right')
                    else:
                        cell.alignment = Alignment(horizontal='center')
                
                current_row += 1
                
            except Exception as e:
                print(f"Warning: Could not analyze {sheet_name}: {e}")
                # Add error row
                error_data = [sheet_name.replace(" (ACF_PACF)", ""), "Error", "N/A", "N/A", "N/A", "N/A", "N/A", f"→ {sheet_name}"]
                for col, value in enumerate(error_data, 1):
                    cell = dashboard.cell(row=current_row, column=col, value=value)
                    cell.font = Font(size=10, color="FF0000")
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center')
                current_row += 1
    
    # === INTERPRETATION GUIDE SECTION ===
    current_row += 2
    dashboard.cell(row=current_row, column=1, value="Statistical Interpretation Guide")
    dashboard.cell(row=current_row, column=1).font = header_font
    dashboard.cell(row=current_row, column=1).fill = header_fill
    dashboard.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    interpretation_text = [
        "• ACF (Autocorrelation Function): Measures linear correlation between observations at different time lags",
        "• PACF (Partial Autocorrelation Function): Measures correlation after removing effects of intermediate lags",
        "• Confidence Intervals: Values outside ±CI bands indicate statistically significant autocorrelation",
        "• Higher |ACF|/|PACF| values suggest stronger temporal patterns in the data",
        "• Daily/Weekly scales show short-term patterns; Monthly/Period scales show long-term trends",
        "• Use ACF/PACF patterns to identify appropriate ARIMA model parameters (p, d, q)"
    ]
    
    for text in interpretation_text:
        dashboard.cell(row=current_row, column=1, value=text)
        dashboard.cell(row=current_row, column=1).font = normal_font
        dashboard.merge_cells(f'A{current_row}:H{current_row}')
        current_row += 1
    
    # === MINI-CHARTS SECTION ===
    current_row += 2
    dashboard.cell(row=current_row, column=1, value="Visual Comparison Charts")
    dashboard.cell(row=current_row, column=1).font = header_font
    dashboard.cell(row=current_row, column=1).fill = header_fill
    dashboard.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    # Create mini-charts for each ACF/PACF sheet
    chart_col = 1
    charts_per_row = 2
    chart_row_start = current_row
    
    for i, sheet_name in enumerate(acf_pacf_sheets):
        if sheet_name in workbook.sheetnames:
            try:
                source_sheet = workbook[sheet_name]
                
                # Create a small line chart showing ACF/PACF patterns
                mini_chart = LineChart()
                mini_chart.title = sheet_name.replace(" (ACF_PACF)", "")
                mini_chart.style = 2
                mini_chart.width = 7  # Smaller width
                mini_chart.height = 5  # Smaller height
                
                # Find ACF/PACF data range in source sheet
                headers_row = [cell.value for cell in source_sheet[3]]
                acf_cols = [h for h in headers_row if h and '_ACF_Lag_' in str(h) and '_Significant' not in str(h)]
                
                if acf_cols:
                    # Find data row
                    data_row = None
                    for row_num in range(4, min(15, source_sheet.max_row + 1)):
                        test_col = None
                        for col_idx, header in enumerate(headers_row, 1):
                            if header == acf_cols[0]:
                                test_col = col_idx
                                break
                        if test_col:
                            test_value = source_sheet.cell(row=row_num, column=test_col).value
                            if isinstance(test_value, (int, float)):
                                data_row = row_num
                                break
                    
                    if data_row:
                        # Add ACF data series
                        acf_start_col = None
                        acf_end_col = None
                        for col_idx, header in enumerate(headers_row, 1):
                            if header in acf_cols:
                                if acf_start_col is None:
                                    acf_start_col = col_idx
                                acf_end_col = col_idx
                        
                        if acf_start_col and acf_end_col:
                            # Create data reference for ACF values
                            acf_data = Reference(source_sheet, 
                                               min_col=acf_start_col, max_col=acf_end_col,
                                               min_row=data_row, max_row=data_row)
                            mini_chart.add_data(acf_data, titles_from_data=False)
                            
                            # Position chart
                            chart_position = f"{get_column_letter(chart_col * 4 - 3)}{chart_row_start + 1}"
                            dashboard.add_chart(mini_chart, chart_position)
                            
                            print(f"[CHART] Added mini-chart for {sheet_name} at {chart_position}")
                
                # Update positioning for next chart
                chart_col += 1
                if chart_col > charts_per_row:
                    chart_col = 1
                    chart_row_start += 8  # Space for chart height
                    
            except Exception as e:
                print(f"Warning: Could not create mini-chart for {sheet_name}: {e}")
    
    # Update current_row to account for charts
    current_row = chart_row_start + 8
    
    # === NAVIGATION SECTION ===
    current_row += 2
    dashboard.cell(row=current_row, column=1, value="Quick Navigation")
    dashboard.cell(row=current_row, column=1).font = header_font
    dashboard.cell(row=current_row, column=1).fill = header_fill
    dashboard.merge_cells(f'A{current_row}:H{current_row}')
    current_row += 1
    
    # Create navigation links
    nav_col = 1
    for sheet_name in acf_pacf_sheets:
        if sheet_name in workbook.sheetnames:
            dashboard.cell(row=current_row, column=nav_col, value=f"→ {sheet_name}")
            dashboard.cell(row=current_row, column=nav_col).font = Font(size=10, color="0000FF", underline="single")
            nav_col += 1
            if nav_col > 4:  # Start new row after 4 links
                nav_col = 1
                current_row += 1
    
    # Auto-adjust column widths
    for col in range(1, 9):
        column_letter = get_column_letter(col)
        dashboard.column_dimensions[column_letter].width = 15
    
    # Special width adjustments
    dashboard.column_dimensions['A'].width = 18  # Time Scale
    dashboard.column_dimensions['H'].width = 20  # Sheet Link
    
    print(f"[OK] Created enhanced ACF/PACF Dashboard with summary statistics and navigation")
    return dashboard


def enhance_acf_pacf_visualization(workbook):
    """
    Enhance ACF/PACF sheets with professional charts and create dashboard.
    
    Args:
        workbook: openpyxl workbook object
    
    Returns:
        List of enhanced sheet names
    """
    
    print("enhance_acf_pacf_visualization called")
    print(f"Workbook type: {type(workbook)}")
    
    enhanced_sheets = []
    acf_pacf_sheet_names = []
    
    print(f"Scanning {len(workbook.sheetnames)} sheets for ACF/PACF sheets")
    
    # Prevent dashboard duplication - remove existing dashboard if it exists
    if "ACF_PACF_Dashboard" in workbook.sheetnames:
        workbook.remove(workbook["ACF_PACF_Dashboard"])
        print("[FIX] Removed existing ACF_PACF_Dashboard to prevent duplication")
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Check if this is an ACF/PACF sheet by looking for ACF/PACF columns
        # Fixed pattern: actual sheets use (ACF_PACF) not (ACF/PACF)
        if "(ACF_PACF)" in sheet_name or ("ACF_PACF" in sheet_name and "Dashboard" not in sheet_name):
            print(f"Found ACF/PACF sheet: {sheet_name}")
            acf_pacf_sheet_names.append(sheet_name)
            worksheet = workbook[sheet_name]
            
            # Determine sheet type
            if "Daily" in sheet_name:
                sheet_type = "daily"
            elif "Weekly" in sheet_name:
                sheet_type = "weekly"
            elif "Monthly" in sheet_name:
                sheet_type = "monthly"
            elif "Biweekly" in sheet_name:
                sheet_type = "biweekly"
            elif "Period" in sheet_name:
                sheet_type = "period"
            else:
                sheet_type = "unknown"
            
            # Find data range
            data_start_row = 2  # After headers
            data_end_row = worksheet.max_row
            
            # Count ACF/PACF columns - headers are in row 3 (consistent with chart generation)
            headers = [cell.value for cell in worksheet[3]]  # Headers are in row 3
            acf_cols = [h for h in headers if h and '_ACF_Lag_' in str(h) and '_Significant' not in str(h)]
            pacf_cols = [h for h in headers if h and '_PACF_Lag_' in str(h) and '_Significant' not in str(h)]
            total_lags = len(set([h.split('_Lag_')[1] for h in acf_cols + pacf_cols if '_Lag_' in str(h)]))
            
            print(f"Found {len(acf_cols)} ACF columns and {len(pacf_cols)} PACF columns")
            if acf_cols or pacf_cols:
                print(f"Sample ACF/PACF columns: {(acf_cols + pacf_cols)[:3]}")
            
            # Check for existing charts to prevent duplication
            existing_chart_count = len(worksheet._charts) if hasattr(worksheet, '_charts') else 0
            if existing_chart_count > 0:
                print(f"[SKIP] Charts already exist in {sheet_name} ({existing_chart_count} charts found) - skipping chart creation to prevent duplication")
                enhanced_sheets.append(sheet_name)
                continue
            
            # Add ACF/PACF chart (only if no existing charts)
            chart = add_acf_pacf_chart(worksheet, data_start_row, data_end_row, sheet_type)
            if chart:
                # Add ARIMA forecast chart
                arima_chart = add_arima_forecast_chart(worksheet, data_start_row, data_end_row, sheet_type)
                
                # Add summary info
                computed_lags = len(acf_cols)  # Approximate
                add_chart_summary_info(worksheet, data_end_row, sheet_type, total_lags, computed_lags)
                enhanced_sheets.append(sheet_name)
    
    # Create dashboard sheet with proper positioning
    if acf_pacf_sheet_names:
        # Calculate proper position: after the last ACF/PACF sheet
        last_acf_pacf_position = 0
        for sheet_name in acf_pacf_sheet_names:
            if sheet_name in workbook.sheetnames:
                position = workbook.sheetnames.index(sheet_name)
                last_acf_pacf_position = max(last_acf_pacf_position, position)
        
        target_position = last_acf_pacf_position + 1
        create_acf_pacf_dashboard_sheet(workbook, acf_pacf_sheet_names, target_position)
    
    print(f"[OK] Enhanced {len(enhanced_sheets)} sheets with ACF/PACF visualizations")
    return enhanced_sheets


def add_forecast_summary_info(worksheet, data_end_row, sheet_type, forecast_quality=None, model_order=None):
    """
{{ ... }}
    
    Args:
        worksheet: openpyxl worksheet object
        data_end_row: Last row of main data
        sheet_type: Type of sheet (daily, weekly, etc.)
        forecast_quality: Quality assessment (Good/Warning/Error)
        model_order: ARIMA model order tuple (p,d,q)
    """
    
    # Position summary below chart area
    summary_start_row = data_end_row + 20  # Below chart
    
    # Summary data
    summary_data = [
        ("[TREND] ARIMA Forecast Summary", ""),
        ("Time Scale:", sheet_type.title()),
        ("Forecast Quality:", forecast_quality or "Unknown"),
        ("Model Order:", f"ARIMA{model_order}" if model_order else "Unknown"),
        ("Forecast Horizon:", f"{_get_forecast_horizon(sheet_type)} periods")
    ]
    
    # Add summary to worksheet
    for i, (label, value) in enumerate(summary_data):
        worksheet.cell(row=summary_start_row + i, column=1, value=label)
        worksheet.cell(row=summary_start_row + i, column=2, value=value)
        
        # Bold the first row (title)
        if i == 0:
            worksheet.cell(row=summary_start_row + i, column=1).font = Font(bold=True)
    
    print(f"[OK] Added ARIMA forecast summary to {worksheet.title}")


def _get_forecast_horizon(sheet_type):
    """
    Get the forecast horizon for a given sheet type.
    
    Args:
        sheet_type: Type of sheet (daily, weekly, etc.)
        
    Returns:
        int: Forecast horizon in periods
    """
    horizons = {
        'daily': 14,
        'weekly': 6,
        'biweekly': 4,
        'monthly': 3,
        'period': 2
    }
    return horizons.get(sheet_type, 14)


def enhance_arima_forecast_visualization(workbook):
    """Add ARIMA forecast charts to sheets with forecast data.
    
    This function looks for sheets with Total_Files and Total_Files_Forecast columns,
    and adds a chart comparing the actual vs. forecast values.
    Only adds charts to sheets that match the enabled time scales in the configuration.
    
    Args:
        workbook: The openpyxl workbook to enhance
        
    Returns:
        List of sheet names that were enhanced with ARIMA charts
    """
    print("[INFO] Adding ARIMA forecast charts to sheets with forecast data...")
    
    # Get enabled time scales from configuration
    import json
    try:
        with open('report_config.json', 'r') as f:
            config = json.load(f)
            forecast_options = config.get('forecast_options', {})
            enabled_time_scales = [scale.lower() for scale in forecast_options.get('time_scales', [])]
            print(f"[INFO] Enabled time scales for ARIMA forecasting: {enabled_time_scales}")
    except Exception as e:
        print(f"[WARNING] Could not read forecast configuration: {e}")
        enabled_time_scales = ['daily', 'weekly']  # Default if config can't be read
    
    charts_added = 0
    enhanced_sheets = []
    
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        
        if ws.max_row <= 3:  # Need at least headers plus one data row
            continue
        
        # Check if this sheet corresponds to an enabled time scale
        sheet_time_scale = None
        if 'Daily' in sheet_name:
            sheet_time_scale = 'daily'
        elif 'Weekly' in sheet_name:
            sheet_time_scale = 'weekly'
        elif 'Biweekly' in sheet_name:
            sheet_time_scale = 'biweekly'
        elif 'Monthly' in sheet_name:
            sheet_time_scale = 'monthly'
        elif 'Period' in sheet_name:
            sheet_time_scale = 'period'
        
        # Only add ARIMA charts to sheets configured for ARIMA forecasts
        from chart_config_helper import should_add_chart
        if not should_add_chart(sheet_name, 'arima_forecast'):
            print(f"[ARIMA] Skipping sheet '{sheet_name}' - ARIMA charts only added to ACF_PACF sheets")
            continue
            
        # FIXED: Headers are in row 3, not row 1
        headers = [str(cell.value or '') for cell in ws[3]]
        print(f"[ARIMA] Checking sheet '{sheet_name}' for forecast columns...")
        
        total_files_col, forecast_col, forecast_lower_col, forecast_upper_col = None, None, None, None
        
        # Find required columns with more flexible matching
        for i, header in enumerate(headers, 1):
            if header and 'Total_Files' == header.strip():
                total_files_col = i
                print(f"[ARIMA] Found Total_Files column at position {i}")
            elif header and 'Total_Files_Forecast' in header:
                forecast_col = i
                print(f"[ARIMA] Found Total_Files_Forecast column at position {i}")
            elif header and 'Total_Files_Forecast_Lower' in header:
                forecast_lower_col = i
                print(f"[ARIMA] Found Total_Files_Forecast_Lower column at position {i}")
            elif header and 'Total_Files_Forecast_Upper' in header:
                forecast_upper_col = i
                print(f"[ARIMA] Found Total_Files_Forecast_Upper column at position {i}")
        
        # Proceed only if both main columns are found
        if total_files_col and forecast_col:
            # FIXED: Data starts at row 4 (after headers in row 3)
            data_start_row = 4
            # Check if the forecast data is numeric
            first_forecast = ws.cell(row=data_start_row, column=forecast_col).value
            
            if isinstance(first_forecast, (int, float)):
                print(f"[ARIMA] Adding ARIMA forecast chart to: {sheet_name}")
                
                # Create chart with proven logic
                chart = LineChart()
                chart.title = f"ARIMA Forecast vs. Actual - {sheet_name}"
                chart.style = 12
                chart.height = 15  # Slightly larger for better visibility
                chart.width = 20
                chart.y_axis.title = "Total Files"
                chart.x_axis.title = "Time"
                
                # Define data series for historical and forecast values
                # FIXED: Use correct row numbers for headers and data
                historical_data = Reference(ws, min_col=total_files_col, min_row=3, max_row=ws.max_row)
                forecast_data = Reference(ws, min_col=forecast_col, min_row=3, max_row=ws.max_row)
                
                chart.add_data(historical_data, titles_from_data=True)
                chart.add_data(forecast_data, titles_from_data=True)
                
                # Add confidence intervals if available
                if forecast_lower_col and forecast_upper_col:
                    lower_ci_data = Reference(ws, min_col=forecast_lower_col, min_row=3, max_row=ws.max_row)
                    upper_ci_data = Reference(ws, min_col=forecast_upper_col, min_row=3, max_row=ws.max_row)
                    chart.add_data(lower_ci_data, titles_from_data=True)
                    chart.add_data(upper_ci_data, titles_from_data=True)
                
                # Style the series for clarity
                if len(chart.series) > 0:  # Actual data
                    chart.series[0].graphicalProperties.line.solidFill = "4F81BD"  # Blue
                    chart.series[0].graphicalProperties.line.width = 25000
                if len(chart.series) > 1:  # Forecast data
                    chart.series[1].graphicalProperties.line.solidFill = "C0504D"  # Red
                    chart.series[1].graphicalProperties.line.dashStyle = "dash"
                    chart.series[1].graphicalProperties.line.width = 25000
                
                # Style confidence intervals if present
                if len(chart.series) > 2:  # Lower CI
                    chart.series[2].graphicalProperties.line.solidFill = "A9A9A9"  # Gray
                    chart.series[2].graphicalProperties.line.dashStyle = "dot"
                    chart.series[2].graphicalProperties.line.width = 15000
                if len(chart.series) > 3:  # Upper CI
                    chart.series[3].graphicalProperties.line.solidFill = "A9A9A9"  # Gray
                    chart.series[3].graphicalProperties.line.dashStyle = "dot"
                    chart.series[3].graphicalProperties.line.width = 15000
                
                # Position chart below existing data
                chart_position = f'A{ws.max_row + 5}'
                ws.add_chart(chart, chart_position)
                charts_added += 1
                enhanced_sheets.append(sheet_name)
                
                # Add forecast summary information below the chart
                forecast_quality_col = next((i for i, h in enumerate(headers, 1) if h and 'Forecast_Quality' in h), None)
                forecast_model_col = next((i for i, h in enumerate(headers, 1) if h and 'Forecast_Model' in h), None)
                
                if forecast_quality_col and forecast_model_col:
                    quality = ws.cell(row=data_start_row, column=forecast_quality_col).value
                    model = ws.cell(row=data_start_row, column=forecast_model_col).value
                    if model and 'ARIMA' in model:
                        # Extract model order (p,d,q) from string like 'ARIMA(2,1,2)'
                        import re
                        model_match = re.search(r'ARIMA\((\d+),(\d+),(\d+)\)', model)
                        if model_match:
                            p, d, q = map(int, model_match.groups())
                            sheet_type = _detect_sheet_type(sheet_name)
                            add_forecast_summary_info(ws, ws.max_row + 20, sheet_type, quality, (p,d,q))
            else:
                print(f"[ARIMA] Skipping {sheet_name}: Forecast data is not numeric ('{first_forecast}')")
    
    if charts_added > 0:
        print(f"[OK] Successfully added {charts_added} ARIMA forecast charts.")
    else:
        print("[WARNING] No ARIMA charts were added. Check if forecast columns contain numeric data.")
    
    return enhanced_sheets


def _detect_sheet_type(sheet_name: str) -> str:
    """
    Detect the sheet type from the sheet name.
    
    Args:
        sheet_name: Name of the Excel sheet
        
    Returns:
        str: Sheet type (daily, weekly, monthly, biweekly, period, unknown)
    """
    lower = sheet_name.lower()
    if "daily" in lower:
        return "daily"
    if "weekly" in lower:
        return "weekly"
    if "monthly" in lower:
        return "monthly"
    if "biweekly" in lower:
        return "biweekly"
    if "period" in lower:
        return "period"
    return "unknown"
