#!/usr/bin/env python3
"""
Debug script to identify why Table 3 (Filter Impact Summary) is not appearing in the report
"""

import openpyxl
import os
import glob
import re

def debug_table3():
    """Debug why Table 3 is missing from the report"""
    
    print("=== Table 3 Debug ===")
    
    try:
        # Find the most recent Excel file
        excel_files = glob.glob("AR_Analysis_Report_*.xlsx")
        if not excel_files:
            print("❌ No Excel files found")
            return
        
        latest_file = max(excel_files, key=os.path.getctime)
        print(f"📊 Examining: {latest_file}")
        
        # Open the workbook
        wb = openpyxl.load_workbook(latest_file)
        
        if "Data Cleaning" not in wb.sheetnames:
            print("❌ No 'Data Cleaning' sheet found")
            return
        
        ws = wb["Data Cleaning"]
        print(f"✅ Found 'Data Cleaning' sheet with {ws.max_row} rows")
        
        # Scan all rows to find all tables and section headers
        print("\n=== All Section Headers ===")
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value:
                # Check if this looks like a section header
                if isinstance(cell_value, str) and (cell_value.startswith("Table") or "Logic" in cell_value or "Notes" in cell_value):
                    print(f"Row {row}: {cell_value}")
                    
                    # Check for formatting to confirm it's a header
                    cell = ws.cell(row=row, column=1)
                    is_bold = cell.font.bold if cell.font else False
                    print(f"   Bold: {is_bold}")
        
        # Check for specific Table 3 text
        print("\n=== Searching for Table 3 ===")
        table3_pattern = re.compile(r"Table\s*3|Filter\s*Impact\s*Summary", re.IGNORECASE)
        for row in range(1, ws.max_row + 1):
            for col in range(1, 5):  # Check first 4 columns
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and isinstance(cell_value, str) and table3_pattern.search(cell_value):
                    print(f"Found potential Table 3 reference at Row {row}, Col {col}: {cell_value}")
        
        # Check if there's a gap or issue after Logic Explanation table
        print("\n=== Checking for Gap After Logic Explanation ===")
        logic_row = None
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Logic Explanation" in str(cell_value):
                logic_row = row
                print(f"Found Logic Explanation at row {row}")
                break
        
        if logic_row:
            # Check the next 20 rows after Logic Explanation
            print(f"\n=== Content After Logic Explanation (rows {logic_row+1} to {logic_row+20}) ===")
            for r in range(logic_row + 1, min(logic_row + 20, ws.max_row + 1)):
                row_content = []
                for col in range(1, 5):
                    cell_value = ws.cell(row=r, column=col).value
                    row_content.append(str(cell_value) if cell_value is not None else "")
                
                if any(row_content):  # Only show non-empty rows
                    print(f"Row {r}: {' | '.join(row_content)}")
        
        # Check for Notes section which should come after Table 3
        print("\n=== Checking for Notes Section ===")
        for row in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "Notes:" in str(cell_value):
                print(f"Found Notes section at row {row}")
                
                # Check content of notes
                for r in range(row + 1, min(row + 10, ws.max_row + 1)):
                    note_content = ws.cell(row=r, column=1).value
                    if note_content:
                        print(f"  Note row {r}: {note_content}")
        
        wb.close()
        
    except Exception as e:
        print(f"❌ Error examining Excel file: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    debug_table3()
