#!/usr/bin/env python3
"""
Quick check of MP3 Duration Analysis sheet content.
"""

import openpyxl

def quick_mp3_check():
    try:
        print("=== QUICK MP3 DURATION ANALYSIS CHECK ===")
        
        # Load the latest report
        wb = openpyxl.load_workbook('AR_Analysis_Report_20250726_173817.xlsx')
        
        if 'MP3 Duration Analysis' in wb.sheetnames:
            print("✅ MP3 Duration Analysis sheet found!")
            
            sheet = wb['MP3 Duration Analysis']
            print(f"Sheet dimensions: {sheet.max_row} rows x {sheet.max_column} columns")
            
            # Check Collection Period table for "Unknown" values
            print("\n=== COLLECTION PERIOD TABLE CHECK ===")
            period_section_found = False
            unknown_count = 0
            valid_period_count = 0
            
            for row in range(1, min(sheet.max_row + 1, 50)):  # Check first 50 rows
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "Collection Period" in str(cell_value):
                    period_section_found = True
                    print(f"Found Collection Period section at row {row}")
                    
                    # Check the next several rows for data
                    for data_row in range(row + 2, min(row + 20, sheet.max_row + 1)):
                        period_val = sheet.cell(row=data_row, column=1).value
                        school_year_val = sheet.cell(row=data_row, column=2).value
                        
                        if period_val and str(period_val) != "Period":  # Skip header
                            if "Unknown" in str(period_val):
                                unknown_count += 1
                            else:
                                valid_period_count += 1
                                print(f"  Row {data_row}: {period_val} | {school_year_val}")
                    break
            
            if period_section_found:
                print(f"Period data: {valid_period_count} valid entries, {unknown_count} unknown entries")
                if unknown_count == 0:
                    print("✅ Collection Period table fixed - no 'Unknown' values!")
                else:
                    print("❌ Collection Period table still has 'Unknown' values")
            
            # Check Monthly table for data
            print("\n=== MONTHLY TABLE CHECK ===")
            monthly_section_found = False
            monthly_data_count = 0
            
            for row in range(1, min(sheet.max_row + 1, 50)):
                cell_value = sheet.cell(row=row, column=1).value
                if cell_value and "Monthly" in str(cell_value):
                    monthly_section_found = True
                    print(f"Found Monthly section at row {row}")
                    
                    # Check for data in the next rows
                    for data_row in range(row + 2, min(row + 15, sheet.max_row + 1)):
                        month_val = sheet.cell(row=data_row, column=1).value
                        files_2122 = sheet.cell(row=data_row, column=2).value
                        files_2223 = sheet.cell(row=data_row, column=4).value
                        
                        if month_val and str(month_val) not in ["Month", ""]:
                            monthly_data_count += 1
                            print(f"  {month_val}: 2021-22={files_2122}, 2022-23={files_2223}")
                    break
            
            if monthly_section_found:
                if monthly_data_count > 0:
                    print(f"✅ Monthly table populated with {monthly_data_count} months of data!")
                else:
                    print("❌ Monthly table still empty")
            
            return True
        else:
            print("❌ MP3 Duration Analysis sheet not found")
            return False
            
    except Exception as e:
        print(f"❌ Error: {e}")
        return False

if __name__ == "__main__":
    quick_mp3_check()
