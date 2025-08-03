#!/usr/bin/env python3
"""
Test script to verify MP3 Duration Analysis sheet fixes.
"""

import sys
import os
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

from report_generator.sheet_creators.specialized import SpecializedSheetCreator
from report_generator.formatters import ExcelFormatter
from db_utils import get_db_connection
import openpyxl

def test_mp3_duration_analysis():
    """Test the MP3 Duration Analysis sheet creation with fixes."""
    
    print("Testing MP3 Duration Analysis sheet fixes...")
    
    # Initialize components
    db = get_db_connection()
    formatter = ExcelFormatter()
    creator = SpecializedSheetCreator(db, formatter)
    
    # Create a test workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    try:
        # Create the MP3 Duration Analysis sheet
        print("[INFO] Creating MP3 Duration Analysis sheet...")
        creator.create_mp3_duration_analysis_sheet(wb)
        
        # Save test file
        test_filename = "test_mp3_duration_fixes.xlsx"
        wb.save(test_filename)
        print(f"[SUCCESS] Test file saved: {test_filename}")
        
        # Verify the sheet content
        if 'MP3 Duration Analysis' in wb.sheetnames:
            ws = wb['MP3 Duration Analysis']
            print(f"[INFO] Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Check for properly formatted totals
            print("\n[INFO] Checking for properly formatted totals...")
            totals_found = False
            for row in range(1, min(21, ws.max_row + 1)):
                for col in range(1, min(9, ws.max_column + 1)):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and 'Total' in str(cell_value):
                        print(f"[FOUND] Totals row at Row {row}: {cell_value}")
                        
                        # Check the formatting of time values in this row
                        for time_col in range(3, 7):  # Columns with time values
                            time_value = ws.cell(row=row, column=time_col).value
                            if time_value and ':' in str(time_value):
                                print(f"  [GOOD] Formatted time in Col {time_col}: {time_value}")
                                totals_found = True
                            elif time_value and isinstance(time_value, (int, float)):
                                print(f"  [BAD] Raw number in Col {time_col}: {time_value}")
            
            if totals_found:
                print("[SUCCESS] Found properly formatted totals!")
            else:
                print("[WARNING] No properly formatted totals found")
                
            # Check for proper school year data
            print("\n[INFO] Checking for proper school year data...")
            for row in range(1, min(21, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and ('2021' in str(cell_value) or '2022' in str(cell_value)):
                    print(f"[GOOD] Found school year data at Row {row}: {cell_value}")
                elif cell_value and 'Unknown' in str(cell_value):
                    print(f"[BAD] Found 'Unknown' at Row {row}: {cell_value}")
        else:
            print("[ERROR] MP3 Duration Analysis sheet not found!")
            
    except Exception as e:
        print(f"[ERROR] Test failed: {e}")
        import traceback
        traceback.print_exc()
    
    finally:
        wb.close()

if __name__ == "__main__":
    test_mp3_duration_analysis()
